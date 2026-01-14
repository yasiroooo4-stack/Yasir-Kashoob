"""
Milk Collection Center ERP - Main Server
نظام ERP لمركز تجميع الحليب - المروج للألبان

REFACTORED: Models imported from models/all_models.py
REFACTORED: Config imported from config.py
REFACTORED: Database imported from database.py
"""
from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
from enum import Enum
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import io
import secrets
import aiosmtplib
import httpx
import asyncio
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Import from refactored modules
from database import db, client
from config import (
    SECRET_KEY, ALGORITHM, ACCESS_TOKEN_EXPIRE_HOURS,
    SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, SMTP_FROM_EMAIL,
    DEFAULT_CENTERS
)

# Import all models from centralized models file
from models.all_models import (
    # Auth models
    UserBase, UserCreate, UserUpdate, PasswordChange, User, UserLogin, Token, PasswordResetToken,
    # Supplier models
    SupplierBase, SupplierCreate, Supplier, SupplierLoginRequest, SupplierModificationRequest,
    SupplierFeedRequestBase, SupplierFeedRequestCreate, SupplierFeedRequest,
    SupplierMessageBase, SupplierMessageCreate, SupplierMessage,
    # Milk models
    QualityTest, MilkReceptionBase, MilkReceptionCreate, MilkReception,
    # Customer models
    CustomerBase, CustomerCreate, Customer,
    # Sale models
    SaleBase, SaleCreate, Sale,
    # Inventory models
    InventoryBase, InventoryUpdate, Inventory,
    # Payment models
    PaymentBase, PaymentCreate, Payment, PaymentApproval,
    # Treasury models
    TreasuryTransaction, TreasuryBalance,
    # Financial models
    AccountType, Account, JournalEntry, JournalEntryLine,
    AccountsPayable, AccountsReceivable, FixedAsset, Budget, BudgetLine, TaxRecord,
    # Employee models
    EmployeeBase, EmployeeCreate, Employee, EmployeeAllowances, EmployeeSalaryStructure,
    SalaryHistoryBase, SalaryHistoryCreate, SalaryHistory,
    # Attendance models
    AttendanceBase, AttendanceCreate, Attendance,
    # Leave models
    LeaveRequestBase, LeaveRequestCreate, LeaveRequest,
    # Excuse models
    ExcuseRequestBase, ExcuseRequestCreate, ExcuseRequest,
    # Expense models
    ExpenseRequestBase, ExpenseRequestCreate, ExpenseRequest,
    # Car contract models
    CarContractBase, CarContractCreate, CarContract,
    # Official letter models
    OfficialLetterBase, OfficialLetterCreate, OfficialLetter,
    # Fingerprint device models
    FingerprintDeviceBase, FingerprintDeviceCreate, FingerprintDevice,
    # Shift models
    ShiftBase, ShiftCreate, Shift, EmployeeShiftBase, EmployeeShiftCreate, EmployeeShift,
    # Overtime models
    OvertimeBase, OvertimeCreate, Overtime,
    # Loan models
    LoanBase, LoanCreate, Loan, LoanPayment,
    # Employee document models
    EmployeeDocumentBase, EmployeeDocumentCreate, EmployeeDocument,
    # Payroll models
    PayrollPeriod, PayrollRecord,
    # Collection center models
    CollectionCenterBase, CollectionCenterCreate, CollectionCenter,
    # Activity log models
    ActivityLog, DeviceSettings,
    # Feed models
    FeedCompanyBase, FeedCompanyCreate, FeedCompany,
    FeedTypeBase, FeedTypeCreate, FeedType,
    FeedPurchaseBase, FeedPurchaseCreate, FeedPurchase,
    # Legal models
    LegalContractBase, LegalContractCreate, LegalContract,
    LegalCaseBase, LegalCaseCreate, LegalCase,
    LegalConsultationBase, LegalConsultationCreate, LegalConsultation,
    LegalDocumentBase, LegalDocumentCreate, LegalDocument,
    # Project models
    ProjectBase, ProjectCreate, Project,
    ProjectTaskBase, ProjectTaskCreate, ProjectTask,
    ProjectTeamMemberBase, ProjectTeamMemberCreate, ProjectTeamMember,
    ProjectMilestoneBase, ProjectMilestoneCreate, ProjectMilestone,
    # Operations models
    DailyOperationBase, DailyOperationCreate, DailyOperation,
    EquipmentBase, EquipmentCreate, Equipment,
    MaintenanceRecordBase, MaintenanceRecordCreate, MaintenanceRecord,
    IncidentReportBase, IncidentReportCreate, IncidentReport,
    VehicleBase, VehicleCreate, Vehicle,
    # Marketing models
    MarketingCampaignBase, MarketingCampaignCreate, MarketingCampaign,
    LeadBase, LeadCreate, Lead,
    SocialMediaPostBase, SocialMediaPostCreate, SocialMediaPost,
    SalesOfferBase, SalesOfferCreate, SalesOffer,
    MarketReturnBase, MarketReturnCreate, MarketReturn,
    MarketSalesSummaryBase, MarketSalesSummaryCreate, MarketSalesSummary,
    # Holiday models
    OfficialHolidayBase, OfficialHolidayCreate, OfficialHoliday,
    EmployeeWeeklyOffBase, EmployeeWeeklyOffCreate,
    PublicHolidayBase, PublicHolidayCreate, PublicHoliday,
    # Analysis models
    AnalysisRequest,
    # Settings models
    UserAppearanceSettings,
    # ZKTeco models
    ZKTecoDeviceBase, ZKTecoDeviceCreate, ZKTecoDevice, ZKTecoSyncSettings,
    # Warning models
    WarningBase, WarningCreate, Warning,
)

# Try to import LLM chat
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    LLM_AVAILABLE = True
except ImportError:
    LLM_AVAILABLE = False
    logging.warning("emergentintegrations not available - AI features disabled")

security = HTTPBearer()

# Create the main app
app = FastAPI(title="Milk Collection Center ERP")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

@app.on_event("startup")
async def startup_event():
    """Initialize default collection centers on startup"""
    try:
        for center_data in DEFAULT_CENTERS:
            existing = await db.collection_centers.find_one({"code": center_data["code"]})
            if not existing:
                center = CollectionCenter(**center_data)
                await db.collection_centers.insert_one(center.model_dump())
                logging.info(f"Created default center: {center_data['name']}")
            else:
                logging.info(f"Center already exists: {center_data['name']}")
        
        # Check and auto-accrue leave balance if end of month
        await check_and_accrue_monthly_leave()
    except Exception as e:
        logging.error(f"Error initializing default centers: {e}")

async def check_and_accrue_monthly_leave():
    """تحقق تلقائي وإضافة رصيد الإجازات في نهاية الشهر"""
    try:
        from calendar import monthrange
        today = datetime.now()
        last_day_of_month = monthrange(today.year, today.month)[1]
        
        # Check if today is the last day of month or first day of new month
        if today.day == last_day_of_month or today.day == 1:
            # Determine which month to accrue
            if today.day == 1:
                # First day of month - accrue for previous month
                if today.month == 1:
                    month = f"{today.year - 1}-12"
                else:
                    month = f"{today.year}-{today.month - 1:02d}"
            else:
                # Last day of month - accrue for current month
                month = today.strftime("%Y-%m")
            
            # Check if already accrued
            existing = await db.leave_balance_logs.find_one({"month": month, "reason": "monthly_accrual"})
            if not existing:
                logging.info(f"Auto-accruing leave balance for month: {month}")
                
                # Get all active employees
                employees = await db.hr_employees.find({"is_active": True}, {"_id": 0}).to_list(500)
                
                from models.all_models import LeaveBalanceLog
                logs = []
                updated_count = 0
                
                for emp in employees:
                    position = emp.get("position", "")
                    rate = get_leave_rate_by_position_helper(position)
                    
                    # Use custom rate if set
                    custom_rate = emp.get("monthly_leave_rate")
                    if custom_rate and custom_rate != 2.6:
                        rate = custom_rate
                    
                    previous_balance = emp.get("leave_balance", 0)
                    new_balance = round(previous_balance + rate, 2)
                    
                    await db.hr_employees.update_one(
                        {"id": emp["id"]},
                        {"$set": {"leave_balance": new_balance, "monthly_leave_rate": rate}}
                    )
                    
                    log = LeaveBalanceLog(
                        employee_id=emp["id"],
                        employee_name=emp["name"],
                        month=month,
                        amount_added=rate,
                        previous_balance=previous_balance,
                        new_balance=new_balance,
                        reason="monthly_accrual"
                    )
                    logs.append(log.model_dump())
                    updated_count += 1
                
                if logs:
                    await db.leave_balance_logs.insert_many(logs)
                
                logging.info(f"Auto-accrued leave balance for {updated_count} employees for month {month}")
            else:
                logging.info(f"Leave balance already accrued for month: {month}")
    except Exception as e:
        logging.error(f"Error in auto leave accrual: {e}")

def get_leave_rate_by_position_helper(position):
    """Helper function to get leave rate by position"""
    if not position:
        return 2.6
    
    position_lower = position.lower() if position else ""
    
    if "مدير عام" in position or "general manager" in position_lower or "director general" in position_lower:
        return 3.5
    if "نائب المدير" in position or "نائب مدير" in position or "deputy" in position_lower:
        return 3.5
    if "مدير الموارد البشرية" in position or "hr manager" in position_lower or "موارد بشرية" in position:
        return 3.0
    if "أمن وسلامة" in position or "أمن" in position or "سلامة" in position or "safety" in position_lower or "security" in position_lower:
        return 3.0
    if "مشرف" in position or "supervisor" in position_lower:
        return 3.0
    
    return 2.6

# ==================== AUTHENTICATION ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(allowed_roles: List[str]):
    async def role_checker(current_user: dict = Depends(get_current_user)):
        if current_user["role"] not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return current_user
    return role_checker

# Activity logging helper
async def log_activity(user_id: str, user_name: str, action: str, entity_type: str = None, 
                       entity_id: str = None, entity_name: str = None, details: str = None,
                       center_id: str = None, center_name: str = None):
    activity = ActivityLog(
        user_id=user_id,
        user_name=user_name,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_name=entity_name,
        details=details,
        center_id=center_id,
        center_name=center_name
    )
    await db.activity_logs.insert_one(activity.model_dump())

# Email sending helper
SMTP_USE_SSL = os.environ.get('SMTP_USE_SSL', 'false').lower() == 'true'

async def send_email(to_email: str, subject: str, html_content: str):
    """Send email using SMTP"""
    try:
        message = MIMEMultipart("alternative")
        message["From"] = SMTP_FROM_EMAIL
        message["To"] = to_email
        message["Subject"] = subject
        
        html_part = MIMEText(html_content, "html", "utf-8")
        message.attach(html_part)
        
        # Use SSL for port 465, TLS for other ports
        if SMTP_USE_SSL or SMTP_PORT == 465:
            await aiosmtplib.send(
                message,
                hostname=SMTP_HOST,
                port=SMTP_PORT,
                username=SMTP_USER,
                password=SMTP_PASSWORD,
                use_tls=True  # SSL/TLS connection
            )
        else:
            await aiosmtplib.send(
                message,
                hostname=SMTP_HOST,
                port=SMTP_PORT,
                username=SMTP_USER,
                password=SMTP_PASSWORD,
                start_tls=True  # STARTTLS
            )
        return True
    except Exception as e:
        logging.error(f"Error sending email: {e}")
        return False

async def send_password_reset_email(email: str, token: str, full_name: str):
    """Send password reset email"""
    reset_link = f"{os.environ.get('FRONTEND_URL', 'https://dairysoft.preview.emergentagent.com')}/reset-password?token={token}"
    
    html_content = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="ar">
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Arial, sans-serif; direction: rtl; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(135deg, #8B4513, #D2691E); color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0; }}
            .content {{ background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px; }}
            .button {{ display: inline-block; background: #8B4513; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; margin: 20px 0; }}
            .footer {{ text-align: center; color: #666; font-size: 12px; margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>المروج للألبان</h1>
            </div>
            <div class="content">
                <h2>مرحباً {full_name}</h2>
                <p>تم طلب استرجاع كلمة المرور لحسابك في نظام المروج للألبان.</p>
                <p>اضغط على الزر أدناه لإعادة تعيين كلمة المرور:</p>
                <p style="text-align: center;">
                    <a href="{reset_link}" class="button">إعادة تعيين كلمة المرور</a>
                </p>
                <p><strong>ملاحظة:</strong> هذا الرابط صالح لمدة ساعة واحدة فقط.</p>
                <p>إذا لم تطلب إعادة تعيين كلمة المرور، يرجى تجاهل هذا البريد.</p>
            </div>
            <div class="footer">
                <p>© 2025 المروج للألبان - جميع الحقوق محفوظة</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return await send_email(email, "إعادة تعيين كلمة المرور - المروج للألبان", html_content)

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"$or": [{"username": user_data.username}, {"email": user_data.email}]})
    if existing:
        raise HTTPException(status_code=400, detail="Username or email already exists")
    
    user = User(**user_data.model_dump(exclude={"password"}))
    user_dict = user.model_dump()
    user_dict["password"] = hash_password(user_data.password)
    
    await db.users.insert_one(user_dict)
    
    token = create_access_token({"sub": user.id, "role": user.role})
    return Token(
        access_token=token,
        token_type="bearer",
        user={"id": user.id, "username": user.username, "email": user.email, "full_name": user.full_name, "role": user.role}
    )

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"username": credentials.username})
    # Check for both password and password_hash fields for compatibility
    password_field = user.get("password_hash") or user.get("password") if user else None
    if not user or not password_field or not verify_password(credentials.password, password_field):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_access_token({"sub": user["id"], "role": user["role"]})
    
    # Log login activity
    await log_activity(
        user_id=user["id"],
        user_name=user["full_name"],
        action="login",
        details="تسجيل دخول للنظام"
    )
    
    return Token(
        access_token=token,
        token_type="bearer",
        user={
            "id": user["id"], 
            "username": user["username"], 
            "email": user["email"], 
            "full_name": user["full_name"], 
            "role": user["role"], 
            "phone": user.get("phone"), 
            "avatar_url": user.get("avatar_url"), 
            "department": user.get("department"),
            "permissions": user.get("permissions", [])
        }
    )

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

@api_router.put("/auth/profile")
async def update_profile(profile_data: UserUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in profile_data.model_dump().items() if v is not None}
    if update_data:
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": update_data}
        )
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password": 0})
    return user

@api_router.put("/auth/password")
async def change_password(password_data: PasswordChange, current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["id"]})
    password_field = user.get("password_hash") or user.get("password") if user else None
    if not password_field or not verify_password(password_data.current_password, password_field):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    new_hash = hash_password(password_data.new_password)
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"password_hash": new_hash, "password": new_hash}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="password_change",
        details="تغيير كلمة المرور"
    )
    
    return {"message": "Password changed successfully"}

# Password Reset Endpoints
@api_router.post("/auth/forgot-password")
async def forgot_password(email: str = Form(...)):
    """Request password reset - sends email with reset link"""
    user = await db.users.find_one({"email": email})
    if not user:
        # Don't reveal if email exists or not for security
        return {"message": "If the email exists, a reset link will be sent"}
    
    # Generate reset token
    token = secrets.token_urlsafe(32)
    expires_at = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
    
    reset_token = PasswordResetToken(
        user_id=user["id"],
        email=email,
        token=token,
        expires_at=expires_at
    )
    
    # Invalidate any existing tokens for this user
    await db.password_reset_tokens.update_many(
        {"user_id": user["id"], "used": False},
        {"$set": {"used": True}}
    )
    
    # Save new token
    await db.password_reset_tokens.insert_one(reset_token.model_dump())
    
    # Send email
    email_sent = await send_password_reset_email(email, token, user["full_name"])
    
    if email_sent:
        await log_activity(
            user_id=user["id"],
            user_name=user["full_name"],
            action="password_reset_request",
            details=f"طلب استرجاع كلمة المرور للبريد: {email}"
        )
    
    return {"message": "If the email exists, a reset link will be sent", "email_sent": email_sent}

@api_router.post("/auth/reset-password")
async def reset_password(token: str = Form(...), new_password: str = Form(...)):
    """Reset password using token from email"""
    reset_token = await db.password_reset_tokens.find_one({
        "token": token,
        "used": False
    })
    
    if not reset_token:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")
    
    # Check if token expired
    expires_at = datetime.fromisoformat(reset_token["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Reset token has expired")
    
    # Update password - update both fields for compatibility
    new_hash = hash_password(new_password)
    await db.users.update_one(
        {"id": reset_token["user_id"]},
        {"$set": {
            "password": new_hash,
            "password_hash": new_hash
        }}
    )
    
    # Mark token as used
    await db.password_reset_tokens.update_one(
        {"token": token},
        {"$set": {"used": True}}
    )
    
    user = await db.users.find_one({"id": reset_token["user_id"]})
    
    await log_activity(
        user_id=reset_token["user_id"],
        user_name=user["full_name"] if user else "Unknown",
        action="password_reset_complete",
        details="تم إعادة تعيين كلمة المرور بنجاح"
    )
    
    return {"message": "Password reset successfully"}

@api_router.get("/auth/verify-reset-token")
async def verify_reset_token(token: str):
    """Verify if reset token is valid"""
    reset_token = await db.password_reset_tokens.find_one({
        "token": token,
        "used": False
    })
    
    if not reset_token:
        return {"valid": False, "message": "Invalid token"}
    
    expires_at = datetime.fromisoformat(reset_token["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        return {"valid": False, "message": "Token expired"}
    
    return {"valid": True, "email": reset_token["email"]}

# ==================== COLLECTION CENTER ROUTES (مراكز التجميع) ====================

@api_router.get("/centers", response_model=List[CollectionCenter])
async def get_centers(current_user: dict = Depends(get_current_user)):
    centers = await db.collection_centers.find({"is_active": True}, {"_id": 0}).to_list(100)
    return centers

@api_router.post("/centers", response_model=CollectionCenter)
async def create_center(center_data: CollectionCenterCreate, current_user: dict = Depends(require_role(["admin"]))):
    center = CollectionCenter(**center_data.model_dump())
    await db.collection_centers.insert_one(center.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_center",
        entity_type="center",
        entity_id=center.id,
        entity_name=center.name,
        details=f"إنشاء مركز تجميع: {center.name}"
    )
    
    return center

@api_router.put("/centers/{center_id}", response_model=CollectionCenter)
async def update_center(center_id: str, center_data: CollectionCenterCreate, current_user: dict = Depends(require_role(["admin"]))):
    result = await db.collection_centers.update_one(
        {"id": center_id},
        {"$set": center_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Center not found")
    center = await db.collection_centers.find_one({"id": center_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_center",
        entity_type="center",
        entity_id=center_id,
        entity_name=center.get("name"),
        details=f"تعديل مركز تجميع: {center.get('name')}"
    )
    
    return center

@api_router.delete("/centers/{center_id}")
async def delete_center(center_id: str, current_user: dict = Depends(require_role(["admin"]))):
    center = await db.collection_centers.find_one({"id": center_id}, {"_id": 0})
    if not center:
        raise HTTPException(status_code=404, detail="Center not found")
    
    result = await db.collection_centers.update_one(
        {"id": center_id},
        {"$set": {"is_active": False}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_center",
        entity_type="center",
        entity_id=center_id,
        entity_name=center.get("name"),
        details=f"حذف مركز تجميع: {center.get('name')}"
    )
    
    return {"message": "Center deleted successfully"}

# ==================== ACTIVITY LOG ROUTES (سجل النشاط) ====================

@api_router.get("/activity-logs")
async def get_activity_logs(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if user_id:
        query["user_id"] = user_id
    if action:
        query["action"] = action
    if start_date:
        query["timestamp"] = {"$gte": start_date}
    if end_date:
        if "timestamp" in query:
            query["timestamp"]["$lte"] = end_date
        else:
            query["timestamp"] = {"$lte": end_date}
    
    logs = await db.activity_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return logs

# ==================== SUPPLIER ROUTES ====================

@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier_data: SupplierCreate, current_user: dict = Depends(get_current_user)):
    supplier = Supplier(**supplier_data.model_dump())
    await db.suppliers.insert_one(supplier.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_supplier",
        entity_type="supplier",
        entity_id=supplier.id,
        entity_name=supplier.name,
        center_id=supplier.center_id,
        center_name=supplier.center_name,
        details=f"إضافة مورد: {supplier.name}"
    )
    
    return supplier

@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers(center_id: Optional[str] = None, search: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"is_active": True}
    if center_id:
        query["center_id"] = center_id
    if search:
        # Search by name, supplier_code, or phone
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"supplier_code": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    suppliers = await db.suppliers.find(query, {"_id": 0}).to_list(5000)
    return suppliers

@api_router.get("/suppliers/{supplier_id}", response_model=Supplier)
async def get_supplier(supplier_id: str, current_user: dict = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return supplier

@api_router.put("/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(supplier_id: str, supplier_data: SupplierCreate, current_user: dict = Depends(get_current_user)):
    result = await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": supplier_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_supplier",
        entity_type="supplier",
        entity_id=supplier_id,
        entity_name=supplier.get("name"),
        details=f"تعديل بيانات مورد: {supplier.get('name')}"
    )
    
    return supplier

@api_router.put("/suppliers/{supplier_id}/transfer-center")
async def transfer_supplier_to_center(
    supplier_id: str,
    new_center: str = Query(..., description="اسم المركز الجديد"),
    current_user: dict = Depends(require_role(["admin"]))
):
    """نقل مورد من مركز إلى مركز آخر - يحتاج موافقة المدير (Yasir Kashoob IT)"""
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    old_center = supplier.get("center_name", "غير محدد")
    
    # Check if current user is the IT manager (Yasir Kashoob)
    is_it_manager = current_user.get("username") == "yasir" or current_user.get("department") == "it"
    
    if is_it_manager:
        # Direct approval for IT manager
        result = await db.suppliers.update_one(
            {"id": supplier_id},
            {"$set": {"center_name": new_center, "address": new_center}}
        )
        
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user["full_name"],
            action="transfer_supplier",
            entity_type="supplier",
            entity_id=supplier_id,
            entity_name=supplier.get("name"),
            details=f"نقل المورد {supplier.get('name')} من {old_center} إلى {new_center}"
        )
        
        return {
            "message": f"تم نقل المورد بنجاح من {old_center} إلى {new_center}",
            "supplier_id": supplier_id,
            "old_center": old_center,
            "new_center": new_center,
            "approved": True
        }
    else:
        # Create approval request for non-IT users
        request = SupplierModificationRequest(
            supplier_id=supplier_id,
            supplier_name=supplier.get("name"),
            supplier_code=supplier.get("supplier_code"),
            request_type="transfer",
            current_data={"center_name": old_center},
            new_data={"center_name": new_center},
            reason=f"نقل من {old_center} إلى {new_center}",
            requested_by=current_user["id"],
            requested_by_name=current_user["full_name"]
        )
        
        await db.supplier_modification_requests.insert_one(request.model_dump())
        
        return {
            "message": "تم إرسال طلب نقل المورد للمدير للموافقة",
            "request_id": request.id,
            "requires_approval": True
        }

# Get all supplier modification requests (for IT manager)
@api_router.get("/admin/supplier-modification-requests")
async def get_supplier_modification_requests(
    status: Optional[str] = None,
    current_user: dict = Depends(require_role(["admin"]))
):
    """جلب جميع طلبات تعديل/نقل الموردين (للمدير)"""
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.supplier_modification_requests.find(query, {"_id": 0}).sort("requested_at", -1).to_list(500)
    return requests

# Approve supplier modification request
@api_router.put("/admin/supplier-modification-requests/{request_id}/approve")
async def approve_supplier_modification(
    request_id: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    """الموافقة على طلب تعديل/نقل المورد"""
    # Only IT manager can approve
    if current_user.get("username") != "yasir" and current_user.get("department") != "it":
        raise HTTPException(status_code=403, detail="فقط مدير تقنية المعلومات يمكنه الموافقة على الطلبات")
    
    mod_request = await db.supplier_modification_requests.find_one({"id": request_id}, {"_id": 0})
    if not mod_request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if mod_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail="تم معالجة هذا الطلب مسبقاً")
    
    # Apply the modification
    supplier_id = mod_request["supplier_id"]
    request_type = mod_request["request_type"]
    new_data = mod_request["new_data"]
    
    if request_type == "transfer":
        await db.suppliers.update_one(
            {"id": supplier_id},
            {"$set": {"center_name": new_data["center_name"], "address": new_data["center_name"]}}
        )
    elif request_type == "update":
        await db.suppliers.update_one(
            {"id": supplier_id},
            {"$set": new_data}
        )
    elif request_type == "delete":
        await db.suppliers.update_one(
            {"id": supplier_id},
            {"$set": {"is_active": False}}
        )
    
    # Update request status
    await db.supplier_modification_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_by_name": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action=f"approve_{request_type}_supplier",
        entity_type="supplier",
        entity_id=supplier_id,
        entity_name=mod_request["supplier_name"],
        details=f"موافقة على طلب {request_type} للمورد {mod_request['supplier_name']}"
    )
    
    return {"message": "تمت الموافقة على الطلب وتنفيذ التعديل"}

# Reject supplier modification request
@api_router.put("/admin/supplier-modification-requests/{request_id}/reject")
async def reject_supplier_modification(
    request_id: str,
    reason: str = "",
    current_user: dict = Depends(require_role(["admin"]))
):
    """رفض طلب تعديل/نقل المورد"""
    if current_user.get("username") != "yasir" and current_user.get("department") != "it":
        raise HTTPException(status_code=403, detail="فقط مدير تقنية المعلومات يمكنه رفض الطلبات")
    
    mod_request = await db.supplier_modification_requests.find_one({"id": request_id}, {"_id": 0})
    if not mod_request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if mod_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail="تم معالجة هذا الطلب مسبقاً")
    
    await db.supplier_modification_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "approved_by": current_user["id"],
            "approved_by_name": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": reason
        }}
    )
    
    return {"message": "تم رفض الطلب"}

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: dict = Depends(require_role(["admin"]))):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    result = await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": {"is_active": False}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_supplier",
        entity_type="supplier",
        entity_id=supplier_id,
        entity_name=supplier.get("name"),
        details=f"حذف مورد: {supplier.get('name')}"
    )
    
    return {"message": "Supplier deleted successfully"}

# ==================== SUPPLIER PORTAL (بوابة الموردين) ====================

# Models imported from models/all_models.py:
# SupplierFeedRequestBase, SupplierFeedRequestCreate, SupplierFeedRequest
# SupplierMessageBase, SupplierMessageCreate, SupplierMessage
# SupplierLoginRequest

@api_router.post("/supplier-portal/login")
async def supplier_portal_login(login_data: SupplierLoginRequest):
    """تسجيل دخول المورد بالكود وكلمة المرور"""
    import hashlib
    
    supplier = await db.suppliers.find_one({"supplier_code": login_data.supplier_code, "is_active": True}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="كود المورد غير صحيح أو غير مفعل")
    
    # Check password
    stored_password_hash = supplier.get("password_hash")
    portal_password = supplier.get("portal_password")  # New simple hash for default password
    password_changed = supplier.get("password_changed", False)
    
    password_valid = False
    
    # First check if using the new portal_password (0000 default)
    if portal_password:
        input_hash = hashlib.sha256(login_data.password.encode()).hexdigest()
        if input_hash == portal_password:
            password_valid = True
    
    # Then check old password_hash if set
    if not password_valid and stored_password_hash:
        if verify_password(login_data.password, stored_password_hash):
            password_valid = True
    
    # Finally check legacy default (phone last 4 digits)
    if not password_valid and not stored_password_hash and not portal_password:
        phone = supplier.get("phone", "")
        default_password = phone[-4:] if len(phone) >= 4 else "1234"
        if login_data.password == default_password:
            password_valid = True
    
    if not password_valid:
        raise HTTPException(status_code=401, detail="كلمة المرور غير صحيحة")
    
    # Create token
    token_data = {
        "sub": supplier["id"],
        "supplier_code": login_data.supplier_code,
        "type": "supplier",
        "exp": datetime.now(timezone.utc) + timedelta(hours=24)
    }
    token = jwt.encode(token_data, SECRET_KEY, algorithm=ALGORITHM)
    
    return {
        "access_token": token,
        "supplier": {
            "id": supplier["id"],
            "name": supplier.get("name"),
            "code": supplier.get("supplier_code"),
            "phone": supplier.get("phone"),
            "balance": supplier.get("balance", 0),
            "total_supplied": supplier.get("total_supplied", 0),
            "milk_type": supplier.get("milk_type"),
            "center_name": supplier.get("center_name"),
            "has_custom_password": password_changed or stored_password_hash is not None,
            "password_changed": password_changed
        }
    }

# Supplier Portal - Recover password by phone
@api_router.post("/supplier-portal/recover-password")
async def supplier_recover_password(supplier_code: str, phone: str):
    """استرجاع كلمة المرور عن طريق رقم الهاتف"""
    supplier = await db.suppliers.find_one({"supplier_code": supplier_code}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="كود المورد غير موجود")
    
    # Verify phone number matches
    stored_phone = supplier.get("phone", "").replace(" ", "").replace("-", "")
    input_phone = phone.replace(" ", "").replace("-", "")
    
    # Check if phone matches (allow partial match for last digits)
    if not (stored_phone == input_phone or stored_phone.endswith(input_phone[-8:]) or input_phone.endswith(stored_phone[-8:])):
        raise HTTPException(status_code=400, detail="رقم الهاتف غير مطابق لسجلات المورد")
    
    # Generate new password (last 4 digits of phone)
    new_password = stored_phone[-4:] if len(stored_phone) >= 4 else "1234"
    
    # Reset password to default (remove custom password)
    await db.suppliers.update_one(
        {"supplier_code": supplier_code},
        {"$unset": {"password_hash": ""}}
    )
    
    return {
        "message": "تم إعادة تعيين كلمة المرور بنجاح",
        "new_password": new_password,
        "hint": f"كلمة المرور الجديدة هي آخر 4 أرقام من رقم هاتفك: {new_password}"
    }

# Supplier Portal - Change password
@api_router.put("/supplier-portal/change-password")
async def supplier_change_password(
    supplier_code: str,
    current_password: str,
    new_password: str
):
    """تغيير كلمة مرور المورد"""
    import hashlib
    
    supplier = await db.suppliers.find_one({"supplier_code": supplier_code}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    # Verify current password
    stored_password_hash = supplier.get("password_hash")
    portal_password = supplier.get("portal_password")
    
    password_valid = False
    
    # Check portal_password (default 0000)
    if portal_password:
        input_hash = hashlib.sha256(current_password.encode()).hexdigest()
        if input_hash == portal_password:
            password_valid = True
    
    # Check old password_hash
    if not password_valid and stored_password_hash:
        if verify_password(current_password, stored_password_hash):
            password_valid = True
    
    # Check legacy default
    if not password_valid and not stored_password_hash and not portal_password:
        phone = supplier.get("phone", "")
        default_password = phone[-4:] if len(phone) >= 4 else "1234"
        if current_password == default_password:
            password_valid = True
    
    if not password_valid:
        raise HTTPException(status_code=401, detail="كلمة المرور الحالية غير صحيحة")
    
    # Validate new password
    if len(new_password) < 4:
        raise HTTPException(status_code=400, detail="كلمة المرور يجب أن تكون 4 أحرف على الأقل")
    
    # Hash and save new password (using both methods for compatibility)
    new_password_hash = hash_password(new_password)
    new_portal_password = hashlib.sha256(new_password.encode()).hexdigest()
    
    await db.suppliers.update_one(
        {"supplier_code": supplier_code},
        {"$set": {
            "password_hash": new_password_hash,
            "portal_password": new_portal_password,
            "password_changed": True
        }}
    )
    
    return {"message": "تم تغيير كلمة المرور بنجاح"}

# Supplier Portal - Send OTP for password recovery
@api_router.post("/supplier-portal/send-otp")
async def send_otp_for_recovery(data: dict):
    """إرسال رمز التحقق لاسترجاع كلمة المرور"""
    import random
    
    phone = data.get("phone", "").strip()
    if not phone:
        raise HTTPException(status_code=400, detail="يرجى إدخال رقم الهاتف")
    
    # Find supplier by phone
    supplier = await db.suppliers.find_one({"phone": phone, "is_active": True}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="رقم الهاتف غير مسجل في النظام")
    
    # Generate 4-digit OTP
    otp = str(random.randint(1000, 9999))
    
    # Store OTP with expiry (5 minutes)
    otp_expiry = datetime.now(timezone.utc) + timedelta(minutes=5)
    await db.otp_codes.update_one(
        {"phone": phone},
        {"$set": {
            "phone": phone,
            "otp": otp,
            "expiry": otp_expiry.isoformat(),
            "supplier_id": supplier["id"],
            "used": False
        }},
        upsert=True
    )
    
    # In production, send SMS here
    # For now, log the OTP (in real app, use SMS gateway like Twilio)
    logging.info(f"OTP for {phone}: {otp}")
    
    return {
        "message": "تم إرسال رمز التحقق إلى هاتفك",
        "otp_sent": True,
        # For testing only - remove in production
        "debug_otp": otp
    }

# Supplier Portal - Verify OTP and reset password
@api_router.post("/supplier-portal/verify-otp-reset")
async def verify_otp_and_reset_password(data: dict):
    """التحقق من رمز OTP وتغيير كلمة المرور"""
    import hashlib
    
    phone = data.get("phone", "").strip()
    otp = data.get("otp", "").strip()
    new_password = data.get("new_password", "")
    
    if not phone or not otp or not new_password:
        raise HTTPException(status_code=400, detail="يرجى ملء جميع الحقول")
    
    if len(new_password) < 4:
        raise HTTPException(status_code=400, detail="كلمة المرور يجب أن تكون 4 أحرف على الأقل")
    
    # Find OTP record
    otp_record = await db.otp_codes.find_one({"phone": phone, "used": False}, {"_id": 0})
    if not otp_record:
        raise HTTPException(status_code=400, detail="لم يتم إرسال رمز التحقق لهذا الرقم")
    
    # Check if OTP expired
    expiry = datetime.fromisoformat(otp_record["expiry"].replace("Z", "+00:00"))
    if datetime.now(timezone.utc) > expiry:
        raise HTTPException(status_code=400, detail="انتهت صلاحية رمز التحقق. يرجى طلب رمز جديد")
    
    # Verify OTP
    if otp_record["otp"] != otp:
        raise HTTPException(status_code=400, detail="رمز التحقق غير صحيح")
    
    # Mark OTP as used
    await db.otp_codes.update_one(
        {"phone": phone},
        {"$set": {"used": True}}
    )
    
    # Update supplier password
    supplier_id = otp_record["supplier_id"]
    new_password_hash = hash_password(new_password)
    new_portal_password = hashlib.sha256(new_password.encode()).hexdigest()
    
    await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": {
            "password_hash": new_password_hash,
            "portal_password": new_portal_password,
            "password_changed": True
        }}
    )
    
    return {"message": "تم تغيير كلمة المرور بنجاح"}

# Supplier Portal - Set password for supplier (Admin only)
@api_router.put("/admin/suppliers/{supplier_id}/set-password")
async def admin_set_supplier_password(
    supplier_id: str,
    new_password: str,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """تعيين كلمة مرور للمورد (للإدارة فقط)"""
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    if len(new_password) < 4:
        raise HTTPException(status_code=400, detail="كلمة المرور يجب أن تكون 4 أحرف على الأقل")
    
    new_password_hash = hash_password(new_password)
    await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": {"password_hash": new_password_hash}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="set_supplier_password",
        entity_type="supplier",
        entity_id=supplier_id,
        entity_name=supplier.get("name"),
        details=f"تعيين كلمة مرور للمورد: {supplier.get('name')}"
    )
    
    return {"message": "تم تعيين كلمة المرور بنجاح"}

# Supplier Portal - Get my data
@api_router.get("/supplier-portal/me")
async def get_supplier_portal_data(authorization: str = None):
    """جلب بيانات المورد الحالي"""
    if not authorization:
        raise HTTPException(status_code=401, detail="غير مصرح")
    
    try:
        token = authorization.replace("Bearer ", "")
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("type") != "supplier":
            raise HTTPException(status_code=401, detail="غير مصرح")
        
        supplier_id = payload.get("sub")
        supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
        if not supplier:
            raise HTTPException(status_code=404, detail="المورد غير موجود")
        
        return supplier
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة")
    except:
        raise HTTPException(status_code=401, detail="غير مصرح")

# Supplier Portal - Get my milk receptions
@api_router.get("/supplier-portal/milk-receptions")
async def get_supplier_milk_receptions(
    supplier_code: str,
    month: Optional[int] = None,
    year: Optional[int] = None
):
    """جلب سجلات استلام الحليب للمورد"""
    supplier = await db.suppliers.find_one({"supplier_code": supplier_code}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    query = {"supplier_id": supplier["id"]}
    
    if month and year:
        # Filter by month and year
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query["date"] = {"$gte": start_date, "$lt": end_date}
    
    receptions = await db.milk_receptions.find(query, {"_id": 0}).sort("date", -1).to_list(100)
    
    # Calculate totals
    total_quantity = sum(r.get("quantity_liters", 0) for r in receptions)
    total_amount = sum(r.get("total_amount", 0) for r in receptions)
    
    return {
        "supplier": {
            "name": supplier.get("name"),
            "code": supplier.get("supplier_code"),
            "balance": supplier.get("balance", 0),
            "total_supplied": supplier.get("total_supplied", 0)
        },
        "receptions": receptions,
        "summary": {
            "total_quantity": total_quantity,
            "total_amount": total_amount,
            "count": len(receptions)
        }
    }

# Supplier Portal - Request feed (تحويل رصيد إلى أعلاف)
@api_router.post("/supplier-portal/feed-requests")
async def create_supplier_feed_request(request_data: SupplierFeedRequestCreate):
    """إنشاء طلب تحويل رصيد إلى أعلاف"""
    # Verify supplier exists and has enough balance
    supplier = await db.suppliers.find_one({"supplier_code": request_data.supplier_code}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    if supplier.get("balance", 0) < request_data.amount_to_deduct:
        raise HTTPException(status_code=400, detail="الرصيد غير كافي")
    
    feed_request = SupplierFeedRequest(**request_data.model_dump())
    await db.supplier_feed_requests.insert_one(feed_request.model_dump())
    
    return {
        "message": "تم إرسال طلب الأعلاف بنجاح وبانتظار موافقة قسم الأعلاف",
        "request_id": feed_request.id,
        "status": "pending"
    }

# Supplier Portal - Get my feed requests
@api_router.get("/supplier-portal/feed-requests")
async def get_supplier_feed_requests(supplier_code: str):
    """جلب طلبات الأعلاف للمورد"""
    requests = await db.supplier_feed_requests.find(
        {"supplier_code": supplier_code}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return requests

# Supplier Portal - Send message
@api_router.post("/supplier-portal/messages")
async def create_supplier_message(message_data: SupplierMessageCreate):
    """إرسال رسالة من المورد"""
    message = SupplierMessage(**message_data.model_dump())
    await db.supplier_messages.insert_one(message.model_dump())
    
    return {
        "message": "تم إرسال الرسالة بنجاح",
        "message_id": message.id
    }

# Supplier Portal - Get my messages
@api_router.get("/supplier-portal/messages")
async def get_supplier_messages(supplier_code: str):
    """جلب رسائل المورد"""
    messages = await db.supplier_messages.find(
        {"supplier_code": supplier_code}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return messages

# Admin - Get all feed requests
@api_router.get("/admin/supplier-feed-requests")
async def get_all_feed_requests(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب جميع طلبات الأعلاف (للإدارة)"""
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.supplier_feed_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return requests

# Admin - Approve feed request
@api_router.put("/admin/supplier-feed-requests/{request_id}/approve")
async def approve_feed_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """الموافقة على طلب الأعلاف وخصم المبلغ"""
    feed_request = await db.supplier_feed_requests.find_one({"id": request_id}, {"_id": 0})
    if not feed_request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if feed_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail="تم معالجة هذا الطلب مسبقاً")
    
    # Deduct from supplier balance
    supplier = await db.suppliers.find_one({"id": feed_request["supplier_id"]}, {"_id": 0})
    if supplier.get("balance", 0) < feed_request["amount_to_deduct"]:
        raise HTTPException(status_code=400, detail="رصيد المورد غير كافي")
    
    await db.suppliers.update_one(
        {"id": feed_request["supplier_id"]},
        {"$inc": {"balance": -feed_request["amount_to_deduct"]}}
    )
    
    # Update request status
    await db.supplier_feed_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_by_name": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="approve_feed_request",
        entity_type="supplier_feed_request",
        entity_id=request_id,
        entity_name=feed_request["supplier_name"],
        details=f"موافقة على طلب أعلاف: {feed_request['supplier_name']} - {feed_request['amount_to_deduct']} ريال"
    )
    
    return {"message": "تمت الموافقة على الطلب وخصم المبلغ من الرصيد"}

# Admin - Reject feed request
@api_router.put("/admin/supplier-feed-requests/{request_id}/reject")
async def reject_feed_request(
    request_id: str, 
    reason: str = "",
    current_user: dict = Depends(get_current_user)
):
    """رفض طلب الأعلاف"""
    feed_request = await db.supplier_feed_requests.find_one({"id": request_id}, {"_id": 0})
    if not feed_request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if feed_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail="تم معالجة هذا الطلب مسبقاً")
    
    await db.supplier_feed_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "approved_by": current_user["id"],
            "approved_by_name": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": reason
        }}
    )
    
    return {"message": "تم رفض الطلب"}

# Admin - Get all supplier messages
@api_router.get("/admin/supplier-messages")
async def get_all_supplier_messages(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب جميع رسائل الموردين (للإدارة)"""
    query = {}
    if status:
        query["status"] = status
    
    messages = await db.supplier_messages.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return messages

# Admin - Reply to supplier message
@api_router.put("/admin/supplier-messages/{message_id}/reply")
async def reply_to_supplier_message(
    message_id: str,
    reply: str,
    current_user: dict = Depends(get_current_user)
):
    """الرد على رسالة المورد"""
    await db.supplier_messages.update_one(
        {"id": message_id},
        {"$set": {
            "status": "replied",
            "reply": reply,
            "replied_by": current_user["full_name"],
            "replied_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "تم إرسال الرد"}

# ==================== MILK RECEPTION ROUTES ====================

@api_router.post("/milk-receptions", response_model=MilkReception)
async def create_milk_reception(reception_data: MilkReceptionCreate, current_user: dict = Depends(get_current_user)):
    reception = MilkReception(**reception_data.model_dump())
    reception.total_amount = reception.quantity_liters * reception.price_per_liter
    reception.created_by = current_user["id"]
    
    await db.milk_receptions.insert_one(reception.model_dump())
    
    # Update supplier's total supplied
    await db.suppliers.update_one(
        {"id": reception.supplier_id},
        {"$inc": {"total_supplied": reception.quantity_liters, "balance": reception.total_amount}}
    )
    
    # Update inventory
    await db.inventory.update_one(
        {"product_type": "raw_milk"},
        {"$inc": {"quantity_liters": reception.quantity_liters}, "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_milk_reception",
        entity_type="milk_reception",
        entity_id=reception.id,
        entity_name=reception.supplier_name,
        details=f"استلام حليب: {reception.quantity_liters} لتر من {reception.supplier_name}"
    )
    
    # === AUTO JOURNAL ENTRY: Milk Purchase ===
    # Dr: مشتريات الحليب (5100) / Cr: الموردين (2110)
    await create_auto_journal_entry(
        description=f"شراء حليب من {reception.supplier_name} - {reception.quantity_liters} لتر",
        lines=[
            {"account_number": "5100", "debit": reception.total_amount, "credit": 0, "description": "تكلفة شراء الحليب"},
            {"account_number": "2110", "debit": 0, "credit": reception.total_amount, "description": f"مستحق للمورد {reception.supplier_name}"}
        ],
        reference_type="milk_purchase",
        reference_id=reception.id,
        created_by_id=current_user["id"],
        created_by_name=current_user["full_name"]
    )
    
    return reception

@api_router.get("/milk-receptions")
async def get_milk_receptions(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    supplier_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if supplier_id:
        query["supplier_id"] = supplier_id
    if start_date:
        query["reception_date"] = {"$gte": start_date}
    if end_date:
        if "reception_date" in query:
            query["reception_date"]["$lte"] = end_date
        else:
            query["reception_date"] = {"$lte": end_date}
    
    receptions = await db.milk_receptions.find(query, {"_id": 0}).sort("reception_date", -1).to_list(2500)
    return receptions

@api_router.get("/milk-receptions/{reception_id}", response_model=MilkReception)
async def get_milk_reception(reception_id: str, current_user: dict = Depends(get_current_user)):
    reception = await db.milk_receptions.find_one({"id": reception_id}, {"_id": 0})
    if not reception:
        raise HTTPException(status_code=404, detail="Milk reception not found")
    return reception

# ==================== CUSTOMER ROUTES ====================

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, current_user: dict = Depends(get_current_user)):
    customer = Customer(**customer_data.model_dump())
    await db.customers.insert_one(customer.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_customer",
        entity_type="customer",
        entity_id=customer.id,
        entity_name=customer.name,
        details=f"إضافة عميل: {customer.name}"
    )
    
    return customer

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(current_user: dict = Depends(get_current_user)):
    customers = await db.customers.find({"is_active": True}, {"_id": 0}).to_list(1000)
    return customers

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer_data: CustomerCreate, current_user: dict = Depends(get_current_user)):
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": customer_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_customer",
        entity_type="customer",
        entity_id=customer_id,
        entity_name=customer.get("name"),
        details=f"تعديل بيانات عميل: {customer.get('name')}"
    )
    
    return customer

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(require_role(["admin"]))):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"is_active": False}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_customer",
        entity_type="customer",
        entity_id=customer_id,
        entity_name=customer.get("name"),
        details=f"حذف عميل: {customer.get('name')}"
    )
    
    return {"message": "Customer deleted successfully"}

# ==================== SALES ROUTES ====================

@api_router.post("/sales", response_model=Sale)
async def create_sale(sale_data: SaleCreate, current_user: dict = Depends(get_current_user)):
    sale = Sale(**sale_data.model_dump())
    sale.total_amount = sale.quantity_liters * sale.price_per_liter
    sale.created_by = current_user["id"]
    sale.is_paid = sale.sale_type == "cash"
    
    await db.sales.insert_one(sale.model_dump())
    
    # Update customer's total purchases
    balance_change = 0 if sale.is_paid else sale.total_amount
    await db.customers.update_one(
        {"id": sale.customer_id},
        {"$inc": {"total_purchases": sale.total_amount, "balance": balance_change}}
    )
    
    # Update inventory
    await db.inventory.update_one(
        {"product_type": "raw_milk"},
        {"$inc": {"quantity_liters": -sale.quantity_liters}, "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_sale",
        entity_type="sale",
        entity_id=sale.id,
        entity_name=sale.customer_name,
        details=f"عملية بيع: {sale.quantity_liters} لتر إلى {sale.customer_name} - {sale.total_amount} ر.ع"
    )
    
    # === AUTO JOURNAL ENTRY: Milk Sale ===
    if sale.is_paid:
        # Cash sale: Dr: الصندوق (1111) / Cr: إيرادات مبيعات الحليب (4100)
        await create_auto_journal_entry(
            description=f"بيع حليب نقدي إلى {sale.customer_name} - {sale.quantity_liters} لتر",
            lines=[
                {"account_number": "1111", "debit": sale.total_amount, "credit": 0, "description": "نقدية من بيع الحليب"},
                {"account_number": "4100", "debit": 0, "credit": sale.total_amount, "description": "إيراد مبيعات الحليب"}
            ],
            reference_type="milk_sale",
            reference_id=sale.id,
            created_by_id=current_user["id"],
            created_by_name=current_user["full_name"]
        )
    else:
        # Credit sale: Dr: العملاء (1120) / Cr: إيرادات مبيعات الحليب (4100)
        await create_auto_journal_entry(
            description=f"بيع حليب آجل إلى {sale.customer_name} - {sale.quantity_liters} لتر",
            lines=[
                {"account_number": "1120", "debit": sale.total_amount, "credit": 0, "description": f"مستحق من العميل {sale.customer_name}"},
                {"account_number": "4100", "debit": 0, "credit": sale.total_amount, "description": "إيراد مبيعات الحليب"}
            ],
            reference_type="milk_sale",
            reference_id=sale.id,
            created_by_id=current_user["id"],
            created_by_name=current_user["full_name"]
        )
    
    return sale

@api_router.get("/sales", response_model=List[Sale])
async def get_sales(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    if start_date:
        query["sale_date"] = {"$gte": start_date}
    if end_date:
        if "sale_date" in query:
            query["sale_date"]["$lte"] = end_date
        else:
            query["sale_date"] = {"$lte": end_date}
    
    sales = await db.sales.find(query, {"_id": 0}).sort("sale_date", -1).to_list(1000)
    return sales

# ==================== INVENTORY ROUTES ====================

@api_router.get("/inventory")
async def get_inventory(current_user: dict = Depends(get_current_user)):
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(100)
    return inventory

@api_router.post("/inventory", response_model=Inventory)
async def create_inventory(inventory_data: InventoryBase, current_user: dict = Depends(require_role(["admin", "employee"]))):
    inventory = Inventory(**inventory_data.model_dump())
    await db.inventory.insert_one(inventory.model_dump())
    return inventory

@api_router.put("/inventory/{inventory_id}")
async def update_inventory(inventory_id: str, inventory_data: InventoryUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in inventory_data.model_dump().items() if v is not None}
    update_data["last_updated"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.inventory.update_one(
        {"id": inventory_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    
    inventory = await db.inventory.find_one({"id": inventory_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_inventory",
        entity_type="inventory",
        entity_id=inventory_id,
        entity_name=inventory.get("product_name", ""),
        details=f"تعديل مخزون: {inventory.get('product_name', '')} - الكمية: {inventory.get('quantity_liters', 0)} لتر"
    )
    
    return inventory

@api_router.delete("/inventory/{inventory_id}")
async def delete_inventory(inventory_id: str, current_user: dict = Depends(require_role(["admin"]))):
    """Delete an inventory item (admin only)"""
    existing = await db.inventory.find_one({"id": inventory_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="المخزون غير موجود")
    
    await db.inventory.delete_one({"id": inventory_id})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_inventory",
        entity_type="inventory",
        entity_id=inventory_id,
        entity_name=existing.get("product_name", ""),
        details=f"حذف مخزون: {existing.get('product_name', '')} - الكمية: {existing.get('quantity_liters', 0)} لتر"
    )
    
    return {"message": "تم حذف المخزون بنجاح"}

# ==================== PAYMENT ROUTES ====================

@api_router.post("/payments", response_model=Payment)
async def create_payment(payment_data: PaymentCreate, current_user: dict = Depends(require_role(["admin", "accountant"]))):
    """Create a payment request (requires approval from admin/IT)"""
    payment = Payment(**payment_data.model_dump())
    payment.created_by = current_user["id"]
    payment.created_by_name = current_user.get("full_name", "")
    payment.status = "pending"  # All payments start as pending
    
    await db.payments.insert_one(payment.model_dump())
    
    # Get entity name for logging
    entity_name = ""
    if payment.payment_type == "supplier_payment":
        supplier = await db.suppliers.find_one({"id": payment.related_id}, {"_id": 0})
        entity_name = supplier.get("name", "") if supplier else ""
    elif payment.payment_type == "customer_receipt":
        customer = await db.customers.find_one({"id": payment.related_id}, {"_id": 0})
        entity_name = customer.get("name", "") if customer else ""
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_payment_request",
        entity_type="payment",
        entity_id=payment.id,
        entity_name=entity_name,
        details=f"طلب دفعة مالية: {payment.amount} ر.ع - {entity_name} (في انتظار الموافقة)"
    )
    
    return payment

@api_router.post("/payments/{payment_id}/approve")
async def approve_payment(payment_id: str, approval: PaymentApproval, current_user: dict = Depends(require_role(["admin"]))):
    """Approve or reject a payment request (admin/IT only)"""
    
    # Get the payment
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="الدفعة غير موجودة")
    
    if payment.get("status") != "pending":
        raise HTTPException(status_code=400, detail="هذه الدفعة تمت معالجتها مسبقاً")
    
    entity_name = payment.get("related_name", "")
    amount = payment.get("amount", 0)
    
    if approval.action == "approve":
        # For supplier payment, check treasury balance
        if payment.get("payment_type") == "supplier_payment":
            treasury = await db.treasury.find_one({"type": "main"}, {"_id": 0})
            treasury_balance = treasury.get("current_balance", 0) if treasury else 0
            if amount > treasury_balance:
                raise HTTPException(
                    status_code=400, 
                    detail=f"رصيد الخزينة غير كافٍ. الرصيد الحالي: {treasury_balance} ر.ع، المطلوب: {amount} ر.ع"
                )
        
        # Update payment status
        await db.payments.update_one(
            {"id": payment_id},
            {
                "$set": {
                    "status": "approved",
                    "approved_by": current_user["id"],
                    "approved_by_name": current_user.get("full_name", ""),
                    "approved_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        # Update balances and treasury after approval
        if payment.get("payment_type") == "supplier_payment":
            # Deduct from supplier balance
            await db.suppliers.update_one(
                {"id": payment.get("related_id")},
                {"$inc": {"balance": -amount}}
            )
            # Deduct from treasury (withdrawal)
            await update_treasury(
                transaction_type="withdrawal",
                amount=amount,
                source_type="supplier_payment",
                description=f"دفعة للمورد: {entity_name}",
                source_id=payment_id,
                user_id=current_user["id"],
                user_name=current_user.get("full_name", "")
            )
            
        elif payment.get("payment_type") == "customer_receipt":
            # Deduct from customer balance (receivables)
            await db.customers.update_one(
                {"id": payment.get("related_id")},
                {"$inc": {"balance": -amount}}
            )
            # Add to treasury (deposit)
            await update_treasury(
                transaction_type="deposit",
                amount=amount,
                source_type="customer_receipt",
                description=f"استلام من العميل: {entity_name}",
                source_id=payment_id,
                user_id=current_user["id"],
                user_name=current_user.get("full_name", "")
            )
        
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user["full_name"],
            action="approve_payment",
            entity_type="payment",
            entity_id=payment_id,
            entity_name=entity_name,
            details=f"تمت الموافقة على دفعة: {amount} ر.ع - {entity_name}"
        )
        
        return {"message": "تمت الموافقة على الدفعة بنجاح", "status": "approved"}
    
    elif approval.action == "reject":
        await db.payments.update_one(
            {"id": payment_id},
            {
                "$set": {
                    "status": "rejected",
                    "approved_by": current_user["id"],
                    "approved_by_name": current_user.get("full_name", ""),
                    "approved_at": datetime.now(timezone.utc).isoformat(),
                    "rejection_reason": approval.reason or "لم يتم تحديد السبب"
                }
            }
        )
        
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user["full_name"],
            action="reject_payment",
            entity_type="payment",
            entity_id=payment_id,
            entity_name=entity_name,
            details=f"تم رفض دفعة: {amount} ر.ع - {entity_name} - السبب: {approval.reason or 'غير محدد'}"
        )
        
        return {"message": "تم رفض الدفعة", "status": "rejected"}
    
    raise HTTPException(status_code=400, detail="الإجراء غير صالح")

@api_router.get("/payments/pending", response_model=List[Payment])
async def get_pending_payments(current_user: dict = Depends(require_role(["admin"]))):
    """Get all pending payments awaiting approval (admin only)"""
    payments = await db.payments.find({"status": "pending"}, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    return payments

@api_router.get("/payments", response_model=List[Payment])
async def get_payments(
    payment_type: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if payment_type:
        query["payment_type"] = payment_type
    if status:
        query["status"] = status
    if start_date:
        query["payment_date"] = {"$gte": start_date}
    if end_date:
        if "payment_date" in query:
            query["payment_date"]["$lte"] = end_date
        else:
            query["payment_date"] = {"$lte": end_date}
    
    payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    return payments

# Payment Receipt PDF Generation
@api_router.get("/payments/{payment_id}/receipt")
async def get_payment_receipt_pdf(payment_id: str, current_user: dict = Depends(get_current_user)):
    """Generate PDF receipt for a supplier payment"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import cm
    from io import BytesIO
    import arabic_reshaper
    from bidi.algorithm import get_display
    
    # Register Arabic font
    font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    try:
        pdfmetrics.registerFont(TTFont('Arabic', font_path))
    except:
        pass
    
    # Get payment details
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Get supplier details
    supplier = await db.suppliers.find_one({"id": payment.get("related_id")}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    def reshape_arabic(text):
        try:
            reshaped = arabic_reshaper.reshape(str(text))
            return get_display(reshaped)
        except:
            return str(text)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontName='Arabic',
        fontSize=24,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontName='Arabic',
        fontSize=14,
        alignment=TA_CENTER,
        spaceAfter=10
    )
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontName='Arabic',
        fontSize=12,
        alignment=TA_RIGHT,
        spaceAfter=5
    )
    
    elements = []
    
    # Company Header
    elements.append(Paragraph(reshape_arabic("المروج للألبان"), title_style))
    elements.append(Paragraph(reshape_arabic("Al-Morooj Dairy"), header_style))
    elements.append(Spacer(1, 20))
    
    # Receipt Title
    elements.append(Paragraph(reshape_arabic("إيصال دفع"), title_style))
    elements.append(Spacer(1, 20))
    
    # Payment Date
    payment_date = payment.get("payment_date", "")[:10]
    elements.append(Paragraph(reshape_arabic(f"التاريخ: {payment_date}"), normal_style))
    elements.append(Spacer(1, 10))
    
    # Supplier Information Table
    supplier_data = [
        [reshape_arabic("القيمة"), reshape_arabic("البيان")],
        [reshape_arabic(supplier.get("name", "")), reshape_arabic("اسم المورد")],
        [reshape_arabic(supplier.get("supplier_code", "-")), reshape_arabic("كود المورد")],
        [reshape_arabic(supplier.get("phone", "-")), reshape_arabic("رقم الهاتف")],
        [reshape_arabic(supplier.get("address", "-")), reshape_arabic("العنوان")],
        [reshape_arabic(supplier.get("bank_account", "-")), reshape_arabic("الحساب البنكي")],
        [reshape_arabic(supplier.get("national_id", "-")), reshape_arabic("رقم الهوية")],
    ]
    
    supplier_table = Table(supplier_data, colWidths=[10*cm, 5*cm])
    supplier_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#e2e8f0")),
    ]))
    elements.append(supplier_table)
    elements.append(Spacer(1, 20))
    
    # Payment Details Table
    payment_method_map = {
        "cash": "نقداً",
        "bank_transfer": "تحويل بنكي",
        "check": "شيك"
    }
    payment_method = payment_method_map.get(payment.get("payment_method", "cash"), payment.get("payment_method", ""))
    
    payment_data = [
        [reshape_arabic("القيمة"), reshape_arabic("تفاصيل الدفع")],
        [reshape_arabic(f"{payment.get('amount', 0):,.2f} ر.ع"), reshape_arabic("المبلغ المدفوع")],
        [reshape_arabic(payment_method), reshape_arabic("طريقة الدفع")],
        [reshape_arabic(payment.get("notes", "-") or "-"), reshape_arabic("ملاحظات")],
    ]
    
    payment_table = Table(payment_data, colWidths=[10*cm, 5*cm])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#059669")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f0fdf4")),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#d1fae5")),
    ]))
    elements.append(payment_table)
    elements.append(Spacer(1, 30))
    
    # Signature Section
    sig_data = [
        [reshape_arabic("توقيع المستلم"), reshape_arabic(""), reshape_arabic("توقيع المسؤول")],
        ["________________", "", "________________"],
    ]
    sig_table = Table(sig_data, colWidths=[5*cm, 5*cm, 5*cm])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
    ]))
    elements.append(sig_table)
    
    # Footer
    elements.append(Spacer(1, 40))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontName='Arabic',
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.gray
    )
    elements.append(Paragraph(reshape_arabic(f"رقم الإيصال: {payment_id[:8].upper()}"), footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Log activity
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="generate_payment_receipt",
        entity_type="payment",
        entity_id=payment_id,
        entity_name=supplier.get("name"),
        details=f"طباعة إيصال دفع للمورد: {supplier.get('name')} - المبلغ: {payment.get('amount')}"
    )
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=payment_receipt_{payment_id[:8]}.pdf"}
    )

# ==================== EMPLOYEE ROUTES ====================

@api_router.post("/employees", response_model=Employee)
async def create_employee(employee_data: EmployeeCreate, current_user: dict = Depends(require_role(["admin"]))):
    employee = Employee(**employee_data.model_dump())
    await db.employees.insert_one(employee.model_dump())
    return employee

@api_router.get("/employees", response_model=List[Employee])
async def get_employees(current_user: dict = Depends(require_role(["admin"]))):
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    return employees

@api_router.put("/employees/{employee_id}", response_model=Employee)
async def update_employee(employee_id: str, employee_data: EmployeeCreate, current_user: dict = Depends(require_role(["admin"]))):
    # Get existing employee to preserve is_active status
    existing_employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not existing_employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Update only the fields from employee_data, preserving is_active
    update_data = employee_data.model_dump()
    update_data["is_active"] = existing_employee.get("is_active", True)
    
    result = await db.employees.update_one(
        {"id": employee_id},
        {"$set": update_data}
    )
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return employee

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current_user: dict = Depends(require_role(["admin"]))):
    result = await db.employees.update_one(
        {"id": employee_id},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee deleted successfully"}

# ==================== FEED COMPANY ROUTES (شركات الأعلاف) ====================

@api_router.post("/feed-companies", response_model=FeedCompany)
async def create_feed_company(company_data: FeedCompanyCreate, current_user: dict = Depends(get_current_user)):
    company = FeedCompany(**company_data.model_dump())
    await db.feed_companies.insert_one(company.model_dump())
    return company

@api_router.get("/feed-companies", response_model=List[FeedCompany])
async def get_feed_companies(current_user: dict = Depends(get_current_user)):
    companies = await db.feed_companies.find({"is_active": True}, {"_id": 0}).to_list(1000)
    return companies

@api_router.put("/feed-companies/{company_id}", response_model=FeedCompany)
async def update_feed_company(company_id: str, company_data: FeedCompanyCreate, current_user: dict = Depends(get_current_user)):
    result = await db.feed_companies.update_one(
        {"id": company_id},
        {"$set": company_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Feed company not found")
    company = await db.feed_companies.find_one({"id": company_id}, {"_id": 0})
    return company

@api_router.delete("/feed-companies/{company_id}")
async def delete_feed_company(company_id: str, current_user: dict = Depends(require_role(["admin"]))):
    result = await db.feed_companies.update_one(
        {"id": company_id},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Feed company not found")
    return {"message": "Feed company deleted successfully"}

# ==================== FEED TYPE ROUTES (أنواع الأعلاف) ====================

@api_router.post("/feed-types", response_model=FeedType)
async def create_feed_type(feed_type_data: FeedTypeCreate, current_user: dict = Depends(get_current_user)):
    feed_type = FeedType(**feed_type_data.model_dump())
    await db.feed_types.insert_one(feed_type.model_dump())
    return feed_type

@api_router.get("/feed-types", response_model=List[FeedType])
async def get_feed_types(company_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"is_active": True}
    if company_id:
        query["company_id"] = company_id
    feed_types = await db.feed_types.find(query, {"_id": 0}).to_list(1000)
    return feed_types

@api_router.put("/feed-types/{feed_type_id}", response_model=FeedType)
async def update_feed_type(feed_type_id: str, feed_type_data: FeedTypeCreate, current_user: dict = Depends(get_current_user)):
    result = await db.feed_types.update_one(
        {"id": feed_type_id},
        {"$set": feed_type_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Feed type not found")
    feed_type = await db.feed_types.find_one({"id": feed_type_id}, {"_id": 0})
    return feed_type

@api_router.delete("/feed-types/{feed_type_id}")
async def delete_feed_type(feed_type_id: str, current_user: dict = Depends(require_role(["admin"]))):
    result = await db.feed_types.update_one(
        {"id": feed_type_id},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Feed type not found")
    return {"message": "Feed type deleted successfully"}

# ==================== FEED PURCHASE ROUTES (مشتريات الأعلاف) ====================

@api_router.post("/feed-purchases", response_model=FeedPurchase)
async def create_feed_purchase(purchase_data: FeedPurchaseCreate, current_user: dict = Depends(get_current_user)):
    # Check supplier balance
    supplier = await db.suppliers.find_one({"id": purchase_data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    total_amount = purchase_data.quantity * purchase_data.price_per_unit
    
    if supplier.get("balance", 0) < total_amount:
        raise HTTPException(status_code=400, detail="Insufficient supplier balance")
    
    # Generate invoice number
    count = await db.feed_purchases.count_documents({})
    year = datetime.now().year
    invoice_number = f"FP-{year}-{count + 1:05d}"
    
    purchase = FeedPurchase(**purchase_data.model_dump())
    purchase.invoice_number = invoice_number
    purchase.total_amount = total_amount
    purchase.created_by = current_user["id"]
    purchase.created_by_name = current_user.get("full_name", "")
    # Add supplier details to invoice
    purchase.supplier_phone = supplier.get("phone", "")
    purchase.supplier_address = supplier.get("address", "")
    
    await db.feed_purchases.insert_one(purchase.model_dump())
    
    # Deduct from supplier balance
    await db.suppliers.update_one(
        {"id": purchase.supplier_id},
        {"$inc": {"balance": -total_amount}}
    )
    
    # ===== AUTO-LINK TO INVENTORY =====
    # Add feed to inventory
    feed_type = await db.feed_types.find_one({"id": purchase_data.feed_type_id}, {"_id": 0})
    feed_inventory = {
        "id": str(uuid.uuid4()),
        "product_type": "feed",
        "product_name": purchase.feed_type_name,
        "quantity": purchase.quantity,
        "unit": purchase.unit,
        "price_per_unit": purchase.price_per_unit,
        "total_value": total_amount,
        "supplier_id": purchase.supplier_id,
        "supplier_name": purchase.supplier_name,
        "company_name": purchase.company_name,
        "purchase_id": purchase.id,
        "invoice_number": invoice_number,
        "last_updated": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.feed_inventory.insert_one(feed_inventory)
    
    # ===== AUTO-LINK TO TREASURY/FINANCE =====
    # Record as expense in treasury (withdrawal from supplier credit to feed)
    await update_treasury(
        transaction_type="withdrawal",
        amount=total_amount,
        source_type="feed_purchase",
        source_id=purchase.id,
        description=f"شراء علف: {purchase.feed_type_name} - {purchase.quantity} {purchase.unit} للمورد {supplier.get('name')} (خصم من رصيد المورد)",
        user_id=current_user["id"],
        user_name=current_user.get("full_name", "")
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_feed_purchase",
        entity_type="feed_purchase",
        entity_id=purchase.id,
        entity_name=supplier.get("name"),
        details=f"فاتورة شراء علف: {invoice_number} - {purchase.feed_type_name} - {total_amount} ر.ع من رصيد {supplier.get('name')} (مرتبط بالمخزون والمالية)"
    )
    
    return purchase

# Approve feed purchase invoice (electronic signature)
@api_router.post("/feed-purchases/{purchase_id}/approve")
async def approve_feed_purchase(purchase_id: str, current_user: dict = Depends(require_role(["admin"]))):
    purchase = await db.feed_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if purchase.get("is_approved"):
        raise HTTPException(status_code=400, detail="Invoice already approved")
    
    # Generate signature code
    import hashlib
    signature_data = f"{purchase_id}-{current_user['id']}-{datetime.now().isoformat()}"
    signature_code = hashlib.sha256(signature_data.encode()).hexdigest()[:16].upper()
    
    await db.feed_purchases.update_one(
        {"id": purchase_id},
        {"$set": {
            "is_approved": True,
            "approved_by": current_user["id"],
            "approved_by_name": current_user.get("full_name", ""),
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "signature_code": signature_code
        }}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="approve_feed_purchase",
        entity_type="feed_purchase",
        entity_id=purchase_id,
        details=f"تصديق فاتورة شراء علف: {purchase.get('invoice_number')} - كود التصديق: {signature_code}"
    )
    
    return {"message": "تم تصديق الفاتورة بنجاح", "signature_code": signature_code}

@api_router.get("/feed-purchases", response_model=List[FeedPurchase])
async def get_feed_purchases(
    supplier_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if supplier_id:
        query["supplier_id"] = supplier_id
    if start_date:
        query["purchase_date"] = {"$gte": start_date}
    if end_date:
        if "purchase_date" in query:
            query["purchase_date"]["$lte"] = end_date
        else:
            query["purchase_date"] = {"$lte": end_date}
    
    purchases = await db.feed_purchases.find(query, {"_id": 0}).sort("purchase_date", -1).to_list(1000)
    return purchases

# Get feed purchase invoice for printing
@api_router.get("/feed-purchases/{purchase_id}/invoice")
async def get_feed_purchase_invoice(purchase_id: str, current_user: dict = Depends(get_current_user)):
    """Get feed purchase invoice details for printing"""
    purchase = await db.feed_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get supplier details
    supplier = await db.suppliers.find_one({"id": purchase.get("supplier_id")}, {"_id": 0})
    
    # Get company info
    company_info = {
        "name": "شركة المروج للألبان",
        "name_en": "Al Morooj Dairy Company",
        "address": "سلطنة عمان",
        "phone": "+968 XXXX XXXX",
        "cr_number": "XXXXXXXX"
    }
    
    return {
        "invoice": purchase,
        "supplier": supplier,
        "company": company_info,
        "print_time": datetime.now(timezone.utc).isoformat()
    }

@api_router.get("/feed-purchases/supplier/{supplier_id}")
async def get_supplier_feed_purchases(supplier_id: str, current_user: dict = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    purchases = await db.feed_purchases.find({"supplier_id": supplier_id}, {"_id": 0}).sort("purchase_date", -1).to_list(100)
    
    return {
        "supplier": supplier,
        "purchases": purchases,
        "total_spent": sum(p.get("total_amount", 0) for p in purchases),
        "available_balance": supplier.get("balance", 0)
    }

@api_router.put("/feed-purchases/{purchase_id}", response_model=FeedPurchase)
async def update_feed_purchase(purchase_id: str, purchase_data: FeedPurchaseCreate, current_user: dict = Depends(get_current_user)):
    # Get existing purchase
    existing = await db.feed_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Feed purchase not found")
    
    # Get supplier
    supplier = await db.suppliers.find_one({"id": purchase_data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Calculate new total
    new_total = purchase_data.quantity * purchase_data.price_per_unit
    old_total = existing.get("total_amount", 0)
    difference = new_total - old_total
    
    # Check if supplier has enough balance for the difference
    if difference > 0 and supplier.get("balance", 0) < difference:
        raise HTTPException(status_code=400, detail="Insufficient supplier balance")
    
    # Update purchase
    update_data = purchase_data.model_dump()
    update_data["total_amount"] = new_total
    
    await db.feed_purchases.update_one(
        {"id": purchase_id},
        {"$set": update_data}
    )
    
    # Update supplier balance
    if difference != 0:
        await db.suppliers.update_one(
            {"id": purchase_data.supplier_id},
            {"$inc": {"balance": -difference}}
        )
    
    purchase = await db.feed_purchases.find_one({"id": purchase_id}, {"_id": 0})
    return purchase

@api_router.delete("/feed-purchases/{purchase_id}")
async def delete_feed_purchase(purchase_id: str, current_user: dict = Depends(get_current_user)):
    # Get existing purchase
    existing = await db.feed_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Feed purchase not found")
    
    # Get supplier name for logging
    supplier = await db.suppliers.find_one({"id": existing["supplier_id"]}, {"_id": 0})
    supplier_name = supplier.get("name", "") if supplier else ""
    
    # Refund supplier balance
    await db.suppliers.update_one(
        {"id": existing["supplier_id"]},
        {"$inc": {"balance": existing.get("total_amount", 0)}}
    )
    
    # ===== REMOVE FROM INVENTORY =====
    await db.feed_inventory.delete_one({"purchase_id": purchase_id})
    
    # ===== REVERSE TREASURY TRANSACTION =====
    # Add back to treasury (deposit) to reverse the withdrawal
    await update_treasury(
        transaction_type="deposit",
        amount=existing.get("total_amount", 0),
        source_type="feed_purchase_reversal",
        source_id=purchase_id,
        description=f"إلغاء شراء علف: {existing.get('feed_type_name', '')} - إرجاع {existing.get('total_amount', 0)} ر.ع لرصيد {supplier_name}",
        user_id=current_user["id"],
        user_name=current_user.get("full_name", "")
    )
    
    # Delete purchase
    await db.feed_purchases.delete_one({"id": purchase_id})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_feed_purchase",
        entity_type="feed_purchase",
        entity_id=purchase_id,
        entity_name=supplier_name,
        details=f"حذف شراء علف وإرجاع {existing.get('total_amount', 0)} ر.ع لرصيد {supplier_name} (تم تحديث المخزون والمالية)"
    )
    
    return {"message": "Feed purchase deleted, amount refunded to supplier, and inventory/treasury updated"}


# ==================== FEED PURCHASE PRINT (طباعة طلب شراء الأعلاف) ====================

@api_router.get("/feed-purchases/{purchase_id}/print")
async def print_feed_purchase(purchase_id: str, current_user: dict = Depends(get_current_user)):
    """Generate printable PDF for feed purchase request"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    import io
    
    # Get purchase data
    purchase = await db.feed_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Get supplier data
    supplier = await db.suppliers.find_one({"id": purchase["supplier_id"]}, {"_id": 0})
    
    # Get all feed types for the table
    feed_types = await db.feed_types.find({}, {"_id": 0}).to_list(100)
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    
    # Try to register Arabic font
    try:
        pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        arabic_font = 'Arabic'
    except:
        arabic_font = 'Helvetica'
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', fontSize=18, alignment=TA_CENTER, fontName=arabic_font,
        spaceAfter=10
    )
    header_style = ParagraphStyle(
        'Header', fontSize=12, alignment=TA_LEFT, fontName=arabic_font,
        spaceAfter=5
    )
    remark_style = ParagraphStyle(
        'Remark', fontSize=9, alignment=TA_LEFT, fontName=arabic_font,
        spaceAfter=3
    )
    
    elements = []
    
    # Title
    elements.append(Paragraph("PURCHASE REQUEST", title_style))
    elements.append(Paragraph("طلب شراء أعلاف", title_style))
    elements.append(Spacer(1, 10*mm))
    
    # Header info table
    header_data = [
        [f"Farmer Name: {supplier.get('name', '')}", f"AMDC/DFI/{purchase.get('invoice_number', '')}"],
        [f"Farmer Code: {supplier.get('id', '')[:8]}", f"Date: {purchase.get('purchase_date', '')[:10]}"],
        [f"Farmer ID: {supplier.get('phone', '')}", ""],
    ]
    header_table = Table(header_data, colWidths=[100*mm, 80*mm])
    header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 10*mm))
    
    # Products table header
    products_header = [
        ["SL", "Product / اسم المنتج", "Weight / الوزن", "Quantity / الكمية"]
    ]
    
    # Build products data - show all feed types with purchased quantity
    products_data = []
    sl = 1
    for ft in feed_types:
        qty = 0
        weight = ft.get('kg_per_unit', 40)
        if ft['id'] == purchase.get('feed_type_id'):
            qty = purchase.get('quantity', 0)
        products_data.append([
            str(sl),
            f"{ft.get('name', '')}",
            f"{weight}",
            str(int(qty)) if qty else "0"
        ])
        sl += 1
    
    # If no feed types, show just the purchased item
    if not products_data:
        products_data.append([
            "1",
            purchase.get('feed_type_name', ''),
            str(purchase.get('unit', 'kg')),
            str(purchase.get('quantity', 0))
        ])
    
    all_products = products_header + products_data
    products_table = Table(all_products, colWidths=[15*mm, 90*mm, 35*mm, 40*mm])
    products_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(products_table)
    elements.append(Spacer(1, 15*mm))
    
    # Total
    total_data = [
        ["", "", "Total / الإجمالي:", f"{purchase.get('total_amount', 0)} OMR"]
    ]
    total_table = Table(total_data, colWidths=[15*mm, 90*mm, 35*mm, 40*mm])
    total_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (2, 0), (3, 0), 'CENTER'),
        ('FONTWEIGHT', (2, 0), (-1, -1), 'BOLD'),
    ]))
    elements.append(total_table)
    elements.append(Spacer(1, 15*mm))
    
    # Remarks
    elements.append(Paragraph("Remark: / ملاحظة:", header_style))
    elements.append(Paragraph(
        "1. The customer agreed to transfer the full feeds as per the purchase request signed",
        remark_style
    ))
    elements.append(Paragraph(
        "1- أنا العميل الموقع أعلاه موافق على شحن الكمية الموضحة في طلب الشراء بالكامل",
        remark_style
    ))
    elements.append(Spacer(1, 3*mm))
    elements.append(Paragraph(
        "2. All farmers should bring ID copy. Without ID proof, feeds will not be issued",
        remark_style
    ))
    elements.append(Paragraph(
        "2- على جميع المربين إحضار نسخة من البطاقة الشخصية. بدون البطاقة الشخصية لن يتم صرف الأعلاف",
        remark_style
    ))
    
    # Signature section
    elements.append(Spacer(1, 20*mm))
    sig_data = [
        ["Signature / التوقيع:", "_________________", "Date / التاريخ:", "_________________"]
    ]
    sig_table = Table(sig_data, colWidths=[40*mm, 50*mm, 40*mm, 50*mm])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    elements.append(sig_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=purchase_request_{purchase_id[:8]}.pdf"
        }
    )


# ==================== FEED INVENTORY & REPORTS (مخزون وتقارير الأعلاف) ====================

@api_router.get("/feed-inventory")
async def get_feed_inventory(current_user: dict = Depends(get_current_user)):
    """Get all feed inventory items"""
    inventory = await db.feed_inventory.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return inventory

@api_router.get("/feed-inventory/summary")
async def get_feed_inventory_summary(current_user: dict = Depends(get_current_user)):
    """Get feed inventory summary by feed type"""
    inventory = await db.feed_inventory.find({}, {"_id": 0}).to_list(1000)
    
    # Group by feed type
    summary = {}
    for item in inventory:
        product_name = item.get("product_name", "Unknown")
        if product_name not in summary:
            summary[product_name] = {
                "product_name": product_name,
                "company_name": item.get("company_name", ""),
                "unit": item.get("unit", "kg"),
                "total_quantity": 0,
                "total_value": 0,
                "purchase_count": 0
            }
        summary[product_name]["total_quantity"] += item.get("quantity", 0)
        summary[product_name]["total_value"] += item.get("total_value", 0)
        summary[product_name]["purchase_count"] += 1
    
    return list(summary.values())

@api_router.get("/reports/feed-purchases")
async def get_feed_purchase_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    supplier_id: Optional[str] = None,
    feed_type_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get detailed feed purchase report with statistics"""
    query = {}
    
    if supplier_id:
        query["supplier_id"] = supplier_id
    if feed_type_id:
        query["feed_type_id"] = feed_type_id
    if start_date:
        query["purchase_date"] = {"$gte": start_date}
    if end_date:
        if "purchase_date" in query:
            query["purchase_date"]["$lte"] = end_date
        else:
            query["purchase_date"] = {"$lte": end_date}
    
    purchases = await db.feed_purchases.find(query, {"_id": 0}).sort("purchase_date", -1).to_list(10000)
    
    # Calculate statistics
    total_purchases = len(purchases)
    total_amount = sum(p.get("total_amount", 0) for p in purchases)
    total_quantity = sum(p.get("quantity", 0) for p in purchases)
    
    # Group by supplier
    by_supplier = {}
    for p in purchases:
        sid = p.get("supplier_id")
        if sid not in by_supplier:
            by_supplier[sid] = {
                "supplier_name": p.get("supplier_name", ""),
                "total_amount": 0,
                "total_quantity": 0,
                "purchase_count": 0
            }
        by_supplier[sid]["total_amount"] += p.get("total_amount", 0)
        by_supplier[sid]["total_quantity"] += p.get("quantity", 0)
        by_supplier[sid]["purchase_count"] += 1
    
    # Group by feed type
    by_feed_type = {}
    for p in purchases:
        ftid = p.get("feed_type_id")
        if ftid not in by_feed_type:
            by_feed_type[ftid] = {
                "feed_type_name": p.get("feed_type_name", ""),
                "company_name": p.get("company_name", ""),
                "total_amount": 0,
                "total_quantity": 0,
                "purchase_count": 0
            }
        by_feed_type[ftid]["total_amount"] += p.get("total_amount", 0)
        by_feed_type[ftid]["total_quantity"] += p.get("quantity", 0)
        by_feed_type[ftid]["purchase_count"] += 1
    
    # Group by month
    by_month = {}
    for p in purchases:
        date_str = p.get("purchase_date", "")[:7]  # YYYY-MM
        if date_str not in by_month:
            by_month[date_str] = {
                "month": date_str,
                "total_amount": 0,
                "total_quantity": 0,
                "purchase_count": 0
            }
        by_month[date_str]["total_amount"] += p.get("total_amount", 0)
        by_month[date_str]["total_quantity"] += p.get("quantity", 0)
        by_month[date_str]["purchase_count"] += 1
    
    return {
        "summary": {
            "total_purchases": total_purchases,
            "total_amount": total_amount,
            "total_quantity": total_quantity,
            "average_purchase_amount": total_amount / total_purchases if total_purchases > 0 else 0
        },
        "by_supplier": list(by_supplier.values()),
        "by_feed_type": list(by_feed_type.values()),
        "by_month": sorted(list(by_month.values()), key=lambda x: x["month"], reverse=True),
        "purchases": purchases[:100]  # Return last 100 for display
    }


# ==================== FEED STOCK ALERTS (تنبيهات مخزون الأعلاف) ====================

@api_router.get("/feed-inventory/alerts")
async def get_feed_inventory_alerts(current_user: dict = Depends(get_current_user)):
    """Get feed inventory alerts for low stock items"""
    # Get all feed types with their min_stock_alert
    feed_types = await db.feed_types.find({}, {"_id": 0}).to_list(1000)
    feed_types_dict = {ft["id"]: ft for ft in feed_types}
    
    # Get current inventory grouped by feed type
    inventory = await db.feed_inventory.find({}, {"_id": 0}).to_list(10000)
    
    # Calculate current stock per feed type
    current_stock = {}
    for item in inventory:
        ft_id = item.get("purchase_id")  # Get feed type from purchase
        ft_name = item.get("product_name", "")
        if ft_name not in current_stock:
            current_stock[ft_name] = {
                "product_name": ft_name,
                "current_quantity": 0,
                "unit": item.get("unit", "kg"),
                "total_value": 0
            }
        current_stock[ft_name]["current_quantity"] += item.get("quantity", 0)
        current_stock[ft_name]["total_value"] += item.get("total_value", 0)
    
    # Check for low stock alerts
    alerts = []
    for ft in feed_types:
        ft_name = ft.get("name", "")
        min_stock = ft.get("min_stock_alert", 0)
        current = current_stock.get(ft_name, {}).get("current_quantity", 0)
        
        if min_stock > 0 and current < min_stock:
            alerts.append({
                "feed_type_id": ft.get("id"),
                "feed_type_name": ft_name,
                "company_name": ft.get("company_name", ""),
                "min_stock_alert": min_stock,
                "current_quantity": current,
                "unit": ft.get("unit", "kg"),
                "shortage": min_stock - current,
                "alert_level": "critical" if current == 0 else "warning"
            })
    
    return {
        "alerts": sorted(alerts, key=lambda x: x["shortage"], reverse=True),
        "total_alerts": len(alerts),
        "critical_count": len([a for a in alerts if a["alert_level"] == "critical"]),
        "warning_count": len([a for a in alerts if a["alert_level"] == "warning"])
    }

@api_router.put("/feed-types/{feed_type_id}/min-stock")
async def update_feed_type_min_stock(
    feed_type_id: str,
    min_stock: float,
    current_user: dict = Depends(get_current_user)
):
    """Update minimum stock alert level for a feed type"""
    result = await db.feed_types.update_one(
        {"id": feed_type_id},
        {"$set": {"min_stock_alert": min_stock}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Feed type not found")
    
    feed_type = await db.feed_types.find_one({"id": feed_type_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_feed_min_stock",
        entity_type="feed_type",
        entity_id=feed_type_id,
        entity_name=feed_type.get("name"),
        details=f"تعديل الحد الأدنى للتنبيه: {feed_type.get('name')} - {min_stock}"
    )
    
    return feed_type

# ==================== TREASURY ROUTES (الخزينة) ====================

@api_router.get("/treasury/balance")
async def get_treasury_balance(current_user: dict = Depends(get_current_user)):
    """Get current treasury balance and summary"""
    # Get or create treasury record
    treasury = await db.treasury.find_one({"type": "main"}, {"_id": 0})
    if not treasury:
        treasury = {
            "type": "main",
            "current_balance": 0.0,
            "total_deposits": 0.0,
            "total_withdrawals": 0.0,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
        await db.treasury.insert_one(treasury)
    
    return treasury

@api_router.get("/treasury/transactions")
async def get_treasury_transactions(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """Get treasury transactions with filters"""
    query = {}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if transaction_type:
        query["transaction_type"] = transaction_type
    
    transactions = await db.treasury_transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return transactions

@api_router.post("/treasury/transaction")
async def create_treasury_transaction(
    transaction_type: str,
    amount: float,
    source_type: str,
    description: str,
    source_id: Optional[str] = None,
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """Create a manual treasury transaction"""
    # Get current balance
    treasury = await db.treasury.find_one({"type": "main"}, {"_id": 0})
    current_balance = treasury.get("current_balance", 0) if treasury else 0
    
    # Calculate new balance
    if transaction_type == "deposit":
        new_balance = current_balance + amount
    else:  # withdrawal
        if amount > current_balance:
            raise HTTPException(status_code=400, detail="رصيد الخزينة غير كافٍ")
        new_balance = current_balance - amount
    
    # Create transaction record
    transaction = TreasuryTransaction(
        transaction_type=transaction_type,
        amount=amount,
        source_type=source_type,
        source_id=source_id,
        description=description,
        balance_after=new_balance,
        created_by=current_user["id"],
        created_by_name=current_user.get("full_name", "")
    )
    
    await db.treasury_transactions.insert_one(transaction.model_dump())
    
    # Update treasury balance
    update_data = {
        "current_balance": new_balance,
        "last_updated": datetime.now(timezone.utc).isoformat()
    }
    if transaction_type == "deposit":
        update_data["total_deposits"] = treasury.get("total_deposits", 0) + amount if treasury else amount
    else:
        update_data["total_withdrawals"] = treasury.get("total_withdrawals", 0) + amount if treasury else amount
    
    await db.treasury.update_one(
        {"type": "main"},
        {"$set": update_data},
        upsert=True
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action=f"treasury_{transaction_type}",
        entity_type="treasury",
        details=f"{'إيداع' if transaction_type == 'deposit' else 'سحب'}: {amount} ر.ع - {description}"
    )
    
    return transaction.model_dump()

@api_router.put("/treasury/transaction/{transaction_id}")
async def update_treasury_transaction(
    transaction_id: str,
    amount: Optional[float] = None,
    description: Optional[str] = None,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Update a treasury transaction (admin only)"""
    # Get existing transaction
    existing = await db.treasury_transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="العملية غير موجودة")
    
    # Calculate balance adjustment
    old_amount = existing.get("amount", 0)
    new_amount = amount if amount is not None else old_amount
    amount_diff = new_amount - old_amount
    
    # Update transaction
    update_data = {}
    if amount is not None:
        update_data["amount"] = amount
    if description is not None:
        update_data["description"] = description
    
    if update_data:
        await db.treasury_transactions.update_one(
            {"id": transaction_id},
            {"$set": update_data}
        )
    
    # Update treasury balance if amount changed
    if amount_diff != 0:
        treasury = await db.treasury.find_one({"type": "main"}, {"_id": 0})
        current_balance = treasury.get("current_balance", 0) if treasury else 0
        
        if existing.get("transaction_type") == "deposit":
            new_balance = current_balance + amount_diff
            new_deposits = treasury.get("total_deposits", 0) + amount_diff
            await db.treasury.update_one(
                {"type": "main"},
                {"$set": {"current_balance": new_balance, "total_deposits": new_deposits, "last_updated": datetime.now(timezone.utc).isoformat()}}
            )
        else:
            new_balance = current_balance - amount_diff
            new_withdrawals = treasury.get("total_withdrawals", 0) + amount_diff
            await db.treasury.update_one(
                {"type": "main"},
                {"$set": {"current_balance": new_balance, "total_withdrawals": new_withdrawals, "last_updated": datetime.now(timezone.utc).isoformat()}}
            )
        
        # Update balance_after for this and subsequent transactions
        await db.treasury_transactions.update_one(
            {"id": transaction_id},
            {"$set": {"balance_after": new_balance if existing.get("transaction_type") == "deposit" else current_balance - amount_diff}}
        )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_treasury_transaction",
        entity_type="treasury",
        entity_id=transaction_id,
        details=f"تعديل عملية خزينة: {new_amount} ر.ع"
    )
    
    updated = await db.treasury_transactions.find_one({"id": transaction_id}, {"_id": 0})
    return updated

@api_router.delete("/treasury/transaction/{transaction_id}")
async def delete_treasury_transaction(
    transaction_id: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Delete a treasury transaction and reverse its effect (admin only)"""
    # Get existing transaction
    existing = await db.treasury_transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="العملية غير موجودة")
    
    amount = existing.get("amount", 0)
    transaction_type = existing.get("transaction_type")
    
    # Reverse the transaction effect on treasury
    treasury = await db.treasury.find_one({"type": "main"}, {"_id": 0})
    current_balance = treasury.get("current_balance", 0) if treasury else 0
    
    if transaction_type == "deposit":
        new_balance = current_balance - amount
        new_deposits = max(0, treasury.get("total_deposits", 0) - amount)
        await db.treasury.update_one(
            {"type": "main"},
            {"$set": {"current_balance": new_balance, "total_deposits": new_deposits, "last_updated": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        new_balance = current_balance + amount
        new_withdrawals = max(0, treasury.get("total_withdrawals", 0) - amount)
        await db.treasury.update_one(
            {"type": "main"},
            {"$set": {"current_balance": new_balance, "total_withdrawals": new_withdrawals, "last_updated": datetime.now(timezone.utc).isoformat()}}
        )
    
    # Delete the transaction
    await db.treasury_transactions.delete_one({"id": transaction_id})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_treasury_transaction",
        entity_type="treasury",
        entity_id=transaction_id,
        details=f"حذف عملية خزينة: {amount} ر.ع - {existing.get('description', '')}"
    )
    
    return {"message": "تم حذف العملية وعكس تأثيرها على الخزينة", "new_balance": new_balance}

# Helper function to update treasury
async def update_treasury(transaction_type: str, amount: float, source_type: str, description: str, source_id: str = None, user_id: str = None, user_name: str = None):
    """Helper to update treasury balance from other operations"""
    treasury = await db.treasury.find_one({"type": "main"}, {"_id": 0})
    current_balance = treasury.get("current_balance", 0) if treasury else 0
    
    if transaction_type == "deposit":
        new_balance = current_balance + amount
    else:
        new_balance = current_balance - amount
    
    transaction = TreasuryTransaction(
        transaction_type=transaction_type,
        amount=amount,
        source_type=source_type,
        source_id=source_id,
        description=description,
        balance_after=new_balance,
        created_by=user_id,
        created_by_name=user_name or ""
    )
    
    await db.treasury_transactions.insert_one(transaction.model_dump())
    
    update_data = {
        "current_balance": new_balance,
        "last_updated": datetime.now(timezone.utc).isoformat()
    }
    if transaction_type == "deposit":
        update_data["total_deposits"] = treasury.get("total_deposits", 0) + amount if treasury else amount
    else:
        update_data["total_withdrawals"] = treasury.get("total_withdrawals", 0) + amount if treasury else amount
    
    await db.treasury.update_one({"type": "main"}, {"$set": update_data}, upsert=True)
    
    return new_balance

# ==================== INTEGRATED FINANCIAL REPORTS (التقارير المالية المتكاملة) ====================

@api_router.get("/reports/financial-summary")
async def get_financial_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get integrated financial summary report"""
    
    # Default to current month if no dates provided
    if not start_date:
        today = datetime.now(timezone.utc)
        start_date = datetime(today.year, today.month, 1, tzinfo=timezone.utc).isoformat()
    if not end_date:
        end_date = datetime.now(timezone.utc).isoformat()
    
    date_query = {"$gte": start_date, "$lte": end_date}
    
    # Get milk receptions (purchases)
    receptions = await db.milk_receptions.find(
        {"reception_date": date_query}, {"_id": 0}
    ).to_list(10000)
    total_milk_purchased_liters = sum(r.get("quantity_liters", 0) for r in receptions)
    total_milk_purchased_amount = sum(r.get("total_amount", 0) for r in receptions)
    
    # Get sales
    sales = await db.sales.find(
        {"sale_date": date_query}, {"_id": 0}
    ).to_list(10000)
    total_milk_sold_liters = sum(s.get("quantity_liters", 0) for s in sales)
    total_sales_amount = sum(s.get("total_amount", 0) for s in sales)
    
    # Get supplier payments (approved only)
    supplier_payments = await db.payments.find(
        {"payment_type": "supplier_payment", "status": "approved", "payment_date": date_query}, {"_id": 0}
    ).to_list(10000)
    total_supplier_payments = sum(p.get("amount", 0) for p in supplier_payments)
    
    # Get customer receipts (approved only)
    customer_receipts = await db.payments.find(
        {"payment_type": "customer_receipt", "status": "approved", "payment_date": date_query}, {"_id": 0}
    ).to_list(10000)
    total_customer_receipts = sum(p.get("amount", 0) for p in customer_receipts)
    
    # Get treasury balance
    treasury = await db.treasury.find_one({"type": "main"}, {"_id": 0})
    treasury_balance = treasury.get("current_balance", 0) if treasury else 0
    
    # Get inventory
    inventory = await db.inventory.find_one({"product_type": "raw_milk"}, {"_id": 0})
    current_stock_liters = inventory.get("quantity_liters", 0) if inventory else 0
    
    # Calculate profit/loss
    gross_profit = total_sales_amount - total_milk_purchased_amount
    net_cash_flow = total_customer_receipts - total_supplier_payments
    
    # Get outstanding balances
    suppliers = await db.suppliers.find({"is_active": True}, {"_id": 0, "balance": 1}).to_list(1000)
    total_supplier_dues = sum(s.get("balance", 0) for s in suppliers)
    
    customers = await db.customers.find({"is_active": True}, {"_id": 0, "balance": 1}).to_list(1000)
    total_customer_dues = sum(c.get("balance", 0) for c in customers)
    
    return {
        "period": {
            "start_date": start_date,
            "end_date": end_date
        },
        "purchases": {
            "total_liters": round(total_milk_purchased_liters, 2),
            "total_amount": round(total_milk_purchased_amount, 2),
            "transactions_count": len(receptions),
            "avg_price_per_liter": round(total_milk_purchased_amount / total_milk_purchased_liters, 3) if total_milk_purchased_liters > 0 else 0
        },
        "sales": {
            "total_liters": round(total_milk_sold_liters, 2),
            "total_amount": round(total_sales_amount, 2),
            "transactions_count": len(sales),
            "avg_price_per_liter": round(total_sales_amount / total_milk_sold_liters, 3) if total_milk_sold_liters > 0 else 0
        },
        "payments": {
            "supplier_payments": round(total_supplier_payments, 2),
            "customer_receipts": round(total_customer_receipts, 2),
            "net_cash_flow": round(net_cash_flow, 2)
        },
        "profit_loss": {
            "gross_profit": round(gross_profit, 2),
            "profit_margin_percentage": round((gross_profit / total_sales_amount) * 100, 2) if total_sales_amount > 0 else 0
        },
        "balances": {
            "treasury_balance": round(treasury_balance, 2),
            "supplier_dues": round(total_supplier_dues, 2),
            "customer_receivables": round(total_customer_dues, 2),
            "inventory_liters": round(current_stock_liters, 2)
        }
    }

# ==================== REPORTS & DASHBOARD ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    # Today's date range
    today = datetime.now(timezone.utc).date()
    today_start = datetime(today.year, today.month, today.day, tzinfo=timezone.utc).isoformat()
    
    # Get counts
    suppliers_count = await db.suppliers.count_documents({"is_active": True})
    customers_count = await db.customers.count_documents({"is_active": True})
    
    # Get today's receptions
    today_receptions = await db.milk_receptions.find(
        {"reception_date": {"$gte": today_start}},
        {"_id": 0}
    ).to_list(1000)
    today_milk_quantity = sum(r.get("quantity_liters", 0) for r in today_receptions)
    today_milk_value = sum(r.get("total_amount", 0) for r in today_receptions)
    
    # Get today's sales
    today_sales = await db.sales.find(
        {"sale_date": {"$gte": today_start}},
        {"_id": 0}
    ).to_list(1000)
    today_sales_quantity = sum(s.get("quantity_liters", 0) for s in today_sales)
    today_sales_value = sum(s.get("total_amount", 0) for s in today_sales)
    
    # Get inventory
    inventory = await db.inventory.find_one({"product_type": "raw_milk"}, {"_id": 0})
    current_stock = inventory.get("quantity_liters", 0) if inventory else 0
    
    # Get average quality from today's receptions
    avg_fat = 0
    avg_protein = 0
    if today_receptions:
        fats = [r.get("quality_test", {}).get("fat_percentage", 0) for r in today_receptions if r.get("quality_test")]
        proteins = [r.get("quality_test", {}).get("protein_percentage", 0) for r in today_receptions if r.get("quality_test")]
        avg_fat = sum(fats) / len(fats) if fats else 0
        avg_protein = sum(proteins) / len(proteins) if proteins else 0
    
    # Get supplier balances (amounts owed)
    suppliers = await db.suppliers.find({"is_active": True}, {"_id": 0, "balance": 1}).to_list(1000)
    total_supplier_dues = sum(s.get("balance", 0) for s in suppliers)
    
    # Get customer balances (amounts receivable)
    customers = await db.customers.find({"is_active": True}, {"_id": 0, "balance": 1}).to_list(1000)
    total_customer_dues = sum(c.get("balance", 0) for c in customers)
    
    return {
        "suppliers_count": suppliers_count,
        "customers_count": customers_count,
        "today_milk_quantity": round(today_milk_quantity, 2),
        "today_milk_value": round(today_milk_value, 2),
        "today_sales_quantity": round(today_sales_quantity, 2),
        "today_sales_value": round(today_sales_value, 2),
        "current_stock": round(current_stock, 2),
        "avg_fat_percentage": round(avg_fat, 2),
        "avg_protein_percentage": round(avg_protein, 2),
        "total_supplier_dues": round(total_supplier_dues, 2),
        "total_customer_dues": round(total_customer_dues, 2)
    }

@api_router.get("/reports/daily")
async def get_daily_report(date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    if not date:
        date = datetime.now(timezone.utc).date().isoformat()
    
    day_start = f"{date}T00:00:00"
    day_end = f"{date}T23:59:59"
    
    receptions = await db.milk_receptions.find(
        {"reception_date": {"$gte": day_start, "$lte": day_end}},
        {"_id": 0}
    ).to_list(1000)
    
    sales = await db.sales.find(
        {"sale_date": {"$gte": day_start, "$lte": day_end}},
        {"_id": 0}
    ).to_list(1000)
    
    payments = await db.payments.find(
        {"payment_date": {"$gte": day_start, "$lte": day_end}},
        {"_id": 0}
    ).to_list(1000)
    
    return {
        "date": date,
        "receptions": {
            "count": len(receptions),
            "total_quantity": sum(r.get("quantity_liters", 0) for r in receptions),
            "total_value": sum(r.get("total_amount", 0) for r in receptions),
            "details": receptions
        },
        "sales": {
            "count": len(sales),
            "total_quantity": sum(s.get("quantity_liters", 0) for s in sales),
            "total_value": sum(s.get("total_amount", 0) for s in sales),
            "details": sales
        },
        "payments": {
            "count": len(payments),
            "total_value": sum(p.get("amount", 0) for p in payments),
            "details": payments
        }
    }

@api_router.get("/reports/monthly")
async def get_monthly_report(year: int, month: int, current_user: dict = Depends(get_current_user)):
    month_start = f"{year}-{month:02d}-01T00:00:00"
    if month == 12:
        month_end = f"{year + 1}-01-01T00:00:00"
    else:
        month_end = f"{year}-{month + 1:02d}-01T00:00:00"
    
    receptions = await db.milk_receptions.find(
        {"reception_date": {"$gte": month_start, "$lt": month_end}},
        {"_id": 0}
    ).to_list(10000)
    
    sales = await db.sales.find(
        {"sale_date": {"$gte": month_start, "$lt": month_end}},
        {"_id": 0}
    ).to_list(10000)
    
    payments = await db.payments.find(
        {"payment_date": {"$gte": month_start, "$lt": month_end}},
        {"_id": 0}
    ).to_list(10000)
    
    # Group by day
    daily_data = {}
    for r in receptions:
        day = r.get("reception_date", "")[:10]
        if day not in daily_data:
            daily_data[day] = {"reception_qty": 0, "reception_value": 0, "sales_qty": 0, "sales_value": 0}
        daily_data[day]["reception_qty"] += r.get("quantity_liters", 0)
        daily_data[day]["reception_value"] += r.get("total_amount", 0)
    
    for s in sales:
        day = s.get("sale_date", "")[:10]
        if day not in daily_data:
            daily_data[day] = {"reception_qty": 0, "reception_value": 0, "sales_qty": 0, "sales_value": 0}
        daily_data[day]["sales_qty"] += s.get("quantity_liters", 0)
        daily_data[day]["sales_value"] += s.get("total_amount", 0)
    
    return {
        "year": year,
        "month": month,
        "summary": {
            "total_reception_quantity": sum(r.get("quantity_liters", 0) for r in receptions),
            "total_reception_value": sum(r.get("total_amount", 0) for r in receptions),
            "total_sales_quantity": sum(s.get("quantity_liters", 0) for s in sales),
            "total_sales_value": sum(s.get("total_amount", 0) for s in sales),
            "total_payments": sum(p.get("amount", 0) for p in payments)
        },
        "daily_data": [{"date": k, **v} for k, v in sorted(daily_data.items())]
    }

@api_router.get("/reports/supplier/{supplier_id}")
async def get_supplier_report(supplier_id: str, current_user: dict = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    receptions = await db.milk_receptions.find(
        {"supplier_id": supplier_id},
        {"_id": 0}
    ).sort("reception_date", -1).to_list(100)
    
    payments = await db.payments.find(
        {"related_id": supplier_id, "payment_type": "supplier_payment"},
        {"_id": 0}
    ).sort("payment_date", -1).to_list(100)
    
    return {
        "supplier": supplier,
        "receptions": receptions,
        "payments": payments,
        "summary": {
            "total_supplied": supplier.get("total_supplied", 0),
            "current_balance": supplier.get("balance", 0),
            "reception_count": len(receptions)
        }
    }

# ==================== HR - EMPLOYEE MANAGEMENT (إدارة الموظفين) ====================

@api_router.post("/hr/employees", response_model=Employee)
async def create_hr_employee(employee_data: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    # Generate employee code if not provided
    if not employee_data.employee_code:
        count = await db.hr_employees.count_documents({})
        employee_data.employee_code = f"EMP{count + 1:04d}"
    
    employee = Employee(**employee_data.model_dump())
    await db.hr_employees.insert_one(employee.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_employee",
        entity_type="employee",
        entity_id=employee.id,
        entity_name=employee.name,
        details=f"إضافة موظف: {employee.name} - {employee.department}"
    )
    
    return employee

@api_router.get("/hr/employees", response_model=List[Employee])
async def get_hr_employees(
    department: Optional[str] = None,
    is_active: bool = True,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_active": is_active}
    if department:
        query["department"] = department
    employees = await db.hr_employees.find(query, {"_id": 0}).to_list(1000)
    return employees

@api_router.get("/hr/employees/work-schedules")
async def get_all_employees_work_schedules(current_user: dict = Depends(get_current_user)):
    """الحصول على جداول عمل جميع الموظفين"""
    employees = await db.hr_employees.find(
        {"is_active": True},
        {"_id": 0, "id": 1, "name": 1, "employee_id": 1, "department": 1, "job_title": 1, "shift_type": 1, "weekly_off_days": 1}
    ).to_list(1000)
    
    # Add default values if not set
    for emp in employees:
        if "shift_type" not in emp:
            emp["shift_type"] = "morning"
        if "weekly_off_days" not in emp:
            emp["weekly_off_days"] = [4, 5]
    
    return employees

@api_router.get("/hr/employees/{employee_id}", response_model=Employee)
async def get_hr_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    return employee

@api_router.put("/hr/employees/{employee_id}", response_model=Employee)
async def update_hr_employee(employee_id: str, employee_data: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    # Get existing employee to preserve important status fields
    existing_employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not existing_employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check if salary changed - record in salary history
    old_salary = existing_employee.get("salary", 0)
    new_salary = employee_data.salary
    
    if old_salary != new_salary:
        # Create salary history record
        salary_history = SalaryHistory(
            employee_id=employee_id,
            employee_name=existing_employee.get("name"),
            old_salary=old_salary,
            new_salary=new_salary,
            change_reason="adjustment",  # Default reason
            effective_date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            changed_by=current_user["id"],
            changed_by_name=current_user["full_name"],
            notes=f"تم تعديل الراتب من {old_salary} إلى {new_salary}"
        )
        await db.salary_history.insert_one(salary_history.model_dump())
        
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user["full_name"],
            action="salary_change",
            entity_type="employee",
            entity_id=employee_id,
            entity_name=existing_employee.get("name"),
            details=f"تغيير راتب: {existing_employee.get('name')} من {old_salary} إلى {new_salary}"
        )
    
    # Update only the fields from employee_data, preserving is_active and can_login
    update_data = employee_data.model_dump()
    # Preserve status fields that should not be changed during regular updates
    update_data["is_active"] = existing_employee.get("is_active", True)
    update_data["can_login"] = existing_employee.get("can_login", False)
    update_data["weekly_off_days"] = existing_employee.get("weekly_off_days", [4, 5])
    
    result = await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": update_data}
    )
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_employee",
        entity_type="employee",
        entity_id=employee_id,
        entity_name=employee.get("name"),
        details=f"تعديل بيانات موظف: {employee.get('name')}"
    )
    
    return employee

@api_router.delete("/hr/employees/{employee_id}")
async def delete_hr_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    result = await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {"is_active": False}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_employee",
        entity_type="employee",
        entity_id=employee_id,
        entity_name=employee.get("name"),
        details=f"إيقاف موظف: {employee.get('name')}"
    )
    
    return {"message": "Employee deactivated successfully"}

# Create user account for employee
@api_router.post("/hr/employees/{employee_id}/create-account")
async def create_employee_account(
    employee_id: str, 
    password: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check if user already exists
    existing = await db.users.find_one({"email": employee.get("email")})
    if existing:
        raise HTTPException(status_code=400, detail="User account already exists")
    
    # Determine role based on department
    department_roles = {
        "admin": "admin",
        "it": "admin",
        "finance": "accountant",
        "purchasing": "employee",
        "milk_reception": "employee",
        "hr": "employee"
    }
    role = department_roles.get(employee.get("department", ""), "employee")
    
    # Create user
    username = employee.get("employee_code", "").lower() or employee.get("name", "").replace(" ", "").lower()
    user = User(
        username=username,
        email=employee.get("email") or f"{username}@company.com",
        full_name=employee.get("name"),
        phone=employee.get("phone"),
        role=role,
        center_id=employee.get("center_id")
    )
    user_dict = user.model_dump()
    user_dict["password"] = hash_password(password)
    user_dict["employee_id"] = employee_id
    user_dict["department"] = employee.get("department")
    user_dict["permissions"] = employee.get("permissions", [])
    
    await db.users.insert_one(user_dict)
    
    # Update employee
    await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {"can_login": True}}
    )
    
    return {"message": "User account created successfully", "username": username}

# ==================== HR - ATTENDANCE (الحضور والانصراف) ====================

# Helper function to calculate total hours and overtime
def calculate_work_hours(check_in: str, check_out: str) -> tuple:
    """Calculate total hours worked and overtime hours"""
    STANDARD_HOURS = 8
    
    if not check_in or not check_out:
        return (None, None)
    
    try:
        # Parse times
        in_time = datetime.fromisoformat(check_in.replace("Z", "+00:00"))
        out_time = datetime.fromisoformat(check_out.replace("Z", "+00:00"))
        
        # Calculate total hours
        total_hours = (out_time - in_time).total_seconds() / 3600
        
        # Calculate overtime (hours > 8)
        overtime_hours = max(0, total_hours - STANDARD_HOURS)
        
        return (round(total_hours, 2), round(overtime_hours, 2))
    except:
        return (None, None)

# Helper function to get employee's work location
async def get_employee_work_location(employee_id: str) -> str:
    """Get employee's work location from database"""
    employee = await db.hr_employees.find_one({"id": employee_id}, {"work_location": 1})
    return employee.get("work_location") if employee else None


@api_router.post("/hr/bulk-update-weekly-off")
async def bulk_set_weekly_off_days(
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """تحديث أيام الإجازة الأسبوعية لجميع الموظفين الذين لا يملكونها"""
    weekly_off_days = data.get("weekly_off_days", [4, 5])  # Default: Friday & Saturday
    
    # Update all employees who don't have weekly_off_days set
    result = await db.hr_employees.update_many(
        {"$or": [
            {"weekly_off_days": None},
            {"weekly_off_days": {"$exists": False}}
        ]},
        {"$set": {"weekly_off_days": weekly_off_days}}
    )
    
    return {
        "message": f"تم تحديث {result.modified_count} موظف",
        "modified_count": result.modified_count
    }


@api_router.post("/hr/attendance", response_model=Attendance)
async def create_attendance(attendance_data: AttendanceCreate, current_user: dict = Depends(get_current_user)):
    # Check if attendance already exists for this employee and date
    existing = await db.hr_attendance.find_one({
        "employee_id": attendance_data.employee_id,
        "date": attendance_data.date
    })
    if existing:
        # Return 409 Conflict for duplicate record
        raise HTTPException(
            status_code=409,
            detail="سجل الحضور موجود مسبقاً لهذا الموظف في هذا التاريخ"
        )
    
    # Calculate work hours and overtime
    total_hours, overtime_hours = calculate_work_hours(
        attendance_data.check_in, 
        attendance_data.check_out
    )
    
    # Get employee's work location
    work_location = attendance_data.work_location
    if not work_location:
        work_location = await get_employee_work_location(attendance_data.employee_id)
    
    attendance_dict = attendance_data.model_dump()
    attendance_dict["total_hours"] = total_hours
    attendance_dict["overtime_hours"] = overtime_hours
    attendance_dict["work_location"] = work_location
    
    attendance = Attendance(**attendance_dict)
    await db.hr_attendance.insert_one(attendance.model_dump())
    return attendance


@api_router.post("/hr/attendance/upsert")
async def upsert_attendance(attendance_data: AttendanceCreate, current_user: dict = Depends(get_current_user)):
    """Create or update attendance record - for sync purposes"""
    # Calculate work hours and overtime
    total_hours, overtime_hours = calculate_work_hours(
        attendance_data.check_in, 
        attendance_data.check_out
    )
    
    # Get employee's work location if not provided
    work_location = attendance_data.work_location
    if not work_location:
        work_location = await get_employee_work_location(attendance_data.employee_id)
    
    # Check if attendance already exists for this employee and date
    existing = await db.hr_attendance.find_one({
        "employee_id": attendance_data.employee_id,
        "date": attendance_data.date
    })
    
    attendance_dict = attendance_data.model_dump()
    attendance_dict["total_hours"] = total_hours
    attendance_dict["overtime_hours"] = overtime_hours
    attendance_dict["work_location"] = work_location
    
    if existing:
        # Update existing record
        await db.hr_attendance.update_one(
            {"id": existing["id"]},
            {"$set": attendance_dict}
        )
        attendance = await db.hr_attendance.find_one({"id": existing["id"]}, {"_id": 0})
        return {"status": "updated", "attendance": attendance}
    
    attendance = Attendance(**attendance_dict)
    await db.hr_attendance.insert_one(attendance.model_dump())
    return {"status": "created", "attendance": attendance.model_dump()}


# ==================== BULK ATTENDANCE SYNC (مزامنة مجمعة للحضور) ====================

@api_router.post("/hr/attendance/bulk-sync")
async def bulk_sync_attendance(
    records: List[dict],
    current_user: dict = Depends(get_current_user)
):
    """
    Bulk sync attendance records from desktop app.
    Accepts a list of attendance records and creates/updates them.
    """
    imported = 0
    updated = 0
    errors = []
    
    for record in records:
        try:
            # البحث عن الموظف بعدة طرق
            employee = None
            
            # 1. البحث بـ fingerprint_id
            if record.get("fingerprint_id"):
                fp_id = str(record["fingerprint_id"])
                # البحث في fingerprint_id أو fingerprint_id_2
                employee = await db.hr_employees.find_one({
                    "$or": [
                        {"fingerprint_id": fp_id},
                        {"fingerprint_id_2": fp_id}
                    ]
                }, {"_id": 0})
            
            # 2. البحث بـ employee_id
            if not employee and record.get("employee_id"):
                employee = await db.hr_employees.find_one({"id": record["employee_id"]}, {"_id": 0})
            
            # 3. البحث بالاسم (تقريبي)
            if not employee and record.get("employee_name"):
                employee = await db.hr_employees.find_one(
                    {"name": {"$regex": record["employee_name"], "$options": "i"}}, 
                    {"_id": 0}
                )
            
            # إذا لم نجد الموظف، نُنشئ سجل مع تحذير
            if not employee:
                # نُنشئ سجل حتى لو لم نجد الموظف (للمراجعة لاحقاً)
                attendance_data = {
                    "id": str(uuid.uuid4()),
                    "employee_id": str(record.get("fingerprint_id", "")),
                    "employee_name": record.get("employee_name", f"Unknown ({record.get('fingerprint_id')})"),
                    "date": record.get("date"),
                    "check_in": record.get("check_in"),
                    "check_out": record.get("check_out"),
                    "status": record.get("status", "present"),
                    "source": "fingerprint",
                    "device_ip": record.get("device_ip", ""),
                    "fingerprint_id": str(record.get("fingerprint_id", "")),
                    "needs_review": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                # التحقق من وجود سجل مسبق بنفس fingerprint_id والتاريخ
                existing = await db.hr_attendance.find_one({
                    "fingerprint_id": str(record.get("fingerprint_id", "")),
                    "date": record.get("date")
                })
                
                if existing:
                    await db.hr_attendance.update_one(
                        {"id": existing["id"]},
                        {"$set": attendance_data}
                    )
                    updated += 1
                else:
                    await db.hr_attendance.insert_one(attendance_data)
                    imported += 1
                continue
            
            # التحقق من وجود سجل مسبق
            existing = await db.hr_attendance.find_one({
                "employee_id": employee["id"],
                "date": record.get("date")
            })
            
            attendance_data = {
                "id": str(uuid.uuid4()) if not existing else existing["id"],
                "employee_id": employee["id"],
                "employee_name": employee.get("name", ""),
                "date": record.get("date"),
                "check_in": record.get("check_in"),
                "check_out": record.get("check_out"),
                "status": record.get("status", "present"),
                "source": "fingerprint",
                "device_ip": record.get("device_ip", ""),
                "fingerprint_id": str(record.get("fingerprint_id", "")),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            if existing:
                # تحديث السجل الموجود - دمج البصمات من أماكن مختلفة
                update_data = {}
                
                # إذا كانت البصمة الجديدة أبكر من check_in الموجود، نحدث check_in
                new_check_in = record.get("check_in")
                new_check_out = record.get("check_out")
                
                if new_check_in:
                    if not existing.get("check_in") or new_check_in < existing.get("check_in"):
                        update_data["check_in"] = new_check_in
                        update_data["check_in_location"] = record.get("device_ip") or record.get("location", "")
                
                # إذا كانت البصمة الجديدة أحدث من check_out الموجود، نحدث check_out
                if new_check_out:
                    if not existing.get("check_out") or new_check_out > existing.get("check_out"):
                        update_data["check_out"] = new_check_out
                        update_data["check_out_location"] = record.get("device_ip") or record.get("location", "")
                
                # إذا كان لدينا check_in جديد بدون check_out موجود
                if new_check_in and not existing.get("check_in"):
                    update_data["check_in"] = new_check_in
                    update_data["check_in_location"] = record.get("device_ip") or record.get("location", "")
                
                # إذا كان لدينا وقت واحد فقط وليس لدينا check_out
                if new_check_in and not existing.get("check_out") and existing.get("check_in"):
                    # إذا كان الوقت الجديد أحدث من check_in، نعتبره check_out
                    if new_check_in > existing.get("check_in"):
                        update_data["check_out"] = new_check_in
                        update_data["check_out_location"] = record.get("device_ip") or record.get("location", "")
                
                # تتبع جميع الأماكن التي بصم منها الموظف
                locations = existing.get("locations", [])
                new_location = record.get("device_ip") or record.get("location", "unknown")
                if new_location and new_location not in locations:
                    locations.append(new_location)
                update_data["locations"] = locations
                update_data["multi_location"] = len(locations) > 1
                
                if update_data:
                    await db.hr_attendance.update_one(
                        {"id": existing["id"]},
                        {"$set": update_data}
                    )
                updated += 1
            else:
                # إنشاء سجل جديد
                attendance_data["locations"] = [record.get("device_ip") or record.get("location", "unknown")]
                attendance_data["check_in_location"] = record.get("device_ip") or record.get("location", "")
                await db.hr_attendance.insert_one(attendance_data)
                imported += 1
                
        except Exception as e:
            errors.append(f"Error processing record: {str(e)}")
    
    return {
        "success": True,
        "imported": imported,
        "updated": updated,
        "errors": errors[:10] if errors else []  # Return first 10 errors only
    }


@api_router.post("/hr/attendance/sync")
async def sync_single_attendance(
    record: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Sync a single attendance record from desktop app.
    Creates or updates the record. Merges records from different locations.
    """
    try:
        # البحث عن الموظف
        employee = None
        if record.get("fingerprint_id"):
            fp_id = str(record["fingerprint_id"])
            # البحث في fingerprint_id أو fingerprint_id_2
            employee = await db.hr_employees.find_one({
                "$or": [
                    {"fingerprint_id": fp_id},
                    {"fingerprint_id_2": fp_id}
                ]
            }, {"_id": 0})
        if not employee and record.get("employee_id"):
            employee = await db.hr_employees.find_one({"id": record["employee_id"]}, {"_id": 0})
        
        if not employee:
            raise HTTPException(status_code=404, detail=f"Employee not found for fingerprint_id: {record.get('fingerprint_id')}")
        
        # التحقق من وجود سجل مسبق
        existing = await db.hr_attendance.find_one({
            "employee_id": employee["id"],
            "date": record.get("date")
        })
        
        new_location = record.get("device_ip") or record.get("location", "unknown")
        
        if existing:
            # دمج البصمات من أماكن مختلفة
            update_data = {}
            
            new_check_in = record.get("check_in")
            new_check_out = record.get("check_out")
            
            # تحديث check_in إذا كان أبكر
            if new_check_in:
                if not existing.get("check_in") or new_check_in < existing.get("check_in"):
                    update_data["check_in"] = new_check_in
                    update_data["check_in_location"] = new_location
                # إذا كان الوقت الجديد أحدث وليس لدينا check_out
                elif not existing.get("check_out") and new_check_in > existing.get("check_in"):
                    update_data["check_out"] = new_check_in
                    update_data["check_out_location"] = new_location
            
            # تحديث check_out إذا كان أحدث
            if new_check_out:
                if not existing.get("check_out") or new_check_out > existing.get("check_out"):
                    update_data["check_out"] = new_check_out
                    update_data["check_out_location"] = new_location
            
            # تتبع جميع الأماكن
            locations = existing.get("locations", [])
            if new_location and new_location not in locations:
                locations.append(new_location)
            update_data["locations"] = locations
            update_data["multi_location"] = len(locations) > 1
            
            if update_data:
                await db.hr_attendance.update_one(
                    {"id": existing["id"]},
                    {"$set": update_data}
                )
            
            updated_record = await db.hr_attendance.find_one({"id": existing["id"]}, {"_id": 0})
            return {"status": "updated", "attendance": updated_record, "multi_location": len(locations) > 1}
        else:
            # إنشاء سجل جديد
            attendance_data = {
                "id": str(uuid.uuid4()),
                "employee_id": employee["id"],
                "employee_name": employee.get("name", ""),
                "date": record.get("date"),
                "check_in": record.get("check_in"),
                "check_out": record.get("check_out"),
                "check_in_location": new_location,
                "status": record.get("status", "present"),
                "source": "fingerprint",
                "device_ip": record.get("device_ip", ""),
                "locations": [new_location],
                "multi_location": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.hr_attendance.insert_one(attendance_data)
            return {"status": "created", "attendance": attendance_data}
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/hr/attendance/{attendance_id}")
async def update_attendance(
    attendance_id: str,
    attendance_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update an attendance record"""
    existing = await db.hr_attendance.find_one({"id": attendance_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    
    # Update fields
    update_data = {
        "check_in": attendance_data.get("check_in", existing.get("check_in")),
        "check_out": attendance_data.get("check_out", existing.get("check_out")),
        "status": attendance_data.get("status", existing.get("status")),
        "notes": attendance_data.get("notes", existing.get("notes", "")),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.hr_attendance.update_one({"id": attendance_id}, {"$set": update_data})
    
    updated = await db.hr_attendance.find_one({"id": attendance_id}, {"_id": 0})
    return updated


@api_router.delete("/hr/attendance/{attendance_id}")
async def delete_attendance(
    attendance_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete an attendance record"""
    existing = await db.hr_attendance.find_one({"id": attendance_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    
    await db.hr_attendance.delete_one({"id": attendance_id})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_attendance",
        entity_type="attendance",
        entity_id=attendance_id,
        entity_name=existing.get("employee_name", ""),
        details=f"حذف سجل حضور: {existing.get('employee_name')} - {existing.get('date')}"
    )
    
    return {"message": "Attendance record deleted", "id": attendance_id}


@api_router.delete("/hr/attendance/employee/{employee_id}/all")
async def delete_all_employee_attendance(
    employee_id: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    """حذف جميع سجلات حضور موظف معين"""
    # Get employee info first
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # Count records before deletion
    count = await db.hr_attendance.count_documents({"employee_id": employee_id})
    
    if count == 0:
        return {"message": "لا توجد سجلات حضور لهذا الموظف", "deleted_count": 0}
    
    # Delete all attendance records
    result = await db.hr_attendance.delete_many({"employee_id": employee_id})
    
    # Log the activity
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_all_attendance",
        entity_type="attendance",
        entity_id=employee_id,
        entity_name=employee.get("name", ""),
        details=f"حذف جميع سجلات حضور: {employee.get('name')} ({result.deleted_count} سجل)"
    )
    
    return {
        "message": f"تم حذف {result.deleted_count} سجل حضور للموظف {employee.get('name')}",
        "deleted_count": result.deleted_count,
        "employee_name": employee.get("name")
    }


@api_router.put("/hr/employees/{employee_id}/clear-fingerprint")
async def clear_employee_fingerprint(
    employee_id: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    """مسح معرف البصمة للموظف"""
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    old_fingerprint = employee.get("fingerprint_id")
    
    await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {"fingerprint_id": None, "fingerprint_cleared_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Log the activity
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="clear_fingerprint",
        entity_type="employee",
        entity_id=employee_id,
        entity_name=employee.get("name", ""),
        details=f"مسح بصمة: {employee.get('name')} (البصمة القديمة: {old_fingerprint})"
    )
    
    return {
        "message": f"تم مسح بصمة الموظف {employee.get('name')}",
        "old_fingerprint_id": old_fingerprint
    }


@api_router.get("/hr/attendance")
async def get_attendance(
    employee_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    attendance = await db.hr_attendance.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Add fingerprint_id from employee data if not in record
    employee_cache = {}
    for record in attendance:
        if not record.get("fingerprint_id") and record.get("employee_id"):
            emp_id = record["employee_id"]
            if emp_id not in employee_cache:
                emp = await db.hr_employees.find_one({"id": emp_id}, {"fingerprint_id": 1})
                employee_cache[emp_id] = emp.get("fingerprint_id") if emp else None
            record["fingerprint_id"] = employee_cache[emp_id]
    
    return attendance

@api_router.get("/hr/attendance/report")
async def get_attendance_report(
    year: int,
    month: int,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    month_start = f"{year}-{month:02d}-01"
    if month == 12:
        month_end = f"{year + 1}-01-01"
    else:
        month_end = f"{year}-{month + 1:02d}-01"
    
    query = {"date": {"$gte": month_start, "$lt": month_end}}
    if employee_id:
        query["employee_id"] = employee_id
    
    attendance = await db.hr_attendance.find(query, {"_id": 0}).to_list(10000)
    
    # Group by employee
    employee_report = {}
    for record in attendance:
        emp_id = record.get("employee_id")
        if emp_id not in employee_report:
            employee_report[emp_id] = {
                "employee_name": record.get("employee_name"),
                "present_days": 0,
                "absent_days": 0,
                "late_days": 0,
                "total_hours": 0,
                "records": []
            }
        
        employee_report[emp_id]["records"].append(record)
        if record.get("check_in"):
            employee_report[emp_id]["present_days"] += 1
            # Calculate hours if both check_in and check_out exist
            if record.get("check_out"):
                try:
                    check_in = datetime.fromisoformat(record["check_in"].replace("Z", "+00:00"))
                    check_out = datetime.fromisoformat(record["check_out"].replace("Z", "+00:00"))
                    hours = (check_out - check_in).total_seconds() / 3600
                    employee_report[emp_id]["total_hours"] += hours
                except:
                    pass
    
    return {
        "year": year,
        "month": month,
        "report": list(employee_report.values())
    }

# ==================== HR - LEAVE REQUESTS (طلبات الإجازة) ====================

@api_router.post("/hr/leave-requests", response_model=LeaveRequest)
async def create_leave_request(request_data: LeaveRequestCreate, current_user: dict = Depends(get_current_user)):
    leave_request = LeaveRequest(**request_data.model_dump())
    await db.hr_leave_requests.insert_one(leave_request.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_leave_request",
        entity_type="leave_request",
        entity_id=leave_request.id,
        entity_name=request_data.employee_name,
        details=f"طلب إجازة: {request_data.employee_name} - {request_data.leave_type}"
    )
    
    return leave_request

@api_router.get("/hr/leave-requests")
async def get_leave_requests(
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if employee_id:
        query["employee_id"] = employee_id
    
    requests = await db.hr_leave_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return requests

@api_router.put("/hr/leave-requests/{request_id}/approve")
async def approve_leave_request(request_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.hr_leave_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Leave request not found")
    
    request = await db.hr_leave_requests.find_one({"id": request_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="approve_leave_request",
        entity_type="leave_request",
        entity_id=request_id,
        entity_name=request.get("employee_name"),
        details=f"الموافقة على إجازة: {request.get('employee_name')}"
    )
    
    return request

@api_router.put("/hr/leave-requests/{request_id}/reject")
async def reject_leave_request(request_id: str, reason: str = "", current_user: dict = Depends(get_current_user)):
    result = await db.hr_leave_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "approved_by": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": reason
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Leave request not found")
    
    request = await db.hr_leave_requests.find_one({"id": request_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="reject_leave_request",
        entity_type="leave_request",
        entity_id=request_id,
        entity_name=request.get("employee_name"),
        details=f"رفض إجازة: {request.get('employee_name')} - {reason}"
    )
    
    return request

# ==================== HR - FILE UPLOAD (رفع الملفات) ====================

import base64
import os

UPLOAD_DIR = "/app/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

@api_router.post("/hr/upload-file")
async def upload_hr_file(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع ملف (صورة أو PDF) للأعذار أو أي غرض آخر"""
    allowed_types = ['.pdf', '.png', '.jpg', '.jpeg', '.gif', '.webp']
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    if file_ext not in allowed_types:
        raise HTTPException(status_code=400, detail=f"نوع الملف غير مدعوم. الأنواع المدعومة: {', '.join(allowed_types)}")
    
    # Check file size (max 10MB)
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="حجم الملف يجب أن يكون أقل من 10 ميجابايت")
    
    # Generate unique filename
    import uuid
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Return URL
    file_url = f"/api/uploads/{unique_filename}"
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="upload_file",
        details=f"رفع ملف: {file.filename}"
    )
    
    return {"url": file_url, "filename": unique_filename, "original_name": file.filename}

@api_router.get("/uploads/{filename}")
async def get_uploaded_file(filename: str):
    """جلب ملف مرفوع"""
    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="الملف غير موجود")
    
    # Determine content type
    ext = os.path.splitext(filename)[1].lower()
    content_types = {
        '.pdf': 'application/pdf',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.gif': 'image/gif',
        '.webp': 'image/webp'
    }
    content_type = content_types.get(ext, 'application/octet-stream')
    
    with open(file_path, "rb") as f:
        content = f.read()
    
    return StreamingResponse(io.BytesIO(content), media_type=content_type)

# ==================== HR - OFFICIAL HOLIDAYS (العطلات الرسمية) ====================

# Models imported from models/all_models.py:
# OfficialHolidayBase, OfficialHolidayCreate, OfficialHoliday

@api_router.post("/hr/official-holidays")
async def create_official_holiday(
    holiday_data: OfficialHolidayCreate,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """إضافة عطلة رسمية"""
    holiday = OfficialHoliday(**holiday_data.model_dump(), created_by=current_user["full_name"])
    await db.hr_official_holidays.insert_one(holiday.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_official_holiday",
        entity_type="official_holiday",
        entity_id=holiday.id,
        entity_name=holiday_data.name,
        details=f"إضافة عطلة رسمية: {holiday_data.name} - {holiday_data.date}"
    )
    
    return holiday

@api_router.get("/hr/official-holidays")
async def get_official_holidays(
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة العطلات الرسمية"""
    query = {}
    if year:
        query["date"] = {"$regex": f"^{year}"}
    
    holidays = await db.hr_official_holidays.find(query, {"_id": 0}).sort("date", 1).to_list(100)
    return holidays

@api_router.delete("/hr/official-holidays/{holiday_id}")
async def delete_official_holiday(
    holiday_id: str,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """حذف عطلة رسمية"""
    result = await db.hr_official_holidays.delete_one({"id": holiday_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="العطلة غير موجودة")
    return {"message": "تم حذف العطلة بنجاح"}

@api_router.get("/hr/check-holiday/{date}")
async def check_if_holiday(date: str, current_user: dict = Depends(get_current_user)):
    """التحقق إذا كان تاريخ معين عطلة رسمية"""
    holiday = await db.hr_official_holidays.find_one({"date": date}, {"_id": 0})
    if holiday:
        return {"is_holiday": True, "holiday": holiday}
    
    # Check day of week (Friday=4, Saturday=5 in Python weekday)
    from datetime import datetime
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        day_of_week = date_obj.weekday()
        # Friday = 4, Saturday = 5
        is_weekend = day_of_week in [4, 5]
        return {"is_holiday": is_weekend, "is_weekend": is_weekend, "day_of_week": day_of_week}
    except:
        return {"is_holiday": False}

# ==================== HR - EMPLOYEE WEEKLY OFF DAYS (أيام الإجازة الأسبوعية للموظف) ====================

# Models imported from models/all_models.py:
# EmployeeWeeklyOffBase, EmployeeWeeklyOffCreate

@api_router.get("/hr/settings/holidays-config")
async def get_holidays_config(current_user: dict = Depends(get_current_user)):
    """
    الحصول على إعدادات الإجازات الموحدة
    يُرجع:
    - العطل الرسمية (من hr_official_holidays)
    - الإعدادات الافتراضية لأيام الراحة الأسبوعية
    """
    # Get all official holidays
    official_holidays = await db.hr_official_holidays.find({}, {"_id": 0}).sort("date", 1).to_list(200)
    
    # Get default weekly off days (from system settings or default)
    settings = await db.system_settings.find_one({"key": "default_weekly_off_days"}, {"_id": 0})
    default_weekly_off_days = settings.get("value", [4, 5]) if settings else [4, 5]  # Friday & Saturday
    
    return {
        "official_holidays": official_holidays,
        "default_weekly_off_days": default_weekly_off_days,
        "day_names": {
            0: "الأحد",
            1: "الإثنين", 
            2: "الثلاثاء",
            3: "الأربعاء",
            4: "الخميس",
            5: "الجمعة",
            6: "السبت"
        },
        "note": "أيام الراحة الأسبوعية يمكن تخصيصها لكل موظف من خلال hr_employees.weekly_off_days"
    }

@api_router.put("/hr/settings/default-weekly-off-days")
async def set_default_weekly_off_days(
    data: dict,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """تعيين أيام الراحة الأسبوعية الافتراضية للموظفين الجدد"""
    weekly_off_days = data.get("weekly_off_days", [4, 5])
    
    await db.system_settings.update_one(
        {"key": "default_weekly_off_days"},
        {"$set": {"key": "default_weekly_off_days", "value": weekly_off_days}},
        upsert=True
    )
    
    return {"message": "تم تحديث أيام الراحة الأسبوعية الافتراضية", "weekly_off_days": weekly_off_days}

@api_router.put("/hr/employees/{employee_id}/weekly-off")
async def set_employee_weekly_off(
    employee_id: str,
    off_data: EmployeeWeeklyOffCreate,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """تعيين أيام الإجازة الأسبوعية للموظف"""
    result = await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {"weekly_off_days": off_data.off_days}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="set_weekly_off",
        entity_type="employee",
        entity_id=employee_id,
        entity_name=off_data.employee_name,
        details=f"تحديد أيام الإجازة الأسبوعية: {off_data.off_days}"
    )
    
    return employee

@api_router.put("/hr/employees/{employee_id}/work-schedule")
async def update_employee_work_schedule(
    employee_id: str,
    shift_type: str = Query(..., description="نوع الوردية: morning, afternoon, night"),
    weekly_off_days: str = Query(..., description="أيام الإجازة (مفصولة بفاصلة) مثل: 4,5"),
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """تحديث جدول عمل الموظف (الوردية وأيام الإجازة)"""
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # Parse weekly off days
    try:
        off_days = [int(d.strip()) for d in weekly_off_days.split(",") if d.strip()]
    except:
        off_days = [4, 5]  # Default Friday, Saturday
    
    # Update employee
    result = await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {
            "shift_type": shift_type,
            "weekly_off_days": off_days
        }}
    )
    
    shift_names = {"morning": "صباحي", "afternoon": "مسائي", "night": "ليلي"}
    day_names = {0: "الأحد", 1: "الإثنين", 2: "الثلاثاء", 3: "الأربعاء", 4: "الخميس", 5: "الجمعة", 6: "السبت"}
    off_day_names = [day_names.get(d, str(d)) for d in off_days]
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_work_schedule",
        entity_type="employee",
        entity_id=employee_id,
        entity_name=employee.get("name"),
        details=f"تحديث جدول العمل: وردية {shift_names.get(shift_type, shift_type)}, إجازة: {', '.join(off_day_names)}"
    )
    
    return {
        "message": "تم تحديث جدول العمل بنجاح",
        "employee_id": employee_id,
        "shift_type": shift_type,
        "weekly_off_days": off_days
    }

@api_router.get("/hr/employees/{employee_id}/check-working-day/{date}")
async def check_employee_working_day(
    employee_id: str,
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """التحقق إذا كان يوم معين يوم عمل للموظف"""
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    from datetime import datetime
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        day_of_week = date_obj.weekday()
    except:
        raise HTTPException(status_code=400, detail="تنسيق التاريخ غير صحيح")
    
    # Get employee's weekly off days
    weekly_off_days = employee.get("weekly_off_days", [4, 5])  # Default: Friday & Saturday
    
    # Check if it's an official holiday
    holiday = await db.hr_official_holidays.find_one({"date": date}, {"_id": 0})
    
    is_off_day = day_of_week in weekly_off_days
    is_official_holiday = holiday is not None
    
    return {
        "date": date,
        "day_of_week": day_of_week,
        "is_weekly_off": is_off_day,
        "is_official_holiday": is_official_holiday,
        "is_working_day": not is_off_day and not is_official_holiday,
        "holiday_name": holiday.get("name") if holiday else None,
        "employee_weekly_off_days": weekly_off_days
    }

# ==================== HR - EXCUSE REQUESTS (طلبات الأعذار) ====================

@api_router.post("/hr/excuse-requests", response_model=ExcuseRequest)
async def create_excuse_request(request_data: ExcuseRequestCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء طلب عذر جديد"""
    excuse_request = ExcuseRequest(**request_data.model_dump())
    await db.hr_excuse_requests.insert_one(excuse_request.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_excuse_request",
        entity_type="excuse_request",
        entity_id=excuse_request.id,
        entity_name=request_data.employee_name,
        details=f"طلب عذر: {request_data.employee_name} - {request_data.excuse_date} - {request_data.excuse_type}"
    )
    
    return excuse_request

@api_router.get("/hr/excuse-requests")
async def get_excuse_requests(
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة طلبات الأعذار"""
    query = {}
    if status:
        query["status"] = status
    if employee_id:
        query["employee_id"] = employee_id
    
    requests = await db.hr_excuse_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return requests

@api_router.get("/hr/excuse-requests/{request_id}")
async def get_excuse_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """جلب تفاصيل طلب عذر محدد"""
    request = await db.hr_excuse_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="Excuse request not found")
    return request

@api_router.put("/hr/excuse-requests/{request_id}/approve")
async def approve_excuse_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """الموافقة على طلب العذر وتسجيل حضور تلقائي"""
    # Get the excuse request
    excuse_request = await db.hr_excuse_requests.find_one({"id": request_id}, {"_id": 0})
    if not excuse_request:
        raise HTTPException(status_code=404, detail="Excuse request not found")
    
    if excuse_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    # Update excuse request status
    await db.hr_excuse_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_by_name": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "attendance_updated": True
        }}
    )
    
    # Create or update attendance record as "excused" presence
    excuse_date = excuse_request.get("excuse_date")
    employee_id = excuse_request.get("employee_id")
    employee_name = excuse_request.get("employee_name")
    
    # Check if attendance exists for this date
    existing_attendance = await db.hr_attendance.find_one({
        "employee_id": employee_id,
        "date": excuse_date
    })
    
    if existing_attendance:
        # Update existing attendance to mark as excused
        await db.hr_attendance.update_one(
            {"id": existing_attendance["id"]},
            {"$set": {
                "source": "excuse_approved",
                "check_in": excuse_request.get("start_time") or "08:00",
                "check_out": excuse_request.get("end_time") or "16:00",
                "notes": f"عذر معتمد: {excuse_request.get('excuse_type')} - {excuse_request.get('reason')}"
            }}
        )
    else:
        # Create new attendance record
        attendance = Attendance(
            employee_id=employee_id,
            employee_name=employee_name,
            date=excuse_date,
            check_in=excuse_request.get("start_time") or "08:00",
            check_out=excuse_request.get("end_time") or "16:00",
            source="excuse_approved",
            total_hours=8.0
        )
        await db.hr_attendance.insert_one(attendance.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="approve_excuse_request",
        entity_type="excuse_request",
        entity_id=request_id,
        entity_name=employee_name,
        details=f"موافقة على عذر: {employee_name} - {excuse_date}"
    )
    
    updated_request = await db.hr_excuse_requests.find_one({"id": request_id}, {"_id": 0})
    return updated_request

@api_router.put("/hr/excuse-requests/{request_id}/reject")
async def reject_excuse_request(
    request_id: str, 
    reason: str = "",
    current_user: dict = Depends(get_current_user)
):
    """رفض طلب العذر وتسجيل غياب"""
    excuse_request = await db.hr_excuse_requests.find_one({"id": request_id}, {"_id": 0})
    if not excuse_request:
        raise HTTPException(status_code=404, detail="Excuse request not found")
    
    if excuse_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    # Update excuse request status
    await db.hr_excuse_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "approved_by": current_user["id"],
            "approved_by_name": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": reason,
            "attendance_updated": True
        }}
    )
    
    # Ensure attendance is marked as absent
    excuse_date = excuse_request.get("excuse_date")
    employee_id = excuse_request.get("employee_id")
    employee_name = excuse_request.get("employee_name")
    
    existing_attendance = await db.hr_attendance.find_one({
        "employee_id": employee_id,
        "date": excuse_date
    })
    
    if not existing_attendance:
        # Create attendance record marked as absent
        attendance = Attendance(
            employee_id=employee_id,
            employee_name=employee_name,
            date=excuse_date,
            check_in=None,
            check_out=None,
            source="excuse_rejected",
            total_hours=0
        )
        await db.hr_attendance.insert_one(attendance.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="reject_excuse_request",
        entity_type="excuse_request",
        entity_id=request_id,
        entity_name=employee_name,
        details=f"رفض عذر: {employee_name} - {excuse_date} - {reason}"
    )
    
    updated_request = await db.hr_excuse_requests.find_one({"id": request_id}, {"_id": 0})
    return updated_request

@api_router.delete("/hr/excuse-requests/{request_id}")
async def delete_excuse_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """حذف طلب عذر"""
    excuse_request = await db.hr_excuse_requests.find_one({"id": request_id}, {"_id": 0})
    if not excuse_request:
        raise HTTPException(status_code=404, detail="Excuse request not found")
    
    await db.hr_excuse_requests.delete_one({"id": request_id})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_excuse_request",
        entity_type="excuse_request",
        entity_id=request_id,
        entity_name=excuse_request.get("employee_name"),
        details=f"حذف طلب عذر: {excuse_request.get('employee_name')}"
    )
    
    return {"message": "Excuse request deleted successfully"}

# ==================== HR - SALARY HISTORY (سجل تغييرات الرواتب) ====================

@api_router.get("/hr/salary-history")
async def get_all_salary_history(
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجل تغييرات الرواتب"""
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    
    history = await db.salary_history.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return history

@api_router.get("/hr/salary-history/{employee_id}")
async def get_employee_salary_history(
    employee_id: str,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجل تغييرات راتب موظف معين"""
    history = await db.salary_history.find(
        {"employee_id": employee_id}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return history

@api_router.post("/hr/salary-history")
async def create_salary_history_record(
    history_data: SalaryHistoryCreate,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """إضافة سجل تغيير راتب يدوياً"""
    salary_history = SalaryHistory(
        **history_data.model_dump(),
        changed_by=current_user["id"],
        changed_by_name=current_user["full_name"]
    )
    await db.salary_history.insert_one(salary_history.model_dump())
    
    # Update employee's current salary if this is the latest record
    await db.hr_employees.update_one(
        {"id": history_data.employee_id},
        {"$set": {"salary": history_data.new_salary}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_salary_history",
        entity_type="salary_history",
        entity_id=salary_history.id,
        entity_name=history_data.employee_name,
        details=f"تسجيل تغيير راتب: {history_data.employee_name} من {history_data.old_salary} إلى {history_data.new_salary}"
    )
    
    return salary_history

@api_router.put("/hr/employees/{employee_id}/salary")
async def update_employee_salary(
    employee_id: str,
    new_salary: float,
    change_reason: str = "adjustment",
    effective_date: Optional[str] = None,
    notes: Optional[str] = None,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """تحديث راتب موظف مع التسجيل في السجل"""
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    old_salary = employee.get("salary", 0)
    
    if old_salary == new_salary:
        raise HTTPException(status_code=400, detail="الراتب الجديد مطابق للراتب الحالي")
    
    # Create salary history record
    salary_history = SalaryHistory(
        employee_id=employee_id,
        employee_name=employee.get("name"),
        old_salary=old_salary,
        new_salary=new_salary,
        change_reason=change_reason,
        effective_date=effective_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        changed_by=current_user["id"],
        changed_by_name=current_user["full_name"],
        notes=notes
    )
    await db.salary_history.insert_one(salary_history.model_dump())
    
    # Update employee salary
    await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {"salary": new_salary}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_salary",
        entity_type="employee",
        entity_id=employee_id,
        entity_name=employee.get("name"),
        details=f"تحديث راتب: {employee.get('name')} من {old_salary} إلى {new_salary} - السبب: {change_reason}"
    )
    
    return {
        "message": "تم تحديث الراتب بنجاح",
        "employee_name": employee.get("name"),
        "old_salary": old_salary,
        "new_salary": new_salary,
        "change_reason": change_reason,
        "effective_date": salary_history.effective_date
    }

# ==================== HR - EXPENSE REQUESTS (طلبات المصاريف) ====================

@api_router.post("/hr/expense-requests", response_model=ExpenseRequest)
async def create_expense_request(request_data: ExpenseRequestCreate, current_user: dict = Depends(get_current_user)):
    expense_request = ExpenseRequest(**request_data.model_dump())
    await db.hr_expense_requests.insert_one(expense_request.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_expense_request",
        entity_type="expense_request",
        entity_id=expense_request.id,
        entity_name=request_data.employee_name,
        details=f"طلب مصاريف: {request_data.employee_name} - {request_data.amount} ر.ع"
    )
    
    return expense_request

@api_router.get("/hr/expense-requests")
async def get_expense_requests(
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if employee_id:
        query["employee_id"] = employee_id
    
    requests = await db.hr_expense_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return requests

@api_router.put("/hr/expense-requests/{request_id}/approve")
async def approve_expense_request(request_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.hr_expense_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Expense request not found")
    
    request = await db.hr_expense_requests.find_one({"id": request_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="approve_expense_request",
        entity_type="expense_request",
        entity_id=request_id,
        entity_name=request.get("employee_name"),
        details=f"الموافقة على مصاريف: {request.get('employee_name')} - {request.get('amount')} ر.ع"
    )
    
    return request

@api_router.put("/hr/expense-requests/{request_id}/reject")
async def reject_expense_request(request_id: str, reason: str = "", current_user: dict = Depends(get_current_user)):
    result = await db.hr_expense_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "approved_by": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": reason
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Expense request not found")
    
    request = await db.hr_expense_requests.find_one({"id": request_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="reject_expense_request",
        entity_type="expense_request",
        entity_id=request_id,
        entity_name=request.get("employee_name"),
        details=f"رفض مصاريف: {request.get('employee_name')} - {reason}"
    )
    
    return request

@api_router.put("/hr/expense-requests/{request_id}/pay")
async def mark_expense_paid(request_id: str, current_user: dict = Depends(require_role(["admin", "accountant"]))):
    result = await db.hr_expense_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "paid",
            "paid_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Expense request not found")
    
    request = await db.hr_expense_requests.find_one({"id": request_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="pay_expense_request",
        entity_type="expense_request",
        entity_id=request_id,
        entity_name=request.get("employee_name"),
        details=f"صرف مصاريف: {request.get('employee_name')} - {request.get('amount')} ر.ع"
    )
    
    return request

# ==================== HR - CAR CONTRACTS (عقود السيارات) ====================

@api_router.post("/hr/car-contracts", response_model=CarContract)
async def create_car_contract(contract_data: CarContractCreate, current_user: dict = Depends(get_current_user)):
    contract = CarContract(**contract_data.model_dump())
    await db.hr_car_contracts.insert_one(contract.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_car_contract",
        entity_type="car_contract",
        entity_id=contract.id,
        entity_name=contract_data.employee_name,
        details=f"عقد سيارة جديد: {contract_data.employee_name} - {contract_data.car_type}"
    )
    
    return contract

@api_router.get("/hr/car-contracts")
async def get_car_contracts(
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if employee_id:
        query["employee_id"] = employee_id
    
    contracts = await db.hr_car_contracts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return contracts

@api_router.put("/hr/car-contracts/{contract_id}", response_model=CarContract)
async def update_car_contract(contract_id: str, contract_data: CarContractCreate, current_user: dict = Depends(get_current_user)):
    result = await db.hr_car_contracts.update_one(
        {"id": contract_id},
        {"$set": contract_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Car contract not found")
    
    contract = await db.hr_car_contracts.find_one({"id": contract_id}, {"_id": 0})
    return contract

@api_router.delete("/hr/car-contracts/{contract_id}")
async def delete_car_contract(contract_id: str, current_user: dict = Depends(get_current_user)):
    contract = await db.hr_car_contracts.find_one({"id": contract_id}, {"_id": 0})
    if not contract:
        raise HTTPException(status_code=404, detail="Car contract not found")
    
    result = await db.hr_car_contracts.update_one(
        {"id": contract_id},
        {"$set": {"status": "cancelled"}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="cancel_car_contract",
        entity_type="car_contract",
        entity_id=contract_id,
        entity_name=contract.get("employee_name"),
        details=f"إلغاء عقد سيارة: {contract.get('employee_name')}"
    )
    
    return {"message": "Car contract cancelled"}

# ==================== HR - OFFICIAL LETTERS (الرسائل الرسمية) ====================

@api_router.post("/hr/official-letters", response_model=OfficialLetter)
async def create_official_letter(letter_data: OfficialLetterCreate, current_user: dict = Depends(get_current_user)):
    # Generate letter number
    count = await db.hr_official_letters.count_documents({})
    year = datetime.now().year
    letter_number = f"LTR-{year}-{count + 1:04d}"
    
    letter = OfficialLetter(**letter_data.model_dump())
    letter_dict = letter.model_dump()
    letter_dict["letter_number"] = letter_number
    
    await db.hr_official_letters.insert_one(letter_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_official_letter",
        entity_type="official_letter",
        entity_id=letter.id,
        entity_name=letter_data.employee_name,
        details=f"رسالة رسمية: {letter_data.letter_type} - {letter_data.employee_name}"
    )
    
    return OfficialLetter(**letter_dict)

@api_router.get("/hr/official-letters")
async def get_official_letters(
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    letter_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if employee_id:
        query["employee_id"] = employee_id
    if letter_type:
        query["letter_type"] = letter_type
    
    letters = await db.hr_official_letters.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return letters

@api_router.put("/hr/official-letters/{letter_id}/issue")
async def issue_official_letter(letter_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.hr_official_letters.update_one(
        {"id": letter_id},
        {"$set": {
            "status": "issued",
            "issued_by": current_user["full_name"],
            "issued_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Official letter not found")
    
    letter = await db.hr_official_letters.find_one({"id": letter_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="issue_official_letter",
        entity_type="official_letter",
        entity_id=letter_id,
        entity_name=letter.get("employee_name"),
        details=f"إصدار رسالة رسمية: {letter.get('letter_number')} - {letter.get('employee_name')}"
    )
    
    return letter

# Approve official letter (electronic signature by HR manager)
@api_router.post("/hr/official-letters/{letter_id}/approve")
async def approve_official_letter(letter_id: str, current_user: dict = Depends(require_role(["admin", "hr_manager"]))):
    letter = await db.hr_official_letters.find_one({"id": letter_id}, {"_id": 0})
    if not letter:
        raise HTTPException(status_code=404, detail="Letter not found")
    
    if letter.get("is_approved"):
        raise HTTPException(status_code=400, detail="Letter already approved")
    
    # Generate electronic signature code
    import hashlib
    signature_data = f"{letter_id}-{current_user['id']}-{datetime.now().isoformat()}"
    signature_code = hashlib.sha256(signature_data.encode()).hexdigest()[:16].upper()
    
    await db.hr_official_letters.update_one(
        {"id": letter_id},
        {"$set": {
            "status": "approved",
            "is_approved": True,
            "approved_by": current_user["id"],
            "approved_by_name": current_user.get("full_name", ""),
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "signature_code": signature_code
        }}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="approve_official_letter",
        entity_type="official_letter",
        entity_id=letter_id,
        entity_name=letter.get("employee_name"),
        details=f"تصديق رسالة رسمية: {letter.get('letter_number')} - كود التصديق: {signature_code}"
    )
    
    return {"message": "تم تصديق الرسالة بنجاح", "signature_code": signature_code}

# Reject official letter
@api_router.post("/hr/official-letters/{letter_id}/reject")
async def reject_official_letter(letter_id: str, reason: str = "", current_user: dict = Depends(require_role(["admin", "hr_manager"]))):
    letter = await db.hr_official_letters.find_one({"id": letter_id}, {"_id": 0})
    if not letter:
        raise HTTPException(status_code=404, detail="Letter not found")
    
    await db.hr_official_letters.update_one(
        {"id": letter_id},
        {"$set": {
            "status": "rejected",
            "rejection_reason": reason,
            "approved_by": current_user["id"],
            "approved_by_name": current_user.get("full_name", ""),
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="reject_official_letter",
        entity_type="official_letter",
        entity_id=letter_id,
        entity_name=letter.get("employee_name"),
        details=f"رفض رسالة رسمية: {letter.get('letter_number')} - السبب: {reason}"
    )
    
    return {"message": "تم رفض الرسالة"}

# Mark letter as printed
@api_router.post("/hr/official-letters/{letter_id}/print")
async def mark_letter_printed(letter_id: str, current_user: dict = Depends(get_current_user)):
    letter = await db.hr_official_letters.find_one({"id": letter_id}, {"_id": 0})
    if not letter:
        raise HTTPException(status_code=404, detail="Letter not found")
    
    if not letter.get("is_approved"):
        raise HTTPException(status_code=400, detail="يجب تصديق الرسالة قبل الطباعة")
    
    await db.hr_official_letters.update_one(
        {"id": letter_id},
        {"$set": {
            "is_printed": True,
            "printed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "تم تسجيل الطباعة"}

# Get employee's own letters
@api_router.get("/hr/my-letters")
async def get_my_letters(current_user: dict = Depends(get_current_user)):
    # Find employee by username or user_id
    employee = await db.hr_employees.find_one({
        "$or": [
            {"user_id": current_user["id"]},
            {"username": current_user["username"]}
        ]
    }, {"_id": 0})
    
    if not employee:
        return []
    
    letters = await db.hr_official_letters.find(
        {"employee_id": employee["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return letters

# ==================== HR - FINGERPRINT DEVICES (أجهزة البصمة) ====================

@api_router.post("/hr/fingerprint-devices", response_model=FingerprintDevice)
async def create_fingerprint_device(device_data: FingerprintDeviceCreate, current_user: dict = Depends(require_role(["admin"]))):
    device = FingerprintDevice(**device_data.model_dump())
    await db.hr_fingerprint_devices.insert_one(device.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_fingerprint_device",
        entity_type="fingerprint_device",
        entity_id=device.id,
        entity_name=device_data.name,
        details=f"إضافة جهاز بصمة: {device_data.name} - {device_data.ip_address}"
    )
    
    return device

@api_router.get("/hr/fingerprint-devices")
async def get_fingerprint_devices(current_user: dict = Depends(get_current_user)):
    devices = await db.hr_fingerprint_devices.find({"is_active": True}, {"_id": 0}).to_list(100)
    return devices

@api_router.put("/hr/fingerprint-devices/{device_id}", response_model=FingerprintDevice)
async def update_fingerprint_device(device_id: str, device_data: FingerprintDeviceCreate, current_user: dict = Depends(require_role(["admin"]))):
    result = await db.hr_fingerprint_devices.update_one(
        {"id": device_id},
        {"$set": device_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Device not found")
    device = await db.hr_fingerprint_devices.find_one({"id": device_id}, {"_id": 0})
    return device

@api_router.delete("/hr/fingerprint-devices/{device_id}")
async def delete_fingerprint_device(device_id: str, current_user: dict = Depends(require_role(["admin"]))):
    device = await db.hr_fingerprint_devices.find_one({"id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    result = await db.hr_fingerprint_devices.update_one(
        {"id": device_id},
        {"$set": {"is_active": False}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_fingerprint_device",
        entity_type="fingerprint_device",
        entity_id=device_id,
        entity_name=device.get("name"),
        details=f"حذف جهاز بصمة: {device.get('name')}"
    )
    
    return {"message": "Device deleted successfully"}

@api_router.post("/hr/fingerprint-devices/{device_id}/sync")
async def sync_fingerprint_device(device_id: str, current_user: dict = Depends(require_role(["admin"]))):
    """Sync attendance data from Hikvision fingerprint device"""
    import aiohttp
    
    device = await db.hr_fingerprint_devices.find_one({"id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    try:
        # Hikvision API integration
        device_url = f"http://{device['ip_address']}/csl/login"
        
        async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=10)) as session:
            # Try to connect to ZKTeco device
            try:
                async with session.get(f"http://{device['ip_address']}/", timeout=5) as response:
                    # Device is reachable
                    pass
            except Exception as conn_error:
                raise HTTPException(
                    status_code=500, 
                    detail=f"لا يمكن الاتصال بجهاز البصمة ({device['ip_address']}). تأكد من:\n1. أن الجهاز متصل بالشبكة\n2. أن عنوان IP صحيح\n3. أن النظام على نفس الشبكة المحلية"
                )
            
            # Login to device
            login_data = {
                "id": device.get("login_id"),
                "password": device.get("password")
            }
            
            try:
                async with session.post(device_url, data=login_data, timeout=10) as response:
                    if response.status != 200:
                        raise HTTPException(status_code=500, detail="فشل تسجيل الدخول للجهاز. تحقق من بيانات الدخول.")
                    
                    # Note: ZKTeco API integration requires specific SDK
                    # This is a placeholder for actual implementation
            except aiohttp.ClientError as e:
                raise HTTPException(status_code=500, detail=f"خطأ في الاتصال: {str(e)}")
        
        # Update last sync time
        await db.hr_fingerprint_devices.update_one(
            {"id": device_id},
            {"$set": {"last_sync": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {
            "message": "تم الاتصال بالجهاز بنجاح. ملاحظة: تحتاج إلى تثبيت ZKTeco SDK لسحب البيانات تلقائياً.", 
            "device": device["name"],
            "note": "يمكنك استخدام خيار 'إضافة حضور' لإدخال البيانات يدوياً"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Fingerprint sync error: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"فشل المزامنة: الجهاز غير متصل أو خارج الشبكة المحلية ({device['ip_address']})"
        )

# Manual attendance import endpoint
@api_router.post("/hr/attendance/import")
async def import_attendance(
    records: List[AttendanceCreate],
    current_user: dict = Depends(require_role(["admin"]))
):
    """Import attendance records manually or from device export"""
    imported = 0
    for record in records:
        existing = await db.hr_attendance.find_one({
            "employee_id": record.employee_id,
            "date": record.date
        })
        if existing:
            await db.hr_attendance.update_one(
                {"id": existing["id"]},
                {"$set": record.model_dump()}
            )
        else:
            attendance = Attendance(**record.model_dump())
            await db.hr_attendance.insert_one(attendance.model_dump())
        imported += 1
    
    return {"message": f"Imported {imported} attendance records"}

# Import attendance from Excel file
@api_router.post("/hr/attendance/import-excel")
async def import_attendance_from_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_role(["admin"]))
):
    """Import attendance records from Excel file"""
    import pandas as pd
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يجب أن يكون الملف بصيغة Excel (.xlsx أو .xls)")
    
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content))
        
        # Map column names (support both Arabic and English)
        column_mapping = {
            'التاريخ': 'date',
            'Date': 'date',
            'اسم الموظف': 'employee_name',
            'Employee Name': 'employee_name',
            'رقم الموظف': 'employee_id',
            'Employee ID': 'employee_id',
            'وقت الحضور': 'check_in',
            'Check In': 'check_in',
            'وقت الانصراف': 'check_out',
            'Check Out': 'check_out',
        }
        
        df = df.rename(columns=column_mapping)
        
        # Check required columns
        required_cols = ['date', 'employee_name']
        for col in required_cols:
            if col not in df.columns:
                raise HTTPException(status_code=400, detail=f"عمود مطلوب غير موجود: {col}")
        
        imported = 0
        updated = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Find employee by name if no ID provided
                employee_id = row.get('employee_id', '')
                employee_name = str(row.get('employee_name', ''))
                
                if not employee_id:
                    employee = await db.hr_employees.find_one({"name": employee_name}, {"_id": 0})
                    if employee:
                        employee_id = employee['id']
                    else:
                        employee_id = employee_name  # Use name as fallback
                
                # Parse date
                date_val = row.get('date')
                if pd.isna(date_val):
                    continue
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)[:10]
                
                # Parse times
                check_in = row.get('check_in', '')
                check_out = row.get('check_out', '')
                
                if pd.isna(check_in):
                    check_in = None
                elif hasattr(check_in, 'strftime'):
                    check_in = check_in.strftime('%H:%M')
                else:
                    check_in = str(check_in)[:5] if check_in else None
                
                if pd.isna(check_out):
                    check_out = None
                elif hasattr(check_out, 'strftime'):
                    check_out = check_out.strftime('%H:%M')
                else:
                    check_out = str(check_out)[:5] if check_out else None
                
                # Check if record exists
                existing = await db.hr_attendance.find_one({
                    "employee_id": employee_id,
                    "date": date_str
                })
                
                if existing:
                    # Update existing record
                    update_data = {"source": "excel_import"}
                    if check_in:
                        update_data["check_in"] = check_in
                    if check_out:
                        update_data["check_out"] = check_out
                    
                    await db.hr_attendance.update_one(
                        {"id": existing["id"]},
                        {"$set": update_data}
                    )
                    updated += 1
                else:
                    # Create new record
                    attendance = Attendance(
                        employee_id=employee_id,
                        employee_name=employee_name,
                        date=date_str,
                        check_in=check_in,
                        check_out=check_out,
                        source="excel_import"
                    )
                    await db.hr_attendance.insert_one(attendance.model_dump())
                    imported += 1
                    
            except Exception as e:
                errors.append(f"خطأ في الصف {idx + 2}: {str(e)}")
        
        # Log activity
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user["full_name"],
            action="import_attendance_excel",
            entity_type="attendance",
            details=f"استيراد {imported} سجل جديد و تحديث {updated} سجل من ملف Excel"
        )
        
        return {
            "message": f"تم استيراد {imported} سجل جديد وتحديث {updated} سجل",
            "imported": imported,
            "updated": updated,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        logging.error(f"Error importing Excel: {e}")
        raise HTTPException(status_code=500, detail=f"خطأ في معالجة الملف: {str(e)}")

# Import attendance from ZKTeco MDB file
@api_router.post("/hr/attendance/import-zkteco")
async def import_attendance_from_zkteco(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_role(["admin"]))
):
    """Import attendance records from ZKTeco MDB database file"""
    import tempfile
    import os
    import subprocess
    import csv
    from io import StringIO
    
    if not file.filename.endswith('.mdb'):
        raise HTTPException(status_code=400, detail="يجب أن يكون الملف بصيغة MDB من جهاز ZKTeco")
    
    try:
        # Save uploaded file temporarily
        content = await file.read()
        with tempfile.NamedTemporaryFile(suffix='.mdb', delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        
        # Helper function to read table using mdb-export CLI
        def read_mdb_table(mdb_path, table_name):
            try:
                result = subprocess.run(
                    ['mdb-export', mdb_path, table_name],
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                if result.returncode != 0:
                    return []
                
                # Parse CSV output
                reader = csv.DictReader(StringIO(result.stdout))
                return list(reader)
            except Exception as e:
                logging.warning(f"Could not read table {table_name}: {e}")
                return []
        
        # Extract users from MDB using mdb-export CLI
        user_map = {}
        try:
            userinfo_rows = read_mdb_table(tmp_path, 'USERINFO')
            for row in userinfo_rows:
                user_id = str(row.get('USERID', ''))
                name = row.get('Name', '') or row.get('Badgenumber', '') or ''
                badge = str(row.get('Badgenumber', ''))
                if user_id:
                    user_map[user_id] = {'name': name, 'badge': badge}
        except Exception as e:
            logging.warning(f"Could not read USERINFO table: {e}")
        
        # Extract attendance records from MDB
        checkinout_rows = read_mdb_table(tmp_path, 'CHECKINOUT')
        if not checkinout_rows:
            os.unlink(tmp_path)
            raise HTTPException(status_code=500, detail="فشل في قراءة جدول CHECKINOUT أو الجدول فارغ")
        
        # Group records by user and date
        attendance_by_day = {}
        for row in checkinout_rows:
            user_id = str(row.get('USERID', ''))
            check_time = row.get('CHECKTIME', None)
            
            if not user_id or not check_time:
                continue
            
            try:
                # Parse check_time string from CSV
                from datetime import datetime as dt
                check_time_dt = None
                
                # Try different formats
                for fmt in ["%m/%d/%y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
                    try:
                        check_time_dt = dt.strptime(check_time, fmt)
                        break
                    except:
                        continue
                
                if not check_time_dt:
                    continue
                    
                date_str = check_time_dt.strftime("%Y-%m-%d")
                time_str = check_time_dt.strftime("%H:%M")
                
                key = f"{user_id}_{date_str}"
                if key not in attendance_by_day:
                    user_info = user_map.get(user_id, {'name': f'User_{user_id}', 'badge': user_id})
                    attendance_by_day[key] = {
                        'user_id': user_id,
                        'employee_name': user_info['name'],
                        'employee_badge': user_info['badge'],
                        'date': date_str,
                        'times': []
                    }
                attendance_by_day[key]['times'].append(time_str)
            except Exception as e:
                continue
        
        # Process and save attendance records
        imported = 0
        updated = 0
        
        for key, record in attendance_by_day.items():
            times = sorted(record['times'])
            check_in = times[0] if times else None
            check_out = times[-1] if len(times) > 1 else None
            
            # Find employee by badge or name
            employee = await db.hr_employees.find_one(
                {"$or": [
                    {"employee_id": record['employee_badge']},
                    {"name": record['employee_name']}
                ]},
                {"_id": 0}
            )
            
            employee_id = employee['id'] if employee else record['employee_badge']
            employee_name = employee['name'] if employee else record['employee_name']
            
            # Check if record exists
            existing = await db.hr_attendance.find_one({
                "employee_id": employee_id,
                "date": record['date']
            })
            
            if existing:
                # Update if new times are different
                update_data = {"source": "zkteco_import"}
                if check_in and (not existing.get('check_in') or check_in < existing.get('check_in', '23:59')):
                    update_data["check_in"] = check_in
                if check_out and (not existing.get('check_out') or check_out > existing.get('check_out', '00:00')):
                    update_data["check_out"] = check_out
                
                if len(update_data) > 1:
                    await db.hr_attendance.update_one(
                        {"id": existing["id"]},
                        {"$set": update_data}
                    )
                    updated += 1
            else:
                # Create new record
                attendance = Attendance(
                    employee_id=employee_id,
                    employee_name=employee_name,
                    date=record['date'],
                    check_in=check_in,
                    check_out=check_out,
                    source="zkteco_import"
                )
                await db.hr_attendance.insert_one(attendance.model_dump())
                imported += 1
        
        # Cleanup temp file
        import os
        os.unlink(tmp_path)
        
        # Log activity
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user["full_name"],
            action="import_attendance_zkteco",
            entity_type="attendance",
            details=f"استيراد {imported} سجل جديد و تحديث {updated} سجل من ملف ZKTeco"
        )
        
        return {
            "message": f"تم استيراد {imported} سجل جديد وتحديث {updated} سجل من جهاز البصمة",
            "imported": imported,
            "updated": updated,
            "total_users": len(user_map),
            "total_days": len(attendance_by_day)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error importing ZKTeco MDB: {e}")
        raise HTTPException(status_code=500, detail=f"خطأ في معالجة الملف: {str(e)}")

# Export attendance to Excel
@api_router.get("/hr/attendance/export/excel")
async def export_attendance_excel(
    year: Optional[int] = None,
    month: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export attendance report to Excel with date range support"""
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment
    
    # Handle date range - prefer start_date/end_date over year/month
    if start_date and end_date:
        date_from = start_date
        date_to = end_date
        filename_suffix = f"{start_date}_to_{end_date}"
        sheet_name = f"الحضور {start_date} - {end_date}"
    elif year and month:
        date_from = f"{year}-{month:02d}-01"
        if month == 12:
            date_to = f"{year + 1}-01-01"
        else:
            date_to = f"{year}-{month + 1:02d}-01"
        filename_suffix = f"{year}_{month}"
        sheet_name = f"الحضور {month}-{year}"
    else:
        # Default to current month
        from datetime import datetime
        now = datetime.now()
        date_from = f"{now.year}-{now.month:02d}-01"
        date_to = now.strftime("%Y-%m-%d")
        filename_suffix = f"{now.year}_{now.month}"
        sheet_name = f"الحضور {now.month}-{now.year}"
    
    query = {"date": {"$gte": date_from, "$lte": date_to}}
    if employee_id:
        query["employee_id"] = employee_id
    
    attendance = await db.hr_attendance.find(query, {"_id": 0}).sort([("employee_name", 1), ("date", 1)]).to_list(10000)
    
    # Get all employees for name/code lookup
    employees_list = await db.hr_employees.find({}, {"_id": 0, "id": 1, "name": 1, "employee_code": 1, "fingerprint_id": 1}).to_list(1000)
    emp_lookup = {}
    for emp in employees_list:
        emp_lookup[emp.get('id')] = emp
        emp_lookup[emp.get('fingerprint_id')] = emp
        emp_lookup[emp.get('employee_code')] = emp
        emp_lookup[emp.get('name')] = emp
    
    if not attendance:
        # Return empty template
        df = pd.DataFrame(columns=['Employee Name', 'Employee Code', 'Fingerprint ID', 'Date', 'Check In', 'Check Out', 'Source'])
    else:
        # Group attendance by employee and add summary - REMOVE DUPLICATE DATES
        from collections import defaultdict
        employee_data = defaultdict(lambda: {"records": {}, "name": "", "code": "", "fingerprint": ""})
        
        for record in attendance:
            emp_id = record.get('employee_id', '')
            emp_name = record.get('employee_name', 'Unknown')
            
            # Try to find employee info
            emp_info = emp_lookup.get(emp_id) or emp_lookup.get(emp_name)
            if emp_info:
                key = emp_info.get('id', emp_id)
                employee_data[key]["name"] = emp_info.get('name', emp_name)
                employee_data[key]["code"] = emp_info.get('employee_code', '')
                employee_data[key]["fingerprint"] = emp_info.get('fingerprint_id', '')
            else:
                key = emp_id or emp_name
                employee_data[key]["name"] = emp_name
            
            # Use date as key to avoid duplicates - keep the first record for each date
            record_date = record.get('date', '')
            if record_date and record_date not in employee_data[key]["records"]:
                employee_data[key]["records"][record_date] = record
        
        # Create structured data with employee headers and summaries
        rows = []
        for emp_key in sorted(employee_data.keys(), key=lambda x: employee_data[x]["name"]):
            emp_info = employee_data[emp_key]
            records = list(emp_info["records"].values())  # Convert dict to list
            days_count = len(records)  # Each date appears only once now
            
            emp_name = emp_info["name"] or "Unknown"
            emp_code = emp_info["code"] or "-"
            emp_fp = emp_info["fingerprint"] or "-"
            
            # Add employee header row
            rows.append({
                'Employee Name': f"📋 {emp_name}",
                'Employee Code': emp_code,
                'Fingerprint ID': emp_fp,
                'Date': f"Attendance Days: {days_count}",
                'Check In': '',
                'Check Out': '',
                'Source': ''
            })
            
            # Add attendance records
            for record in sorted(records, key=lambda x: x.get('date', '')):
                rows.append({
                    'Employee Name': emp_name,
                    'Employee Code': emp_code,
                    'Fingerprint ID': emp_fp,
                    'Date': record.get('date', ''),
                    'Check In': record.get('check_in', '-'),
                    'Check Out': record.get('check_out', '-'),
                    'Source': record.get('source', 'manual')
                })
            
            # Add empty row after each employee
            rows.append({
                'Employee Name': '',
                'Employee Code': '',
                'Fingerprint ID': '',
                'Date': '',
                'Check In': '',
                'Check Out': '',
                'Source': ''
            })
        
        df = pd.DataFrame(rows)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Excel sheet name max 31 chars
        
        worksheet = writer.sheets[sheet_name[:31]]
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        employee_header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        employee_header_font = Font(bold=True, color='FFFFFF')
        
        # Style header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Style employee header rows (rows starting with 📋)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if row[0].value and str(row[0].value).startswith('📋'):
                for cell in row:
                    cell.fill = employee_header_fill
                    cell.font = employee_header_font
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 5, 50)
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_{filename_suffix}.xlsx"}
    )

# Export attendance to PDF
@api_router.get("/hr/attendance/export/pdf")
async def export_attendance_pdf(
    year: Optional[int] = None,
    month: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export attendance report to PDF with date range support and Arabic font"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import inch, cm
    import arabic_reshaper
    from bidi.algorithm import get_display
    import os
    
    # Register Arabic font
    font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'NotoSansArabic.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('Arabic', font_path))
        arabic_font = 'Arabic'
    else:
        arabic_font = 'Helvetica'
    
    def reshape_arabic(text):
        """Reshape Arabic text for proper display"""
        if not text:
            return text
        try:
            reshaped = arabic_reshaper.reshape(str(text))
            return get_display(reshaped)
        except:
            return text
    
    # Handle date range - prefer start_date/end_date over year/month
    if start_date and end_date:
        date_from = start_date
        date_to = end_date
        filename_suffix = f"{start_date}_to_{end_date}"
        period_text = f"Period: {start_date} to {end_date}"
    elif year and month:
        date_from = f"{year}-{month:02d}-01"
        if month == 12:
            date_to = f"{year + 1}-01-01"
        else:
            date_to = f"{year}-{month + 1:02d}-01"
        filename_suffix = f"{year}_{month}"
        period_text = f"Month: {month}/{year}"
    else:
        from datetime import datetime
        now = datetime.now()
        date_from = f"{now.year}-{now.month:02d}-01"
        date_to = now.strftime("%Y-%m-%d")
        filename_suffix = f"{now.year}_{now.month}"
        period_text = f"Month: {now.month}/{now.year}"
    
    query = {"date": {"$gte": date_from, "$lte": date_to}}
    if employee_id:
        query["employee_id"] = employee_id
    
    attendance = await db.hr_attendance.find(query, {"_id": 0}).sort([("employee_name", 1), ("date", 1)]).to_list(10000)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Create header table with logo on RIGHT side and company name on LEFT
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.png')
    
    # Company name styles
    company_name_ar = reshape_arabic("المروج للألبان")
    company_name_en = "Almorooj Dairy"
    
    # Header with logo on right, text on left
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=1.2*inch, height=1.2*inch)
        
        # Create header table: [Company Text | Logo]
        header_data = [[
            Paragraph(f'''<para align="left">
                <font name="{arabic_font}" size="16" color="#1a5f2a">{company_name_ar}</font><br/>
                <font size="12" color="#1a5f2a">{company_name_en}</font><br/><br/>
                <font name="{arabic_font}" size="14">{reshape_arabic("تقرير الحضور والانصراف")}</font><br/>
                <font size="10">Attendance Report</font><br/>
                <font size="9">{period_text}</font>
            </para>''', styles['Normal']),
            logo
        ]]
        
        header_table = Table(header_data, colWidths=[500, 100])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(header_table)
    else:
        # No logo - just text
        company_style = ParagraphStyle('Company', fontName=arabic_font, fontSize=18, alignment=TA_CENTER, textColor=colors.HexColor('#1a5f2a'))
        elements.append(Paragraph(company_name_ar, company_style))
        elements.append(Paragraph(company_name_en, ParagraphStyle('CompanyEn', fontSize=14, alignment=TA_CENTER, textColor=colors.HexColor('#1a5f2a'))))
        elements.append(Spacer(1, 15))
        title_style = ParagraphStyle('Title', fontName=arabic_font, fontSize=16, alignment=TA_CENTER)
        elements.append(Paragraph(reshape_arabic("تقرير الحضور والانصراف"), title_style))
        elements.append(Paragraph("Attendance Report", ParagraphStyle('TitleEn', fontSize=12, alignment=TA_CENTER)))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(period_text, ParagraphStyle('Date', alignment=TA_CENTER)))
    
    elements.append(Spacer(1, 20))
    
    # Get all employees for name/code lookup
    employees_list = await db.hr_employees.find({}, {"_id": 0, "id": 1, "name": 1, "employee_code": 1, "fingerprint_id": 1}).to_list(1000)
    emp_lookup = {}
    for emp in employees_list:
        emp_lookup[emp.get('id')] = emp
        emp_lookup[emp.get('fingerprint_id')] = emp
        emp_lookup[emp.get('employee_code')] = emp
    
    # Group by employee - REMOVE DUPLICATE DATES
    from collections import defaultdict
    employee_data = defaultdict(lambda: {"records": {}, "name": "", "code": "", "fingerprint": ""})
    
    for record in attendance:
        emp_id = record.get('employee_id', '')
        emp_name = record.get('employee_name', 'Unknown')
        
        # Try to find employee info
        emp_info = emp_lookup.get(emp_id) or emp_lookup.get(emp_name)
        if emp_info:
            key = emp_info.get('id', emp_id)
            employee_data[key]["name"] = emp_info.get('name', emp_name)
            employee_data[key]["code"] = emp_info.get('employee_code', '')
            employee_data[key]["fingerprint"] = emp_info.get('fingerprint_id', '')
        else:
            key = emp_id or emp_name
            employee_data[key]["name"] = emp_name
        
        # Use date as key to avoid duplicates
        record_date = record.get('date', '')
        if record_date and record_date not in employee_data[key]["records"]:
            employee_data[key]["records"][record_date] = record
    
    # Create tables for each employee
    emp_count = 0
    total_records = 0
    for emp_key in sorted(employee_data.keys(), key=lambda x: employee_data[x]["name"]):
        emp_count += 1
        emp_info = employee_data[emp_key]
        records = list(emp_info["records"].values())  # Convert dict to list
        total_records += len(records)
        days_count = len(records)
        
        # Employee header box with full details
        emp_name = emp_info["name"] or "Unknown"
        emp_code = emp_info["code"] or "-"
        emp_fp = emp_info["fingerprint"] or "-"
        
        # Create employee info table with name displayed properly
        name_para = Paragraph(f'<b>{emp_name}</b>', ParagraphStyle('EmpName', fontSize=10, alignment=TA_CENTER))
        emp_header_data = [
            ['Employee Name', 'Code', 'Fingerprint ID', 'Days'],
            [name_para, emp_code, emp_fp, str(days_count)]
        ]
        
        emp_table = Table(emp_header_data, colWidths=[250, 80, 100, 70])
        emp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#E8F5E9')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, 1), 10),
            ('FONTNAME', (0, 1), (0, 1), arabic_font),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, 1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E7D32')),
        ]))
        elements.append(emp_table)
        elements.append(Spacer(1, 5))
        
        # Attendance records table
        headers = ['#', 'Date', 'Check In', 'Check Out', 'Source']
        data = [headers]
        
        for idx, record in enumerate(sorted(records, key=lambda x: x.get('date', '')), 1):
            data.append([
                str(idx),
                record.get('date', ''),
                record.get('check_in', '-'),
                record.get('check_out', '-'),
                record.get('source', 'manual')
            ])
        
        table = Table(data, repeatRows=1, colWidths=[30, 80, 80, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    # Summary footer
    summary_style = ParagraphStyle('Summary', fontSize=10, alignment=TA_CENTER)
    elements.append(Paragraph(f"Total Employees: {emp_count} | Total Records: {total_records}", summary_style))
    
    if not employee_data:
        elements.append(Paragraph("No attendance records", ParagraphStyle('NoData', alignment=TA_CENTER)))
    
    doc.build(elements)
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=attendance_{filename_suffix}.pdf"}
    )

# ==================== HR - SHIFTS (الورديات) ====================

@api_router.get("/hr/shifts")
async def get_shifts(current_user: dict = Depends(get_current_user)):
    """Get all shifts"""
    shifts = await db.hr_shifts.find({"is_active": True}, {"_id": 0}).to_list(100)
    return shifts

@api_router.post("/hr/shifts", response_model=Shift)
async def create_shift(shift_data: ShiftCreate, current_user: dict = Depends(require_role(["admin", "hr"]))):
    """Create a new shift"""
    shift = Shift(**shift_data.model_dump())
    await db.hr_shifts.insert_one(shift.model_dump())
    return shift

@api_router.put("/hr/shifts/{shift_id}", response_model=Shift)
async def update_shift(shift_id: str, shift_data: ShiftCreate, current_user: dict = Depends(require_role(["admin", "hr"]))):
    """Update a shift"""
    await db.hr_shifts.update_one(
        {"id": shift_id},
        {"$set": shift_data.model_dump()}
    )
    shift = await db.hr_shifts.find_one({"id": shift_id}, {"_id": 0})
    return shift

@api_router.delete("/hr/shifts/{shift_id}")
async def delete_shift(shift_id: str, current_user: dict = Depends(require_role(["admin", "hr"]))):
    """Delete (deactivate) a shift"""
    await db.hr_shifts.update_one(
        {"id": shift_id},
        {"$set": {"is_active": False}}
    )
    return {"message": "تم حذف الوردية بنجاح"}

# Employee Shift Assignments
@api_router.get("/hr/employee-shifts")
async def get_employee_shifts(
    employee_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get employee shift assignments"""
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    
    shifts = await db.hr_employee_shifts.find(query, {"_id": 0}).to_list(1000)
    return shifts

@api_router.post("/hr/employee-shifts", response_model=EmployeeShift)
async def assign_employee_shift(
    shift_data: EmployeeShiftCreate,
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Assign shift to employee"""
    assignment = EmployeeShift(**shift_data.model_dump(), created_by=current_user["full_name"])
    await db.hr_employee_shifts.insert_one(assignment.model_dump())
    return assignment

@api_router.post("/hr/employee-shifts/bulk")
async def bulk_assign_shifts(
    assignments: List[EmployeeShiftCreate],
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Bulk assign shifts to multiple employees"""
    created = []
    for shift_data in assignments:
        assignment = EmployeeShift(**shift_data.model_dump(), created_by=current_user["full_name"])
        await db.hr_employee_shifts.insert_one(assignment.model_dump())
        created.append(assignment.model_dump())
    return {"message": f"تم تعيين {len(created)} وردية بنجاح", "assignments": created}

@api_router.delete("/hr/employee-shifts/{assignment_id}")
async def delete_employee_shift(assignment_id: str, current_user: dict = Depends(require_role(["admin", "hr"]))):
    """Delete shift assignment"""
    await db.hr_employee_shifts.delete_one({"id": assignment_id})
    return {"message": "تم حذف التعيين بنجاح"}

# ==================== HR - OVERTIME (العمل الإضافي) ====================

@api_router.get("/hr/overtime")
async def get_overtime(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get overtime records"""
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if status:
        query["status"] = status
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    overtime = await db.hr_overtime.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return overtime

@api_router.post("/hr/overtime", response_model=Overtime)
async def create_overtime(
    overtime_data: OvertimeCreate,
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Create overtime record"""
    # Calculate total amount if hourly_rate provided
    data = overtime_data.model_dump()
    if data.get("hourly_rate") and data.get("hours"):
        data["total_amount"] = data["hourly_rate"] * data["hours"] * data.get("rate", 1.5)
    
    overtime = Overtime(**data)
    await db.hr_overtime.insert_one(overtime.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_overtime",
        entity_type="overtime",
        details=f"تسجيل {overtime.hours} ساعات عمل إضافي للموظف {overtime.employee_name}"
    )
    
    return overtime

@api_router.put("/hr/overtime/{overtime_id}/approve")
async def approve_overtime(
    overtime_id: str,
    approved: bool = True,
    rejection_reason: Optional[str] = None,
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Approve or reject overtime"""
    update_data = {
        "status": "approved" if approved else "rejected",
        "approved_by": current_user["full_name"],
        "approved_at": datetime.now(timezone.utc).isoformat()
    }
    if not approved and rejection_reason:
        update_data["rejection_reason"] = rejection_reason
    
    await db.hr_overtime.update_one(
        {"id": overtime_id},
        {"$set": update_data}
    )
    
    overtime = await db.hr_overtime.find_one({"id": overtime_id}, {"_id": 0})
    return overtime

@api_router.delete("/hr/overtime/{overtime_id}")
async def delete_overtime(overtime_id: str, current_user: dict = Depends(require_role(["admin", "hr"]))):
    """Delete overtime record"""
    await db.hr_overtime.delete_one({"id": overtime_id})
    return {"message": "تم حذف سجل العمل الإضافي"}

# Get overtime summary for payroll
@api_router.get("/hr/overtime/summary/{employee_id}")
async def get_overtime_summary(
    employee_id: str,
    start_date: str,
    end_date: str,
    current_user: dict = Depends(get_current_user)
):
    """Get overtime summary for an employee in a date range"""
    overtime_records = await db.hr_overtime.find({
        "employee_id": employee_id,
        "status": "approved",
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(100)
    
    total_hours = sum(r.get("hours", 0) for r in overtime_records)
    total_amount = sum(r.get("total_amount", 0) for r in overtime_records)
    
    return {
        "employee_id": employee_id,
        "period": f"{start_date} - {end_date}",
        "total_hours": total_hours,
        "total_amount": total_amount,
        "records_count": len(overtime_records),
        "records": overtime_records
    }

# ==================== HR - LOANS & ADVANCES (السلف والقروض) ====================

@api_router.get("/hr/loans")
async def get_loans(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    loan_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get loans and advances"""
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if status:
        query["status"] = status
    if loan_type:
        query["loan_type"] = loan_type
    
    loans = await db.hr_loans.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return loans

@api_router.post("/hr/loans", response_model=Loan)
async def create_loan(
    loan_data: LoanCreate,
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Create loan or advance request"""
    data = loan_data.model_dump()
    
    # Calculate installment amount if not provided
    if data.get("installments") and data.get("amount"):
        data["installment_amount"] = data["amount"] / data["installments"]
    data["remaining_amount"] = data["amount"]
    
    loan = Loan(**data)
    await db.hr_loans.insert_one(loan.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_loan",
        entity_type="loan",
        details=f"إنشاء طلب {'سلفة' if loan.loan_type == 'advance' else 'قرض'} بمبلغ {loan.amount} للموظف {loan.employee_name}"
    )
    
    return loan

@api_router.put("/hr/loans/{loan_id}/approve")
async def approve_loan(
    loan_id: str,
    approved: bool = True,
    rejection_reason: Optional[str] = None,
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Approve or reject loan"""
    update_data = {
        "status": "approved" if approved else "rejected",
        "approved_by": current_user["full_name"],
        "approved_at": datetime.now(timezone.utc).isoformat()
    }
    if approved:
        update_data["status"] = "active"  # Active means approved and deduction can start
    if not approved and rejection_reason:
        update_data["rejection_reason"] = rejection_reason
    
    await db.hr_loans.update_one(
        {"id": loan_id},
        {"$set": update_data}
    )
    
    loan = await db.hr_loans.find_one({"id": loan_id}, {"_id": 0})
    return loan

@api_router.post("/hr/loans/{loan_id}/payment")
async def record_loan_payment(
    loan_id: str,
    amount: float,
    payment_method: str = "salary_deduction",
    notes: Optional[str] = None,
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Record a loan payment"""
    loan = await db.hr_loans.find_one({"id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="القرض غير موجود")
    
    # Create payment record
    payment = LoanPayment(
        loan_id=loan_id,
        employee_id=loan["employee_id"],
        amount=amount,
        payment_date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        payment_method=payment_method,
        notes=notes
    )
    await db.hr_loan_payments.insert_one(payment.model_dump())
    
    # Update loan
    new_paid = loan.get("paid_amount", 0) + amount
    new_remaining = loan["amount"] - new_paid
    new_installments = loan.get("paid_installments", 0) + 1
    
    update_data = {
        "paid_amount": new_paid,
        "remaining_amount": new_remaining,
        "paid_installments": new_installments
    }
    
    # Check if loan is fully paid
    if new_remaining <= 0:
        update_data["status"] = "completed"
        update_data["remaining_amount"] = 0
    
    await db.hr_loans.update_one({"id": loan_id}, {"$set": update_data})
    
    return {"message": "تم تسجيل الدفعة بنجاح", "payment": payment.model_dump()}

@api_router.get("/hr/loans/{loan_id}/payments")
async def get_loan_payments(loan_id: str, current_user: dict = Depends(get_current_user)):
    """Get payment history for a loan"""
    payments = await db.hr_loan_payments.find({"loan_id": loan_id}, {"_id": 0}).to_list(100)
    return payments

@api_router.get("/hr/loans/employee/{employee_id}/active")
async def get_employee_active_loans(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get active loans for an employee (for payroll deduction)"""
    loans = await db.hr_loans.find({
        "employee_id": employee_id,
        "status": "active"
    }, {"_id": 0}).to_list(100)
    
    total_deduction = sum(l.get("installment_amount", 0) for l in loans)
    
    return {
        "employee_id": employee_id,
        "active_loans": loans,
        "monthly_deduction": total_deduction
    }

# ==================== HR - EMPLOYEE DOCUMENTS (وثائق الموظفين) ====================

@api_router.get("/hr/documents")
async def get_employee_documents(
    employee_id: Optional[str] = None,
    document_type: Optional[str] = None,
    expiring_soon: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Get employee documents"""
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if document_type:
        query["document_type"] = document_type
    
    documents = await db.hr_documents.find(query, {"_id": 0}).to_list(1000)
    
    # Update expiry status
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for doc in documents:
        if doc.get("expiry_date"):
            expiry = doc["expiry_date"]
            doc["is_expired"] = expiry < today
            # Calculate days to expiry
            from datetime import datetime as dt
            try:
                expiry_dt = dt.strptime(expiry, "%Y-%m-%d")
                today_dt = dt.strptime(today, "%Y-%m-%d")
                doc["days_to_expiry"] = (expiry_dt - today_dt).days
            except:
                doc["days_to_expiry"] = None
    
    if expiring_soon:
        # Filter documents expiring in next 30 days
        documents = [d for d in documents if d.get("days_to_expiry") is not None and 0 < d["days_to_expiry"] <= 30]
    
    return documents

@api_router.post("/hr/documents", response_model=EmployeeDocument)
async def create_employee_document(
    document_data: EmployeeDocumentCreate,
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Create employee document record"""
    document = EmployeeDocument(**document_data.model_dump(), uploaded_by=current_user["full_name"])
    await db.hr_documents.insert_one(document.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_document",
        entity_type="document",
        details=f"إضافة وثيقة {document.document_name} للموظف {document.employee_name}"
    )
    
    return document

@api_router.put("/hr/documents/{document_id}", response_model=EmployeeDocument)
async def update_employee_document(
    document_id: str,
    document_data: EmployeeDocumentCreate,
    current_user: dict = Depends(require_role(["admin", "hr"]))
):
    """Update employee document"""
    await db.hr_documents.update_one(
        {"id": document_id},
        {"$set": document_data.model_dump()}
    )
    document = await db.hr_documents.find_one({"id": document_id}, {"_id": 0})
    return document

@api_router.delete("/hr/documents/{document_id}")
async def delete_employee_document(document_id: str, current_user: dict = Depends(require_role(["admin", "hr"]))):
    """Delete employee document"""
    await db.hr_documents.delete_one({"id": document_id})
    return {"message": "تم حذف الوثيقة بنجاح"}

@api_router.get("/hr/documents/expiring")
async def get_expiring_documents(days: int = 30, current_user: dict = Depends(get_current_user)):
    """Get documents expiring within specified days"""
    documents = await db.hr_documents.find({}, {"_id": 0}).to_list(5000)
    
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    from datetime import datetime as dt, timedelta
    future_date = (dt.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    
    expiring = []
    for doc in documents:
        if doc.get("expiry_date"):
            expiry = doc["expiry_date"]
            if today <= expiry <= future_date:
                try:
                    expiry_dt = dt.strptime(expiry, "%Y-%m-%d")
                    today_dt = dt.strptime(today, "%Y-%m-%d")
                    doc["days_to_expiry"] = (expiry_dt - today_dt).days
                    expiring.append(doc)
                except:
                    pass
    
    # Sort by days to expiry
    expiring.sort(key=lambda x: x.get("days_to_expiry", 999))
    
    return expiring

# ==================== HR - DEPARTMENTS & PERMISSIONS ====================

DEPARTMENTS = [
    {"id": "admin", "name": "الإدارة", "name_en": "Administration"},
    {"id": "it", "name": "تقنية المعلومات", "name_en": "IT"},
    {"id": "hr", "name": "الموارد البشرية", "name_en": "Human Resources"},
    {"id": "finance", "name": "المالية", "name_en": "Finance"},
    {"id": "purchasing", "name": "المشتريات", "name_en": "Purchasing"},
    {"id": "milk_reception", "name": "استلام الحليب", "name_en": "Milk Reception"},
    {"id": "sales", "name": "المبيعات", "name_en": "Sales"},
    {"id": "inventory", "name": "المخازن", "name_en": "Inventory"},
    {"id": "legal", "name": "القسم القانوني", "name_en": "Legal"},
    {"id": "projects", "name": "المشاريع", "name_en": "Projects"},
    {"id": "operations", "name": "العمليات", "name_en": "Operations"},
    {"id": "marketing", "name": "التسويق", "name_en": "Marketing"},
]

PERMISSIONS = {
    "admin": ["all"],
    "it": ["all"],
    "hr": ["hr", "employees", "attendance", "leave", "expense", "car_contracts", "letters"],
    "finance": ["finance", "payments", "reports", "expense"],
    "purchasing": ["suppliers", "feed_purchases", "inventory"],
    "milk_reception": ["milk_reception", "suppliers", "quality"],
    "sales": ["sales", "customers", "inventory"],
    "inventory": ["inventory", "reports"],
    "legal": ["legal", "contracts", "cases", "consultations", "documents"],
    "projects": ["projects", "tasks", "milestones", "team_members"],
    "operations": ["operations", "equipment", "maintenance", "incidents", "vehicles"],
    "marketing": ["marketing", "campaigns", "leads", "offers", "returns", "social"]
}

# قائمة الصلاحيات المتاحة
AVAILABLE_PERMISSIONS = [
    {"id": "dashboard", "name": "لوحة التحكم", "name_en": "Dashboard"},
    {"id": "suppliers", "name": "الموردين", "name_en": "Suppliers"},
    {"id": "milk_reception", "name": "استلام الحليب", "name_en": "Milk Reception"},
    {"id": "customers", "name": "العملاء", "name_en": "Customers"},
    {"id": "sales", "name": "المبيعات", "name_en": "Sales"},
    {"id": "feed_purchases", "name": "مشتريات الأعلاف", "name_en": "Feed Purchases"},
    {"id": "inventory", "name": "المخزون", "name_en": "Inventory"},
    {"id": "finance", "name": "المالية", "name_en": "Finance"},
    {"id": "hr", "name": "الموارد البشرية", "name_en": "Human Resources"},
    {"id": "employees", "name": "الموظفين", "name_en": "Employees"},
    {"id": "reports", "name": "التقارير", "name_en": "Reports"},
    {"id": "settings", "name": "الإعدادات", "name_en": "Settings"},
    {"id": "attendance", "name": "الحضور والانصراف", "name_en": "Attendance"},
    {"id": "leave", "name": "الإجازات", "name_en": "Leave Requests"},
    {"id": "expense", "name": "المصاريف", "name_en": "Expenses"},
    {"id": "car_contracts", "name": "عقود السيارات", "name_en": "Car Contracts"},
    {"id": "letters", "name": "الرسائل الرسمية", "name_en": "Official Letters"},
    {"id": "quality", "name": "فحص الجودة", "name_en": "Quality Testing"},
    {"id": "payments", "name": "المدفوعات", "name_en": "Payments"},
    # Legal permissions
    {"id": "legal", "name": "القسم القانوني", "name_en": "Legal"},
    {"id": "contracts", "name": "العقود القانونية", "name_en": "Legal Contracts"},
    {"id": "cases", "name": "القضايا", "name_en": "Legal Cases"},
    {"id": "consultations", "name": "الاستشارات القانونية", "name_en": "Legal Consultations"},
    {"id": "documents", "name": "المستندات القانونية", "name_en": "Legal Documents"},
    # Projects permissions
    {"id": "projects", "name": "المشاريع", "name_en": "Projects"},
    {"id": "tasks", "name": "مهام المشاريع", "name_en": "Project Tasks"},
    {"id": "milestones", "name": "المراحل", "name_en": "Milestones"},
    {"id": "team_members", "name": "أعضاء الفريق", "name_en": "Team Members"},
    # Operations permissions
    {"id": "operations", "name": "العمليات", "name_en": "Operations"},
    {"id": "equipment", "name": "المعدات", "name_en": "Equipment"},
    {"id": "maintenance", "name": "الصيانة", "name_en": "Maintenance"},
    {"id": "incidents", "name": "الحوادث", "name_en": "Incidents"},
    {"id": "vehicles", "name": "المركبات", "name_en": "Vehicles"},
    # Marketing permissions
    {"id": "marketing", "name": "التسويق", "name_en": "Marketing"},
    {"id": "campaigns", "name": "الحملات التسويقية", "name_en": "Marketing Campaigns"},
    {"id": "leads", "name": "العملاء المحتملين", "name_en": "Leads"},
    {"id": "offers", "name": "العروض", "name_en": "Sales Offers"},
    {"id": "returns", "name": "المرتجعات", "name_en": "Returns"},
    {"id": "social", "name": "وسائل التواصل", "name_en": "Social Media"},
    {"id": "all", "name": "جميع الصلاحيات", "name_en": "All Permissions"},
]

@api_router.get("/hr/departments")
async def get_departments():
    return DEPARTMENTS

@api_router.get("/hr/available-permissions")
async def get_available_permissions():
    """Get list of all available permissions"""
    return AVAILABLE_PERMISSIONS

@api_router.get("/hr/permissions/{department}")
async def get_department_permissions(department: str):
    return {"department": department, "permissions": PERMISSIONS.get(department, [])}

@api_router.get("/hr/managers")
async def get_managers(current_user: dict = Depends(get_current_user)):
    """Get list of employees who can be managers (department heads)"""
    managers = await db.hr_employees.find(
        {"is_active": True, "position": {"$regex": "مدير|مسؤول|رئيس|Manager|Head|Supervisor", "$options": "i"}},
        {"_id": 0}
    ).to_list(100)
    
    # Also include employees from admin/it/hr departments
    dept_heads = await db.hr_employees.find(
        {"is_active": True, "department": {"$in": ["admin", "it", "hr"]}},
        {"_id": 0}
    ).to_list(100)
    
    # Merge and deduplicate
    all_managers = {m["id"]: m for m in managers}
    for m in dept_heads:
        all_managers[m["id"]] = m
    
    return list(all_managers.values())

# ==================== HR - DASHBOARD ====================

@api_router.get("/hr/dashboard")
async def get_hr_dashboard(current_user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Count employees
    total_employees = await db.hr_employees.count_documents({"is_active": True})
    
    # Today's attendance
    today_attendance = await db.hr_attendance.count_documents({"date": today})
    
    # Today's absent (total employees - today attendance)
    today_absent = max(0, total_employees - today_attendance)
    
    # Pending leave requests
    pending_leaves = await db.hr_leave_requests.count_documents({"status": "pending"})
    
    # Pending expense requests
    pending_expenses = await db.hr_expense_requests.count_documents({"status": "pending"})
    
    # Active car contracts
    active_contracts = await db.hr_car_contracts.count_documents({"status": "active"})
    
    # Active warnings
    active_warnings = await db.hr_warnings.count_documents({"status": "active"})
    
    # Recent activities
    recent_leaves = await db.hr_leave_requests.find(
        {}, {"_id": 0}
    ).sort("created_at", -1).to_list(5)
    
    recent_expenses = await db.hr_expense_requests.find(
        {}, {"_id": 0}
    ).sort("created_at", -1).to_list(5)
    
    return {
        "total_employees": total_employees,
        "today_attendance": today_attendance,
        "today_absent": today_absent,
        "pending_leaves": pending_leaves,
        "pending_expenses": pending_expenses,
        "active_car_contracts": active_contracts,
        "active_warnings": active_warnings,
        "recent_leave_requests": recent_leaves,
        "recent_expense_requests": recent_expenses
    }

# ==================== AUTO JOURNAL ENTRY HELPER (إنشاء قيود محاسبية آلية) ====================

async def create_auto_journal_entry(
    description: str,
    lines: list,  # [{account_number, debit, credit}]
    reference_type: str,  # milk_purchase, milk_sale, feed_purchase, payment
    reference_id: str,
    created_by_id: str,
    created_by_name: str
):
    """
    إنشاء قيد يومية آلي للعمليات المالية
    Auto-create journal entry for financial transactions
    """
    try:
        # Get account IDs from account numbers
        entry_lines = []
        total_debit = 0
        total_credit = 0
        
        for line in lines:
            account = await db.chart_of_accounts.find_one(
                {"account_number": line["account_number"], "is_active": True},
                {"_id": 0}
            )
            if not account:
                logging.warning(f"Account not found: {line['account_number']}")
                continue
            
            debit = line.get("debit", 0)
            credit = line.get("credit", 0)
            total_debit += debit
            total_credit += credit
            
            entry_lines.append({
                "id": str(uuid.uuid4()),
                "account_id": account["id"],
                "account_number": account["account_number"],
                "account_name": account["name"],
                "debit": debit,
                "credit": credit,
                "description": line.get("description", "")
            })
        
        if not entry_lines:
            logging.warning(f"No valid accounts found for auto journal entry: {description}")
            return None
        
        # Verify balance
        if abs(total_debit - total_credit) > 0.01:
            logging.error(f"Unbalanced auto journal entry: debit={total_debit}, credit={total_credit}")
            return None
        
        # Generate entry number
        count = await db.journal_entries.count_documents({})
        entry_number = f"JV-{datetime.now().year}-{count + 1:05d}"
        
        entry = {
            "id": str(uuid.uuid4()),
            "entry_number": entry_number,
            "entry_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "description": description,
            "reference_type": reference_type,
            "reference_id": reference_id,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "status": "posted",  # Auto entries are auto-posted
            "created_by": created_by_id,
            "created_by_name": created_by_name,
            "posted_at": datetime.now(timezone.utc).isoformat(),
            "posted_by": "النظام (آلي)",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.journal_entries.insert_one(entry)
        
        # Insert lines
        for line in entry_lines:
            line["journal_entry_id"] = entry["id"]
            await db.journal_entry_lines.insert_one(line)
        
        # Update account balances
        for line in entry_lines:
            balance_change = line["debit"] - line["credit"]
            await db.chart_of_accounts.update_one(
                {"id": line["account_id"]},
                {"$inc": {"balance": balance_change}}
            )
        
        logging.info(f"Auto journal entry created: {entry_number} - {description}")
        return entry
        
    except Exception as e:
        logging.error(f"Error creating auto journal entry: {e}")
        return None

# ==================== FINANCIAL SYSTEM (النظام المالي) ====================

# ---------- Chart of Accounts (شجرة الحسابات) ----------

@api_router.get("/finance/accounts")
async def get_all_accounts(
    account_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب شجرة الحسابات"""
    query = {"is_active": True}
    if account_type:
        query["account_type"] = account_type
    
    accounts = await db.chart_of_accounts.find(query, {"_id": 0}).sort("account_number", 1).to_list(1000)
    return accounts

@api_router.post("/finance/accounts")
async def create_account(
    account: Account,
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """إنشاء حساب جديد"""
    # Check if account number exists
    existing = await db.chart_of_accounts.find_one({"account_number": account.account_number})
    if existing:
        raise HTTPException(status_code=400, detail="رقم الحساب موجود مسبقاً")
    
    await db.chart_of_accounts.insert_one(account.model_dump())
    return {"message": "تم إنشاء الحساب بنجاح", "account": account.model_dump()}

@api_router.put("/finance/accounts/{account_id}")
async def update_account(
    account_id: str,
    data: dict,
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """تحديث حساب"""
    await db.chart_of_accounts.update_one(
        {"id": account_id},
        {"$set": data}
    )
    return {"message": "تم تحديث الحساب بنجاح"}

# ==================== BANK ACCOUNTS (الحسابات البنكية) ====================

@api_router.get("/finance/bank-accounts")
async def get_bank_accounts(current_user: dict = Depends(get_current_user)):
    """جلب جميع الحسابات البنكية"""
    accounts = await db.bank_accounts.find({"is_active": True}, {"_id": 0}).sort("bank_name", 1).to_list(100)
    return accounts

@api_router.get("/finance/bank-accounts/{account_id}")
async def get_bank_account(account_id: str, current_user: dict = Depends(get_current_user)):
    """جلب حساب بنكي محدد"""
    account = await db.bank_accounts.find_one({"id": account_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="الحساب غير موجود")
    return account

@api_router.post("/finance/bank-accounts")
async def create_bank_account(
    data: dict,
    current_user: dict = Depends(require_role(["admin", "accountant", "finance_manager"]))
):
    """إنشاء حساب بنكي جديد"""
    from models.all_models import BankAccount
    
    # Check if bank account number exists
    existing = await db.bank_accounts.find_one({"bank_account_number": data.get("bank_account_number")})
    if existing:
        raise HTTPException(status_code=400, detail="رقم الحساب البنكي موجود مسبقاً")
    
    # If marked as default, unset other defaults
    if data.get("is_default"):
        await db.bank_accounts.update_many({}, {"$set": {"is_default": False}})
    
    bank_account = BankAccount(**data)
    bank_account_dict = bank_account.model_dump()
    bank_account_dict["current_balance"] = data.get("opening_balance", 0.0)
    
    await db.bank_accounts.insert_one(bank_account_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_bank_account",
        entity_type="bank_account",
        entity_id=bank_account.id,
        entity_name=f"{data.get('bank_name')} - {data.get('bank_account_number')}",
        details=f"إنشاء حساب بنكي: {data.get('bank_name')}"
    )
    
    # Remove _id before returning
    bank_account_dict.pop("_id", None)
    return {"message": "تم إنشاء الحساب البنكي بنجاح", "account": bank_account_dict}

@api_router.put("/finance/bank-accounts/{account_id}")
async def update_bank_account(
    account_id: str,
    data: dict,
    current_user: dict = Depends(require_role(["admin", "accountant", "finance_manager"]))
):
    """تحديث حساب بنكي"""
    existing = await db.bank_accounts.find_one({"id": account_id})
    if not existing:
        raise HTTPException(status_code=404, detail="الحساب غير موجود")
    
    # If marking as default, unset other defaults
    if data.get("is_default"):
        await db.bank_accounts.update_many({"id": {"$ne": account_id}}, {"$set": {"is_default": False}})
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.bank_accounts.update_one(
        {"id": account_id},
        {"$set": data}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_bank_account",
        entity_type="bank_account",
        entity_id=account_id,
        entity_name=data.get("bank_name", existing.get("bank_name")),
        details=f"تحديث حساب بنكي"
    )
    
    return {"message": "تم تحديث الحساب البنكي بنجاح"}

@api_router.delete("/finance/bank-accounts/{account_id}")
async def delete_bank_account(
    account_id: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    """حذف (إلغاء تفعيل) حساب بنكي"""
    result = await db.bank_accounts.update_one(
        {"id": account_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الحساب غير موجود")
    return {"message": "تم حذف الحساب البنكي بنجاح"}

@api_router.get("/finance/bank-accounts/default")
async def get_default_bank_account(current_user: dict = Depends(get_current_user)):
    """جلب الحساب البنكي الافتراضي"""
    account = await db.bank_accounts.find_one({"is_default": True, "is_active": True}, {"_id": 0})
    if not account:
        # Return first active account if no default
        account = await db.bank_accounts.find_one({"is_active": True}, {"_id": 0})
    return account

# Initialize default accounts if none exist
@api_router.post("/finance/accounts/initialize")
async def initialize_chart_of_accounts(
    current_user: dict = Depends(require_role(["admin"]))
):
    """تهيئة شجرة الحسابات الافتراضية"""
    existing = await db.chart_of_accounts.count_documents({})
    if existing > 0:
        return {"message": "شجرة الحسابات موجودة مسبقاً", "count": existing}
    
    default_accounts = [
        # Assets - الأصول
        {"account_number": "1000", "name": "الأصول", "account_type": "asset", "parent_id": None},
        {"account_number": "1100", "name": "الأصول المتداولة", "account_type": "asset", "parent_id": "1000"},
        {"account_number": "1110", "name": "النقدية والبنوك", "account_type": "asset", "parent_id": "1100"},
        {"account_number": "1111", "name": "الصندوق", "account_type": "asset", "parent_id": "1110"},
        {"account_number": "1112", "name": "البنك", "account_type": "asset", "parent_id": "1110"},
        {"account_number": "1120", "name": "العملاء (المدينون)", "account_type": "asset", "parent_id": "1100"},
        {"account_number": "1130", "name": "المخزون", "account_type": "asset", "parent_id": "1100"},
        {"account_number": "1200", "name": "الأصول الثابتة", "account_type": "asset", "parent_id": "1000"},
        {"account_number": "1210", "name": "المباني", "account_type": "asset", "parent_id": "1200"},
        {"account_number": "1220", "name": "المعدات والآلات", "account_type": "asset", "parent_id": "1200"},
        {"account_number": "1230", "name": "السيارات", "account_type": "asset", "parent_id": "1200"},
        {"account_number": "1240", "name": "الأثاث والتجهيزات", "account_type": "asset", "parent_id": "1200"},
        {"account_number": "1290", "name": "مجمع الإهلاك", "account_type": "asset", "parent_id": "1200"},
        
        # Liabilities - الخصوم
        {"account_number": "2000", "name": "الخصوم", "account_type": "liability", "parent_id": None},
        {"account_number": "2100", "name": "الخصوم المتداولة", "account_type": "liability", "parent_id": "2000"},
        {"account_number": "2110", "name": "الموردون (الدائنون)", "account_type": "liability", "parent_id": "2100"},
        {"account_number": "2120", "name": "الرواتب المستحقة", "account_type": "liability", "parent_id": "2100"},
        {"account_number": "2130", "name": "الضرائب المستحقة", "account_type": "liability", "parent_id": "2100"},
        {"account_number": "2200", "name": "الخصوم طويلة الأجل", "account_type": "liability", "parent_id": "2000"},
        {"account_number": "2210", "name": "القروض", "account_type": "liability", "parent_id": "2200"},
        
        # Equity - حقوق الملكية
        {"account_number": "3000", "name": "حقوق الملكية", "account_type": "equity", "parent_id": None},
        {"account_number": "3100", "name": "رأس المال", "account_type": "equity", "parent_id": "3000"},
        {"account_number": "3200", "name": "الأرباح المحتجزة", "account_type": "equity", "parent_id": "3000"},
        
        # Revenue - الإيرادات
        {"account_number": "4000", "name": "الإيرادات", "account_type": "revenue", "parent_id": None},
        {"account_number": "4100", "name": "إيرادات مبيعات الحليب", "account_type": "revenue", "parent_id": "4000"},
        {"account_number": "4200", "name": "إيرادات أخرى", "account_type": "revenue", "parent_id": "4000"},
        
        # Expenses - المصروفات
        {"account_number": "5000", "name": "المصروفات", "account_type": "expense", "parent_id": None},
        {"account_number": "5100", "name": "تكلفة شراء الحليب", "account_type": "expense", "parent_id": "5000"},
        {"account_number": "5200", "name": "الرواتب والأجور", "account_type": "expense", "parent_id": "5000"},
        {"account_number": "5300", "name": "مصاريف التشغيل", "account_type": "expense", "parent_id": "5000"},
        {"account_number": "5310", "name": "الكهرباء والماء", "account_type": "expense", "parent_id": "5300"},
        {"account_number": "5320", "name": "الوقود والنقل", "account_type": "expense", "parent_id": "5300"},
        {"account_number": "5330", "name": "الصيانة والإصلاح", "account_type": "expense", "parent_id": "5300"},
        {"account_number": "5400", "name": "مصاريف إدارية", "account_type": "expense", "parent_id": "5000"},
        {"account_number": "5500", "name": "مصاريف الإهلاك", "account_type": "expense", "parent_id": "5000"},
    ]
    
    for acc in default_accounts:
        account = Account(
            account_number=acc["account_number"],
            name=acc["name"],
            account_type=acc["account_type"],
            parent_id=acc.get("parent_id")
        )
        await db.chart_of_accounts.insert_one(account.model_dump())
    
    return {"message": "تم إنشاء شجرة الحسابات بنجاح", "count": len(default_accounts)}

# ---------- Journal Entries (القيود اليومية) ----------

@api_router.get("/finance/journal-entries")
async def get_journal_entries(
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب القيود اليومية"""
    query = {}
    if status:
        query["status"] = status
    if start_date:
        query["entry_date"] = {"$gte": start_date}
    if end_date:
        if "entry_date" in query:
            query["entry_date"]["$lte"] = end_date
        else:
            query["entry_date"] = {"$lte": end_date}
    
    entries = await db.journal_entries.find(query, {"_id": 0}).sort("entry_date", -1).to_list(500)
    
    # Get lines for each entry
    for entry in entries:
        lines = await db.journal_entry_lines.find({"journal_entry_id": entry["id"]}, {"_id": 0}).to_list(100)
        entry["lines"] = lines
    
    return entries

@api_router.post("/finance/journal-entries")
async def create_journal_entry(
    data: dict,
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """إنشاء قيد يومية جديد"""
    # Generate entry number
    count = await db.journal_entries.count_documents({})
    entry_number = f"JV-{datetime.now().year}-{count + 1:05d}"
    
    lines = data.pop("lines", [])
    
    # Calculate totals
    total_debit = sum(line.get("debit", 0) for line in lines)
    total_credit = sum(line.get("credit", 0) for line in lines)
    
    if abs(total_debit - total_credit) > 0.01:
        raise HTTPException(status_code=400, detail="القيد غير متوازن - المدين لا يساوي الدائن")
    
    entry = JournalEntry(
        entry_number=entry_number,
        entry_date=data.get("entry_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        description=data.get("description", ""),
        reference_type=data.get("reference_type"),
        reference_id=data.get("reference_id"),
        total_debit=total_debit,
        total_credit=total_credit,
        status="draft",
        created_by=current_user["id"],
        created_by_name=current_user["full_name"]
    )
    
    await db.journal_entries.insert_one(entry.model_dump())
    
    # Insert lines
    for line in lines:
        entry_line = JournalEntryLine(
            journal_entry_id=entry.id,
            account_id=line["account_id"],
            account_number=line["account_number"],
            account_name=line["account_name"],
            debit=line.get("debit", 0),
            credit=line.get("credit", 0),
            description=line.get("description")
        )
        await db.journal_entry_lines.insert_one(entry_line.model_dump())
    
    return {"message": "تم إنشاء القيد بنجاح", "entry_number": entry_number, "id": entry.id}

@api_router.put("/finance/journal-entries/{entry_id}/post")
async def post_journal_entry(
    entry_id: str,
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """ترحيل قيد يومية"""
    entry = await db.journal_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="القيد غير موجود")
    
    if entry["status"] == "posted":
        raise HTTPException(status_code=400, detail="القيد مرحل مسبقاً")
    
    # Update account balances
    lines = await db.journal_entry_lines.find({"journal_entry_id": entry_id}, {"_id": 0}).to_list(100)
    for line in lines:
        balance_change = line["debit"] - line["credit"]
        await db.chart_of_accounts.update_one(
            {"id": line["account_id"]},
            {"$inc": {"balance": balance_change}}
        )
    
    await db.journal_entries.update_one(
        {"id": entry_id},
        {"$set": {
            "status": "posted",
            "posted_at": datetime.now(timezone.utc).isoformat(),
            "posted_by": current_user["full_name"]
        }}
    )
    
    return {"message": "تم ترحيل القيد بنجاح"}

# ---------- Fixed Assets (الأصول الثابتة) ----------

@api_router.get("/finance/fixed-assets")
async def get_fixed_assets(
    category: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب الأصول الثابتة"""
    query = {}
    if category:
        query["category"] = category
    if status:
        query["status"] = status
    
    assets = await db.fixed_assets.find(query, {"_id": 0}).sort("purchase_date", -1).to_list(500)
    return assets

@api_router.post("/finance/fixed-assets")
async def create_fixed_asset(
    data: dict,
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """إضافة أصل ثابت جديد"""
    # Generate asset number
    count = await db.fixed_assets.count_documents({})
    asset_number = f"FA-{count + 1:04d}"
    
    asset = FixedAsset(
        asset_number=asset_number,
        name=data["name"],
        category=data["category"],
        purchase_date=data["purchase_date"],
        purchase_cost=data["purchase_cost"],
        useful_life_years=data.get("useful_life_years", 5),
        salvage_value=data.get("salvage_value", 0),
        depreciation_method=data.get("depreciation_method", "straight_line"),
        current_value=data["purchase_cost"],
        location=data.get("location"),
        assigned_to=data.get("assigned_to"),
        notes=data.get("notes")
    )
    
    await db.fixed_assets.insert_one(asset.model_dump())
    return {"message": "تم إضافة الأصل بنجاح", "asset_number": asset_number}

@api_router.post("/finance/fixed-assets/calculate-depreciation")
async def calculate_depreciation(
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """حساب الإهلاك الشهري لجميع الأصول"""
    assets = await db.fixed_assets.find({"status": "active"}, {"_id": 0}).to_list(500)
    
    depreciation_entries = []
    
    for asset in assets:
        if asset["depreciation_method"] == "straight_line":
            # القسط الثابت
            annual_depreciation = (asset["purchase_cost"] - asset["salvage_value"]) / asset["useful_life_years"]
            monthly_depreciation = annual_depreciation / 12
            
            new_accumulated = asset["accumulated_depreciation"] + monthly_depreciation
            new_current_value = asset["purchase_cost"] - new_accumulated
            
            if new_current_value >= asset["salvage_value"]:
                await db.fixed_assets.update_one(
                    {"id": asset["id"]},
                    {"$set": {
                        "accumulated_depreciation": new_accumulated,
                        "current_value": new_current_value
                    }}
                )
                depreciation_entries.append({
                    "asset_name": asset["name"],
                    "depreciation": monthly_depreciation
                })
    
    return {
        "message": f"تم حساب الإهلاك لـ {len(depreciation_entries)} أصل",
        "entries": depreciation_entries
    }

# ---------- Budgets (الميزانيات) ----------

@api_router.get("/finance/budgets")
async def get_budgets(
    fiscal_year: Optional[int] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب الميزانيات"""
    query = {}
    if fiscal_year:
        query["fiscal_year"] = fiscal_year
    if status:
        query["status"] = status
    
    budgets = await db.budgets.find(query, {"_id": 0}).sort("fiscal_year", -1).to_list(100)
    
    for budget in budgets:
        lines = await db.budget_lines.find({"budget_id": budget["id"]}, {"_id": 0}).to_list(100)
        budget["lines"] = lines
        budget["total_budgeted"] = sum(l.get("budgeted_amount", 0) for l in lines)
        budget["total_actual"] = sum(l.get("actual_amount", 0) for l in lines)
    
    return budgets

@api_router.post("/finance/budgets")
async def create_budget(
    data: dict,
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """إنشاء ميزانية جديدة"""
    budget = Budget(
        name=data["name"],
        fiscal_year=data["fiscal_year"],
        start_date=data["start_date"],
        end_date=data["end_date"],
        created_by=current_user["id"]
    )
    
    await db.budgets.insert_one(budget.model_dump())
    
    # Create budget lines if provided
    lines = data.get("lines", [])
    for line in lines:
        budget_line = BudgetLine(
            budget_id=budget.id,
            account_id=line["account_id"],
            account_name=line["account_name"],
            budgeted_amount=line.get("budgeted_amount", 0)
        )
        await db.budget_lines.insert_one(budget_line.model_dump())
    
    return {"message": "تم إنشاء الميزانية بنجاح", "id": budget.id}

# ---------- Tax Records (الضرائب) ----------

@api_router.get("/finance/taxes")
async def get_tax_records(
    tax_type: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجلات الضرائب"""
    query = {}
    if tax_type:
        query["tax_type"] = tax_type
    if status:
        query["status"] = status
    
    records = await db.tax_records.find(query, {"_id": 0}).sort("period_end", -1).to_list(200)
    return records

@api_router.post("/finance/taxes")
async def create_tax_record(
    data: dict,
    current_user: dict = Depends(require_role(["admin", "accountant"]))
):
    """إنشاء سجل ضريبي"""
    tax_amount = data["taxable_amount"] * (data["tax_rate"] / 100)
    
    record = TaxRecord(
        tax_type=data["tax_type"],
        period=data["period"],
        period_start=data["period_start"],
        period_end=data["period_end"],
        taxable_amount=data["taxable_amount"],
        tax_rate=data["tax_rate"],
        tax_amount=tax_amount,
        due_date=data.get("due_date"),
        notes=data.get("notes")
    )
    
    await db.tax_records.insert_one(record.model_dump())
    return {"message": "تم إنشاء السجل الضريبي بنجاح", "id": record.id, "tax_amount": tax_amount}

# ---------- Accounts Payable (الحسابات الدائنة) ----------

@api_router.get("/finance/accounts-payable")
async def get_accounts_payable(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب الحسابات الدائنة (مستحقات الموردين)"""
    query = {}
    if status:
        query["status"] = status
    
    records = await db.accounts_payable.find(query, {"_id": 0}).sort("due_date", 1).to_list(500)
    return records

@api_router.get("/finance/accounts-payable/summary")
async def get_accounts_payable_summary(current_user: dict = Depends(get_current_user)):
    """ملخص الحسابات الدائنة"""
    records = await db.accounts_payable.find({}, {"_id": 0}).to_list(1000)
    
    total_payable = sum(r.get("balance", 0) for r in records)
    overdue = [r for r in records if r["status"] != "paid" and r["due_date"] < datetime.now(timezone.utc).strftime("%Y-%m-%d")]
    
    return {
        "total_payable": total_payable,
        "total_records": len(records),
        "unpaid_count": len([r for r in records if r["status"] == "unpaid"]),
        "partial_count": len([r for r in records if r["status"] == "partial"]),
        "overdue_count": len(overdue),
        "overdue_amount": sum(r.get("balance", 0) for r in overdue)
    }

# ---------- Accounts Receivable (الحسابات المدينة) ----------

@api_router.get("/finance/accounts-receivable")
async def get_accounts_receivable(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب الحسابات المدينة (مستحقات من العملاء)"""
    query = {}
    if status:
        query["status"] = status
    
    records = await db.accounts_receivable.find(query, {"_id": 0}).sort("due_date", 1).to_list(500)
    return records

@api_router.get("/finance/accounts-receivable/summary")
async def get_accounts_receivable_summary(current_user: dict = Depends(get_current_user)):
    """ملخص الحسابات المدينة"""
    records = await db.accounts_receivable.find({}, {"_id": 0}).to_list(1000)
    
    total_receivable = sum(r.get("balance", 0) for r in records)
    overdue = [r for r in records if r["status"] != "paid" and r["due_date"] < datetime.now(timezone.utc).strftime("%Y-%m-%d")]
    
    return {
        "total_receivable": total_receivable,
        "total_records": len(records),
        "unpaid_count": len([r for r in records if r["status"] == "unpaid"]),
        "partial_count": len([r for r in records if r["status"] == "partial"]),
        "overdue_count": len(overdue),
        "overdue_amount": sum(r.get("balance", 0) for r in overdue)
    }

# ---------- Financial Reports (التقارير المالية) ----------

@api_router.get("/finance/reports/trial-balance")
async def get_trial_balance(
    as_of_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """ميزان المراجعة"""
    accounts = await db.chart_of_accounts.find({"is_active": True}, {"_id": 0}).sort("account_number", 1).to_list(1000)
    
    total_debit = 0
    total_credit = 0
    
    for acc in accounts:
        balance = acc.get("balance", 0)
        if balance >= 0:
            acc["debit_balance"] = balance
            acc["credit_balance"] = 0
            total_debit += balance
        else:
            acc["debit_balance"] = 0
            acc["credit_balance"] = abs(balance)
            total_credit += abs(balance)
    
    return {
        "accounts": accounts,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "is_balanced": abs(total_debit - total_credit) < 0.01
    }

@api_router.get("/finance/reports/income-statement")
async def get_income_statement(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """قائمة الدخل"""
    revenue_accounts = await db.chart_of_accounts.find(
        {"account_type": "revenue", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    expense_accounts = await db.chart_of_accounts.find(
        {"account_type": "expense", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    total_revenue = sum(abs(acc.get("balance", 0)) for acc in revenue_accounts)
    total_expenses = sum(abs(acc.get("balance", 0)) for acc in expense_accounts)
    net_income = total_revenue - total_expenses
    
    return {
        "revenue": revenue_accounts,
        "total_revenue": total_revenue,
        "expenses": expense_accounts,
        "total_expenses": total_expenses,
        "net_income": net_income,
        "period": {"start": start_date, "end": end_date}
    }

@api_router.get("/finance/reports/balance-sheet")
async def get_balance_sheet(
    as_of_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الميزانية العمومية"""
    assets = await db.chart_of_accounts.find(
        {"account_type": "asset", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    liabilities = await db.chart_of_accounts.find(
        {"account_type": "liability", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    equity = await db.chart_of_accounts.find(
        {"account_type": "equity", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    total_assets = sum(acc.get("balance", 0) for acc in assets)
    total_liabilities = sum(abs(acc.get("balance", 0)) for acc in liabilities)
    total_equity = sum(abs(acc.get("balance", 0)) for acc in equity)
    
    return {
        "assets": assets,
        "total_assets": total_assets,
        "liabilities": liabilities,
        "total_liabilities": total_liabilities,
        "equity": equity,
        "total_equity": total_equity,
        "is_balanced": abs(total_assets - (total_liabilities + total_equity)) < 0.01,
        "as_of_date": as_of_date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    }

@api_router.get("/finance/dashboard")
async def get_finance_dashboard(current_user: dict = Depends(get_current_user)):
    """لوحة تحكم النظام المالي"""
    # Get account summaries
    assets = await db.chart_of_accounts.find({"account_type": "asset"}, {"_id": 0}).to_list(100)
    liabilities = await db.chart_of_accounts.find({"account_type": "liability"}, {"_id": 0}).to_list(100)
    revenue = await db.chart_of_accounts.find({"account_type": "revenue"}, {"_id": 0}).to_list(100)
    expenses = await db.chart_of_accounts.find({"account_type": "expense"}, {"_id": 0}).to_list(100)
    
    # Get recent transactions
    recent_entries = await db.journal_entries.find(
        {"status": "posted"}, {"_id": 0}
    ).sort("posted_at", -1).limit(10).to_list(10)
    
    # Get AP/AR summaries
    ap_total = await db.accounts_payable.aggregate([
        {"$match": {"status": {"$ne": "paid"}}},
        {"$group": {"_id": None, "total": {"$sum": "$balance"}}}
    ]).to_list(1)
    
    ar_total = await db.accounts_receivable.aggregate([
        {"$match": {"status": {"$ne": "paid"}}},
        {"$group": {"_id": None, "total": {"$sum": "$balance"}}}
    ]).to_list(1)
    
    # Get fixed assets summary
    fixed_assets = await db.fixed_assets.find({"status": "active"}, {"_id": 0}).to_list(500)
    
    return {
        "summary": {
            "total_assets": sum(a.get("balance", 0) for a in assets),
            "total_liabilities": sum(abs(l.get("balance", 0)) for l in liabilities),
            "total_revenue": sum(abs(r.get("balance", 0)) for r in revenue),
            "total_expenses": sum(abs(e.get("balance", 0)) for e in expenses),
            "accounts_payable": ap_total[0]["total"] if ap_total else 0,
            "accounts_receivable": ar_total[0]["total"] if ar_total else 0,
            "fixed_assets_value": sum(a.get("current_value", 0) for a in fixed_assets),
            "fixed_assets_count": len(fixed_assets)
        },
        "recent_entries": recent_entries
    }

# ==================== REPORTS EXPORT (تصدير التقارير) ====================

@api_router.get("/reports/export/suppliers/excel")
async def export_suppliers_excel(current_user: dict = Depends(get_current_user)):
    """Export suppliers report to Excel"""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    suppliers = await db.suppliers.find({"is_active": True}, {"_id": 0}).to_list(1000)
    
    if not suppliers:
        raise HTTPException(status_code=404, detail="No suppliers found")
    
    # Create DataFrame
    df = pd.DataFrame(suppliers)
    columns_map = {
        'name': 'اسم المورد',
        'code': 'الكود',
        'phone': 'الهاتف',
        'bank_account': 'رقم الحساب البنكي',
        'balance': 'الرصيد',
        'total_supplied': 'إجمالي التوريد',
        'center_name': 'المركز'
    }
    
    # Select and rename columns
    available_cols = [col for col in columns_map.keys() if col in df.columns]
    df = df[available_cols]
    df = df.rename(columns={k: v for k, v in columns_map.items() if k in available_cols})
    
    # Create Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='الموردين', index=False)
        
        # Style the worksheet
        workbook = writer.book
        worksheet = writer.sheets['الموردين']
        
        # Style header
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = max_length + 5
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=suppliers_report.xlsx"}
    )

@api_router.get("/reports/export/milk-receptions/excel")
async def export_milk_receptions_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    supplier_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export milk receptions report to Excel"""
    import pandas as pd
    
    query = {}
    if start_date:
        query["reception_date"] = {"$gte": start_date}
    if end_date:
        if "reception_date" in query:
            query["reception_date"]["$lte"] = end_date
        else:
            query["reception_date"] = {"$lte": end_date}
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    receptions = await db.milk_receptions.find(query, {"_id": 0}).sort("reception_date", -1).to_list(10000)
    
    if not receptions:
        raise HTTPException(status_code=404, detail="No receptions found")
    
    df = pd.DataFrame(receptions)
    columns_map = {
        'reception_date': 'تاريخ الاستلام',
        'supplier_name': 'اسم المورد',
        'quantity_liters': 'الكمية (لتر)',
        'price_per_liter': 'سعر اللتر',
        'total_amount': 'المبلغ الإجمالي',
        'fat_percentage': 'نسبة الدهون',
        'protein_percentage': 'نسبة البروتين'
    }
    
    available_cols = [col for col in columns_map.keys() if col in df.columns]
    df = df[available_cols]
    df = df.rename(columns={k: v for k, v in columns_map.items() if k in available_cols})
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='استلام الحليب', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['استلام الحليب']
        
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = max_length + 5
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=milk_receptions_report.xlsx"}
    )

@api_router.get("/reports/export/hr/employees/excel")
async def export_employees_excel(current_user: dict = Depends(get_current_user)):
    """Export HR employees report to Excel"""
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment
    
    employees = await db.hr_employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    
    if not employees:
        raise HTTPException(status_code=404, detail="No employees found")
    
    df = pd.DataFrame(employees)
    columns_map = {
        'employee_code': 'كود الموظف',
        'name': 'اسم الموظف',
        'department': 'القسم',
        'position': 'المنصب',
        'phone': 'الهاتف',
        'email': 'البريد الإلكتروني',
        'salary': 'الراتب',
        'hire_date': 'تاريخ التعيين'
    }
    
    available_cols = [col for col in columns_map.keys() if col in df.columns]
    df = df[available_cols]
    df = df.rename(columns={k: v for k, v in columns_map.items() if k in available_cols})
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='الموظفين', index=False)
        
        worksheet = writer.sheets['الموظفين']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = max_length + 5
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees_report.xlsx"}
    )

@api_router.get("/reports/export/finance/excel")
async def export_finance_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export finance report to Excel"""
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment
    
    query = {}
    if start_date:
        query["payment_date"] = {"$gte": start_date}
    if end_date:
        if "payment_date" in query:
            query["payment_date"]["$lte"] = end_date
        else:
            query["payment_date"] = {"$lte": end_date}
    
    payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(10000)
    
    if not payments:
        raise HTTPException(status_code=404, detail="No payments found")
    
    df = pd.DataFrame(payments)
    columns_map = {
        'payment_date': 'تاريخ الدفع',
        'payment_type': 'نوع الدفع',
        'related_name': 'الاسم',
        'amount': 'المبلغ',
        'payment_method': 'طريقة الدفع',
        'bank_account': 'الحساب البنكي',
        'notes': 'ملاحظات'
    }
    
    available_cols = [col for col in columns_map.keys() if col in df.columns]
    df = df[available_cols]
    df = df.rename(columns={k: v for k, v in columns_map.items() if k in available_cols})
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='المدفوعات', index=False)
        
        worksheet = writer.sheets['المدفوعات']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = max_length + 5
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finance_report.xlsx"}
    )

@api_router.get("/reports/export/suppliers/pdf")
async def export_suppliers_pdf(current_user: dict = Depends(get_current_user)):
    """Export suppliers report to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    
    suppliers = await db.suppliers.find({"is_active": True}, {"_id": 0}).to_list(1000)
    
    if not suppliers:
        raise HTTPException(status_code=404, detail="No suppliers found")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=18)
    elements.append(Paragraph("تقرير الموردين - Suppliers Report", title_style))
    elements.append(Spacer(1, 20))
    
    # Date
    date_style = ParagraphStyle('Date', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)
    elements.append(Paragraph(f"التاريخ: {datetime.now().strftime('%Y-%m-%d')}", date_style))
    elements.append(Spacer(1, 20))
    
    # Table data
    headers = ['Code', 'Name', 'Phone', 'Bank Account', 'Balance', 'Total Supplied', 'Center']
    data = [headers]
    
    for s in suppliers:
        row = [
            s.get('code', ''),
            s.get('name', ''),
            s.get('phone', ''),
            s.get('bank_account', ''),
            f"{s.get('balance', 0):.3f}",
            f"{s.get('total_supplied', 0):.2f}",
            s.get('center_name', '')
        ]
        data.append(row)
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E9ECF1')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=suppliers_report.pdf"}
    )

@api_router.get("/reports/export/daily/pdf")
async def export_daily_report_pdf(
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Export daily report to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    
    # Get daily data
    receptions = await db.milk_receptions.find({"reception_date": date}, {"_id": 0}).to_list(1000)
    sales = await db.sales.find({"sale_date": date}, {"_id": 0}).to_list(1000)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
    elements.append(Paragraph(f"التقرير اليومي - Daily Report", title_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"التاريخ: {date}", ParagraphStyle('Date', alignment=TA_CENTER)))
    elements.append(Spacer(1, 20))
    
    # Summary
    total_milk = sum(r.get('quantity_liters', 0) for r in receptions)
    total_sales = sum(s.get('total_amount', 0) for s in sales)
    
    summary_data = [
        ['الوصف', 'القيمة'],
        ['إجمالي الحليب المستلم (لتر)', f'{total_milk:.2f}'],
        ['عدد عمليات الاستلام', str(len(receptions))],
        ['إجمالي المبيعات (ر.ع)', f'{total_sales:.3f}'],
        ['عدد عمليات البيع', str(len(sales))],
    ]
    
    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Receptions detail
    if receptions:
        elements.append(Paragraph("تفاصيل استلام الحليب - Milk Receptions", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        rec_headers = ['Supplier', 'Quantity (L)', 'Price/L', 'Total', 'Fat %']
        rec_data = [rec_headers]
        for r in receptions:
            rec_data.append([
                r.get('supplier_name', ''),
                f"{r.get('quantity_liters', 0):.2f}",
                f"{r.get('price_per_liter', 0):.3f}",
                f"{r.get('total_amount', 0):.3f}",
                f"{r.get('fat_percentage', 0):.1f}"
            ])
        
        rec_table = Table(rec_data, repeatRows=1)
        rec_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(rec_table)
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=daily_report_{date}.pdf"}
    )

# ==================== LEGAL MODULE ROUTES (قسم القانون) ====================

# Legal Contracts (العقود القانونية)
@api_router.post("/legal/contracts", response_model=LegalContract)
async def create_legal_contract(contract_data: LegalContractCreate, current_user: dict = Depends(get_current_user)):
    # Generate contract number
    count = await db.legal_contracts.count_documents({})
    year = datetime.now().year
    contract_number = f"CTR-{year}-{count + 1:04d}"
    
    contract = LegalContract(**contract_data.model_dump())
    contract_dict = contract.model_dump()
    contract_dict["contract_number"] = contract_number
    contract_dict["created_by"] = current_user["id"]
    
    await db.legal_contracts.insert_one(contract_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_legal_contract",
        entity_type="legal_contract",
        entity_id=contract.id,
        entity_name=contract_data.title,
        details=f"عقد قانوني: {contract_data.title} - {contract_data.party_name}"
    )
    
    return LegalContract(**contract_dict)

@api_router.get("/legal/contracts")
async def get_legal_contracts(
    status: Optional[str] = None,
    contract_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if contract_type:
        query["contract_type"] = contract_type
    
    contracts = await db.legal_contracts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return contracts

@api_router.get("/legal/contracts/{contract_id}")
async def get_legal_contract(contract_id: str, current_user: dict = Depends(get_current_user)):
    contract = await db.legal_contracts.find_one({"id": contract_id}, {"_id": 0})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    return contract

@api_router.put("/legal/contracts/{contract_id}", response_model=LegalContract)
async def update_legal_contract(contract_id: str, contract_data: LegalContractCreate, current_user: dict = Depends(get_current_user)):
    result = await db.legal_contracts.update_one(
        {"id": contract_id},
        {"$set": contract_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    contract = await db.legal_contracts.find_one({"id": contract_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_legal_contract",
        entity_type="legal_contract",
        entity_id=contract_id,
        entity_name=contract.get("title"),
        details=f"تعديل عقد: {contract.get('title')}"
    )
    
    return contract

@api_router.delete("/legal/contracts/{contract_id}")
async def delete_legal_contract(contract_id: str, current_user: dict = Depends(get_current_user)):
    contract = await db.legal_contracts.find_one({"id": contract_id}, {"_id": 0})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    await db.legal_contracts.update_one(
        {"id": contract_id},
        {"$set": {"status": "terminated"}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="terminate_legal_contract",
        entity_type="legal_contract",
        entity_id=contract_id,
        entity_name=contract.get("title"),
        details=f"إنهاء عقد: {contract.get('title')}"
    )
    
    return {"message": "Contract terminated successfully"}

# Legal Cases (القضايا القانونية)
@api_router.post("/legal/cases", response_model=LegalCase)
async def create_legal_case(case_data: LegalCaseCreate, current_user: dict = Depends(get_current_user)):
    count = await db.legal_cases.count_documents({})
    year = datetime.now().year
    case_number = f"CASE-{year}-{count + 1:04d}"
    
    case = LegalCase(**case_data.model_dump())
    case_dict = case.model_dump()
    case_dict["case_number"] = case_number
    case_dict["created_by"] = current_user["id"]
    
    await db.legal_cases.insert_one(case_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_legal_case",
        entity_type="legal_case",
        entity_id=case.id,
        entity_name=case_data.title,
        details=f"قضية قانونية: {case_data.title}"
    )
    
    return LegalCase(**case_dict)

@api_router.get("/legal/cases")
async def get_legal_cases(
    status: Optional[str] = None,
    case_type: Optional[str] = None,
    priority: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if case_type:
        query["case_type"] = case_type
    if priority:
        query["priority"] = priority
    
    cases = await db.legal_cases.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return cases

@api_router.put("/legal/cases/{case_id}", response_model=LegalCase)
async def update_legal_case(case_id: str, case_data: LegalCaseCreate, current_user: dict = Depends(get_current_user)):
    result = await db.legal_cases.update_one(
        {"id": case_id},
        {"$set": case_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Case not found")
    
    case = await db.legal_cases.find_one({"id": case_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_legal_case",
        entity_type="legal_case",
        entity_id=case_id,
        entity_name=case.get("title"),
        details=f"تعديل قضية: {case.get('title')}"
    )
    
    return case

@api_router.put("/legal/cases/{case_id}/close")
async def close_legal_case(case_id: str, outcome: str, settlement_amount: Optional[float] = None, current_user: dict = Depends(get_current_user)):
    result = await db.legal_cases.update_one(
        {"id": case_id},
        {"$set": {
            "status": "closed",
            "outcome": outcome,
            "settlement_amount": settlement_amount,
            "closed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Case not found")
    
    case = await db.legal_cases.find_one({"id": case_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="close_legal_case",
        entity_type="legal_case",
        entity_id=case_id,
        entity_name=case.get("title"),
        details=f"إغلاق قضية: {case.get('title')} - {outcome}"
    )
    
    return case

# Legal Consultations (الاستشارات القانونية)
@api_router.post("/legal/consultations", response_model=LegalConsultation)
async def create_legal_consultation(consultation_data: LegalConsultationCreate, current_user: dict = Depends(get_current_user)):
    consultation = LegalConsultation(**consultation_data.model_dump())
    await db.legal_consultations.insert_one(consultation.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_legal_consultation",
        entity_type="legal_consultation",
        entity_id=consultation.id,
        entity_name=consultation_data.subject,
        details=f"استشارة قانونية: {consultation_data.subject}"
    )
    
    return consultation

@api_router.get("/legal/consultations")
async def get_legal_consultations(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    
    consultations = await db.legal_consultations.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return consultations

@api_router.put("/legal/consultations/{consultation_id}/respond")
async def respond_to_consultation(consultation_id: str, response: str, current_user: dict = Depends(get_current_user)):
    result = await db.legal_consultations.update_one(
        {"id": consultation_id},
        {"$set": {
            "status": "completed",
            "response": response,
            "responded_by": current_user["full_name"],
            "responded_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Consultation not found")
    
    consultation = await db.legal_consultations.find_one({"id": consultation_id}, {"_id": 0})
    return consultation

# Legal Documents (المستندات القانونية)
@api_router.post("/legal/documents", response_model=LegalDocument)
async def create_legal_document(document_data: LegalDocumentCreate, current_user: dict = Depends(get_current_user)):
    document = LegalDocument(**document_data.model_dump())
    document_dict = document.model_dump()
    document_dict["created_by"] = current_user["id"]
    
    await db.legal_documents.insert_one(document_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_legal_document",
        entity_type="legal_document",
        entity_id=document.id,
        entity_name=document_data.title,
        details=f"مستند قانوني: {document_data.title}"
    )
    
    return LegalDocument(**document_dict)

@api_router.get("/legal/documents")
async def get_legal_documents(
    document_type: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if document_type:
        query["document_type"] = document_type
    if status:
        query["status"] = status
    
    documents = await db.legal_documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return documents

# Legal Dashboard Stats
@api_router.get("/legal/dashboard")
async def get_legal_dashboard(current_user: dict = Depends(get_current_user)):
    contracts_active = await db.legal_contracts.count_documents({"status": "active"})
    contracts_expiring = await db.legal_contracts.count_documents({
        "status": "active",
        "end_date": {"$lte": (datetime.now() + timedelta(days=30)).isoformat()}
    })
    cases_open = await db.legal_cases.count_documents({"status": {"$in": ["open", "in_progress"]}})
    consultations_pending = await db.legal_consultations.count_documents({"status": "pending"})
    
    return {
        "contracts_active": contracts_active,
        "contracts_expiring_soon": contracts_expiring,
        "cases_open": cases_open,
        "consultations_pending": consultations_pending
    }


# ==================== LEGAL REVIEWS (المراجعات القانونية) ====================

@api_router.get("/legal/reviews")
async def get_legal_reviews(current_user: dict = Depends(get_current_user)):
    """جلب جميع المراجعات القانونية"""
    reviews = await db.legal_reviews.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return reviews

@api_router.post("/legal/reviews")
async def create_legal_review(data: dict, current_user: dict = Depends(get_current_user)):
    """إنشاء مراجعة قانونية جديدة"""
    review = {
        "id": str(uuid.uuid4()),
        "review_type": data.get("review_type"),
        "title": data.get("title"),
        "description": data.get("description", ""),
        "reviewer_name": data.get("reviewer_name"),
        "review_date": data.get("review_date"),
        "status": data.get("status", "pending"),
        "findings": data.get("findings", ""),
        "recommendations": data.get("recommendations", ""),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.legal_reviews.insert_one(review)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_legal_review",
        entity_type="legal_review",
        entity_id=review["id"],
        entity_name=review["title"],
        details=f"مراجعة جديدة: {review['title']}"
    )
    
    return {"message": "تم إنشاء المراجعة بنجاح", "id": review["id"]}

@api_router.put("/legal/reviews/{review_id}")
async def update_legal_review(review_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """تحديث مراجعة قانونية"""
    update_data = {k: v for k, v in data.items() if k not in ["id", "created_at", "created_by"]}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["id"]
    
    result = await db.legal_reviews.update_one({"id": review_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="المراجعة غير موجودة")
    
    return {"message": "تم تحديث المراجعة بنجاح"}


# ==================== SUPPLIER WAIVERS (تنازلات الموردين) ====================

@api_router.get("/legal/waivers")
async def get_supplier_waivers(current_user: dict = Depends(get_current_user)):
    """جلب جميع تنازلات الموردين"""
    waivers = await db.supplier_waivers.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return waivers

@api_router.post("/legal/waivers")
async def create_supplier_waiver(data: dict, current_user: dict = Depends(get_current_user)):
    """إنشاء تنازل جديد"""
    waiver = {
        "id": str(uuid.uuid4()),
        "from_supplier_id": data.get("from_supplier_id"),
        "from_supplier_name": data.get("from_supplier_name"),
        "to_supplier_id": data.get("to_supplier_id"),
        "to_supplier_name": data.get("to_supplier_name"),
        "quota_amount": float(data.get("quota_amount", 0)),
        "waiver_date": data.get("waiver_date"),
        "reason": data.get("reason"),
        "notes": data.get("notes", ""),
        "documents": data.get("documents", []),
        "status": data.get("status", "pending"),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.supplier_waivers.insert_one(waiver)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_supplier_waiver",
        entity_type="supplier_waiver",
        entity_id=waiver["id"],
        entity_name=f"{waiver['from_supplier_name']} → {waiver['to_supplier_name']}",
        details=f"تنازل جديد: {waiver['from_supplier_name']} → {waiver['to_supplier_name']} ({waiver['quota_amount']} لتر)"
    )
    
    return {"message": "تم إنشاء التنازل بنجاح", "id": waiver["id"]}

@api_router.put("/legal/waivers/{waiver_id}")
async def update_supplier_waiver(waiver_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """تحديث تنازل"""
    update_data = {k: v for k, v in data.items() if k not in ["id", "created_at", "created_by"]}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["id"]
    
    result = await db.supplier_waivers.update_one({"id": waiver_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="التنازل غير موجود")
    
    # If approved, update supplier quotas
    if data.get("status") == "approved":
        # Update from_supplier quota
        await db.suppliers.update_one(
            {"id": data.get("from_supplier_id")},
            {"$inc": {"quota": -float(data.get("quota_amount", 0))}}
        )
        # Update to_supplier quota
        await db.suppliers.update_one(
            {"id": data.get("to_supplier_id")},
            {"$inc": {"quota": float(data.get("quota_amount", 0))}}
        )
    
    return {"message": "تم تحديث التنازل بنجاح"}

@api_router.get("/legal/waivers/{waiver_id}")
async def get_supplier_waiver(waiver_id: str, current_user: dict = Depends(get_current_user)):
    """جلب تفاصيل تنازل"""
    waiver = await db.supplier_waivers.find_one({"id": waiver_id}, {"_id": 0})
    if not waiver:
        raise HTTPException(status_code=404, detail="التنازل غير موجود")
    return waiver


# ==================== PROJECTS MODULE ROUTES (قسم المشاريع) ====================

# Projects
@api_router.post("/projects", response_model=Project)
async def create_project(project_data: ProjectCreate, current_user: dict = Depends(get_current_user)):
    count = await db.projects.count_documents({})
    year = datetime.now().year
    project_code = f"PRJ-{year}-{count + 1:04d}"
    
    project = Project(**project_data.model_dump())
    project_dict = project.model_dump()
    project_dict["project_code"] = project_code
    project_dict["created_by"] = current_user["id"]
    
    await db.projects.insert_one(project_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_project",
        entity_type="project",
        entity_id=project.id,
        entity_name=project_data.name,
        details=f"مشروع جديد: {project_data.name}"
    )
    
    return Project(**project_dict)

@api_router.get("/projects")
async def get_projects(
    status: Optional[str] = None,
    project_type: Optional[str] = None,
    manager_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if project_type:
        query["project_type"] = project_type
    if manager_id:
        query["manager_id"] = manager_id
    
    projects = await db.projects.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return projects

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str, current_user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project

@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, project_data: ProjectCreate, current_user: dict = Depends(get_current_user)):
    result = await db.projects.update_one(
        {"id": project_id},
        {"$set": project_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_project",
        entity_type="project",
        entity_id=project_id,
        entity_name=project.get("name"),
        details=f"تعديل مشروع: {project.get('name')}"
    )
    
    return project

@api_router.put("/projects/{project_id}/status")
async def update_project_status(project_id: str, status: str, progress_percentage: Optional[float] = None, current_user: dict = Depends(get_current_user)):
    update_data = {"status": status}
    if progress_percentage is not None:
        update_data["progress_percentage"] = progress_percentage
    
    result = await db.projects.update_one(
        {"id": project_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_project_status",
        entity_type="project",
        entity_id=project_id,
        entity_name=project.get("name"),
        details=f"تحديث حالة مشروع: {project.get('name')} - {status}"
    )
    
    return project

# Project Tasks
@api_router.post("/projects/tasks", response_model=ProjectTask)
async def create_project_task(task_data: ProjectTaskCreate, current_user: dict = Depends(get_current_user)):
    task = ProjectTask(**task_data.model_dump())
    await db.project_tasks.insert_one(task.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_project_task",
        entity_type="project_task",
        entity_id=task.id,
        entity_name=task_data.task_name,
        details=f"مهمة جديدة: {task_data.task_name} - {task_data.project_name}"
    )
    
    return task

@api_router.get("/projects/{project_id}/tasks")
async def get_project_tasks(project_id: str, status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"project_id": project_id}
    if status:
        query["status"] = status
    
    tasks = await db.project_tasks.find(query, {"_id": 0}).sort("due_date", 1).to_list(1000)
    return tasks

@api_router.put("/projects/tasks/{task_id}", response_model=ProjectTask)
async def update_project_task(task_id: str, task_data: ProjectTaskCreate, current_user: dict = Depends(get_current_user)):
    result = await db.project_tasks.update_one(
        {"id": task_id},
        {"$set": task_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task = await db.project_tasks.find_one({"id": task_id}, {"_id": 0})
    return task

@api_router.put("/projects/tasks/{task_id}/complete")
async def complete_project_task(task_id: str, actual_hours: float = 0, current_user: dict = Depends(get_current_user)):
    result = await db.project_tasks.update_one(
        {"id": task_id},
        {"$set": {
            "status": "completed",
            "progress_percentage": 100,
            "actual_hours": actual_hours,
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task = await db.project_tasks.find_one({"id": task_id}, {"_id": 0})
    return task

# Project Team Members
@api_router.post("/projects/team", response_model=ProjectTeamMember)
async def add_project_team_member(member_data: ProjectTeamMemberCreate, current_user: dict = Depends(get_current_user)):
    member = ProjectTeamMember(**member_data.model_dump())
    await db.project_team_members.insert_one(member.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="add_project_team_member",
        entity_type="project_team_member",
        entity_id=member.id,
        entity_name=member_data.employee_name,
        details=f"إضافة عضو للمشروع: {member_data.employee_name} - {member_data.project_name}"
    )
    
    return member

@api_router.get("/projects/{project_id}/team")
async def get_project_team(project_id: str, current_user: dict = Depends(get_current_user)):
    members = await db.project_team_members.find({"project_id": project_id, "is_active": True}, {"_id": 0}).to_list(100)
    return members

@api_router.delete("/projects/team/{member_id}")
async def remove_project_team_member(member_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.project_team_members.update_one(
        {"id": member_id},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Team member not found")
    return {"message": "Team member removed"}

# Project Milestones
@api_router.post("/projects/milestones", response_model=ProjectMilestone)
async def create_project_milestone(milestone_data: ProjectMilestoneCreate, current_user: dict = Depends(get_current_user)):
    milestone = ProjectMilestone(**milestone_data.model_dump())
    await db.project_milestones.insert_one(milestone.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_project_milestone",
        entity_type="project_milestone",
        entity_id=milestone.id,
        entity_name=milestone_data.name,
        details=f"مرحلة جديدة: {milestone_data.name} - {milestone_data.project_name}"
    )
    
    return milestone

@api_router.get("/projects/{project_id}/milestones")
async def get_project_milestones(project_id: str, current_user: dict = Depends(get_current_user)):
    milestones = await db.project_milestones.find({"project_id": project_id}, {"_id": 0}).sort("due_date", 1).to_list(100)
    return milestones

@api_router.put("/projects/milestones/{milestone_id}/achieve")
async def achieve_milestone(milestone_id: str, notes: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    result = await db.project_milestones.update_one(
        {"id": milestone_id},
        {"$set": {
            "status": "achieved",
            "achieved_date": datetime.now(timezone.utc).isoformat(),
            "notes": notes
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Milestone not found")
    
    milestone = await db.project_milestones.find_one({"id": milestone_id}, {"_id": 0})
    return milestone

# Projects Dashboard
@api_router.get("/projects/dashboard/stats")
async def get_projects_dashboard(current_user: dict = Depends(get_current_user)):
    total_projects = await db.projects.count_documents({})
    active_projects = await db.projects.count_documents({"status": "in_progress"})
    completed_projects = await db.projects.count_documents({"status": "completed"})
    overdue_tasks = await db.project_tasks.count_documents({
        "status": {"$ne": "completed"},
        "due_date": {"$lt": datetime.now(timezone.utc).isoformat()}
    })
    
    # Get projects with budget
    projects = await db.projects.find({}, {"_id": 0, "budget": 1, "actual_cost": 1}).to_list(1000)
    total_budget = sum(p.get("budget", 0) for p in projects)
    total_actual_cost = sum(p.get("actual_cost", 0) for p in projects)
    
    return {
        "total_projects": total_projects,
        "active_projects": active_projects,
        "completed_projects": completed_projects,
        "overdue_tasks": overdue_tasks,
        "total_budget": total_budget,
        "total_actual_cost": total_actual_cost
    }

# ==================== OPERATIONS MODULE ROUTES ====================
# REFACTORED: Operations routes moved to routes/operations_routes.py
# ================================================================

# ==================== LEAVE BALANCE (رصيد الإجازات) ====================

@api_router.post("/hr/leave-balance/accrue-monthly")
async def accrue_monthly_leave(
    month: Optional[str] = None,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """
    إضافة رصيد الإجازات الشهري لجميع الموظفين (تراكمي)
    المعدلات حسب المنصب:
    - مدير عام: 3.5 يوم
    - نائب المدير: 3.5 يوم
    - مشرف: 3 أيام
    - مدير الموارد البشرية: 3 أيام
    - مسؤول أمن وسلامة: 3 أيام
    - باقي الموظفين: 2.6 يوم
    """
    from models.all_models import LeaveBalanceLog
    
    if not month:
        month = datetime.now().strftime("%Y-%m")
    
    # Check if already accrued for this month
    existing = await db.leave_balance_logs.find_one({"month": month, "reason": "monthly_accrual"})
    if existing:
        return {"message": f"تم إضافة رصيد الإجازات لشهر {month} مسبقاً", "already_processed": True}
    
    # Get all active employees
    employees = await db.hr_employees.find({"is_active": True}, {"_id": 0}).to_list(500)
    
    updated_count = 0
    logs = []
    
    def get_leave_rate_by_position(position):
        """تحديد معدل الإجازات الشهري حسب المنصب"""
        if not position:
            return 2.6
        
        position_lower = position.lower()
        
        # مدير عام - 3.5 يوم
        if "مدير عام" in position or "general manager" in position_lower or "director general" in position_lower:
            return 3.5
        
        # نائب المدير - 3.5 يوم
        if "نائب المدير" in position or "نائب مدير" in position or "deputy" in position_lower:
            return 3.5
        
        # مدير الموارد البشرية - 3 أيام
        if "مدير الموارد البشرية" in position or "hr manager" in position_lower or "موارد بشرية" in position:
            return 3.0
        
        # مسؤول أمن وسلامة - 3 أيام
        if "أمن وسلامة" in position or "أمن" in position or "سلامة" in position or "safety" in position_lower or "security" in position_lower:
            return 3.0
        
        # مشرف - 3 أيام
        if "مشرف" in position or "supervisor" in position_lower:
            return 3.0
        
        # باقي الموظفين - 2.6 يوم
        return 2.6
    
    for emp in employees:
        position = emp.get("position", "")
        
        # Get rate based on position or use custom rate if set
        custom_rate = emp.get("monthly_leave_rate")
        if custom_rate and custom_rate != 2.6:
            rate = custom_rate  # استخدم المعدل المخصص إذا تم تعيينه
        else:
            rate = get_leave_rate_by_position(position)
        
        previous_balance = emp.get("leave_balance", 0)
        new_balance = round(previous_balance + rate, 2)  # تراكمي
        
        # Update employee
        await db.hr_employees.update_one(
            {"id": emp["id"]},
            {"$set": {"leave_balance": new_balance, "monthly_leave_rate": rate}}
        )
        
        # Create log
        log = LeaveBalanceLog(
            employee_id=emp["id"],
            employee_name=emp["name"],
            month=month,
            amount_added=rate,
            previous_balance=previous_balance,
            new_balance=new_balance,
            reason="monthly_accrual"
        )
        logs.append(log.model_dump())
        updated_count += 1
    
    # Insert all logs
    if logs:
        await db.leave_balance_logs.insert_many(logs)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", ""),
        action="accrue_monthly_leave",
        entity_type="leave_balance",
        entity_id=month,
        entity_name=f"إضافة رصيد إجازات شهر {month}",
        details=f"تم إضافة رصيد الإجازات لـ {updated_count} موظف"
    )
    
    return {
        "message": f"تم إضافة رصيد الإجازات لـ {updated_count} موظف",
        "month": month,
        "employees_updated": updated_count
    }

@api_router.get("/hr/leave-balance/rates")
async def get_leave_balance_rates(current_user: dict = Depends(get_current_user)):
    """جلب معدلات الإجازات الشهرية حسب المنصب"""
    return {
        "rates": [
            {"position": "مدير عام", "rate": 3.5, "position_en": "General Manager"},
            {"position": "نائب المدير", "rate": 3.5, "position_en": "Deputy Director"},
            {"position": "مشرف", "rate": 3.0, "position_en": "Supervisor"},
            {"position": "مدير الموارد البشرية", "rate": 3.0, "position_en": "HR Manager"},
            {"position": "مسؤول أمن وسلامة", "rate": 3.0, "position_en": "Safety Officer"},
            {"position": "موظف (افتراضي)", "rate": 2.6, "position_en": "Employee (Default)"}
        ],
        "note": "يتم إضافة الرصيد تلقائياً في نهاية كل شهر بشكل تراكمي"
    }

@api_router.put("/hr/employees/{employee_id}/leave-rate")
async def update_employee_leave_rate(
    employee_id: str,
    data: dict,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """تحديث معدل الإجازة الشهرية للموظف"""
    rate = data.get("monthly_leave_rate", 2.6)
    rate_type = data.get("leave_rate_type", "auto")
    
    result = await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {"monthly_leave_rate": rate, "leave_rate_type": rate_type}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    return {"message": "تم تحديث معدل الإجازة بنجاح", "new_rate": rate, "rate_type": rate_type}

@api_router.put("/hr/employees/{employee_id}/leave-balance")
async def adjust_employee_leave_balance(
    employee_id: str,
    data: dict,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """تعديل رصيد الإجازات للموظف"""
    from models.all_models import LeaveBalanceLog
    
    adjustment = data.get("adjustment", 0)
    reason = data.get("reason", "manual_adjustment")
    
    emp = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    previous_balance = emp.get("leave_balance", 0)
    new_balance = round(previous_balance + adjustment, 2)
    
    await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {"leave_balance": new_balance}}
    )
    
    # Create log
    log = LeaveBalanceLog(
        employee_id=employee_id,
        employee_name=emp["name"],
        month=datetime.now().strftime("%Y-%m"),
        amount_added=adjustment,
        previous_balance=previous_balance,
        new_balance=new_balance,
        reason=reason
    )
    await db.leave_balance_logs.insert_one(log.model_dump())
    
    return {
        "message": "تم تعديل رصيد الإجازات بنجاح",
        "previous_balance": previous_balance,
        "adjustment": adjustment,
        "new_balance": new_balance
    }

@api_router.get("/hr/leave-balance/logs")
async def get_leave_balance_logs(
    employee_id: Optional[str] = None,
    month: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجل تغييرات رصيد الإجازات"""
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if month:
        query["month"] = month
    
    logs = await db.leave_balance_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return logs

# ==================== MARKETING MODULE ROUTES (قسم التسويق) ====================

# Marketing Campaigns
@api_router.post("/marketing/campaigns", response_model=MarketingCampaign)
async def create_marketing_campaign(campaign_data: MarketingCampaignCreate, current_user: dict = Depends(get_current_user)):
    count = await db.marketing_campaigns.count_documents({})
    year = datetime.now().year
    campaign_code = f"CMP-{year}-{count + 1:04d}"
    
    campaign = MarketingCampaign(**campaign_data.model_dump())
    campaign_dict = campaign.model_dump()
    campaign_dict["campaign_code"] = campaign_code
    campaign_dict["created_by"] = current_user["id"]
    
    await db.marketing_campaigns.insert_one(campaign_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_campaign",
        entity_type="marketing_campaign",
        entity_id=campaign.id,
        entity_name=campaign_data.name,
        details=f"حملة تسويقية جديدة: {campaign_data.name}"
    )
    
    return MarketingCampaign(**campaign_dict)

@api_router.get("/marketing/campaigns")
async def get_marketing_campaigns(
    status: Optional[str] = None,
    campaign_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if campaign_type:
        query["campaign_type"] = campaign_type
    
    campaigns = await db.marketing_campaigns.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return campaigns

@api_router.put("/marketing/campaigns/{campaign_id}", response_model=MarketingCampaign)
async def update_marketing_campaign(campaign_id: str, campaign_data: MarketingCampaignCreate, current_user: dict = Depends(get_current_user)):
    result = await db.marketing_campaigns.update_one(
        {"id": campaign_id},
        {"$set": campaign_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    campaign = await db.marketing_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    return campaign

@api_router.put("/marketing/campaigns/{campaign_id}/status")
async def update_campaign_status(campaign_id: str, status: str, current_user: dict = Depends(get_current_user)):
    result = await db.marketing_campaigns.update_one(
        {"id": campaign_id},
        {"$set": {"status": status}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    campaign = await db.marketing_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    return campaign

# Leads (العملاء المحتملين)
@api_router.post("/marketing/leads", response_model=Lead)
async def create_lead(lead_data: LeadCreate, current_user: dict = Depends(get_current_user)):
    count = await db.marketing_leads.count_documents({})
    lead_code = f"LEAD-{count + 1:05d}"
    
    lead = Lead(**lead_data.model_dump())
    lead_dict = lead.model_dump()
    lead_dict["lead_code"] = lead_code
    lead_dict["created_by"] = current_user["id"]
    
    await db.marketing_leads.insert_one(lead_dict)
    
    # Increment campaign leads if associated
    if lead_data.campaign_id:
        await db.marketing_campaigns.update_one(
            {"id": lead_data.campaign_id},
            {"$inc": {"leads_generated": 1}}
        )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_lead",
        entity_type="lead",
        entity_id=lead.id,
        entity_name=lead_data.name,
        details=f"عميل محتمل جديد: {lead_data.name}"
    )
    
    return Lead(**lead_dict)

@api_router.get("/marketing/leads")
async def get_leads(
    status: Optional[str] = None,
    lead_source: Optional[str] = None,
    assigned_to_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if lead_source:
        query["lead_source"] = lead_source
    if assigned_to_id:
        query["assigned_to_id"] = assigned_to_id
    
    leads = await db.marketing_leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return leads

@api_router.put("/marketing/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, lead_data: LeadCreate, current_user: dict = Depends(get_current_user)):
    result = await db.marketing_leads.update_one(
        {"id": lead_id},
        {"$set": lead_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    lead = await db.marketing_leads.find_one({"id": lead_id}, {"_id": 0})
    return lead

@api_router.put("/marketing/leads/{lead_id}/status")
async def update_lead_status(lead_id: str, status: str, notes: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    update_data = {"status": status, "last_contact_date": datetime.now(timezone.utc).isoformat()}
    
    if status == "won":
        update_data["conversion_date"] = datetime.now(timezone.utc).isoformat()
        # Update campaign conversions
        lead = await db.marketing_leads.find_one({"id": lead_id})
        if lead and lead.get("campaign_id"):
            await db.marketing_campaigns.update_one(
                {"id": lead["campaign_id"]},
                {"$inc": {"conversions": 1}}
            )
    elif status == "lost" and notes:
        update_data["lost_reason"] = notes
    
    result = await db.marketing_leads.update_one(
        {"id": lead_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    lead = await db.marketing_leads.find_one({"id": lead_id}, {"_id": 0})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_lead_status",
        entity_type="lead",
        entity_id=lead_id,
        entity_name=lead.get("name"),
        details=f"تحديث حالة عميل محتمل: {lead.get('name')} - {status}"
    )
    
    return lead

# Social Media Posts
@api_router.post("/marketing/social-posts", response_model=SocialMediaPost)
async def create_social_post(post_data: SocialMediaPostCreate, current_user: dict = Depends(get_current_user)):
    post = SocialMediaPost(**post_data.model_dump())
    post_dict = post.model_dump()
    post_dict["created_by"] = current_user["id"]
    
    await db.social_media_posts.insert_one(post_dict)
    
    return SocialMediaPost(**post_dict)

@api_router.get("/marketing/social-posts")
async def get_social_posts(
    platform: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if platform:
        query["platform"] = platform
    if status:
        query["status"] = status
    
    posts = await db.social_media_posts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return posts

@api_router.put("/marketing/social-posts/{post_id}/publish")
async def publish_social_post(post_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.social_media_posts.update_one(
        {"id": post_id},
        {"$set": {
            "status": "published",
            "published_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Post not found")
    
    post = await db.social_media_posts.find_one({"id": post_id}, {"_id": 0})
    return post

# Sales Offers
@api_router.post("/marketing/offers", response_model=SalesOffer)
async def create_sales_offer(offer_data: SalesOfferCreate, current_user: dict = Depends(get_current_user)):
    count = await db.sales_offers.count_documents({})
    offer_code = f"OFFER-{count + 1:04d}"
    
    offer = SalesOffer(**offer_data.model_dump())
    offer_dict = offer.model_dump()
    offer_dict["offer_code"] = offer_code
    offer_dict["created_by"] = current_user["id"]
    
    await db.sales_offers.insert_one(offer_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_offer",
        entity_type="sales_offer",
        entity_id=offer.id,
        entity_name=offer_data.title,
        details=f"عرض مبيعات جديد: {offer_data.title}"
    )
    
    return SalesOffer(**offer_dict)

@api_router.get("/marketing/offers")
async def get_sales_offers(
    status: Optional[str] = None,
    offer_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if offer_type:
        query["offer_type"] = offer_type
    
    offers = await db.sales_offers.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return offers

@api_router.put("/marketing/offers/{offer_id}/activate")
async def activate_offer(offer_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.sales_offers.update_one(
        {"id": offer_id},
        {"$set": {"status": "active"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Offer not found")
    
    offer = await db.sales_offers.find_one({"id": offer_id}, {"_id": 0})
    return offer

# Market Returns (مرتجعات السوق)
@api_router.post("/marketing/returns", response_model=MarketReturn)
async def create_market_return(return_data: MarketReturnCreate, current_user: dict = Depends(get_current_user)):
    count = await db.market_returns.count_documents({})
    year = datetime.now().year
    return_code = f"RTN-{year}-{count + 1:04d}"
    
    market_return = MarketReturn(**return_data.model_dump())
    return_dict = market_return.model_dump()
    return_dict["return_code"] = return_code
    return_dict["created_by"] = current_user["id"]
    
    await db.market_returns.insert_one(return_dict)
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_market_return",
        entity_type="market_return",
        entity_id=market_return.id,
        entity_name=return_data.customer_name,
        details=f"مرتجع سوق: {return_data.quantity_liters} لتر من {return_data.customer_name}"
    )
    
    return MarketReturn(**return_dict)

@api_router.get("/marketing/returns")
async def get_market_returns(
    status: Optional[str] = None,
    center_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if center_id:
        query["center_id"] = center_id
    if start_date:
        query["return_date"] = {"$gte": start_date}
    if end_date:
        if "return_date" in query:
            query["return_date"]["$lte"] = end_date
        else:
            query["return_date"] = {"$lte": end_date}
    
    returns = await db.market_returns.find(query, {"_id": 0}).sort("return_date", -1).to_list(1000)
    return returns

@api_router.put("/marketing/returns/{return_id}/approve")
async def approve_market_return(return_id: str, disposal_method: str, current_user: dict = Depends(get_current_user)):
    result = await db.market_returns.update_one(
        {"id": return_id},
        {"$set": {
            "status": "approved",
            "disposal_method": disposal_method,
            "approved_by": current_user["full_name"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Return not found")
    
    market_return = await db.market_returns.find_one({"id": return_id}, {"_id": 0})
    return market_return

# Market Sales Summary
@api_router.post("/marketing/sales-summary", response_model=MarketSalesSummary)
async def create_market_sales_summary(summary_data: MarketSalesSummaryCreate, current_user: dict = Depends(get_current_user)):
    summary = MarketSalesSummary(**summary_data.model_dump())
    summary_dict = summary.model_dump()
    summary_dict["created_by"] = current_user["id"]
    
    # Calculate net values
    summary_dict["net_quantity"] = summary_data.total_quantity_sold - summary_data.total_returns
    summary_dict["net_revenue"] = summary_data.total_revenue - (summary_data.total_returns * 0.5)  # Adjust based on return policy
    
    await db.market_sales_summaries.insert_one(summary_dict)
    
    return MarketSalesSummary(**summary_dict)

@api_router.get("/marketing/sales-summary")
async def get_market_sales_summaries(
    center_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if center_id:
        query["center_id"] = center_id
    if start_date:
        query["report_date"] = {"$gte": start_date}
    if end_date:
        if "report_date" in query:
            query["report_date"]["$lte"] = end_date
        else:
            query["report_date"] = {"$lte": end_date}
    
    summaries = await db.market_sales_summaries.find(query, {"_id": 0}).sort("report_date", -1).to_list(1000)
    return summaries

# Marketing Dashboard
@api_router.get("/marketing/dashboard")
async def get_marketing_dashboard(current_user: dict = Depends(get_current_user)):
    # Campaigns stats
    active_campaigns = await db.marketing_campaigns.count_documents({"status": "active"})
    total_campaigns = await db.marketing_campaigns.count_documents({})
    
    # Leads stats
    total_leads = await db.marketing_leads.count_documents({})
    new_leads = await db.marketing_leads.count_documents({"status": "new"})
    qualified_leads = await db.marketing_leads.count_documents({"status": "qualified"})
    converted_leads = await db.marketing_leads.count_documents({"status": "won"})
    
    # Calculate conversion rate
    conversion_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0
    
    # Active offers
    active_offers = await db.sales_offers.count_documents({"status": "active"})
    
    # Returns stats - this month
    this_month_start = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    monthly_returns = await db.market_returns.find(
        {"return_date": {"$gte": this_month_start}},
        {"_id": 0, "quantity_liters": 1, "refund_amount": 1}
    ).to_list(1000)
    
    total_return_quantity = sum(r.get("quantity_liters", 0) for r in monthly_returns)
    total_refund_amount = sum(r.get("refund_amount", 0) or 0 for r in monthly_returns)
    
    # Campaign budget vs actual
    campaigns = await db.marketing_campaigns.find({}, {"_id": 0, "budget": 1, "actual_cost": 1}).to_list(1000)
    total_budget = sum(c.get("budget", 0) for c in campaigns)
    total_actual_cost = sum(c.get("actual_cost", 0) for c in campaigns)
    
    return {
        "campaigns": {
            "total": total_campaigns,
            "active": active_campaigns,
            "total_budget": total_budget,
            "actual_cost": total_actual_cost
        },
        "leads": {
            "total": total_leads,
            "new": new_leads,
            "qualified": qualified_leads,
            "converted": converted_leads,
            "conversion_rate": round(conversion_rate, 2)
        },
        "offers": {
            "active": active_offers
        },
        "returns": {
            "monthly_quantity": total_return_quantity,
            "monthly_refund": total_refund_amount
        }
    }

# ==================== CENTRAL DASHBOARD (لوحة التحكم المركزية) ====================

@api_router.get("/dashboard/central")
async def get_central_dashboard(current_user: dict = Depends(get_current_user)):
    """Central dashboard showing data from all centers"""
    
    # Get all centers
    centers = await db.collection_centers.find({"is_active": True}, {"_id": 0}).to_list(100)
    
    # Today's date
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    this_month_start = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    
    center_stats = []
    total_milk_today = 0
    total_milk_month = 0
    total_sales_today = 0
    total_suppliers = 0
    
    for center in centers:
        center_id = center["id"]
        
        # Today's milk reception for this center
        today_milk = await db.milk_receptions.find(
            {"center_id": center_id, "reception_date": {"$regex": f"^{today}"}},
            {"_id": 0, "quantity_liters": 1, "total_amount": 1}
        ).to_list(1000)
        
        center_milk_today = sum(m.get("quantity_liters", 0) for m in today_milk)
        center_amount_today = sum(m.get("total_amount", 0) for m in today_milk)
        
        # Monthly milk for this center
        monthly_milk = await db.milk_receptions.find(
            {"center_id": center_id, "reception_date": {"$gte": this_month_start}},
            {"_id": 0, "quantity_liters": 1}
        ).to_list(10000)
        
        center_milk_month = sum(m.get("quantity_liters", 0) for m in monthly_milk)
        
        # Suppliers count for this center
        center_suppliers = await db.suppliers.count_documents({"center_id": center_id, "is_active": True})
        
        center_stats.append({
            "center_id": center_id,
            "center_name": center["name"],
            "center_code": center.get("code", ""),
            "today_milk_liters": center_milk_today,
            "today_amount": center_amount_today,
            "monthly_milk_liters": center_milk_month,
            "suppliers_count": center_suppliers
        })
        
        total_milk_today += center_milk_today
        total_milk_month += center_milk_month
        total_suppliers += center_suppliers
    
    # Total sales today
    today_sales = await db.sales.find(
        {"sale_date": {"$regex": f"^{today}"}},
        {"_id": 0, "total_amount": 1, "quantity_liters": 1}
    ).to_list(1000)
    
    total_sales_amount = sum(s.get("total_amount", 0) for s in today_sales)
    total_sales_liters = sum(s.get("quantity_liters", 0) for s in today_sales)
    
    # Inventory status
    inventory = await db.inventory.find_one({"product_type": "raw_milk"}, {"_id": 0})
    current_stock = inventory.get("quantity_liters", 0) if inventory else 0
    
    # HR stats
    total_employees = await db.hr_employees.count_documents({"is_active": True})
    present_today = await db.hr_attendance.count_documents({"date": today, "check_in": {"$ne": None}})
    
    # Pending approvals
    pending_leaves = await db.hr_leave_requests.count_documents({"status": "pending"})
    pending_expenses = await db.hr_expense_requests.count_documents({"status": "pending"})
    
    # Financial summary
    monthly_payments = await db.payments.find(
        {"payment_date": {"$gte": this_month_start}},
        {"_id": 0, "amount": 1, "payment_type": 1}
    ).to_list(10000)
    
    supplier_payments = sum(p.get("amount", 0) for p in monthly_payments if p.get("payment_type") == "supplier_payment")
    customer_receipts = sum(p.get("amount", 0) for p in monthly_payments if p.get("payment_type") == "customer_receipt")
    
    return {
        "summary": {
            "total_centers": len(centers),
            "total_suppliers": total_suppliers,
            "total_employees": total_employees,
            "present_today": present_today
        },
        "milk": {
            "today_liters": total_milk_today,
            "monthly_liters": total_milk_month,
            "current_stock": current_stock
        },
        "sales": {
            "today_liters": total_sales_liters,
            "today_amount": total_sales_amount
        },
        "financial": {
            "monthly_supplier_payments": supplier_payments,
            "monthly_customer_receipts": customer_receipts
        },
        "pending_approvals": {
            "leaves": pending_leaves,
            "expenses": pending_expenses
        },
        "centers": center_stats
    }

# ==================== ATTENDANCE IMPORT FROM EXCEL ====================

@api_router.post("/hr/attendance/import")
async def import_attendance_from_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import attendance records from Excel file"""
    import openpyxl
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        # Expected columns: employee_code, date, check_in, check_out
        headers = [cell.value for cell in ws[1]]
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            try:
                employee_code = str(row[0]) if row[0] else None
                date = str(row[1]) if row[1] else None
                check_in = str(row[2]) if row[2] else None
                check_out = str(row[3]) if row[3] else None
                
                if not employee_code or not date:
                    errors.append(f"Row {row_num}: Missing employee code or date")
                    continue
                
                # Find employee
                employee = await db.hr_employees.find_one(
                    {"$or": [{"employee_code": employee_code}, {"id": employee_code}]},
                    {"_id": 0}
                )
                
                if not employee:
                    errors.append(f"Row {row_num}: Employee {employee_code} not found")
                    continue
                
                # Check if attendance already exists
                existing = await db.hr_attendance.find_one({
                    "employee_id": employee["id"],
                    "date": date
                })
                
                if existing:
                    # Update existing record
                    update_data = {}
                    if check_in:
                        update_data["check_in"] = check_in
                    if check_out:
                        update_data["check_out"] = check_out
                    update_data["source"] = "excel_import"
                    
                    await db.hr_attendance.update_one(
                        {"id": existing["id"]},
                        {"$set": update_data}
                    )
                else:
                    # Create new record
                    attendance = {
                        "id": str(uuid.uuid4()),
                        "employee_id": employee["id"],
                        "employee_name": employee["name"],
                        "date": date,
                        "check_in": check_in,
                        "check_out": check_out,
                        "source": "excel_import",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.hr_attendance.insert_one(attendance)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user["full_name"],
            action="import_attendance",
            details=f"استيراد {imported_count} سجل حضور من Excel"
        )
        
        return {
            "success": True,
            "imported_count": imported_count,
            "errors": errors[:10] if errors else [],  # Return first 10 errors
            "total_errors": len(errors)
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

# ==================== PAYROLL ROUTES (الرواتب) ====================

@api_router.post("/hr/payroll/periods")
async def create_payroll_period(
    name: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """Create a new payroll period (e.g., 16 Nov - 16 Dec for 31 days)
    
    The period is automatically adjusted to ensure 31 working days:
    - From 16th to 15th of next month = 30 days → adjusted to 16th
    - From 16th to 16th of next month = 31 days → kept as is
    """
    from datetime import datetime as dt
    
    start = dt.strptime(start_date, "%Y-%m-%d")
    end = dt.strptime(end_date, "%Y-%m-%d")
    total_days = (end - start).days + 1
    
    # Auto-adjust if less than 31 days
    if total_days < 31:
        from datetime import timedelta
        end = start + timedelta(days=30)  # 31 days including start
        end_date = end.strftime("%Y-%m-%d")
        total_days = 31
    
    period = PayrollPeriod(
        name=name,
        start_date=start_date,
        end_date=end_date,
        total_days=total_days
    )
    
    await db.payroll_periods.insert_one(period.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_payroll_period",
        entity_type="payroll",
        entity_id=period.id,
        entity_name=name,
        details=f"إنشاء فترة رواتب: {name} ({total_days} يوم)"
    )
    
    return period.model_dump()

@api_router.get("/hr/payroll/periods")
async def get_payroll_periods(current_user: dict = Depends(get_current_user)):
    """Get all payroll periods"""
    periods = await db.payroll_periods.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return periods

@api_router.get("/hr/payroll/periods/{period_id}")
async def get_payroll_period(period_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific payroll period with its records"""
    period = await db.payroll_periods.find_one({"id": period_id}, {"_id": 0})
    if not period:
        raise HTTPException(status_code=404, detail="Payroll period not found")
    
    records = await db.payroll_records.find({"period_id": period_id}, {"_id": 0}).to_list(1000)
    
    return {
        "period": period,
        "records": records,
        "summary": {
            "total_employees": len(records),
            "total_gross": sum(r.get("gross_salary", 0) for r in records),
            "total_deductions": sum(r.get("deductions", 0) for r in records),
            "total_net": sum(r.get("net_salary", 0) for r in records)
        }
    }

@api_router.post("/hr/payroll/periods/{period_id}/calculate")
async def calculate_payroll(period_id: str, current_user: dict = Depends(get_current_user)):
    """Calculate payroll for all employees based on attendance with holiday and weekend logic"""
    period = await db.payroll_periods.find_one({"id": period_id}, {"_id": 0})
    if not period:
        raise HTTPException(status_code=404, detail="Payroll period not found")
    
    # Get all active employees (exclude those marked as exclude_from_payroll)
    all_employees = await db.hr_employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    employees = [e for e in all_employees if not e.get("exclude_from_payroll", False)]
    
    # Get attendance records for the period
    start_date = period["start_date"]
    end_date = period["end_date"]
    
    attendance_records = await db.hr_attendance.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(10000)
    
    # Get official holidays for the period (unified source: hr_official_holidays)
    official_holidays = await db.hr_official_holidays.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(100)
    
    # Use hr_official_holidays as the single source of truth for holidays
    holiday_dates = {h["date"] for h in official_holidays}
    
    # Get all salary structures for efficient lookup
    salary_structures = await db.employee_salary_structures.find(
        {"is_active": True}, {"_id": 0}
    ).to_list(1000)
    salary_structure_map = {s["employee_id"]: s for s in salary_structures}
    
    # Delete existing payroll records for this period
    await db.payroll_records.delete_many({"period_id": period_id})
    
    payroll_records = []
    
    # Constants for overtime calculation
    STANDARD_HOURS_PER_DAY = 8
    OVERTIME_MULTIPLIER = 1.5
    WEEKEND_WORK_MULTIPLIER = 1.5  # مضاعف العمل في أيام الإجازة الأسبوعية
    HOLIDAY_WORK_MULTIPLIER = 2.0   # مضاعف العمل في العطل الرسمية
    
    from datetime import datetime as dt, timedelta
    
    for emp in employees:
        # Get employee's weekly off days (default Friday & Saturday)
        employee_weekly_off_days = emp.get("weekly_off_days", [4, 5])  # 4=Friday, 5=Saturday
        
        # Get salary structure for this employee
        salary_struct = salary_structure_map.get(emp.get("id"), {})
        
        # Filter attendance for this employee
        emp_attendance = [a for a in attendance_records if a.get("employee_id") == emp.get("id") or a.get("employee_name") == emp.get("name")]
        
        # Create a set of dates with attendance records for quick lookup
        attendance_by_date = {a["date"]: a for a in emp_attendance}
        
        # Count attendance types
        working_days = 0
        day_off = 0
        sick_leave = 0
        compensation_leave = 0
        annual_leave = 0
        public_holiday = 0
        emergency_leave = 0
        on_duty = 0
        exam_leave = 0
        father_leave = 0
        accompanying_leave = 0
        absent_days = 0
        unpaid_leave = 0
        otp_days = 0
        weekend_work_days = 0  # أيام عمل في الإجازة الأسبوعية
        holiday_work_days = 0  # أيام عمل في العطل الرسمية
        
        # Calculate overtime hours from attendance records
        total_overtime_hours = 0.0
        weekend_work_pay = 0.0
        holiday_work_pay = 0.0
        
        # Iterate through each day in the period
        current_date = dt.strptime(start_date, "%Y-%m-%d")
        end_date_dt = dt.strptime(end_date, "%Y-%m-%d")
        
        while current_date <= end_date_dt:
            date_str = current_date.strftime("%Y-%m-%d")
            day_of_week = current_date.weekday()  # 0=Monday, 6=Sunday
            
            is_official_holiday = date_str in holiday_dates
            is_weekly_off = day_of_week in employee_weekly_off_days
            
            attendance = attendance_by_date.get(date_str)
            
            if attendance:
                status = attendance.get("status", "present")
                overtime_hrs = attendance.get("overtime_hours", 0) or 0
                
                if status == "present":
                    # Check if working on a holiday or weekend
                    if is_official_holiday:
                        holiday_work_days += 1
                    elif is_weekly_off:
                        weekend_work_days += 1
                    else:
                        working_days += 1
                    total_overtime_hours += overtime_hrs
                    
                elif status in ["off", "weekend"]:
                    day_off += 1
                elif status == "sick_leave":
                    sick_leave += 1
                elif status == "compensation_leave":
                    compensation_leave += 1
                elif status == "annual_leave":
                    annual_leave += 1
                elif status == "public_holiday":
                    public_holiday += 1
                elif status == "emergency_leave":
                    emergency_leave += 1
                elif status == "on_duty":
                    on_duty += 1
                elif status == "exam_leave":
                    exam_leave += 1
                elif status == "father_leave":
                    father_leave += 1
                elif status == "accompanying_leave":
                    accompanying_leave += 1
                elif status == "absent":
                    absent_days += 1
                elif status == "unpaid_leave":
                    unpaid_leave += 1
                elif status == "otp":
                    otp_days += 1
            else:
                # No attendance record - check if it's a holiday/weekend
                if is_official_holiday:
                    public_holiday += 1
                elif is_weekly_off:
                    day_off += 1
                else:
                    absent_days += 1  # No record = absent
            
            current_date += timedelta(days=1)
        
        # Get salary components from structure or employee record
        basic_salary = salary_struct.get("basic_salary", emp.get("salary", 0))
        
        # Get detailed allowances
        allowances_data = salary_struct.get("allowances", {})
        housing_allowance = allowances_data.get("housing_allowance", 0) if isinstance(allowances_data, dict) else 0
        transportation_allowance = allowances_data.get("transportation_allowance", 0) if isinstance(allowances_data, dict) else 0
        food_allowance = allowances_data.get("food_allowance", 0) if isinstance(allowances_data, dict) else 0
        phone_allowance = allowances_data.get("phone_allowance", 0) if isinstance(allowances_data, dict) else 0
        fuel_allowance = allowances_data.get("fuel_allowance", 0) if isinstance(allowances_data, dict) else 0
        education_allowance = allowances_data.get("education_allowance", 0) if isinstance(allowances_data, dict) else 0
        medical_allowance = allowances_data.get("medical_allowance", 0) if isinstance(allowances_data, dict) else 0
        special_allowance = allowances_data.get("special_allowance", 0) if isinstance(allowances_data, dict) else 0
        other_allowance = allowances_data.get("other_allowance", 0) if isinstance(allowances_data, dict) else 0
        
        total_allowances = (
            housing_allowance + transportation_allowance + food_allowance +
            phone_allowance + fuel_allowance + education_allowance +
            medical_allowance + special_allowance + other_allowance
        )
        
        # Calculate rates - use 31 days as per company policy (16th to 15th = ~31 days)
        total_monthly_salary = basic_salary + total_allowances
        daily_rate = total_monthly_salary / 31 if total_monthly_salary > 0 else 0
        hourly_rate = daily_rate / STANDARD_HOURS_PER_DAY if daily_rate > 0 else 0
        
        # Total pay days = working + all paid leaves
        total_pay_days = (
            working_days + day_off + sick_leave + compensation_leave + annual_leave + 
            public_holiday + emergency_leave + on_duty + exam_leave + 
            father_leave + accompanying_leave + weekend_work_days + holiday_work_days
        )
        
        # Gross salary for regular work
        gross_salary = daily_rate * total_pay_days
        
        # Weekend work pay (1.5x)
        weekend_work_pay = daily_rate * weekend_work_days * (WEEKEND_WORK_MULTIPLIER - 1)  # Extra pay only
        
        # Holiday work pay (2x)
        holiday_work_pay = daily_rate * holiday_work_days * (HOLIDAY_WORK_MULTIPLIER - 1)  # Extra pay only
        
        # Overtime pay (1.5x hourly rate)
        overtime_pay = total_overtime_hours * hourly_rate * OVERTIME_MULTIPLIER
        
        # Get active loans for this employee to calculate loan deduction
        employee_loans = await db.hr_loans.find({
            "employee_id": emp.get("id"),
            "status": "approved"
        }, {"_id": 0}).to_list(100)
        
        loan_deduction = 0.0
        for loan in employee_loans:
            remaining = loan.get("remaining_amount", loan.get("amount", 0))
            if remaining > 0:
                installment = loan.get("installment_amount", 0)
                if installment > 0:
                    loan_deduction += installment
        
        # Deductions
        absence_deduction = daily_rate * absent_days
        unpaid_deduction = daily_rate * unpaid_leave
        otp_deduction = daily_rate * otp_days * 0.5  # Half day deduction for OTP issues
        total_deductions = absence_deduction + unpaid_deduction + otp_deduction + loan_deduction
        
        # Net salary
        net_salary = gross_salary + overtime_pay + weekend_work_pay + holiday_work_pay - total_deductions
        
        record = PayrollRecord(
            period_id=period_id,
            employee_id=emp.get("id"),
            employee_name=emp.get("name"),
            employee_code=emp.get("employee_code") or emp.get("employee_id"),
            department=emp.get("department"),
            position=emp.get("position"),
            work_location=emp.get("work_location"),
            nationality=emp.get("nationality"),
            working_days=working_days,
            day_off=day_off,
            sick_leave=sick_leave,
            compensation_leave=compensation_leave,
            annual_leave=annual_leave,
            public_holiday=public_holiday,
            emergency_leave=emergency_leave,
            on_duty=on_duty,
            exam_leave=exam_leave,
            father_leave=father_leave,
            accompanying_leave=accompanying_leave,
            absent_days=absent_days,
            unpaid_leave=unpaid_leave,
            otp_days=otp_days,
            total_overtime_hours=round(total_overtime_hours, 2),
            basic_salary=round(basic_salary, 3),
            daily_rate=round(daily_rate, 3),
            hourly_rate=round(hourly_rate, 3),
            total_pay_days=total_pay_days,
            # Detailed allowances
            housing_allowance=round(housing_allowance, 3),
            transportation_allowance=round(transportation_allowance, 3),
            food_allowance=round(food_allowance, 3),
            phone_allowance=round(phone_allowance, 3),
            fuel_allowance=round(fuel_allowance, 3),
            education_allowance=round(education_allowance, 3),
            medical_allowance=round(medical_allowance, 3),
            special_allowance=round(special_allowance, 3),
            other_allowance=round(other_allowance, 3),
            total_allowances=round(total_allowances, 3),
            allowances=round(total_allowances, 3),  # Legacy field
            gross_salary=round(gross_salary, 3),
            overtime_pay=round(overtime_pay, 3),
            absence_deduction=round(absence_deduction, 3),
            loan_deduction=round(loan_deduction, 3),
            other_deduction=round(otp_deduction, 3),
            total_deductions=round(total_deductions, 3),
            deductions=round(total_deductions, 3),  # Legacy field
            net_salary=round(net_salary, 3)
        )
        
        # Add extra fields for weekend/holiday work
        record_dict = record.model_dump()
        record_dict["weekend_work_days"] = weekend_work_days
        record_dict["weekend_work_pay"] = round(weekend_work_pay, 3)
        record_dict["holiday_work_days"] = holiday_work_days
        record_dict["holiday_work_pay"] = round(holiday_work_pay, 3)
        record_dict["unpaid_deduction"] = round(unpaid_deduction, 3)
        
        await db.payroll_records.insert_one(record_dict)
        payroll_records.append(record_dict)
    
    # Update period status
    await db.payroll_periods.update_one(
        {"id": period_id},
        {"$set": {
            "status": "calculated",
            "calculated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="calculate_payroll",
        entity_type="payroll",
        entity_id=period_id,
        entity_name=period["name"],
        details=f"حساب رواتب {len(payroll_records)} موظف"
    )
    
    return {
        "message": f"تم حساب رواتب {len(payroll_records)} موظف",
        "period_id": period_id,
        "records_count": len(payroll_records)
    }

@api_router.get("/hr/payroll/records")
async def get_payroll_records(
    period_id: Optional[str] = None,
    employee_id: Optional[str] = None,
    work_location: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get payroll records with optional filters"""
    query = {}
    if period_id:
        query["period_id"] = period_id
    if employee_id:
        query["employee_id"] = employee_id
    if work_location:
        query["work_location"] = work_location
    
    records = await db.payroll_records.find(query, {"_id": 0}).to_list(1000)
    return records

@api_router.get("/hr/payroll/by-location/{period_id}")
async def get_payroll_by_location(period_id: str, current_user: dict = Depends(get_current_user)):
    """Get payroll records grouped by work location"""
    records = await db.payroll_records.find({"period_id": period_id}, {"_id": 0}).to_list(1000)
    
    # Group by location
    locations = {}
    for record in records:
        loc = record.get("work_location") or "غير محدد"
        if loc not in locations:
            locations[loc] = {
                "location": loc,
                "employees": [],
                "total_employees": 0,
                "total_working_days": 0,
                "total_overtime_hours": 0,
                "total_gross_salary": 0,
                "total_overtime_pay": 0,
                "total_deductions": 0,
                "total_net_salary": 0
            }
        locations[loc]["employees"].append(record)
        locations[loc]["total_employees"] += 1
        locations[loc]["total_working_days"] += record.get("working_days", 0)
        locations[loc]["total_overtime_hours"] += record.get("total_overtime_hours", 0)
        locations[loc]["total_gross_salary"] += record.get("gross_salary", 0)
        locations[loc]["total_overtime_pay"] += record.get("overtime_pay", 0)
        locations[loc]["total_deductions"] += record.get("deductions", 0)
        locations[loc]["total_net_salary"] += record.get("net_salary", 0)
    
    # Round totals
    for loc in locations.values():
        loc["total_gross_salary"] = round(loc["total_gross_salary"], 3)
        loc["total_overtime_pay"] = round(loc["total_overtime_pay"], 3)
        loc["total_deductions"] = round(loc["total_deductions"], 3)
        loc["total_net_salary"] = round(loc["total_net_salary"], 3)
        loc["total_overtime_hours"] = round(loc["total_overtime_hours"], 2)
    
    return {
        "period_id": period_id,
        "locations": list(locations.values()),
        "grand_total": {
            "employees": len(records),
            "gross_salary": round(sum(r.get("gross_salary", 0) for r in records), 3),
            "overtime_pay": round(sum(r.get("overtime_pay", 0) for r in records), 3),
            "deductions": round(sum(r.get("deductions", 0) for r in records), 3),
            "net_salary": round(sum(r.get("net_salary", 0) for r in records), 3)
        }
    }

# Work locations list
WORK_LOCATIONS = ["حجيف", "غدو", "زيك", "الإدارة", "ثمريت", "طاقة", "مرباط"]

@api_router.get("/hr/work-locations")
async def get_work_locations(current_user: dict = Depends(get_current_user)):
    """Get list of available work locations"""
    return {"locations": WORK_LOCATIONS}

@api_router.put("/hr/payroll/records/{record_id}")
async def update_payroll_record(
    record_id: str,
    deductions: Optional[float] = Form(None),
    overtime_pay: Optional[float] = Form(None),
    allowances: Optional[float] = Form(None),
    notes: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Update a payroll record with manual adjustments"""
    record = await db.payroll_records.find_one({"id": record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Payroll record not found")
    
    update_data = {}
    if deductions is not None:
        update_data["deductions"] = deductions
    if overtime_pay is not None:
        update_data["overtime_pay"] = overtime_pay
    if allowances is not None:
        update_data["allowances"] = allowances
    if notes is not None:
        update_data["notes"] = notes
    
    if update_data:
        # Recalculate net salary
        gross = record.get("gross_salary", 0)
        ded = update_data.get("deductions", record.get("deductions", 0))
        ot = update_data.get("overtime_pay", record.get("overtime_pay", 0))
        allow = update_data.get("allowances", record.get("allowances", 0))
        update_data["net_salary"] = round(gross - ded + ot + allow, 3)
        
        await db.payroll_records.update_one(
            {"id": record_id},
            {"$set": update_data}
        )
    
    updated_record = await db.payroll_records.find_one({"id": record_id}, {"_id": 0})
    return updated_record

@api_router.post("/hr/payroll/periods/{period_id}/approve")
async def approve_payroll(period_id: str, current_user: dict = Depends(get_current_user)):
    """Approve a payroll period - Sequential approval (HR -> Finance -> GM)"""
    period = await db.payroll_periods.find_one({"id": period_id}, {"_id": 0})
    if not period:
        raise HTTPException(status_code=404, detail="Payroll period not found")
    
    current_status = period.get("status", "draft")
    user_role = current_user.get("role", "")
    user_department = current_user.get("department", "")
    user_position = current_user.get("position", "")
    user_id = current_user.get("id", "")
    user_name = current_user.get("full_name", "")
    now = datetime.now(timezone.utc).isoformat()
    
    # Check permissions and determine next status
    new_status = None
    update_fields = {}
    approval_message = ""
    
    # مرحلة 1: موافقة الموارد البشرية
    if current_status in ["draft", "calculated", "pending_hr"]:
        # Check if user is HR
        is_hr = (user_department in ["الموارد البشرية", "hr", "HR"] or 
                 user_role in ["hr_manager", "admin"] or
                 "الموارد البشرية" in user_position or "HR" in user_position.upper())
        
        if is_hr or user_role == "admin":
            new_status = "pending_finance"
            update_fields = {
                "status": new_status,
                "hr_approved_at": now,
                "hr_approved_by": user_id,
                "hr_approved_by_name": user_name
            }
            approval_message = "تمت موافقة الموارد البشرية - بانتظار موافقة المالية"
        else:
            raise HTTPException(status_code=403, detail="يجب موافقة الموارد البشرية أولاً")
    
    # مرحلة 2: موافقة المالية
    elif current_status == "pending_finance":
        is_finance = (user_department in ["المالية", "finance", "Finance"] or
                      user_role in ["finance_manager", "admin"] or
                      "المالية" in user_position or "Finance" in user_position)
        
        if is_finance or user_role == "admin":
            new_status = "pending_gm"
            update_fields = {
                "status": new_status,
                "finance_approved_at": now,
                "finance_approved_by": user_id,
                "finance_approved_by_name": user_name
            }
            approval_message = "تمت موافقة المالية - بانتظار موافقة المدير العام"
        else:
            raise HTTPException(status_code=403, detail="هذه المرحلة تتطلب موافقة المالية")
    
    # مرحلة 3: موافقة المدير العام (الموافقة النهائية)
    elif current_status == "pending_gm":
        is_gm = ("المدير العام" in user_position or 
                 "نائب المدير العام" in user_position or
                 "General Manager" in user_position or
                 user_role == "admin")
        
        if is_gm or user_role == "admin":
            new_status = "approved"
            update_fields = {
                "status": new_status,
                "gm_approved_at": now,
                "gm_approved_by": user_id,
                "gm_approved_by_name": user_name,
                # Legacy fields for backward compatibility
                "approved_at": now,
                "approved_by": user_name
            }
            approval_message = "تمت الموافقة النهائية من المدير العام - كشف الرواتب جاهز للصرف"
        else:
            raise HTTPException(status_code=403, detail="هذه المرحلة تتطلب موافقة المدير العام")
    
    elif current_status == "approved":
        raise HTTPException(status_code=400, detail="كشف الرواتب معتمد مسبقاً")
    
    elif current_status == "disbursed":
        raise HTTPException(status_code=400, detail="تم صرف كشف الرواتب مسبقاً")
    
    else:
        raise HTTPException(status_code=400, detail=f"حالة غير صالحة: {current_status}")
    
    # تحديث كشف الرواتب
    await db.payroll_periods.update_one(
        {"id": period_id},
        {"$set": update_fields}
    )
    
    await log_activity(
        user_id=user_id,
        user_name=user_name,
        action="approve_payroll",
        entity_type="payroll",
        entity_id=period_id,
        entity_name=period["name"],
        details=f"{approval_message} - {period['name']}"
    )
    
    return {
        "message": approval_message,
        "new_status": new_status,
        "approved_by": user_name,
        "approved_at": now
    }


@api_router.get("/hr/payroll/periods/{period_id}/approval-status")
async def get_payroll_approval_status(period_id: str, current_user: dict = Depends(get_current_user)):
    """Get detailed approval status for a payroll period"""
    period = await db.payroll_periods.find_one({"id": period_id}, {"_id": 0})
    if not period:
        raise HTTPException(status_code=404, detail="Payroll period not found")
    
    status = period.get("status", "draft")
    
    # تحديد المراحل
    stages = [
        {
            "stage": 1,
            "name": "موافقة الموارد البشرية",
            "name_en": "HR Approval",
            "status": "completed" if period.get("hr_approved_at") else ("current" if status in ["draft", "calculated", "pending_hr"] else "pending"),
            "approved_at": period.get("hr_approved_at"),
            "approved_by": period.get("hr_approved_by_name"),
            "required_role": "hr_manager"
        },
        {
            "stage": 2,
            "name": "موافقة المالية",
            "name_en": "Finance Approval",
            "status": "completed" if period.get("finance_approved_at") else ("current" if status == "pending_finance" else "pending"),
            "approved_at": period.get("finance_approved_at"),
            "approved_by": period.get("finance_approved_by_name"),
            "required_role": "finance_manager"
        },
        {
            "stage": 3,
            "name": "موافقة المدير العام",
            "name_en": "GM Approval",
            "status": "completed" if period.get("gm_approved_at") else ("current" if status == "pending_gm" else "pending"),
            "approved_at": period.get("gm_approved_at"),
            "approved_by": period.get("gm_approved_by_name"),
            "required_role": "general_manager"
        }
    ]
    
    return {
        "period_id": period_id,
        "period_name": period.get("name"),
        "current_status": status,
        "stages": stages,
        "is_fully_approved": status == "approved" or status == "disbursed",
        "can_disburse": status == "approved"
    }



@api_router.post("/hr/payroll/periods/{period_id}/disburse")
async def disburse_payroll(
    period_id: str, 
    from_account: str = "1112",
    to_account: str = "حساب الموظفين",
    current_user: dict = Depends(require_role(["admin", "hr_manager", "finance_manager"]))
):
    """Disburse/Pay approved payroll - Creates automatic journal entry
    
    Args:
        from_account: Source account number (default: 1112 البنك)
        to_account: Description for destination (default: حساب الموظفين)
    """
    period = await db.payroll_periods.find_one({"id": period_id}, {"_id": 0})
    if not period:
        raise HTTPException(status_code=404, detail="Payroll period not found")
    
    if period.get("status") != "approved":
        raise HTTPException(status_code=400, detail="يجب اعتماد كشف الرواتب أولاً قبل الصرف")
    
    if period.get("status") == "disbursed":
        raise HTTPException(status_code=400, detail="تم صرف هذه الرواتب مسبقاً")
    
    # Get all payroll records for this period
    records = await db.payroll_records.find({"period_id": period_id}, {"_id": 0}).to_list(1000)
    
    total_net_salary = sum(r.get("net_salary", 0) for r in records)
    total_deductions = sum(r.get("total_deductions", 0) for r in records)
    total_gross_for_journal = round(total_net_salary + total_deductions, 3)  # Calculate gross from net + deductions for balanced entry
    
    # Create automatic journal entry for payroll disbursement
    # Dr: مصروفات الرواتب (5200) = صافي الراتب + الخصومات
    # Cr: الرواتب المستحقة (2120) = الخصومات (إن وجدت)
    # Cr: الحساب المحدد (from_account) = صافي الراتب
    journal_lines = [
        {"account_number": "5200", "debit": total_gross_for_journal, "credit": 0, "description": "مصروفات الرواتب والبدلات"},
    ]
    
    if total_deductions > 0:
        journal_lines.append({"account_number": "2120", "debit": 0, "credit": round(total_deductions, 3), "description": "خصومات الموظفين"})
    
    journal_lines.append({"account_number": from_account, "debit": 0, "credit": round(total_net_salary, 3), "description": f"صرف رواتب {period['name']} - من حساب {from_account} إلى {to_account}"})
    
    await create_auto_journal_entry(
        description=f"صرف رواتب - {period['name']} - {len(records)} موظف - من {from_account}",
        lines=journal_lines,
        reference_type="payroll_disbursement",
        reference_id=period_id,
        created_by_id=current_user["id"],
        created_by_name=current_user["full_name"]
    )
    
    # Update period status with account info
    await db.payroll_periods.update_one(
        {"id": period_id},
        {"$set": {
            "status": "disbursed",
            "disbursed_at": datetime.now(timezone.utc).isoformat(),
            "disbursed_by": current_user["full_name"],
            "total_disbursed": round(total_net_salary, 3),
            "from_account": from_account,
            "to_account": to_account
        }}
    )
    
    # Mark all records as paid
    await db.payroll_records.update_many(
        {"period_id": period_id},
        {"$set": {"is_paid": True, "paid_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="disburse_payroll",
        entity_type="payroll",
        entity_id=period_id,
        entity_name=period["name"],
        details=f"صرف رواتب {len(records)} موظف بإجمالي {total_net_salary:.3f} ر.ع من حساب {from_account}"
    )
    
    return {
        "message": f"تم صرف الرواتب بنجاح - {len(records)} موظف",
        "total_disbursed": round(total_net_salary, 3),
        "from_account": from_account,
        "to_account": to_account,
        "journal_entry_created": True
    }

@api_router.delete("/hr/payroll/periods/{period_id}")
async def delete_payroll_period(period_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a payroll period and its records"""
    period = await db.payroll_periods.find_one({"id": period_id}, {"_id": 0})
    if not period:
        raise HTTPException(status_code=404, detail="Payroll period not found")
    
    if period.get("status") == "approved":
        raise HTTPException(status_code=400, detail="لا يمكن حذف كشف رواتب معتمد")
    
    await db.payroll_records.delete_many({"period_id": period_id})
    await db.payroll_periods.delete_one({"id": period_id})
    
    return {"message": "تم حذف فترة الرواتب بنجاح"}

# ==================== EMPLOYEE SALARY STRUCTURE (هيكل راتب الموظف) ====================

@api_router.get("/hr/salary-structures")
async def get_all_salary_structures(
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all salary structures or filter by employee"""
    query = {"is_active": True}
    if employee_id:
        query["employee_id"] = employee_id
    
    structures = await db.employee_salary_structures.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return structures

@api_router.get("/hr/salary-structures/{employee_id}")
async def get_employee_salary_structure(
    employee_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get active salary structure for an employee"""
    structure = await db.employee_salary_structures.find_one(
        {"employee_id": employee_id, "is_active": True},
        {"_id": 0}
    )
    if not structure:
        # Return default structure if none exists
        employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
        if employee:
            return {
                "employee_id": employee_id,
                "employee_name": employee.get("name", ""),
                "basic_salary": employee.get("salary", 0),
                "allowances": {
                    "housing_allowance": 0,
                    "transportation_allowance": 0,
                    "food_allowance": 0,
                    "phone_allowance": 0,
                    "fuel_allowance": 0,
                    "education_allowance": 0,
                    "medical_allowance": 0,
                    "special_allowance": 0,
                    "other_allowance": 0
                },
                "total_salary": employee.get("salary", 0)
            }
    return structure

@api_router.post("/hr/salary-structures")
async def create_salary_structure(
    data: dict,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """Create or update salary structure for an employee"""
    employee_id = data.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="employee_id is required")
    
    employee = await db.hr_employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # Deactivate existing structure
    await db.employee_salary_structures.update_many(
        {"employee_id": employee_id, "is_active": True},
        {"$set": {"is_active": False}}
    )
    
    # Parse allowances
    allowances_data = data.get("allowances", {})
    allowances = EmployeeAllowances(
        housing_allowance=float(allowances_data.get("housing_allowance", 0)),
        transportation_allowance=float(allowances_data.get("transportation_allowance", 0)),
        food_allowance=float(allowances_data.get("food_allowance", 0)),
        phone_allowance=float(allowances_data.get("phone_allowance", 0)),
        fuel_allowance=float(allowances_data.get("fuel_allowance", 0)),
        education_allowance=float(allowances_data.get("education_allowance", 0)),
        medical_allowance=float(allowances_data.get("medical_allowance", 0)),
        special_allowance=float(allowances_data.get("special_allowance", 0)),
        other_allowance=float(allowances_data.get("other_allowance", 0))
    )
    
    basic_salary = float(data.get("basic_salary", employee.get("salary", 0)))
    total_allowances = sum([
        allowances.housing_allowance,
        allowances.transportation_allowance,
        allowances.food_allowance,
        allowances.phone_allowance,
        allowances.fuel_allowance,
        allowances.education_allowance,
        allowances.medical_allowance,
        allowances.special_allowance,
        allowances.other_allowance
    ])
    
    structure = EmployeeSalaryStructure(
        employee_id=employee_id,
        employee_name=employee.get("name", ""),
        basic_salary=basic_salary,
        allowances=allowances,
        total_salary=basic_salary + total_allowances,
        effective_date=data.get("effective_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        notes=data.get("notes"),
        created_by=current_user["id"]
    )
    
    # Add bank information to structure
    structure_dict = structure.model_dump()
    if data.get("bank_name"):
        structure_dict["bank_name"] = data.get("bank_name")
    if data.get("bank_account"):
        structure_dict["bank_account"] = data.get("bank_account")
    
    await db.employee_salary_structures.insert_one(structure_dict)
    
    # Update employee salary field
    await db.hr_employees.update_one(
        {"id": employee_id},
        {"$set": {"salary": basic_salary + total_allowances}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_salary_structure",
        entity_type="salary_structure",
        entity_id=structure.id,
        entity_name=employee.get("name"),
        details=f"تحديث هيكل راتب {employee.get('name')} - الأساسي: {basic_salary}, البدلات: {total_allowances}"
    )
    
    return {"message": "تم تحديث هيكل الراتب بنجاح", "structure": structure.model_dump()}

@api_router.put("/hr/salary-structures/{structure_id}")
async def update_salary_structure(
    structure_id: str,
    data: dict,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """Update an existing salary structure"""
    structure = await db.employee_salary_structures.find_one({"id": structure_id}, {"_id": 0})
    if not structure:
        raise HTTPException(status_code=404, detail="هيكل الراتب غير موجود")
    
    update_data = {}
    if "basic_salary" in data:
        update_data["basic_salary"] = float(data["basic_salary"])
    
    if "allowances" in data:
        allowances_data = data["allowances"]
        update_data["allowances"] = {
            "housing_allowance": float(allowances_data.get("housing_allowance", 0)),
            "transportation_allowance": float(allowances_data.get("transportation_allowance", 0)),
            "food_allowance": float(allowances_data.get("food_allowance", 0)),
            "phone_allowance": float(allowances_data.get("phone_allowance", 0)),
            "fuel_allowance": float(allowances_data.get("fuel_allowance", 0)),
            "education_allowance": float(allowances_data.get("education_allowance", 0)),
            "medical_allowance": float(allowances_data.get("medical_allowance", 0)),
            "special_allowance": float(allowances_data.get("special_allowance", 0)),
            "other_allowance": float(allowances_data.get("other_allowance", 0))
        }
    
    if update_data:
        # Recalculate total
        basic = update_data.get("basic_salary", structure.get("basic_salary", 0))
        allowances = update_data.get("allowances", structure.get("allowances", {}))
        total_allowances = sum(allowances.values()) if isinstance(allowances, dict) else 0
        update_data["total_salary"] = basic + total_allowances
        
        await db.employee_salary_structures.update_one(
            {"id": structure_id},
            {"$set": update_data}
        )
    
    return {"message": "تم تحديث هيكل الراتب بنجاح"}

@api_router.post("/hr/salary-structures/sync-all")
async def sync_all_salary_structures(
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """Create salary structures for all employees who don't have one, using their current salary"""
    employees = await db.hr_employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    
    created = 0
    skipped = 0
    
    for emp in employees:
        emp_id = emp.get("id")
        emp_salary = emp.get("salary", 0)
        
        # Check if employee already has an active structure
        existing = await db.employee_salary_structures.find_one(
            {"employee_id": emp_id, "is_active": True}
        )
        
        if existing:
            skipped += 1
            continue
        
        if emp_salary <= 0:
            skipped += 1
            continue
        
        # Create new structure with basic salary = employee salary
        structure = EmployeeSalaryStructure(
            employee_id=emp_id,
            employee_name=emp.get("name", ""),
            basic_salary=emp_salary,
            allowances=EmployeeAllowances(),  # Default all allowances to 0
            total_salary=emp_salary
        )
        
        await db.employee_salary_structures.insert_one(structure.model_dump())
        created += 1
    
    return {
        "message": f"تم إنشاء {created} هيكل راتب جديد، تم تخطي {skipped} موظف",
        "created": created,
        "skipped": skipped
    }



# ==================== PUBLIC HOLIDAYS (العطل الرسمية) ====================

@api_router.get("/hr/public-holidays")
async def get_public_holidays(
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get public holidays - redirects to official holidays (unified source)"""
    # Redirect to hr_official_holidays as the single source of truth
    query = {}
    if year:
        query["date"] = {"$regex": f"^{year}"}
    
    holidays = await db.hr_official_holidays.find(query, {"_id": 0}).sort("date", 1).to_list(100)
    return holidays

@api_router.post("/hr/public-holidays")
async def create_public_holiday(
    data: dict,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """Create a public holiday - adds to official holidays (unified source)"""
    from models.all_models import OfficialHoliday
    
    holiday = OfficialHoliday(
        name=data["name"],
        date=data["date"],
        applies_to="all",
        is_recurring=data.get("is_recurring", False),
        notes=data.get("notes"),
        created_by=current_user.get("full_name")
    )
    
    await db.hr_official_holidays.insert_one(holiday.model_dump())
    return {"message": "تم إضافة العطلة الرسمية بنجاح", "holiday": holiday.model_dump()}

@api_router.delete("/hr/public-holidays/{holiday_id}")
async def delete_public_holiday(
    holiday_id: str,
    current_user: dict = Depends(require_role(["admin", "hr_manager"]))
):
    """Delete a public holiday - removes from official holidays (unified source)"""
    await db.hr_official_holidays.delete_one({"id": holiday_id})
    return {"message": "تم حذف العطلة الرسمية بنجاح"}

# ==================== AI ANALYSIS (التحليل الذكي) ====================

# Model imported from models/all_models.py: AnalysisRequest

# Initialize Gemini chat with Emergent key
def get_llm_chat(session_id: str = "analysis"):
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="EMERGENT_LLM_KEY not configured")
    
    system_message = """أنت مساعد تحليل بيانات ذكي لنظام ERP لمركز تجميع الحليب "المروج للألبان".
    
مهمتك هي تحليل البيانات والإجابة على أسئلة المستخدم بشكل واضح ومفيد.

البيانات المتاحة:
- بيانات الموظفين والحضور والانصراف
- بيانات المبيعات والعملاء
- بيانات استلام الحليب من الموردين
- بيانات الرواتب والمدفوعات

قواعد مهمة:
1. أجب باللغة العربية إذا كان السؤال بالعربية
2. قدم إحصائيات وأرقام محددة عند توفرها
3. اقترح تحسينات إذا كانت مناسبة
4. كن موجزاً ومفيداً
5. استخدم التنسيق المناسب (قوائم، جداول) لتوضيح البيانات"""

    chat = LlmChat(
        api_key=api_key,
        session_id=session_id,
        system_message=system_message
    ).with_model("gemini", "gemini-2.5-flash")
    
    return chat

@api_router.post("/analysis/query")
async def analyze_query(request: AnalysisRequest, current_user: dict = Depends(get_current_user)):
    """Analyze natural language query and return data insights using Gemini 2.5 Flash"""
    
    try:
        # Get data context based on category
        context_data = {}
        
        # Fetch relevant data for context
        if request.category in ["general", "hr", "attendance"]:
            # Get attendance summary
            attendance = await db.hr_attendance.find({}).to_list(1000)
            employees = await db.hr_employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
            
            present_count = len([a for a in attendance if a.get("status") == "present"])
            absent_count = len([a for a in attendance if a.get("status") == "absent"])
            leave_count = len([a for a in attendance if a.get("status") == "leave"])
            
            context_data["attendance"] = {
                "total_employees": len(employees),
                "present_days_total": present_count,
                "absent_days_total": absent_count,
                "leave_days_total": leave_count,
                "departments": list(set([e.get("department", "unknown") for e in employees]))
            }
        
        if request.category in ["general", "sales"]:
            # Get sales summary
            sales = await db.sales.find({}, {"_id": 0}).to_list(1000)
            customers = await db.customers.find({"is_active": True}, {"_id": 0}).to_list(1000)
            context_data["sales"] = {
                "total_sales": len(sales),
                "total_amount": sum([s.get("total_amount", 0) for s in sales]),
                "total_customers": len(customers),
            }
        
        if request.category in ["general", "milk"]:
            # Get milk reception summary
            receptions = await db.milk_receptions.find({}, {"_id": 0}).to_list(1000)
            suppliers = await db.suppliers.find({"is_active": True}, {"_id": 0}).to_list(1000)
            context_data["milk"] = {
                "total_receptions": len(receptions),
                "total_quantity_liters": sum([r.get("quantity", 0) for r in receptions]),
                "total_amount": sum([r.get("total_amount", 0) for r in receptions]),
                "total_suppliers": len(suppliers),
            }
        
        if request.category in ["general", "hr"]:
            # Get payroll summary
            payroll_records = await db.payroll_records.find({}, {"_id": 0}).to_list(1000)
            context_data["payroll"] = {
                "total_records": len(payroll_records),
                "total_gross": sum([p.get("gross_salary", 0) for p in payroll_records]),
                "total_net": sum([p.get("net_salary", 0) for p in payroll_records]),
            }
        
        # Create user message with context
        user_prompt = f"""السؤال: {request.question}

البيانات الحالية من النظام:
{context_data}

قدم إجابة تحليلية مفصلة بناءً على البيانات المتاحة."""

        # Call Gemini API using emergentintegrations
        chat = get_llm_chat(session_id=f"analysis_{current_user['id']}")
        user_message = UserMessage(text=user_prompt)
        answer = await chat.send_message(user_message)
        
        # Log the analysis
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user["full_name"],
            action="ai_analysis",
            details=f"سؤال: {request.question[:100]}..."
        )
        
        return {
            "question": request.question,
            "answer": answer,
            "category": request.category,
            "data_summary": context_data,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
        
    except Exception as e:
        logging.error(f"Analysis error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"خطأ في التحليل: {str(e)}")

@api_router.get("/analysis/summary")
async def get_analysis_summary(current_user: dict = Depends(get_current_user)):
    """Get quick summary statistics for analysis dashboard"""
    
    # Attendance stats
    employees = await db.hr_employees.find({"is_active": True}, {"_id": 0}).to_list(1000)
    attendance = await db.hr_attendance.find({}).to_list(5000)
    
    present_count = len([a for a in attendance if a.get("status") == "present"])
    absent_count = len([a for a in attendance if a.get("status") == "absent"])
    
    # Sales stats
    sales = await db.sales.find({}, {"_id": 0}).to_list(1000)
    total_sales_amount = sum([s.get("total_amount", 0) for s in sales])
    
    # Milk reception stats
    receptions = await db.milk_receptions.find({}, {"_id": 0}).to_list(1000)
    total_milk = sum([r.get("quantity", 0) for r in receptions])
    
    # Supplier stats
    suppliers = await db.suppliers.find({"is_active": True}, {"_id": 0}).to_list(500)
    
    # Department breakdown
    departments = {}
    for emp in employees:
        dept = emp.get("department", "unknown")
        departments[dept] = departments.get(dept, 0) + 1
    
    return {
        "employees": {
            "total": len(employees),
            "by_department": departments
        },
        "attendance": {
            "present_total": present_count,
            "absent_total": absent_count,
            "attendance_rate": round(present_count / (present_count + absent_count) * 100, 2) if (present_count + absent_count) > 0 else 0
        },
        "sales": {
            "total_transactions": len(sales),
            "total_amount": total_sales_amount
        },
        "milk": {
            "total_receptions": len(receptions),
            "total_quantity": total_milk
        },
        "suppliers": {
            "total_active": len(suppliers)
        }
    }

# ==================== USER SETTINGS / APPEARANCE ====================

# System background images - imported from utils/helpers.py
from utils.helpers import SYSTEM_BACKGROUNDS

# Model imported from models/all_models.py: UserAppearanceSettings

@api_router.get("/user/settings")
async def get_user_settings(current_user: dict = Depends(get_current_user)):
    """Get user appearance settings"""
    settings = await db.user_settings.find_one({"user_id": current_user["id"]}, {"_id": 0})
    if not settings:
        # Return default settings
        default_bg = SYSTEM_BACKGROUNDS[0]
        return {
            "user_id": current_user["id"],
            "background_id": "bg1",
            "background_url": default_bg["url"],
            "theme": "light",
            "sidebar_collapsed": False
        }
    return settings

@api_router.put("/user/settings")
async def update_user_settings(settings: UserAppearanceSettings, current_user: dict = Depends(get_current_user)):
    """Update user appearance settings"""
    
    # Get background URL from ID
    background_url = settings.background_url
    if settings.background_id:
        for bg in SYSTEM_BACKGROUNDS:
            if bg["id"] == settings.background_id:
                background_url = bg["url"]
                break
    
    settings_data = {
        "user_id": current_user["id"],
        "background_id": settings.background_id,
        "background_url": background_url,
        "theme": settings.theme,
        "sidebar_collapsed": settings.sidebar_collapsed,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.user_settings.update_one(
        {"user_id": current_user["id"]},
        {"$set": settings_data},
        upsert=True
    )
    
    return settings_data

@api_router.get("/system/backgrounds")
async def get_system_backgrounds(current_user: dict = Depends(get_current_user)):
    """Get available system background images"""
    return SYSTEM_BACKGROUNDS

# ==================== ZKTeco Sync Manager APIs ====================

# Models imported from models/all_models.py:
# ZKTecoDeviceBase, ZKTecoDeviceCreate, ZKTecoDevice, ZKTecoSyncSettings

@api_router.get("/hr/zkteco/devices")
async def get_zkteco_devices(current_user: dict = Depends(get_current_user)):
    """Get all ZKTeco devices and sync settings"""
    devices = await db.zkteco_devices.find({"is_active": True}, {"_id": 0}).to_list(100)
    
    # Get sync settings
    settings = await db.zkteco_settings.find_one({"type": "sync_settings"}, {"_id": 0})
    
    return {
        "devices": devices,
        "auto_sync_enabled": settings.get("auto_sync_enabled", False) if settings else False,
        "sync_interval": settings.get("sync_interval", 60) if settings else 60,
        "last_sync": settings.get("last_sync") if settings else None
    }

@api_router.post("/hr/zkteco/devices", response_model=ZKTecoDevice)
async def create_zkteco_device(device_data: ZKTecoDeviceCreate, current_user: dict = Depends(require_role(["admin"]))):
    """Add a new ZKTeco device"""
    device = ZKTecoDevice(**device_data.model_dump())
    await db.zkteco_devices.insert_one(device.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="add_zkteco_device",
        entity_type="zkteco_device",
        entity_id=device.id,
        entity_name=device.name,
        details=f"إضافة جهاز بصمة: {device.name} ({device.ip_address})"
    )
    
    return device

@api_router.delete("/hr/zkteco/devices/{device_id}")
async def delete_zkteco_device(device_id: str, current_user: dict = Depends(require_role(["admin"]))):
    """Delete a ZKTeco device"""
    device = await db.zkteco_devices.find_one({"id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    await db.zkteco_devices.update_one(
        {"id": device_id},
        {"$set": {"is_active": False}}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_zkteco_device",
        entity_type="zkteco_device",
        entity_id=device_id,
        entity_name=device.get("name"),
        details=f"حذف جهاز بصمة: {device.get('name')}"
    )
    
    return {"message": "Device deleted successfully"}

@api_router.post("/hr/zkteco/devices/{device_id}/test")
async def test_zkteco_device(device_id: str, current_user: dict = Depends(get_current_user)):
    """Test connection to a ZKTeco device"""
    device = await db.zkteco_devices.find_one({"id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    import socket
    
    try:
        # Try to connect to the device
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(5)
        result = sock.connect_ex((device["ip_address"], device["port"]))
        sock.close()
        
        if result == 0:
            # Update device status
            await db.zkteco_devices.update_one(
                {"id": device_id},
                {"$set": {"is_online": True}}
            )
            
            return {
                "success": True,
                "serial_number": "ZKTeco-" + device["ip_address"].replace(".", ""),
                "users_count": 0,
                "records_count": 0,
                "message": "اتصال ناجح - الجهاز متصل"
            }
        else:
            await db.zkteco_devices.update_one(
                {"id": device_id},
                {"$set": {"is_online": False}}
            )
            return {
                "success": False,
                "error": "لا يمكن الاتصال بالجهاز - تأكد من العنوان والمنفذ"
            }
    except socket.timeout:
        return {
            "success": False,
            "error": "انتهت مهلة الاتصال - الجهاز غير متاح"
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"خطأ في الاتصال: {str(e)}"
        }

@api_router.put("/hr/zkteco/settings")
async def update_zkteco_settings(settings: ZKTecoSyncSettings, current_user: dict = Depends(require_role(["admin"]))):
    """Update ZKTeco sync settings"""
    await db.zkteco_settings.update_one(
        {"type": "sync_settings"},
        {"$set": {
            "type": "sync_settings",
            "auto_sync_enabled": settings.auto_sync_enabled,
            "sync_interval": settings.sync_interval,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user["id"]
        }},
        upsert=True
    )
    
    return {"message": "Settings updated successfully"}

@api_router.post("/hr/zkteco/sync")
async def sync_zkteco_attendance(current_user: dict = Depends(require_role(["admin"]))):
    """Sync attendance data from all ZKTeco devices"""
    devices = await db.zkteco_devices.find({"is_active": True}, {"_id": 0}).to_list(100)
    
    if not devices:
        raise HTTPException(status_code=400, detail="لا توجد أجهزة مضافة")
    
    total_imported = 0
    total_updated = 0
    errors = []
    
    for device in devices:
        try:
            # Here you would integrate with actual ZKTeco SDK
            # For now, we'll simulate the sync process
            
            # Update device last sync time
            await db.zkteco_devices.update_one(
                {"id": device["id"]},
                {"$set": {
                    "last_sync": datetime.now(timezone.utc).isoformat(),
                    "is_online": True
                }}
            )
            
            logging.info(f"تم مزامنة الجهاز: {device['name']} ({device['ip_address']})")
            
        except Exception as e:
            errors.append(f"{device['name']}: {str(e)}")
            logging.error(f"خطأ في مزامنة الجهاز {device['name']}: {e}")
    
    # Update last sync time in settings
    await db.zkteco_settings.update_one(
        {"type": "sync_settings"},
        {"$set": {"last_sync": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="zkteco_sync",
        entity_type="attendance",
        details=f"مزامنة البصمات: {total_imported} جديد، {total_updated} محدث"
    )
    
    return {
        "success": True,
        "imported": total_imported,
        "updated": total_updated,
        "errors": errors if errors else None,
        "message": f"تم مزامنة {len(devices)} جهاز"
    }

# ==================== HR - WARNINGS (الإنذارات) ====================

# Models imported from models/all_models.py:
# WarningBase, WarningCreate, Warning

@api_router.get("/hr/warnings")
async def get_warnings(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all warnings"""
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if status:
        query["status"] = status
    
    warnings = await db.hr_warnings.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return warnings

@api_router.post("/hr/warnings", response_model=Warning)
async def create_warning(warning_data: WarningCreate, current_user: dict = Depends(require_role(["admin", "hr_manager"]))):
    """Create a new warning"""
    warning = Warning(**warning_data.model_dump())
    warning.issued_by = current_user["id"]
    warning.issued_by_name = current_user["full_name"]
    
    await db.hr_warnings.insert_one(warning.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="create_warning",
        entity_type="warning",
        entity_id=warning.id,
        entity_name=warning_data.employee_name,
        details=f"إنذار {warning_data.warning_type} للموظف: {warning_data.employee_name}"
    )
    
    return warning

@api_router.put("/hr/warnings/{warning_id}")
async def update_warning(warning_id: str, warning_data: WarningCreate, current_user: dict = Depends(require_role(["admin", "hr_manager"]))):
    """Update a warning"""
    existing = await db.hr_warnings.find_one({"id": warning_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Warning not found")
    
    await db.hr_warnings.update_one(
        {"id": warning_id},
        {"$set": warning_data.model_dump()}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_warning",
        entity_type="warning",
        entity_id=warning_id,
        entity_name=warning_data.employee_name,
        details=f"تحديث إنذار الموظف: {warning_data.employee_name}"
    )
    
    return {"message": "Warning updated successfully"}

@api_router.delete("/hr/warnings/{warning_id}")
async def delete_warning(warning_id: str, current_user: dict = Depends(require_role(["admin", "hr_manager"]))):
    """Delete a warning"""
    warning = await db.hr_warnings.find_one({"id": warning_id}, {"_id": 0})
    if not warning:
        raise HTTPException(status_code=404, detail="Warning not found")
    
    await db.hr_warnings.delete_one({"id": warning_id})
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="delete_warning",
        entity_type="warning",
        entity_id=warning_id,
        entity_name=warning.get("employee_name"),
        details=f"حذف إنذار الموظف: {warning.get('employee_name')}"
    )
    
    return {"message": "Warning deleted successfully"}

# ==================== ADVANCED REPORTS (التقارير المتقدمة) ====================

@api_router.get("/reports/payroll/comparison")
async def get_payroll_comparison_report(
    period1_id: str,
    period2_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    تقرير مقارنة الرواتب بين فترتين
    Payroll comparison report between two periods
    """
    # Get period details
    period1 = await db.payroll_periods.find_one({"id": period1_id}, {"_id": 0})
    period2 = await db.payroll_periods.find_one({"id": period2_id}, {"_id": 0})
    
    if not period1 or not period2:
        raise HTTPException(status_code=404, detail="فترة واحدة أو أكثر غير موجودة")
    
    # Get records for both periods
    records1 = await db.payroll_records.find({"period_id": period1_id}, {"_id": 0}).to_list(1000)
    records2 = await db.payroll_records.find({"period_id": period2_id}, {"_id": 0}).to_list(1000)
    
    # Create lookup maps
    records1_map = {r["employee_id"]: r for r in records1}
    records2_map = {r["employee_id"]: r for r in records2}
    
    # All employee IDs
    all_employees = set(records1_map.keys()) | set(records2_map.keys())
    
    # Build comparison
    comparisons = []
    summary = {
        "period1_total_gross": 0,
        "period2_total_gross": 0,
        "period1_total_net": 0,
        "period2_total_net": 0,
        "period1_total_allowances": 0,
        "period2_total_allowances": 0,
        "period1_total_deductions": 0,
        "period2_total_deductions": 0,
        "employees_with_increase": 0,
        "employees_with_decrease": 0,
        "employees_unchanged": 0,
        "new_employees": 0,
        "removed_employees": 0
    }
    
    for emp_id in all_employees:
        r1 = records1_map.get(emp_id, {})
        r2 = records2_map.get(emp_id, {})
        
        period1_net = r1.get("net_salary", 0)
        period2_net = r2.get("net_salary", 0)
        
        # Calculate changes
        net_change = period2_net - period1_net
        basic_change = r2.get("basic_salary", 0) - r1.get("basic_salary", 0)
        allowances_change = r2.get("total_allowances", r2.get("allowances", 0)) - r1.get("total_allowances", r1.get("allowances", 0))
        deductions_change = r2.get("total_deductions", r2.get("deductions", 0)) - r1.get("total_deductions", r1.get("deductions", 0))
        
        status = "unchanged"
        if not r1:
            status = "new"
            summary["new_employees"] += 1
        elif not r2:
            status = "removed"
            summary["removed_employees"] += 1
        elif net_change > 0:
            status = "increase"
            summary["employees_with_increase"] += 1
        elif net_change < 0:
            status = "decrease"
            summary["employees_with_decrease"] += 1
        else:
            summary["employees_unchanged"] += 1
        
        comparisons.append({
            "employee_id": emp_id,
            "employee_name": r2.get("employee_name") or r1.get("employee_name"),
            "employee_code": r2.get("employee_code") or r1.get("employee_code"),
            "department": r2.get("department") or r1.get("department"),
            "period1": {
                "basic_salary": r1.get("basic_salary", 0),
                "allowances": r1.get("total_allowances", r1.get("allowances", 0)),
                "deductions": r1.get("total_deductions", r1.get("deductions", 0)),
                "overtime_pay": r1.get("overtime_pay", 0),
                "gross_salary": r1.get("gross_salary", 0),
                "net_salary": period1_net,
                "working_days": r1.get("working_days", 0),
                "absent_days": r1.get("absent_days", 0)
            },
            "period2": {
                "basic_salary": r2.get("basic_salary", 0),
                "allowances": r2.get("total_allowances", r2.get("allowances", 0)),
                "deductions": r2.get("total_deductions", r2.get("deductions", 0)),
                "overtime_pay": r2.get("overtime_pay", 0),
                "gross_salary": r2.get("gross_salary", 0),
                "net_salary": period2_net,
                "working_days": r2.get("working_days", 0),
                "absent_days": r2.get("absent_days", 0)
            },
            "changes": {
                "basic_salary": basic_change,
                "allowances": allowances_change,
                "deductions": deductions_change,
                "net_salary": net_change,
                "percentage": round((net_change / period1_net * 100) if period1_net > 0 else 0, 2)
            },
            "status": status
        })
        
        summary["period1_total_gross"] += r1.get("gross_salary", 0)
        summary["period2_total_gross"] += r2.get("gross_salary", 0)
        summary["period1_total_net"] += period1_net
        summary["period2_total_net"] += period2_net
        summary["period1_total_allowances"] += r1.get("total_allowances", r1.get("allowances", 0))
        summary["period2_total_allowances"] += r2.get("total_allowances", r2.get("allowances", 0))
        summary["period1_total_deductions"] += r1.get("total_deductions", r1.get("deductions", 0))
        summary["period2_total_deductions"] += r2.get("total_deductions", r2.get("deductions", 0))
    
    # Calculate net change
    summary["net_change"] = summary["period2_total_net"] - summary["period1_total_net"]
    summary["percentage_change"] = round(
        (summary["net_change"] / summary["period1_total_net"] * 100) if summary["period1_total_net"] > 0 else 0, 2
    )
    
    return {
        "period1": {
            "id": period1_id,
            "name": period1.get("name"),
            "start_date": period1.get("start_date"),
            "end_date": period1.get("end_date")
        },
        "period2": {
            "id": period2_id,
            "name": period2.get("name"),
            "start_date": period2.get("start_date"),
            "end_date": period2.get("end_date")
        },
        "summary": summary,
        "comparisons": sorted(comparisons, key=lambda x: abs(x["changes"]["net_salary"]), reverse=True)
    }

@api_router.get("/reports/financial/monthly")
async def get_monthly_financial_report(
    year: int,
    month: int,
    current_user: dict = Depends(get_current_user)
):
    """
    تقرير مالي شهري شامل
    Comprehensive monthly financial report
    """
    from calendar import monthrange
    
    start_date = f"{year}-{month:02d}-01"
    last_day = monthrange(year, month)[1]
    end_date = f"{year}-{month:02d}-{last_day}"
    
    # Get milk receptions (purchases)
    milk_receptions = await db.milk_receptions.find({
        "reception_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(5000)
    
    # Get sales
    sales = await db.sales.find({
        "sale_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(5000)
    
    # Get payroll records for this period
    payroll_period = await db.payroll_periods.find_one({
        "start_date": {"$gte": start_date},
        "end_date": {"$lte": end_date}
    }, {"_id": 0})
    
    payroll_records = []
    if payroll_period:
        payroll_records = await db.payroll_records.find({
            "period_id": payroll_period["id"]
        }, {"_id": 0}).to_list(1000)
    
    # Get journal entries
    journal_entries = await db.journal_entries.find({
        "entry_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(1000)
    
    # Calculate milk purchases by center
    purchases_by_center = {}
    total_milk_quantity = 0
    total_milk_amount = 0
    
    for r in milk_receptions:
        center = r.get("center_name", "غير محدد")
        if center not in purchases_by_center:
            purchases_by_center[center] = {"quantity": 0, "amount": 0, "count": 0}
        purchases_by_center[center]["quantity"] += r.get("quantity_liters", 0)
        purchases_by_center[center]["amount"] += r.get("total_amount", 0)
        purchases_by_center[center]["count"] += 1
        total_milk_quantity += r.get("quantity_liters", 0)
        total_milk_amount += r.get("total_amount", 0)
    
    # Calculate sales summary
    total_sales_quantity = sum(s.get("quantity_liters", 0) for s in sales)
    total_sales_amount = sum(s.get("total_amount", 0) for s in sales)
    cash_sales = sum(s.get("total_amount", 0) for s in sales if s.get("is_paid"))
    credit_sales = sum(s.get("total_amount", 0) for s in sales if not s.get("is_paid"))
    
    # Calculate payroll summary
    total_gross_salary = sum(p.get("gross_salary", 0) for p in payroll_records)
    total_net_salary = sum(p.get("net_salary", 0) for p in payroll_records)
    total_allowances = sum(p.get("total_allowances", p.get("allowances", 0)) for p in payroll_records)
    total_deductions = sum(p.get("total_deductions", p.get("deductions", 0)) for p in payroll_records)
    
    # Gross profit
    gross_profit = total_sales_amount - total_milk_amount
    net_profit = gross_profit - total_net_salary
    
    return {
        "period": {
            "year": year,
            "month": month,
            "start_date": start_date,
            "end_date": end_date
        },
        "revenue": {
            "total_sales": total_sales_amount,
            "cash_sales": cash_sales,
            "credit_sales": credit_sales,
            "quantity_sold_liters": total_sales_quantity,
            "sales_count": len(sales)
        },
        "cost_of_goods": {
            "total_purchases": total_milk_amount,
            "quantity_purchased_liters": total_milk_quantity,
            "purchases_count": len(milk_receptions),
            "by_center": purchases_by_center
        },
        "operating_expenses": {
            "salaries_and_wages": total_net_salary,
            "gross_salaries": total_gross_salary,
            "allowances": total_allowances,
            "deductions": total_deductions,
            "employee_count": len(payroll_records)
        },
        "profitability": {
            "gross_profit": gross_profit,
            "gross_margin_percentage": round((gross_profit / total_sales_amount * 100) if total_sales_amount > 0 else 0, 2),
            "net_profit": net_profit,
            "net_margin_percentage": round((net_profit / total_sales_amount * 100) if total_sales_amount > 0 else 0, 2)
        },
        "journal_entries_count": len(journal_entries)
    }

@api_router.get("/reports/centers/performance")
async def get_centers_performance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    تقرير أداء مراكز التجميع
    Collection centers performance report
    """
    query = {}
    if start_date:
        query["reception_date"] = {"$gte": start_date}
    if end_date:
        if "reception_date" in query:
            query["reception_date"]["$lte"] = end_date
        else:
            query["reception_date"] = {"$lte": end_date}
    
    # Get milk receptions
    receptions = await db.milk_receptions.find(query, {"_id": 0}).to_list(10000)
    
    # Get suppliers by center
    suppliers = await db.suppliers.find({"is_active": True}, {"_id": 0}).to_list(5000)
    suppliers_by_center = {}
    for s in suppliers:
        center = s.get("center_name", "غير محدد")
        if center not in suppliers_by_center:
            suppliers_by_center[center] = 0
        suppliers_by_center[center] += 1
    
    # Group by center
    centers_data = {}
    for r in receptions:
        center = r.get("center_name", "غير محدد")
        if center not in centers_data:
            centers_data[center] = {
                "name": center,
                "total_quantity": 0,
                "total_amount": 0,
                "reception_count": 0,
                "suppliers_count": suppliers_by_center.get(center, 0),
                "avg_fat": [],
                "avg_snf": [],
                "camel_milk": 0,
                "cow_milk": 0,
                "daily_data": {}
            }
        
        centers_data[center]["total_quantity"] += r.get("quantity_liters", 0)
        centers_data[center]["total_amount"] += r.get("total_amount", 0)
        centers_data[center]["reception_count"] += 1
        
        if r.get("fat_percentage"):
            centers_data[center]["avg_fat"].append(r["fat_percentage"])
        if r.get("snf_percentage"):
            centers_data[center]["avg_snf"].append(r["snf_percentage"])
        
        milk_type = r.get("milk_type", "")
        if "إبل" in milk_type or "camel" in milk_type.lower():
            centers_data[center]["camel_milk"] += r.get("quantity_liters", 0)
        else:
            centers_data[center]["cow_milk"] += r.get("quantity_liters", 0)
        
        # Daily breakdown
        date = r.get("reception_date", "").split("T")[0]
        if date:
            if date not in centers_data[center]["daily_data"]:
                centers_data[center]["daily_data"][date] = {"quantity": 0, "amount": 0}
            centers_data[center]["daily_data"][date]["quantity"] += r.get("quantity_liters", 0)
            centers_data[center]["daily_data"][date]["amount"] += r.get("total_amount", 0)
    
    # Calculate averages and rankings
    result = []
    for center, data in centers_data.items():
        avg_fat = round(sum(data["avg_fat"]) / len(data["avg_fat"]), 2) if data["avg_fat"] else 0
        avg_snf = round(sum(data["avg_snf"]) / len(data["avg_snf"]), 2) if data["avg_snf"] else 0
        avg_price = round(data["total_amount"] / data["total_quantity"], 3) if data["total_quantity"] > 0 else 0
        
        result.append({
            "center_name": center,
            "total_quantity": round(data["total_quantity"], 2),
            "total_amount": round(data["total_amount"], 3),
            "reception_count": data["reception_count"],
            "suppliers_count": data["suppliers_count"],
            "avg_fat_percentage": avg_fat,
            "avg_snf_percentage": avg_snf,
            "avg_price_per_liter": avg_price,
            "camel_milk_liters": round(data["camel_milk"], 2),
            "cow_milk_liters": round(data["cow_milk"], 2),
            "camel_percentage": round((data["camel_milk"] / data["total_quantity"] * 100) if data["total_quantity"] > 0 else 0, 1),
            "daily_average_quantity": round(data["total_quantity"] / len(data["daily_data"]) if data["daily_data"] else 0, 2),
            "active_days": len(data["daily_data"])
        })
    
    # Sort by total quantity (highest first)
    result = sorted(result, key=lambda x: x["total_quantity"], reverse=True)
    
    # Add ranking
    for i, r in enumerate(result):
        r["rank"] = i + 1
    
    # Calculate totals
    totals = {
        "total_quantity": sum(r["total_quantity"] for r in result),
        "total_amount": sum(r["total_amount"] for r in result),
        "total_receptions": sum(r["reception_count"] for r in result),
        "total_suppliers": sum(r["suppliers_count"] for r in result),
        "centers_count": len(result)
    }
    
    return {
        "period": {"start_date": start_date, "end_date": end_date},
        "totals": totals,
        "centers": result
    }

@api_router.get("/reports/inventory/alerts")
async def get_inventory_alerts(
    current_user: dict = Depends(get_current_user)
):
    """
    تنبيهات المخزون المنخفض
    Low inventory alerts
    """
    # Get inventory items
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(100)
    
    # Get inventory settings (thresholds)
    settings = await db.system_settings.find_one({"type": "inventory"}, {"_id": 0})
    default_threshold = 100  # Default threshold for milk in liters
    
    alerts = []
    for item in inventory:
        product_type = item.get("product_type", "unknown")
        quantity = item.get("quantity_liters", item.get("quantity", 0))
        
        # Get threshold for this product
        threshold = default_threshold
        if settings and settings.get("thresholds"):
            threshold = settings["thresholds"].get(product_type, default_threshold)
        
        if quantity <= threshold:
            alerts.append({
                "product_type": product_type,
                "current_quantity": quantity,
                "threshold": threshold,
                "deficit": threshold - quantity,
                "severity": "critical" if quantity <= threshold * 0.5 else "warning",
                "last_updated": item.get("last_updated")
            })
    
    # Get feed types inventory
    feed_types = await db.feed_types.find({}, {"_id": 0}).to_list(100)
    for feed in feed_types:
        stock = feed.get("stock_quantity", 0)
        min_stock = feed.get("min_stock_quantity", 10)
        
        if stock <= min_stock:
            alerts.append({
                "product_type": f"feed_{feed.get('name')}",
                "product_name": feed.get("name"),
                "current_quantity": stock,
                "threshold": min_stock,
                "deficit": min_stock - stock,
                "unit": feed.get("unit", "كيس"),
                "severity": "critical" if stock <= min_stock * 0.5 else "warning",
                "category": "feed"
            })
    
    # Sort by severity (critical first)
    alerts = sorted(alerts, key=lambda x: (0 if x["severity"] == "critical" else 1, -x.get("deficit", 0)))
    
    return {
        "alerts_count": len(alerts),
        "critical_count": len([a for a in alerts if a["severity"] == "critical"]),
        "warning_count": len([a for a in alerts if a["severity"] == "warning"]),
        "alerts": alerts
    }

@api_router.post("/reports/inventory/set-threshold")
async def set_inventory_threshold(
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """
    تعيين حد أدنى للمخزون
    Set minimum inventory threshold
    """
    product_type = data.get("product_type")
    threshold = data.get("threshold")
    
    if not product_type or threshold is None:
        raise HTTPException(status_code=400, detail="product_type and threshold are required")
    
    # Update or create settings
    await db.system_settings.update_one(
        {"type": "inventory"},
        {
            "$set": {f"thresholds.{product_type}": threshold},
            "$setOnInsert": {"type": "inventory", "created_at": datetime.now(timezone.utc).isoformat()}
        },
        upsert=True
    )
    
    return {"message": f"تم تعيين الحد الأدنى لـ {product_type} إلى {threshold}"}

@api_router.post("/reports/inventory/send-alerts")
async def send_inventory_alerts_email(
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """
    إرسال تنبيهات المخزون المنخفض عبر البريد الإلكتروني
    Send low inventory alerts via email
    """
    email = data.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مطلوب")
    
    # Get alerts
    alerts_response = await get_inventory_alerts(current_user)
    alerts = alerts_response.get("alerts", [])
    
    if not alerts:
        return {"message": "لا توجد تنبيهات للإرسال", "sent": False}
    
    # Build email content
    critical_alerts = [a for a in alerts if a["severity"] == "critical"]
    warning_alerts = [a for a in alerts if a["severity"] == "warning"]
    
    html_content = f"""
    <html dir="rtl">
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; direction: rtl; }}
            .header {{ background-color: #8B4513; color: white; padding: 20px; text-align: center; }}
            .content {{ padding: 20px; }}
            .critical {{ background-color: #fee2e2; border-right: 4px solid #dc2626; padding: 10px; margin: 10px 0; }}
            .warning {{ background-color: #fef3c7; border-right: 4px solid #f59e0b; padding: 10px; margin: 10px 0; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: right; }}
            th {{ background-color: #8B4513; color: white; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🚨 تنبيه المخزون - المروج للألبان</h1>
        </div>
        <div class="content">
            <p>السلام عليكم،</p>
            <p>هناك <strong>{len(alerts)}</strong> تنبيه للمخزون المنخفض:</p>
            
            <h3>📊 ملخص التنبيهات:</h3>
            <ul>
                <li>🔴 تنبيهات حرجة: {len(critical_alerts)}</li>
                <li>🟡 تنبيهات تحذيرية: {len(warning_alerts)}</li>
            </ul>
            
            <table>
                <tr>
                    <th>المنتج</th>
                    <th>الكمية الحالية</th>
                    <th>الحد الأدنى</th>
                    <th>النقص</th>
                    <th>الحالة</th>
                </tr>
    """
    
    for alert in alerts:
        severity_class = "critical" if alert["severity"] == "critical" else "warning"
        severity_text = "🔴 حرج" if alert["severity"] == "critical" else "🟡 تحذير"
        product_name = alert.get("product_name", alert.get("product_type", "غير معروف"))
        
        html_content += f"""
                <tr class="{severity_class}">
                    <td>{product_name}</td>
                    <td>{alert.get('current_quantity', 0)} {alert.get('unit', 'لتر')}</td>
                    <td>{alert.get('threshold', 0)}</td>
                    <td>{alert.get('deficit', 0)}</td>
                    <td>{severity_text}</td>
                </tr>
        """
    
    html_content += """
            </table>
            
            <p style="margin-top: 20px;">يرجى اتخاذ الإجراءات اللازمة لتجديد المخزون.</p>
            <hr>
            <p style="color: #666; font-size: 12px;">
                هذا البريد مرسل تلقائياً من نظام ERP المروج للألبان
            </p>
        </div>
    </body>
    </html>
    """
    
    # Try to send email
    try:
        smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
        smtp_port = int(os.environ.get("SMTP_PORT", 587))
        smtp_user = os.environ.get("SMTP_USER")
        smtp_pass = os.environ.get("SMTP_PASSWORD")
        
        if not smtp_user or not smtp_pass:
            # Store alert for manual notification
            alert_record = {
                "id": str(uuid.uuid4()),
                "type": "inventory_alert",
                "email": email,
                "alerts_count": len(alerts),
                "content": html_content,
                "status": "pending",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user["id"]
            }
            await db.pending_notifications.insert_one(alert_record)
            
            return {
                "message": "تم حفظ التنبيه - إعدادات البريد الإلكتروني غير مكتملة",
                "sent": False,
                "alerts_count": len(alerts),
                "note": "يرجى إضافة إعدادات SMTP في ملف .env"
            }
        
        message = MIMEMultipart("alternative")
        message["Subject"] = f"🚨 تنبيه المخزون - {len(alerts)} منتج يحتاج تجديد"
        message["From"] = smtp_user
        message["To"] = email
        
        message.attach(MIMEText(html_content, "html", "utf-8"))
        
        await aiosmtplib.send(
            message,
            hostname=smtp_host,
            port=smtp_port,
            username=smtp_user,
            password=smtp_pass,
            start_tls=True
        )
        
        # Log the notification
        await db.notification_logs.insert_one({
            "id": str(uuid.uuid4()),
            "type": "inventory_alert",
            "email": email,
            "alerts_count": len(alerts),
            "status": "sent",
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "sent_by": current_user["full_name"]
        })
        
        return {
            "message": f"تم إرسال {len(alerts)} تنبيه إلى {email}",
            "sent": True,
            "alerts_count": len(alerts)
        }
        
    except Exception as e:
        logging.error(f"Failed to send email: {e}")
        return {
            "message": f"فشل إرسال البريد: {str(e)}",
            "sent": False,
            "error": str(e)
        }

@api_router.get("/notifications/settings")
async def get_notification_settings(
    current_user: dict = Depends(get_current_user)
):
    """جلب إعدادات الإشعارات"""
    settings = await db.system_settings.find_one({"type": "notifications"}, {"_id": 0})
    return settings or {
        "type": "notifications",
        "email_alerts_enabled": False,
        "alert_email": "",
        "inventory_threshold_alert": True,
        "daily_report_email": False
    }

@api_router.post("/notifications/settings")
async def update_notification_settings(
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """تحديث إعدادات الإشعارات"""
    await db.system_settings.update_one(
        {"type": "notifications"},
        {
            "$set": {
                "email_alerts_enabled": data.get("email_alerts_enabled", False),
                "alert_email": data.get("alert_email", ""),
                "inventory_threshold_alert": data.get("inventory_threshold_alert", True),
                "daily_report_email": data.get("daily_report_email", False),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$setOnInsert": {"type": "notifications", "created_at": datetime.now(timezone.utc).isoformat()}
        },
        upsert=True
    )
    return {"message": "تم تحديث إعدادات الإشعارات"}

# ==================== SMS INTEGRATION (تكامل الرسائل النصية) ====================

async def send_sms_tamimah(phone: str, message: str) -> dict:
    """
    إرسال رسالة SMS - يستخدم الدالة الجديدة من sms_routes
    Wrapper for backward compatibility
    """
    from routes.sms_routes import send_sms_oman
    return await send_sms_oman(phone, message)

@api_router.get("/sms/settings")
async def get_sms_settings(current_user: dict = Depends(require_role(["admin"]))):
    """Get SMS provider settings"""
    from routes.sms_routes import SMS_PROVIDERS
    settings = await db.system_settings.find_one({"type": "sms"}, {"_id": 0})
    if settings and "password" in settings:
        settings["password"] = "********"  # Hide password
    
    default_settings = {
        "type": "sms",
        "provider": "ismart",
        "api_url": "",
        "api_key": "",
        "username": "",
        "sender_id": "MAROOJ",
        "is_configured": False,
        "available_providers": list(SMS_PROVIDERS.keys()),
        "providers_info": SMS_PROVIDERS
    }
    
    if settings:
        default_settings.update(settings)
    
    return default_settings

@api_router.post("/sms/settings")
async def update_sms_settings(
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Update SMS provider settings"""
    update_data = {
        "provider": data.get("provider", "ismart"),
        "api_url": data.get("api_url", ""),
        "api_key": data.get("api_key", ""),
        "username": data.get("username", ""),
        "sender_id": data.get("sender_id", "MAROOJ"),
        "is_configured": bool(data.get("api_url") or data.get("api_key")),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Only update password if provided
    if data.get("password") and data["password"] != "********":
        update_data["password"] = data["password"]
    
    await db.system_settings.update_one(
        {"type": "sms"},
        {
            "$set": update_data,
            "$setOnInsert": {"type": "sms", "created_at": datetime.now(timezone.utc).isoformat()}
        },
        upsert=True
    )
    return {"message": "تم تحديث إعدادات SMS بنجاح"}

@api_router.post("/sms/send")
async def send_sms(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Send SMS message"""
    phone = data.get("phone")
    message = data.get("message")
    
    if not phone or not message:
        raise HTTPException(status_code=400, detail="رقم الهاتف والرسالة مطلوبان")
    
    result = await send_sms_tamimah(phone, message)
    
    # Log the SMS
    sms_log = {
        "id": str(uuid.uuid4()),
        "phone": phone,
        "message": message,
        "status": "sent" if result["success"] else "failed",
        "error": result.get("error"),
        "sent_by": current_user["full_name"],
        "sent_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sms_logs.insert_one(sms_log)
    
    if result["success"]:
        return {"message": "تم إرسال الرسالة بنجاح", "success": True}
    else:
        return {"message": f"فشل الإرسال: {result.get('error')}", "success": False}

@api_router.post("/sms/send-otp")
async def send_otp_sms(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Send OTP via SMS for password reset"""
    phone = data.get("phone")
    
    if not phone:
        raise HTTPException(status_code=400, detail="رقم الهاتف مطلوب")
    
    # Generate OTP
    otp = str(secrets.randbelow(900000) + 100000)  # 6-digit OTP
    
    # Store OTP with expiry (5 minutes)
    otp_record = {
        "id": str(uuid.uuid4()),
        "phone": phone,
        "otp": otp,
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.otp_codes.insert_one(otp_record)
    
    # Send SMS
    message = f"رمز التحقق الخاص بك في المروج للألبان: {otp}\nصالح لمدة 5 دقائق"
    result = await send_sms_tamimah(phone, message)
    
    if result["success"]:
        return {"message": "تم إرسال رمز التحقق", "success": True}
    else:
        # Even if SMS fails, return success for security (don't reveal if SMS works)
        # But log the error
        logging.warning(f"OTP SMS failed for {phone}: {result.get('error')}")
        return {"message": "تم إرسال رمز التحقق", "success": True, "note": "يرجى التواصل مع الدعم إذا لم تصلك الرسالة"}

@api_router.get("/sms/logs")
async def get_sms_logs(
    limit: int = 50,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Get SMS sending logs"""
    logs = await db.sms_logs.find({}, {"_id": 0}).sort("sent_at", -1).limit(limit).to_list(limit)
    return logs

# ==================== SCHEDULED REPORTS (التقارير المجدولة) ====================

@api_router.get("/reports/schedules")
async def get_report_schedules(current_user: dict = Depends(require_role(["admin"]))):
    """Get all scheduled reports"""
    schedules = await db.report_schedules.find({}, {"_id": 0}).to_list(100)
    return schedules

@api_router.post("/reports/schedules")
async def create_report_schedule(
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Create a scheduled report"""
    schedule = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", "تقرير مجدول"),
        "report_type": data.get("report_type"),  # daily_summary, weekly_summary, monthly_financial, inventory_alerts
        "frequency": data.get("frequency", "daily"),  # daily, weekly, monthly
        "day_of_week": data.get("day_of_week", 0),  # 0=Monday, 6=Sunday (for weekly)
        "day_of_month": data.get("day_of_month", 1),  # 1-28 (for monthly)
        "time": data.get("time", "08:00"),  # HH:MM
        "recipients": data.get("recipients", []),  # List of emails
        "is_active": data.get("is_active", True),
        "last_sent": None,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.report_schedules.insert_one(dict(schedule))
    return {"message": "تم إنشاء الجدول بنجاح", "schedule": schedule}

@api_router.put("/reports/schedules/{schedule_id}")
async def update_report_schedule(
    schedule_id: str,
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Update a scheduled report"""
    update_data = {}
    for field in ["name", "report_type", "frequency", "day_of_week", "day_of_month", "time", "recipients", "is_active"]:
        if field in data:
            update_data[field] = data[field]
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.report_schedules.update_one({"id": schedule_id}, {"$set": update_data})
    
    return {"message": "تم تحديث الجدول بنجاح"}

@api_router.delete("/reports/schedules/{schedule_id}")
async def delete_report_schedule(
    schedule_id: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Delete a scheduled report"""
    await db.report_schedules.delete_one({"id": schedule_id})
    return {"message": "تم حذف الجدول بنجاح"}

@api_router.post("/reports/schedules/{schedule_id}/run")
async def run_scheduled_report(
    schedule_id: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Manually run a scheduled report"""
    schedule = await db.report_schedules.find_one({"id": schedule_id}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="الجدول غير موجود")
    
    result = await generate_and_send_scheduled_report(schedule, current_user)
    return result

async def generate_and_send_scheduled_report(schedule: dict, user: dict = None):
    """Generate and send a scheduled report via email"""
    report_type = schedule.get("report_type")
    recipients = schedule.get("recipients", [])
    
    if not recipients:
        return {"success": False, "error": "لا يوجد مستلمين"}
    
    # Generate report content based on type
    today = datetime.now(timezone.utc)
    html_content = ""
    subject = ""
    
    if report_type == "daily_summary":
        subject = f"📊 التقرير اليومي - {today.strftime('%Y-%m-%d')}"
        
        # Get today's data
        today_str = today.strftime("%Y-%m-%d")
        
        receptions = await db.milk_receptions.find({"reception_date": {"$regex": f"^{today_str}"}}, {"_id": 0}).to_list(1000)
        sales = await db.sales.find({"sale_date": {"$regex": f"^{today_str}"}}, {"_id": 0}).to_list(1000)
        
        total_milk = sum(r.get("quantity_liters", 0) for r in receptions)
        total_milk_amount = sum(r.get("total_amount", 0) for r in receptions)
        total_sales = sum(s.get("total_amount", 0) for s in sales)
        
        html_content = f"""
        <html dir="rtl">
        <head><style>
            body {{ font-family: Arial; direction: rtl; }}
            .header {{ background: #8B4513; color: white; padding: 20px; text-align: center; }}
            .card {{ background: #f5f5f5; padding: 15px; margin: 10px; border-radius: 8px; }}
            .value {{ font-size: 24px; font-weight: bold; color: #8B4513; }}
        </style></head>
        <body>
            <div class="header"><h1>📊 التقرير اليومي</h1><p>{today_str}</p></div>
            <div style="padding: 20px;">
                <div class="card">
                    <h3>🥛 استلام الحليب</h3>
                    <p class="value">{total_milk:,.0f} لتر</p>
                    <p>المبلغ: {total_milk_amount:,.3f} ر.ع</p>
                    <p>عدد الاستلامات: {len(receptions)}</p>
                </div>
                <div class="card">
                    <h3>💰 المبيعات</h3>
                    <p class="value">{total_sales:,.3f} ر.ع</p>
                    <p>عدد العمليات: {len(sales)}</p>
                </div>
            </div>
        </body>
        </html>
        """
    
    elif report_type == "weekly_summary":
        subject = f"📈 التقرير الأسبوعي - الأسبوع {today.isocalendar()[1]}"
        
        week_start = today - timedelta(days=today.weekday())
        week_start_str = week_start.strftime("%Y-%m-%d")
        
        receptions = await db.milk_receptions.find({"reception_date": {"$gte": week_start_str}}, {"_id": 0}).to_list(5000)
        sales = await db.sales.find({"sale_date": {"$gte": week_start_str}}, {"_id": 0}).to_list(5000)
        
        total_milk = sum(r.get("quantity_liters", 0) for r in receptions)
        total_milk_amount = sum(r.get("total_amount", 0) for r in receptions)
        total_sales = sum(s.get("total_amount", 0) for s in sales)
        
        html_content = f"""
        <html dir="rtl">
        <head><style>
            body {{ font-family: Arial; direction: rtl; }}
            .header {{ background: #8B4513; color: white; padding: 20px; text-align: center; }}
            .card {{ background: #f5f5f5; padding: 15px; margin: 10px; border-radius: 8px; display: inline-block; width: 45%; }}
            .value {{ font-size: 24px; font-weight: bold; color: #8B4513; }}
        </style></head>
        <body>
            <div class="header"><h1>📈 التقرير الأسبوعي</h1><p>من {week_start_str}</p></div>
            <div style="padding: 20px;">
                <div class="card">
                    <h3>🥛 إجمالي الحليب</h3>
                    <p class="value">{total_milk:,.0f} لتر</p>
                    <p>المبلغ: {total_milk_amount:,.3f} ر.ع</p>
                </div>
                <div class="card">
                    <h3>💰 إجمالي المبيعات</h3>
                    <p class="value">{total_sales:,.3f} ر.ع</p>
                </div>
            </div>
        </body>
        </html>
        """
    
    elif report_type == "monthly_financial":
        subject = f"📊 التقرير المالي الشهري - {today.strftime('%Y-%m')}"
        
        # Get financial data for current month
        year = today.year
        month = today.month
        start_date = f"{year}-{month:02d}-01"
        
        receptions = await db.milk_receptions.find({"reception_date": {"$regex": f"^{year}-{month:02d}"}}, {"_id": 0}).to_list(5000)
        sales = await db.sales.find({"sale_date": {"$regex": f"^{year}-{month:02d}"}}, {"_id": 0}).to_list(5000)
        
        total_purchases = sum(r.get("total_amount", 0) for r in receptions)
        total_sales_amount = sum(s.get("total_amount", 0) for s in sales)
        gross_profit = total_sales_amount - total_purchases
        
        html_content = f"""
        <html dir="rtl">
        <head><style>
            body {{ font-family: Arial; direction: rtl; }}
            .header {{ background: #8B4513; color: white; padding: 20px; text-align: center; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: right; }}
            th {{ background: #8B4513; color: white; }}
            .profit {{ color: {"green" if gross_profit >= 0 else "red"}; font-weight: bold; }}
        </style></head>
        <body>
            <div class="header"><h1>📊 التقرير المالي الشهري</h1><p>{today.strftime('%B %Y')}</p></div>
            <table>
                <tr><th>البند</th><th>المبلغ (ر.ع)</th></tr>
                <tr><td>إجمالي المبيعات</td><td>{total_sales_amount:,.3f}</td></tr>
                <tr><td>تكلفة المشتريات</td><td>{total_purchases:,.3f}</td></tr>
                <tr><td class="profit">إجمالي الربح</td><td class="profit">{gross_profit:,.3f}</td></tr>
            </table>
        </body>
        </html>
        """
    
    elif report_type == "inventory_alerts":
        subject = f"🚨 تنبيهات المخزون - {today.strftime('%Y-%m-%d')}"
        
        # Get inventory alerts
        feed_types = await db.feed_types.find({}, {"_id": 0}).to_list(100)
        alerts = []
        for feed in feed_types:
            stock = feed.get("stock_quantity", 0)
            min_stock = feed.get("min_stock_quantity", 10)
            if stock <= min_stock:
                alerts.append({
                    "name": feed.get("name"),
                    "stock": stock,
                    "min_stock": min_stock,
                    "severity": "critical" if stock <= min_stock * 0.5 else "warning"
                })
        
        if not alerts:
            return {"success": True, "message": "لا توجد تنبيهات", "skipped": True}
        
        alerts_html = ""
        for alert in alerts:
            color = "#dc2626" if alert["severity"] == "critical" else "#f59e0b"
            alerts_html += f"""
            <tr style="background-color: {'#fee2e2' if alert['severity'] == 'critical' else '#fef3c7'}">
                <td>{alert['name']}</td>
                <td>{alert['stock']}</td>
                <td>{alert['min_stock']}</td>
                <td style="color: {color}; font-weight: bold;">{"🔴 حرج" if alert['severity'] == 'critical' else "🟡 تحذير"}</td>
            </tr>
            """
        
        html_content = f"""
        <html dir="rtl">
        <head><style>
            body {{ font-family: Arial; direction: rtl; }}
            .header {{ background: #dc2626; color: white; padding: 20px; text-align: center; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: right; }}
            th {{ background: #8B4513; color: white; }}
        </style></head>
        <body>
            <div class="header"><h1>🚨 تنبيهات المخزون</h1><p>{len(alerts)} منتج يحتاج تجديد</p></div>
            <table>
                <tr><th>المنتج</th><th>الكمية الحالية</th><th>الحد الأدنى</th><th>الحالة</th></tr>
                {alerts_html}
            </table>
        </body>
        </html>
        """
    
    else:
        return {"success": False, "error": f"نوع التقرير غير معروف: {report_type}"}
    
    # Send email to all recipients
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", 587))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASSWORD")
    
    if not smtp_user or not smtp_pass:
        # Save for later
        await db.pending_reports.insert_one({
            "id": str(uuid.uuid4()),
            "schedule_id": schedule.get("id"),
            "subject": subject,
            "content": html_content,
            "recipients": recipients,
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        return {"success": False, "error": "إعدادات SMTP غير مكتملة - تم حفظ التقرير"}
    
    try:
        for recipient in recipients:
            message = MIMEMultipart("alternative")
            message["Subject"] = subject
            message["From"] = smtp_user
            message["To"] = recipient
            message.attach(MIMEText(html_content, "html", "utf-8"))
            
            await aiosmtplib.send(
                message,
                hostname=smtp_host,
                port=smtp_port,
                username=smtp_user,
                password=smtp_pass,
                start_tls=True
            )
        
        # Update last_sent
        await db.report_schedules.update_one(
            {"id": schedule.get("id")},
            {"$set": {"last_sent": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Log
        await db.report_logs.insert_one({
            "id": str(uuid.uuid4()),
            "schedule_id": schedule.get("id"),
            "report_type": report_type,
            "recipients": recipients,
            "status": "sent",
            "sent_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"success": True, "message": f"تم إرسال التقرير إلى {len(recipients)} مستلم"}
        
    except Exception as e:
        logging.error(f"Failed to send scheduled report: {e}")
        return {"success": False, "error": str(e)}

@api_router.get("/reports/logs")
async def get_report_logs(
    limit: int = 50,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Get report sending logs"""
    logs = await db.report_logs.find({}, {"_id": 0}).sort("sent_at", -1).limit(limit).to_list(limit)
    return logs

# ===================== MILK PRICES SETTINGS =====================

@api_router.get("/settings/milk-prices")
async def get_milk_prices(current_user: dict = Depends(get_current_user)):
    """Get milk prices settings"""
    # Default prices
    default_prices = [
        {"id": "camel", "name": "حليب الإبل", "price": 0.350, "is_active": True},
        {"id": "cow", "name": "حليب الأبقار", "price": 0.250, "is_active": True},
        {"id": "sheep", "name": "حليب الأغنام", "price": 0.300, "is_active": True},
    ]
    
    # Get saved prices from DB
    saved_prices = await db.milk_prices.find({}, {"_id": 0}).to_list(100)
    
    # Merge saved with defaults (use saved if exists, otherwise use default)
    result = []
    for default in default_prices:
        saved = next((p for p in saved_prices if p.get("id") == default["id"]), None)
        if saved:
            result.append(saved)
        else:
            result.append(default)
    
    # Add any extra prices not in defaults
    saved_ids = [p.get("id") for p in saved_prices]
    default_ids = [p["id"] for p in default_prices]
    for saved in saved_prices:
        if saved.get("id") not in default_ids:
            result.append(saved)
    
    return result

@api_router.post("/settings/milk-prices")
async def save_milk_price(
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """Save or update milk price"""
    milk_type = data.get("milk_type")
    if not milk_type:
        raise HTTPException(status_code=400, detail="نوع الحليب مطلوب")
    
    price_data = {
        "id": milk_type,
        "name": data.get("name", ""),
        "price": float(data.get("price", 0)),
        "is_active": data.get("is_active", True),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["full_name"]
    }
    
    await db.milk_prices.update_one(
        {"id": milk_type},
        {"$set": price_data},
        upsert=True
    )
    
    # Log the activity
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user["full_name"],
        action="update_milk_price",
        details=f"تحديث سعر {data.get('name')}"
    )
    
    return {"message": "تم حفظ السعر بنجاح", "data": price_data}

@api_router.get("/")
async def root():
    return {"message": "Milk Collection Center ERP API", "version": "1.0.0"}

# Include the router in the main app
app.include_router(api_router)

# Include new modular routers
from routes.customers_routes import router as customers_router
from routes.sales_routes import router as sales_router
from routes.inventory_routes import router as inventory_router
from routes.milk_routes import router as milk_router
from routes.permissions_routes import router as permissions_router

app.include_router(customers_router, prefix="/api")
app.include_router(sales_router, prefix="/api")
app.include_router(inventory_router, prefix="/api")
app.include_router(milk_router, prefix="/api")
app.include_router(permissions_router, prefix="/api")

# Include CCTV router
from routes.cctv_routes import router as cctv_router
app.include_router(cctv_router)

# Include Hik-Connect router
from routes.hikconnect_routes import router as hikconnect_router
app.include_router(hikconnect_router)

# Include SMS router
from routes.sms_routes import router as sms_router
app.include_router(sms_router)

# Include Procurement router
from routes.procurement_routes import router as procurement_router
app.include_router(procurement_router)

# Include Operations router
from routes.operations_routes import router as operations_router
app.include_router(operations_router)

# Mount static files for exports
static_exports_path = Path(__file__).parent / "static" / "exports"
if static_exports_path.exists():
    app.mount("/api/exports", StaticFiles(directory=str(static_exports_path)), name="exports")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
