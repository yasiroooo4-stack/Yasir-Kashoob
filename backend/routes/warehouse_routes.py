"""
Warehouse Management Routes - مسارات إدارة المخازن الشاملة
مع التكامل المالي والمبيعات
"""
from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db
from models.all_models import (
    Warehouse, WarehouseBase,
    ProductCategory, ProductCategoryBase,
    WarehouseProduct, WarehouseProductBase,
    ProductStock, ProductStockBase,
    StockMovement, StockMovementBase,
    LabSolution, LabSolutionBase,
    SolutionConsumption, SolutionConsumptionBase,
    PurchaseRequest, PurchaseRequestBase,
    StockAlert, StockAlertBase,
    AutoReorderRequest, InventorySettings
)
from routes.base import get_current_user, require_role, log_activity
import uuid
import io
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

router = APIRouter(prefix="/warehouse", tags=["Warehouse Management"])

# ==================== إعدادات الحسابات المالية ====================

# أرقام الحسابات الافتراضية للتكامل المالي
FINANCE_ACCOUNTS = {
    "inventory": "1300",          # حساب المخزون (أصل)
    "cogs": "5100",               # تكلفة البضاعة المباعة (مصروف)
    "accounts_payable": "2100",   # الدائنون (التزام)
    "expenses_supplies": "6200",  # مصروفات المستلزمات
    "expenses_feed": "6201",      # مصروفات الأعلاف
    "expenses_maintenance": "6202", # مصروفات الصيانة
    "expenses_lab": "6203",       # مصروفات المختبر
    "expenses_cleaning": "6204",  # مصروفات التنظيف
    "cash": "1111",               # الصندوق (نقدية)
}

# ربط تصنيف المخزن بحساب المصروفات
CATEGORY_TO_EXPENSE_ACCOUNT = {
    "feed": "6201",        # الأعلاف
    "maintenance": "6202",  # الصيانة
    "lab": "6203",         # المختبر
    "cleaning": "6204",    # التنظيف
    "ppe": "6205",         # معدات الحماية
    "equipment": "6206",   # المعدات
    "supplies": "6200",    # المستلزمات
}


# ==================== وظائف التكامل المالي ====================

async def create_warehouse_journal_entry(
    description: str,
    lines: list,
    reference_type: str,
    reference_id: str,
    created_by_id: str,
    created_by_name: str
):
    """
    إنشاء قيد يومية آلي لحركات المستودعات
    """
    try:
        entry_lines = []
        total_debit = 0
        total_credit = 0
        
        for line in lines:
            account = await db.chart_of_accounts.find_one(
                {"account_number": line["account_number"], "is_active": True},
                {"_id": 0}
            )
            if not account:
                logging.warning(f"Warehouse JV: Account not found: {line['account_number']}")
                continue
            
            debit = line.get("debit", 0)
            credit = line.get("credit", 0)
            total_debit += debit
            total_credit += credit
            
            entry_lines.append({
                "id": str(uuid.uuid4()),
                "account_id": account["id"],
                "account_number": account["account_number"],
                "account_name": account["name"],
                "debit": debit,
                "credit": credit,
                "description": line.get("description", "")
            })
        
        if not entry_lines:
            logging.warning(f"Warehouse JV: No valid accounts found for: {description}")
            return None
        
        # التحقق من توازن القيد
        if abs(total_debit - total_credit) > 0.01:
            logging.error(f"Warehouse JV: Unbalanced entry: debit={total_debit}, credit={total_credit}")
            return None
        
        # توليد رقم القيد
        count = await db.journal_entries.count_documents({})
        entry_number = f"WH-{datetime.now().year}-{count + 1:05d}"
        
        entry = {
            "id": str(uuid.uuid4()),
            "entry_number": entry_number,
            "entry_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "description": description,
            "reference_type": reference_type,
            "reference_id": reference_id,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "status": "posted",
            "created_by": created_by_id,
            "created_by_name": created_by_name,
            "posted_at": datetime.now(timezone.utc).isoformat(),
            "posted_by": "النظام (آلي - المستودعات)",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.journal_entries.insert_one(entry)
        
        # إدراج بنود القيد
        for line in entry_lines:
            line["journal_entry_id"] = entry["id"]
            await db.journal_entry_lines.insert_one(line)
        
        # تحديث أرصدة الحسابات
        for line in entry_lines:
            balance_change = line["debit"] - line["credit"]
            await db.chart_of_accounts.update_one(
                {"id": line["account_id"]},
                {"$inc": {"balance": balance_change}}
            )
        
        logging.info(f"Warehouse JV created: {entry_number} - {description}")
        return entry
        
    except Exception as e:
        logging.error(f"Error creating warehouse journal entry: {e}")
        return None


async def get_expense_account_for_warehouse(warehouse_id: str):
    """الحصول على حساب المصروفات المناسب للمخزن"""
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    if not warehouse:
        return FINANCE_ACCOUNTS["expenses_supplies"]
    
    category = warehouse.get("warehouse_category", "")
    return CATEGORY_TO_EXPENSE_ACCOUNT.get(category, FINANCE_ACCOUNTS["expenses_supplies"])


# ==================== المراكز والمخازن الافتراضية ====================

CENTERS = ["زيك", "حجيف", "غدو", "طاقة", "ثمريت", "مرباط"]

WAREHOUSE_CATEGORIES = {
    "internal": {
        "name_ar": "مخزن داخلي",
        "name_en": "Internal Warehouse",
        "sub_warehouses": [
            {"category": "lab", "name_ar": "مخزن المختبر", "name_en": "Lab Warehouse", "temp_controlled": True, "expiry_tracking": True},
            {"category": "maintenance", "name_ar": "مخزن الصيانة", "name_en": "Maintenance Warehouse", "temp_controlled": False, "expiry_tracking": False},
            {"category": "cleaning", "name_ar": "مخزن مواد التنظيف", "name_en": "Cleaning Materials Warehouse", "temp_controlled": False, "expiry_tracking": True},
            {"category": "ppe", "name_ar": "مخزن معدات الحماية", "name_en": "PPE Warehouse", "temp_controlled": False, "expiry_tracking": False},
        ]
    },
    "external": {
        "name_ar": "مخزن خارجي",
        "name_en": "External Warehouse",
        "sub_warehouses": [
            {"category": "feed", "name_ar": "مخزن الأعلاف", "name_en": "Feed Warehouse", "temp_controlled": False, "expiry_tracking": True},
            {"category": "equipment", "name_ar": "مخزن المعدات وقطع الغيار", "name_en": "Equipment & Spare Parts", "temp_controlled": False, "expiry_tracking": False},
            {"category": "supplies", "name_ar": "مخزن مستلزمات الموردين", "name_en": "Supplier Supplies", "temp_controlled": False, "expiry_tracking": False},
        ]
    }
}


# ==================== وظائف التنبيهات ====================

async def send_email_alert(to_email: str, subject: str, body: str):
    """إرسال تنبيه بالبريد الإلكتروني"""
    try:
        smtp_host = os.environ.get("SMTP_HOST", "")
        smtp_port = int(os.environ.get("SMTP_PORT", 465))
        smtp_user = os.environ.get("SMTP_USER", "")
        smtp_password = os.environ.get("SMTP_PASSWORD", "")
        smtp_from = os.environ.get("SMTP_FROM_EMAIL", smtp_user)
        
        if not all([smtp_host, smtp_user, smtp_password, to_email]):
            print(f"Email config missing, skipping email to {to_email}")
            return False
        
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = smtp_from
        msg["To"] = to_email
        
        html_body = f"""
        <html dir="rtl">
        <body style="font-family: Arial, sans-serif; direction: rtl;">
            {body}
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, "plain", "utf-8"))
        msg.attach(MIMEText(html_body, "html", "utf-8"))
        
        with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_from, to_email, msg.as_string())
        
        print(f"Email sent to {to_email}")
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False


async def send_sms_alert(phone: str, message: str):
    """إرسال تنبيه SMS"""
    try:
        # TODO: تكامل مع مزود SMS
        # يمكن استخدام Twilio أو أي مزود آخر
        print(f"SMS to {phone}: {message}")
        return True
    except Exception as e:
        print(f"Failed to send SMS: {e}")
        return False


async def create_stock_alert(
    alert_type: str,
    product_id: str,
    product_name: str,
    product_code: str,
    warehouse_id: str,
    warehouse_name: str,
    center_name: str,
    current_quantity: float,
    min_quantity: float,
    expiry_date: str = None,
    days_to_expiry: int = None
):
    """إنشاء تنبيه مخزون"""
    
    # تحديد الرسالة والأولوية
    if alert_type == "low_stock":
        if current_quantity == 0:
            message = f"⚠️ نفاد المخزون: {product_name} في {warehouse_name} ({center_name})"
            priority = "critical"
        elif current_quantity <= min_quantity * 0.5:
            message = f"🔴 مخزون منخفض جداً: {product_name} - الكمية: {current_quantity} (الحد الأدنى: {min_quantity})"
            priority = "high"
        else:
            message = f"🟡 مخزون منخفض: {product_name} - الكمية: {current_quantity} (الحد الأدنى: {min_quantity})"
            priority = "medium"
    elif alert_type == "expiry_warning":
        message = f"⏰ قرب انتهاء الصلاحية: {product_name} في {warehouse_name} - ينتهي في {days_to_expiry} يوم ({expiry_date})"
        priority = "high" if days_to_expiry <= 7 else "medium"
    elif alert_type == "expired":
        message = f"🚫 منتهي الصلاحية: {product_name} في {warehouse_name} - انتهى في {expiry_date}"
        priority = "critical"
    else:
        message = f"تنبيه مخزون: {product_name}"
        priority = "low"
    
    # التحقق من عدم وجود تنبيه مشابه غير محلول
    existing = await db.stock_alerts.find_one({
        "alert_type": alert_type,
        "product_id": product_id,
        "warehouse_id": warehouse_id,
        "is_resolved": False
    }, {"_id": 0})
    
    if existing:
        # تحديث التنبيه الموجود
        await db.stock_alerts.update_one(
            {"id": existing["id"]},
            {"$set": {
                "current_quantity": current_quantity,
                "message": message,
                "priority": priority,
                "created_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        return existing["id"]
    
    # إنشاء تنبيه جديد
    alert = StockAlert(
        alert_type=alert_type,
        product_id=product_id,
        product_name=product_name,
        product_code=product_code,
        warehouse_id=warehouse_id,
        warehouse_name=warehouse_name,
        center_name=center_name,
        current_quantity=current_quantity,
        min_quantity=min_quantity,
        expiry_date=expiry_date,
        days_to_expiry=days_to_expiry,
        message=message,
        priority=priority
    )
    
    await db.stock_alerts.insert_one(alert.model_dump())
    
    # إرسال التنبيهات
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    if warehouse:
        # إرسال لمشرف المركز
        if warehouse.get("supervisor_email"):
            await send_email_alert(
                warehouse["supervisor_email"],
                f"تنبيه مخزون - {center_name}",
                f"""
                <h2>تنبيه مخزون</h2>
                <p><strong>{message}</strong></p>
                <table border="1" cellpadding="10" style="border-collapse: collapse;">
                    <tr><td>المنتج</td><td>{product_name}</td></tr>
                    <tr><td>الرمز</td><td>{product_code}</td></tr>
                    <tr><td>المخزن</td><td>{warehouse_name}</td></tr>
                    <tr><td>المركز</td><td>{center_name}</td></tr>
                    <tr><td>الكمية الحالية</td><td>{current_quantity}</td></tr>
                    <tr><td>الحد الأدنى</td><td>{min_quantity}</td></tr>
                </table>
                """
            )
            await db.stock_alerts.update_one({"id": alert.id}, {"$set": {"notified_via_email": True}})
        
        if warehouse.get("supervisor_phone"):
            sms_msg = f"تنبيه: {product_name} - الكمية: {current_quantity} ({warehouse_name})"
            await send_sms_alert(warehouse["supervisor_phone"], sms_msg)
            await db.stock_alerts.update_one({"id": alert.id}, {"$set": {"notified_via_sms": True}})
        
        # إرسال لمسؤول المخازن
        if warehouse.get("warehouse_manager_email"):
            await send_email_alert(
                warehouse["warehouse_manager_email"],
                f"تنبيه مخزون - {center_name}",
                f"<h2>{message}</h2><p>المنتج: {product_name} | المخزن: {warehouse_name} | الكمية: {current_quantity}</p>"
            )
    
    return alert.id


async def check_stock_alerts():
    """فحص المخزون وإنشاء التنبيهات"""
    # فحص المخزون المنخفض
    products = await db.warehouse_products.find({"status": "active"}, {"_id": 0}).to_list(1000)
    product_map = {p["id"]: p for p in products}
    
    stock_items = await db.product_stock.find({}, {"_id": 0}).to_list(5000)
    
    for item in stock_items:
        product = product_map.get(item.get("product_id"))
        if not product:
            continue
        
        min_qty = product.get("min_quantity", 0)
        current_qty = item.get("quantity", 0)
        
        # تنبيه نقص المخزون
        if min_qty > 0 and current_qty <= min_qty:
            warehouse = await db.warehouses.find_one({"id": item.get("warehouse_id")}, {"_id": 0})
            await create_stock_alert(
                alert_type="low_stock",
                product_id=item.get("product_id"),
                product_name=item.get("product_name"),
                product_code=item.get("product_code"),
                warehouse_id=item.get("warehouse_id"),
                warehouse_name=item.get("warehouse_name"),
                center_name=warehouse.get("center_name", "") if warehouse else "",
                current_quantity=current_qty,
                min_quantity=min_qty
            )
        
        # تنبيه انتهاء الصلاحية
        if item.get("expiry_date") and product.get("expiry_tracking"):
            try:
                expiry = datetime.strptime(item["expiry_date"], "%Y-%m-%d")
                today = datetime.now()
                days_to_expiry = (expiry - today).days
                
                warehouse = await db.warehouses.find_one({"id": item.get("warehouse_id")}, {"_id": 0})
                alert_days = warehouse.get("expiry_alert_days", 30) if warehouse else 30
                
                if days_to_expiry <= 0:
                    await create_stock_alert(
                        alert_type="expired",
                        product_id=item.get("product_id"),
                        product_name=item.get("product_name"),
                        product_code=item.get("product_code"),
                        warehouse_id=item.get("warehouse_id"),
                        warehouse_name=item.get("warehouse_name"),
                        center_name=warehouse.get("center_name", "") if warehouse else "",
                        current_quantity=current_qty,
                        min_quantity=min_qty,
                        expiry_date=item["expiry_date"],
                        days_to_expiry=days_to_expiry
                    )
                elif days_to_expiry <= alert_days:
                    await create_stock_alert(
                        alert_type="expiry_warning",
                        product_id=item.get("product_id"),
                        product_name=item.get("product_name"),
                        product_code=item.get("product_code"),
                        warehouse_id=item.get("warehouse_id"),
                        warehouse_name=item.get("warehouse_name"),
                        center_name=warehouse.get("center_name", "") if warehouse else "",
                        current_quantity=current_qty,
                        min_quantity=min_qty,
                        expiry_date=item["expiry_date"],
                        days_to_expiry=days_to_expiry
                    )
            except:
                pass


# ==================== المخازن ====================

@router.get("/warehouses")
async def get_warehouses(
    status: Optional[str] = None,
    warehouse_type: Optional[str] = None,
    center_name: Optional[str] = None,
    warehouse_category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على قائمة المخازن"""
    query = {}
    if status:
        query["status"] = status
    if warehouse_type:
        query["warehouse_type"] = warehouse_type
    if center_name:
        query["center_name"] = center_name
    if warehouse_category:
        query["warehouse_category"] = warehouse_category
    
    warehouses = await db.warehouses.find(query, {"_id": 0}).sort([("center_name", 1), ("warehouse_type", 1), ("name", 1)]).to_list(200)
    return warehouses


@router.get("/warehouses/by-center")
async def get_warehouses_by_center(current_user: dict = Depends(get_current_user)):
    """الحصول على المخازن مجمعة حسب المركز"""
    warehouses = await db.warehouses.find({"status": "active"}, {"_id": 0}).to_list(200)
    
    result = {}
    for center in CENTERS:
        result[center] = {
            "internal": [],
            "external": []
        }
    
    for wh in warehouses:
        center = wh.get("center_name")
        wh_type = wh.get("warehouse_type", "internal")
        if center in result:
            if wh_type in result[center]:
                result[center][wh_type].append(wh)
    
    return result


@router.post("/warehouses/initialize-all")
async def initialize_all_warehouses(
    current_user: dict = Depends(require_role(["admin"]))
):
    """إنشاء جميع المخازن لكل المراكز"""
    created_count = 0
    skipped_count = 0
    
    for center in CENTERS:
        # الحصول على معلومات المركز
        center_doc = await db.centers.find_one({"name": center}, {"_id": 0})
        center_id = center_doc.get("id") if center_doc else None
        
        for wh_type, wh_config in WAREHOUSE_CATEGORIES.items():
            # إنشاء المخزن الرئيسي (داخلي/خارجي)
            main_code = f"{center[:3].upper()}-{wh_type[:3].upper()}"
            existing_main = await db.warehouses.find_one({"code": main_code})
            
            main_warehouse_id = None
            if not existing_main:
                main_warehouse = Warehouse(
                    name=f"{wh_config['name_ar']} - {center}",
                    code=main_code,
                    location=center,
                    warehouse_type=wh_type,
                    center_id=center_id,
                    center_name=center,
                    alert_on_low_stock=True,
                    alert_on_expiry=True
                )
                await db.warehouses.insert_one(main_warehouse.model_dump())
                main_warehouse_id = main_warehouse.id
                created_count += 1
            else:
                main_warehouse_id = existing_main.get("id")
                skipped_count += 1
            
            # إنشاء المخازن الفرعية
            for sub_wh in wh_config["sub_warehouses"]:
                sub_code = f"{center[:3].upper()}-{sub_wh['category'][:4].upper()}"
                existing_sub = await db.warehouses.find_one({"code": sub_code})
                
                if not existing_sub:
                    sub_warehouse = Warehouse(
                        name=f"{sub_wh['name_ar']} - {center}",
                        code=sub_code,
                        location=center,
                        warehouse_type=wh_type,
                        warehouse_category=sub_wh["category"],
                        center_id=center_id,
                        center_name=center,
                        parent_warehouse_id=main_warehouse_id,
                        parent_warehouse_name=f"{wh_config['name_ar']} - {center}",
                        temperature_controlled=sub_wh.get("temp_controlled", False),
                        alert_on_low_stock=True,
                        alert_on_expiry=sub_wh.get("expiry_tracking", False)
                    )
                    await db.warehouses.insert_one(sub_warehouse.model_dump())
                    created_count += 1
                else:
                    skipped_count += 1
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", ""),
        action="initialize_warehouses",
        entity_type="warehouse",
        entity_id="all",
        entity_name="جميع المخازن",
        details=f"تم إنشاء {created_count} مخزن جديد، تم تخطي {skipped_count} موجود"
    )
    
    return {
        "message": f"تم إنشاء {created_count} مخزن بنجاح",
        "created": created_count,
        "skipped": skipped_count,
        "centers": CENTERS
    }


@router.post("/warehouses")
async def create_warehouse(
    data: WarehouseBase,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء مخزن جديد"""
    # التحقق من عدم تكرار الرمز
    existing = await db.warehouses.find_one({"code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail="رمز المخزن موجود مسبقاً")
    
    warehouse = Warehouse(**data.model_dump())
    await db.warehouses.insert_one(warehouse.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", ""),
        action="create_warehouse",
        entity_type="warehouse",
        entity_id=warehouse.id,
        entity_name=warehouse.name,
        details=f"إنشاء مخزن: {warehouse.name}"
    )
    
    return warehouse.model_dump()


@router.put("/warehouses/{warehouse_id}")
async def update_warehouse(
    warehouse_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تحديث مخزن"""
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.warehouses.update_one(
        {"id": warehouse_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="المخزن غير موجود")
    
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    return warehouse


@router.delete("/warehouses/{warehouse_id}")
async def delete_warehouse(
    warehouse_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف مخزن"""
    # التحقق من عدم وجود مخزون
    stock_count = await db.product_stock.count_documents({"warehouse_id": warehouse_id})
    if stock_count > 0:
        raise HTTPException(status_code=400, detail="لا يمكن حذف المخزن - يوجد مخزون مرتبط")
    
    await db.warehouses.delete_one({"id": warehouse_id})
    return {"message": "تم حذف المخزن بنجاح"}


# ==================== فئات المنتجات ====================

@router.get("/categories")
async def get_categories(current_user: dict = Depends(get_current_user)):
    """الحصول على فئات المنتجات"""
    categories = await db.product_categories.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return categories


@router.post("/categories")
async def create_category(
    data: ProductCategoryBase,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء فئة منتج"""
    category = ProductCategory(**data.model_dump())
    await db.product_categories.insert_one(category.model_dump())
    return category.model_dump()


# ==================== المنتجات ====================

@router.get("/products")
async def get_products(
    category_id: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على قائمة المنتجات"""
    query = {}
    if category_id:
        query["category_id"] = category_id
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}}
        ]
    
    products = await db.warehouse_products.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return products


@router.post("/products")
async def create_product(
    data: WarehouseProductBase,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء منتج جديد"""
    existing = await db.warehouse_products.find_one({"code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail="رمز المنتج موجود مسبقاً")
    
    product = WarehouseProduct(**data.model_dump())
    await db.warehouse_products.insert_one(product.model_dump())
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", ""),
        action="create_product",
        entity_type="product",
        entity_id=product.id,
        entity_name=product.name,
        details=f"إنشاء منتج: {product.name}"
    )
    
    return product.model_dump()


@router.put("/products/{product_id}")
async def update_product(
    product_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تحديث منتج"""
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.warehouse_products.update_one(
        {"id": product_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    product = await db.warehouse_products.find_one({"id": product_id}, {"_id": 0})
    return product


@router.delete("/products/{product_id}")
async def delete_product(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف منتج"""
    stock_count = await db.product_stock.count_documents({"product_id": product_id, "quantity": {"$gt": 0}})
    if stock_count > 0:
        raise HTTPException(status_code=400, detail="لا يمكن حذف المنتج - يوجد مخزون")
    
    await db.warehouse_products.delete_one({"id": product_id})
    return {"message": "تم حذف المنتج بنجاح"}


# ==================== مخزون المنتجات ====================

@router.get("/stock")
async def get_stock(
    warehouse_id: Optional[str] = None,
    product_id: Optional[str] = None,
    low_stock: Optional[bool] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على مخزون المنتجات"""
    query = {}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    if product_id:
        query["product_id"] = product_id
    if search:
        query["$or"] = [
            {"product_name": {"$regex": search, "$options": "i"}},
            {"product_code": {"$regex": search, "$options": "i"}}
        ]
    
    stock = await db.product_stock.find(query, {"_id": 0}).sort("product_name", 1).to_list(1000)
    
    # Filter low stock if requested
    if low_stock:
        # Get min quantities from products
        products = await db.warehouse_products.find({}, {"_id": 0, "id": 1, "min_quantity": 1}).to_list(500)
        min_qty_map = {p["id"]: p.get("min_quantity", 0) for p in products}
        stock = [s for s in stock if s["quantity"] <= min_qty_map.get(s["product_id"], 0)]
    
    return stock


@router.get("/stock/summary")
async def get_stock_summary(current_user: dict = Depends(get_current_user)):
    """ملخص المخزون"""
    # إجمالي المنتجات
    total_products = await db.warehouse_products.count_documents({})
    
    # إجمالي المخازن
    total_warehouses = await db.warehouses.count_documents({})
    
    # المنتجات منخفضة المخزون
    products = await db.warehouse_products.find({}, {"_id": 0, "id": 1, "min_quantity": 1}).to_list(500)
    min_qty_map = {p["id"]: p.get("min_quantity", 0) for p in products}
    
    stock = await db.product_stock.find({}, {"_id": 0}).to_list(1000)
    low_stock_count = sum(1 for s in stock if s["quantity"] <= min_qty_map.get(s["product_id"], 0))
    
    # إجمالي قيمة المخزون
    total_value = sum(s.get("quantity", 0) * s.get("unit_price", 0) for s in stock if s.get("unit_price"))
    
    # المنتجات منتهية الصلاحية
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    expired_count = await db.product_stock.count_documents({
        "expiry_date": {"$lt": today, "$ne": None}
    })
    
    # حركات اليوم
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0).isoformat()
    today_movements = await db.stock_movements.count_documents({
        "movement_date": {"$gte": today_start}
    })
    
    return {
        "total_products": total_products,
        "total_warehouses": total_warehouses,
        "low_stock_count": low_stock_count,
        "total_value": round(total_value, 2),
        "expired_count": expired_count,
        "today_movements": today_movements
    }


@router.post("/stock/adjust")
async def adjust_stock(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تعديل مخزون (جرد)"""
    product_id = data.get("product_id")
    warehouse_id = data.get("warehouse_id")
    new_quantity = data.get("new_quantity")
    reason = data.get("reason", "تعديل جرد")
    
    # البحث عن المخزون الحالي
    stock = await db.product_stock.find_one({
        "product_id": product_id,
        "warehouse_id": warehouse_id
    }, {"_id": 0})
    
    if not stock:
        raise HTTPException(status_code=404, detail="المخزون غير موجود")
    
    old_quantity = stock.get("quantity", 0)
    difference = new_quantity - old_quantity
    
    # تحديث المخزون
    await db.product_stock.update_one(
        {"id": stock["id"]},
        {"$set": {
            "quantity": new_quantity,
            "available_quantity": new_quantity - stock.get("reserved_quantity", 0),
            "last_updated": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # إنشاء حركة تعديل
    movement = StockMovement(
        movement_type="adjust",
        movement_number=f"ADJ-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        product_id=product_id,
        product_name=stock.get("product_name", ""),
        product_code=stock.get("product_code", ""),
        quantity=abs(difference),
        from_warehouse_id=warehouse_id if difference < 0 else None,
        from_warehouse_name=stock.get("warehouse_name") if difference < 0 else None,
        to_warehouse_id=warehouse_id if difference > 0 else None,
        to_warehouse_name=stock.get("warehouse_name") if difference > 0 else None,
        notes=f"{reason} - الكمية القديمة: {old_quantity}, الكمية الجديدة: {new_quantity}",
        created_by=current_user["id"],
        created_by_name=current_user.get("full_name", "")
    )
    
    await db.stock_movements.insert_one(movement.model_dump())
    
    return {"message": "تم تعديل المخزون بنجاح", "old_quantity": old_quantity, "new_quantity": new_quantity}


# ==================== حركات المخزون ====================

@router.get("/movements")
async def get_movements(
    movement_type: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    product_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على حركات المخزون"""
    query = {}
    if movement_type:
        query["movement_type"] = movement_type
    if warehouse_id:
        query["$or"] = [
            {"from_warehouse_id": warehouse_id},
            {"to_warehouse_id": warehouse_id}
        ]
    if product_id:
        query["product_id"] = product_id
    if start_date:
        query.setdefault("movement_date", {})["$gte"] = start_date
    if end_date:
        query.setdefault("movement_date", {})["$lte"] = end_date + "T23:59:59"
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("movement_date", -1).to_list(limit)
    return movements


@router.post("/movements/receive")
async def receive_stock(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    استلام بضاعة مع تكامل مالي
    قيد مالي: من حـ/ المخزون إلى حـ/ الدائنون (أو النقدية)
    """
    product_id = data.get("product_id")
    warehouse_id = data.get("warehouse_id")
    quantity = data.get("quantity")
    unit_price = data.get("unit_price", 0)
    supplier_id = data.get("supplier_id")
    supplier_name = data.get("supplier_name")
    batch_number = data.get("batch_number")
    expiry_date = data.get("expiry_date")
    reference_number = data.get("reference_number")
    notes = data.get("notes")
    payment_type = data.get("payment_type", "credit")  # credit=آجل, cash=نقدي
    create_journal = data.get("create_journal", True)  # إنشاء قيد مالي
    
    # الحصول على بيانات المنتج والمخزن
    product = await db.warehouse_products.find_one({"id": product_id}, {"_id": 0})
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    if not warehouse:
        raise HTTPException(status_code=404, detail="المخزن غير موجود")
    
    total_value = quantity * unit_price
    
    # تحديث أو إنشاء سجل المخزون
    stock = await db.product_stock.find_one({
        "product_id": product_id,
        "warehouse_id": warehouse_id
    }, {"_id": 0})
    
    if stock:
        new_quantity = stock.get("quantity", 0) + quantity
        # تحديث سعر التكلفة بطريقة المتوسط المرجح
        old_value = stock.get("quantity", 0) * stock.get("unit_price", unit_price)
        new_value = old_value + total_value
        avg_price = new_value / new_quantity if new_quantity > 0 else unit_price
        
        await db.product_stock.update_one(
            {"id": stock["id"]},
            {"$set": {
                "quantity": new_quantity,
                "available_quantity": new_quantity - stock.get("reserved_quantity", 0),
                "unit_price": avg_price,
                "batch_number": batch_number or stock.get("batch_number"),
                "expiry_date": expiry_date or stock.get("expiry_date"),
                "last_updated": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        new_stock = ProductStock(
            product_id=product_id,
            product_name=product.get("name", ""),
            product_code=product.get("code", ""),
            warehouse_id=warehouse_id,
            warehouse_name=warehouse.get("name", ""),
            quantity=quantity,
            available_quantity=quantity,
            batch_number=batch_number,
            expiry_date=expiry_date
        )
        new_stock_dict = new_stock.model_dump()
        new_stock_dict["unit_price"] = unit_price
        await db.product_stock.insert_one(new_stock_dict)
    
    # إنشاء حركة استلام
    movement = StockMovement(
        movement_type="receive",
        movement_number=f"RCV-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        product_id=product_id,
        product_name=product.get("name", ""),
        product_code=product.get("code", ""),
        quantity=quantity,
        unit_price=unit_price,
        total_value=total_value,
        to_warehouse_id=warehouse_id,
        to_warehouse_name=warehouse.get("name", ""),
        reference_type="purchase",
        reference_number=reference_number,
        batch_number=batch_number,
        expiry_date=expiry_date,
        supplier_id=supplier_id,
        supplier_name=supplier_name,
        notes=notes,
        created_by=current_user["id"],
        created_by_name=current_user.get("full_name", "")
    )
    
    await db.stock_movements.insert_one(movement.model_dump())
    
    # إنشاء قيد مالي آلي (إذا كانت القيمة > 0)
    journal_entry = None
    if create_journal and total_value > 0:
        credit_account = FINANCE_ACCOUNTS["cash"] if payment_type == "cash" else FINANCE_ACCOUNTS["accounts_payable"]
        
        journal_entry = await create_warehouse_journal_entry(
            description=f"شراء مخزون: {product.get('name', '')} - {quantity} {product.get('unit', 'قطعة')} من {supplier_name or 'مورد'}",
            lines=[
                {"account_number": FINANCE_ACCOUNTS["inventory"], "debit": total_value, "credit": 0, "description": f"استلام مخزون - {product.get('name', '')}"},
                {"account_number": credit_account, "debit": 0, "credit": total_value, "description": f"مشتريات مخزون - {supplier_name or 'مورد'}"}
            ],
            reference_type="warehouse_receive",
            reference_id=movement.id,
            created_by_id=current_user["id"],
            created_by_name=current_user.get("full_name", "")
        )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", ""),
        action="receive_stock",
        entity_type="stock",
        entity_id=product_id,
        entity_name=product.get("name", ""),
        details=f"استلام {quantity} {product.get('unit', 'قطعة')} من {product.get('name', '')} في مخزن {warehouse.get('name', '')} بقيمة {total_value}"
    )
    
    return {
        "message": "تم استلام البضاعة بنجاح",
        "movement": movement.model_dump(),
        "journal_entry": journal_entry.get("entry_number") if journal_entry else None,
        "total_value": total_value
    }


@router.post("/movements/issue")
async def issue_stock(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    صرف بضاعة مع تكامل مالي
    قيد مالي: من حـ/ المصروفات (حسب نوع المخزن) إلى حـ/ المخزون
    للمبيعات: من حـ/ تكلفة البضاعة المباعة إلى حـ/ المخزون
    """
    product_id = data.get("product_id")
    warehouse_id = data.get("warehouse_id")
    quantity = data.get("quantity")
    unit_price = data.get("unit_price", 0)
    customer_id = data.get("customer_id")
    customer_name = data.get("customer_name")
    reference_number = data.get("reference_number")
    notes = data.get("notes")
    issue_type = data.get("issue_type", "consumption")  # consumption=استهلاك داخلي, sales=مبيعات
    create_journal = data.get("create_journal", True)
    
    # الحصول على المخزون الحالي
    stock = await db.product_stock.find_one({
        "product_id": product_id,
        "warehouse_id": warehouse_id
    }, {"_id": 0})
    
    if not stock:
        raise HTTPException(status_code=404, detail="المخزون غير موجود")
    
    if stock.get("available_quantity", 0) < quantity:
        raise HTTPException(status_code=400, detail="الكمية المطلوبة غير متوفرة")
    
    # استخدام سعر التكلفة من المخزون إذا لم يُحدد
    cost_price = stock.get("unit_price", unit_price) or unit_price
    total_value = quantity * cost_price
    
    # تحديث المخزون
    new_quantity = stock.get("quantity", 0) - quantity
    await db.product_stock.update_one(
        {"id": stock["id"]},
        {"$set": {
            "quantity": new_quantity,
            "available_quantity": new_quantity - stock.get("reserved_quantity", 0),
            "last_updated": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    
    # إنشاء حركة صرف
    movement = StockMovement(
        movement_type="issue",
        movement_number=f"ISS-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        product_id=product_id,
        product_name=stock.get("product_name", ""),
        product_code=stock.get("product_code", ""),
        quantity=quantity,
        unit_price=cost_price,
        total_value=total_value,
        from_warehouse_id=warehouse_id,
        from_warehouse_name=warehouse.get("name", "") if warehouse else "",
        reference_type="sales" if issue_type == "sales" else "consumption",
        reference_number=reference_number,
        customer_id=customer_id,
        customer_name=customer_name,
        notes=notes,
        created_by=current_user["id"],
        created_by_name=current_user.get("full_name", "")
    )
    
    await db.stock_movements.insert_one(movement.model_dump())
    
    # إنشاء قيد مالي آلي
    journal_entry = None
    if create_journal and total_value > 0:
        if issue_type == "sales":
            # صرف للمبيعات: تكلفة البضاعة المباعة
            debit_account = FINANCE_ACCOUNTS["cogs"]
            description = f"تكلفة بيع: {stock.get('product_name', '')} - {quantity} للعميل {customer_name or ''}"
        else:
            # صرف للاستهلاك الداخلي: حسب نوع المخزن
            debit_account = await get_expense_account_for_warehouse(warehouse_id)
            description = f"صرف مخزون: {stock.get('product_name', '')} - {quantity} من {warehouse.get('name', '')}"
        
        journal_entry = await create_warehouse_journal_entry(
            description=description,
            lines=[
                {"account_number": debit_account, "debit": total_value, "credit": 0, "description": f"صرف مخزون - {stock.get('product_name', '')}"},
                {"account_number": FINANCE_ACCOUNTS["inventory"], "debit": 0, "credit": total_value, "description": f"تخفيض المخزون"}
            ],
            reference_type="warehouse_issue",
            reference_id=movement.id,
            created_by_id=current_user["id"],
            created_by_name=current_user.get("full_name", "")
        )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", ""),
        action="issue_stock",
        entity_type="stock",
        entity_id=product_id,
        entity_name=stock.get("product_name", ""),
        details=f"صرف {quantity} من {stock.get('product_name', '')} بقيمة {total_value}"
    )
    
    return {
        "message": "تم صرف البضاعة بنجاح",
        "movement": movement.model_dump(),
        "journal_entry": journal_entry.get("entry_number") if journal_entry else None,
        "total_value": total_value
    }


@router.post("/movements/transfer")
async def transfer_stock(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تحويل بضاعة بين المخازن"""
    product_id = data.get("product_id")
    from_warehouse_id = data.get("from_warehouse_id")
    to_warehouse_id = data.get("to_warehouse_id")
    quantity = data.get("quantity")
    notes = data.get("notes")
    
    if from_warehouse_id == to_warehouse_id:
        raise HTTPException(status_code=400, detail="لا يمكن التحويل لنفس المخزن")
    
    # الحصول على المخزون المصدر
    from_stock = await db.product_stock.find_one({
        "product_id": product_id,
        "warehouse_id": from_warehouse_id
    }, {"_id": 0})
    
    if not from_stock:
        raise HTTPException(status_code=404, detail="المخزون المصدر غير موجود")
    
    if from_stock.get("available_quantity", 0) < quantity:
        raise HTTPException(status_code=400, detail="الكمية المطلوبة غير متوفرة")
    
    # الحصول على بيانات المخازن
    from_warehouse = await db.warehouses.find_one({"id": from_warehouse_id}, {"_id": 0})
    to_warehouse = await db.warehouses.find_one({"id": to_warehouse_id}, {"_id": 0})
    product = await db.warehouse_products.find_one({"id": product_id}, {"_id": 0})
    
    # تحديث المخزون المصدر (نقص)
    new_from_qty = from_stock.get("quantity", 0) - quantity
    await db.product_stock.update_one(
        {"id": from_stock["id"]},
        {"$set": {
            "quantity": new_from_qty,
            "available_quantity": new_from_qty - from_stock.get("reserved_quantity", 0),
            "last_updated": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # تحديث أو إنشاء المخزون الوجهة
    to_stock = await db.product_stock.find_one({
        "product_id": product_id,
        "warehouse_id": to_warehouse_id
    }, {"_id": 0})
    
    if to_stock:
        new_to_qty = to_stock.get("quantity", 0) + quantity
        await db.product_stock.update_one(
            {"id": to_stock["id"]},
            {"$set": {
                "quantity": new_to_qty,
                "available_quantity": new_to_qty - to_stock.get("reserved_quantity", 0),
                "last_updated": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        new_stock = ProductStock(
            product_id=product_id,
            product_name=from_stock.get("product_name", ""),
            product_code=from_stock.get("product_code", ""),
            warehouse_id=to_warehouse_id,
            warehouse_name=to_warehouse.get("name", "") if to_warehouse else "",
            quantity=quantity,
            available_quantity=quantity,
            batch_number=from_stock.get("batch_number"),
            expiry_date=from_stock.get("expiry_date")
        )
        await db.product_stock.insert_one(new_stock.model_dump())
    
    # إنشاء حركة تحويل
    movement = StockMovement(
        movement_type="transfer",
        movement_number=f"TRF-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        product_id=product_id,
        product_name=from_stock.get("product_name", ""),
        product_code=from_stock.get("product_code", ""),
        quantity=quantity,
        from_warehouse_id=from_warehouse_id,
        from_warehouse_name=from_warehouse.get("name", "") if from_warehouse else "",
        to_warehouse_id=to_warehouse_id,
        to_warehouse_name=to_warehouse.get("name", "") if to_warehouse else "",
        reference_type="transfer",
        notes=notes,
        created_by=current_user["id"],
        created_by_name=current_user.get("full_name", "")
    )
    
    await db.stock_movements.insert_one(movement.model_dump())
    
    return {"message": "تم التحويل بنجاح", "movement": movement.model_dump()}


# ==================== المحاليل والفحوصات ====================

@router.get("/solutions")
async def get_solutions(
    solution_type: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    low_stock: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على قائمة المحاليل"""
    query = {}
    if solution_type:
        query["solution_type"] = solution_type
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    
    solutions = await db.lab_solutions.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    
    if low_stock:
        solutions = [s for s in solutions if s.get("current_quantity", 0) <= s.get("min_quantity", 0)]
    
    return solutions


@router.post("/solutions")
async def create_solution(
    data: LabSolutionBase,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء محلول جديد"""
    existing = await db.lab_solutions.find_one({"code": data.code})
    if existing:
        raise HTTPException(status_code=400, detail="رمز المحلول موجود مسبقاً")
    
    solution = LabSolution(**data.model_dump())
    await db.lab_solutions.insert_one(solution.model_dump())
    return solution.model_dump()


@router.put("/solutions/{solution_id}")
async def update_solution(
    solution_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تحديث محلول"""
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.lab_solutions.update_one(
        {"id": solution_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="المحلول غير موجود")
    
    solution = await db.lab_solutions.find_one({"id": solution_id}, {"_id": 0})
    return solution


@router.delete("/solutions/{solution_id}")
async def delete_solution(
    solution_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف محلول"""
    await db.lab_solutions.delete_one({"id": solution_id})
    return {"message": "تم حذف المحلول بنجاح"}


# ==================== استهلاك المحاليل ====================

@router.get("/solutions/consumption")
async def get_solution_consumption(
    solution_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على سجلات استهلاك المحاليل"""
    query = {}
    if solution_id:
        query["solution_id"] = solution_id
    if start_date:
        query.setdefault("consumption_date", {})["$gte"] = start_date
    if end_date:
        query.setdefault("consumption_date", {})["$lte"] = end_date
    
    consumption = await db.solution_consumption.find(query, {"_id": 0}).sort("consumption_date", -1).to_list(500)
    return consumption


@router.post("/solutions/consumption")
async def record_consumption(
    data: SolutionConsumptionBase,
    current_user: dict = Depends(get_current_user)
):
    """تسجيل استهلاك محلول"""
    # التحقق من وجود المحلول
    solution = await db.lab_solutions.find_one({"id": data.solution_id}, {"_id": 0})
    if not solution:
        raise HTTPException(status_code=404, detail="المحلول غير موجود")
    
    # التحقق من توفر الكمية
    if solution.get("current_quantity", 0) < data.quantity_consumed:
        raise HTTPException(status_code=400, detail="الكمية المطلوبة غير متوفرة")
    
    # تحديث كمية المحلول
    new_quantity = solution.get("current_quantity", 0) - data.quantity_consumed
    await db.lab_solutions.update_one(
        {"id": data.solution_id},
        {"$set": {
            "current_quantity": new_quantity,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # إنشاء سجل الاستهلاك
    consumption = SolutionConsumption(
        **data.model_dump(),
        created_by=current_user["id"],
        created_by_name=current_user.get("full_name", "")
    )
    await db.solution_consumption.insert_one(consumption.model_dump())
    
    return consumption.model_dump()


@router.get("/solutions/consumption/daily-summary")
async def get_daily_consumption_summary(
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """ملخص الاستهلاك اليومي للمحاليل"""
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    consumption = await db.solution_consumption.find(
        {"consumption_date": date},
        {"_id": 0}
    ).to_list(500)
    
    # تجميع حسب المحلول
    summary = {}
    for c in consumption:
        sol_id = c.get("solution_id")
        if sol_id not in summary:
            summary[sol_id] = {
                "solution_id": sol_id,
                "solution_name": c.get("solution_name"),
                "solution_code": c.get("solution_code"),
                "total_consumed": 0,
                "total_tests": 0,
                "records": []
            }
        summary[sol_id]["total_consumed"] += c.get("quantity_consumed", 0)
        summary[sol_id]["total_tests"] += c.get("test_count", 0)
        summary[sol_id]["records"].append(c)
    
    return {
        "date": date,
        "solutions": list(summary.values()),
        "total_consumption_records": len(consumption)
    }


# ==================== التقارير والتصدير ====================

@router.get("/reports/stock-value")
async def get_stock_value_report(
    warehouse_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير قيمة المخزون"""
    query = {}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    
    stock = await db.product_stock.find(query, {"_id": 0}).to_list(1000)
    products = await db.warehouse_products.find({}, {"_id": 0}).to_list(500)
    
    product_prices = {p["id"]: p.get("cost_price", 0) for p in products}
    
    report = []
    total_value = 0
    
    for s in stock:
        unit_price = product_prices.get(s["product_id"], 0)
        value = s.get("quantity", 0) * unit_price
        total_value += value
        
        report.append({
            "product_id": s.get("product_id"),
            "product_name": s.get("product_name"),
            "product_code": s.get("product_code"),
            "warehouse_name": s.get("warehouse_name"),
            "quantity": s.get("quantity", 0),
            "unit_price": unit_price,
            "total_value": round(value, 2)
        })
    
    return {
        "report": report,
        "total_value": round(total_value, 2),
        "total_items": len(report)
    }


@router.get("/reports/movements-summary")
async def get_movements_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """ملخص حركات المخزون"""
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    query = {
        "movement_date": {"$gte": start_date, "$lte": end_date + "T23:59:59"}
    }
    
    movements = await db.stock_movements.find(query, {"_id": 0}).to_list(5000)
    
    summary = {
        "receive": {"count": 0, "quantity": 0, "value": 0},
        "issue": {"count": 0, "quantity": 0, "value": 0},
        "transfer": {"count": 0, "quantity": 0, "value": 0},
        "adjust": {"count": 0, "quantity": 0, "value": 0},
        "return": {"count": 0, "quantity": 0, "value": 0}
    }
    
    for m in movements:
        mt = m.get("movement_type", "")
        if mt in summary:
            summary[mt]["count"] += 1
            summary[mt]["quantity"] += m.get("quantity", 0)
            summary[mt]["value"] += m.get("total_value", 0)
    
    return {
        "start_date": start_date,
        "end_date": end_date,
        "summary": summary,
        "total_movements": len(movements)
    }


@router.get("/export/stock/excel")
async def export_stock_excel(
    warehouse_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير المخزون إلى Excel"""
    query = {}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    
    stock = await db.product_stock.find(query, {"_id": 0}).sort("product_name", 1).to_list(1000)
    products = await db.warehouse_products.find({}, {"_id": 0}).to_list(500)
    product_prices = {p["id"]: p.get("cost_price", 0) for p in products}
    
    # إنشاء DataFrame
    data = []
    for s in stock:
        unit_price = product_prices.get(s["product_id"], 0)
        data.append({
            "رمز المنتج": s.get("product_code", ""),
            "اسم المنتج": s.get("product_name", ""),
            "المخزن": s.get("warehouse_name", ""),
            "الكمية": s.get("quantity", 0),
            "الكمية المتاحة": s.get("available_quantity", 0),
            "الكمية المحجوزة": s.get("reserved_quantity", 0),
            "رقم الدفعة": s.get("batch_number", ""),
            "تاريخ الصلاحية": s.get("expiry_date", ""),
            "سعر الوحدة": unit_price,
            "القيمة الإجمالية": round(s.get("quantity", 0) * unit_price, 2)
        })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='المخزون', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['المخزون']
        
        # تنسيق العناوين
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # تعديل عرض الأعمدة
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    output.seek(0)
    
    filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/movements/excel")
async def export_movements_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    movement_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير حركات المخزون إلى Excel"""
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    query = {
        "movement_date": {"$gte": start_date, "$lte": end_date + "T23:59:59"}
    }
    if movement_type:
        query["movement_type"] = movement_type
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("movement_date", -1).to_list(5000)
    
    # ترجمة أنواع الحركات
    movement_types = {
        "receive": "استلام",
        "issue": "صرف",
        "transfer": "تحويل",
        "adjust": "تعديل",
        "return": "إرجاع"
    }
    
    data = []
    for m in movements:
        data.append({
            "رقم الحركة": m.get("movement_number", ""),
            "التاريخ": m.get("movement_date", "")[:10] if m.get("movement_date") else "",
            "نوع الحركة": movement_types.get(m.get("movement_type", ""), m.get("movement_type", "")),
            "رمز المنتج": m.get("product_code", ""),
            "اسم المنتج": m.get("product_name", ""),
            "الكمية": m.get("quantity", 0),
            "سعر الوحدة": m.get("unit_price", 0),
            "القيمة": m.get("total_value", 0),
            "من مخزن": m.get("from_warehouse_name", ""),
            "إلى مخزن": m.get("to_warehouse_name", ""),
            "المورد": m.get("supplier_name", ""),
            "العميل": m.get("customer_name", ""),
            "ملاحظات": m.get("notes", ""),
            "المسؤول": m.get("created_by_name", "")
        })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='حركات المخزون', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['حركات المخزون']
        
        header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    output.seek(0)
    
    filename = f"movements_report_{start_date}_to_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/solutions/excel")
async def export_solutions_excel(
    current_user: dict = Depends(get_current_user)
):
    """تصدير المحاليل إلى Excel"""
    solutions = await db.lab_solutions.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    
    solution_types = {
        "reagent": "كاشف",
        "buffer": "محلول منظم",
        "standard": "محلول قياسي",
        "cleaning": "محلول تنظيف"
    }
    
    data = []
    for s in solutions:
        data.append({
            "رمز المحلول": s.get("code", ""),
            "اسم المحلول": s.get("name", ""),
            "النوع": solution_types.get(s.get("solution_type", ""), s.get("solution_type", "")),
            "الوحدة": s.get("unit", ""),
            "الكمية الحالية": s.get("current_quantity", 0),
            "الحد الأدنى": s.get("min_quantity", 0),
            "المخزن": s.get("warehouse_name", ""),
            "تاريخ الصلاحية": s.get("expiry_date", ""),
            "رقم الدفعة": s.get("batch_number", ""),
            "المورد": s.get("supplier_name", ""),
            "التكلفة/وحدة": s.get("cost_per_unit", 0),
            "الحالة": "نشط" if s.get("status") == "active" else "غير نشط"
        })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='المحاليل', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['المحاليل']
        
        header_fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    output.seek(0)
    
    filename = f"solutions_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ==================== التنبيهات ====================

@router.get("/alerts")
async def get_stock_alerts(
    alert_type: Optional[str] = None,
    priority: Optional[str] = None,
    is_resolved: Optional[bool] = None,
    center_name: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على تنبيهات المخزون"""
    query = {}
    if alert_type:
        query["alert_type"] = alert_type
    if priority:
        query["priority"] = priority
    if is_resolved is not None:
        query["is_resolved"] = is_resolved
    if center_name:
        query["center_name"] = center_name
    
    alerts = await db.stock_alerts.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return alerts


@router.get("/alerts/summary")
async def get_alerts_summary(current_user: dict = Depends(get_current_user)):
    """ملخص التنبيهات"""
    pipeline = [
        {"$match": {"is_resolved": False}},
        {"$group": {
            "_id": {"alert_type": "$alert_type", "priority": "$priority"},
            "count": {"$sum": 1}
        }}
    ]
    
    results = await db.stock_alerts.aggregate(pipeline).to_list(100)
    
    summary = {
        "total_unresolved": 0,
        "by_type": {},
        "by_priority": {"critical": 0, "high": 0, "medium": 0, "low": 0}
    }
    
    for r in results:
        count = r["count"]
        alert_type = r["_id"]["alert_type"]
        priority = r["_id"]["priority"]
        
        summary["total_unresolved"] += count
        
        if alert_type not in summary["by_type"]:
            summary["by_type"][alert_type] = 0
        summary["by_type"][alert_type] += count
        
        if priority in summary["by_priority"]:
            summary["by_priority"][priority] += count
    
    return summary


@router.post("/alerts/{alert_id}/resolve")
async def resolve_alert(
    alert_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حل/إغلاق تنبيه"""
    result = await db.stock_alerts.update_one(
        {"id": alert_id},
        {"$set": {
            "is_resolved": True,
            "resolved_at": datetime.now(timezone.utc).isoformat(),
            "resolved_by": current_user.get("full_name", "")
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="التنبيه غير موجود")
    
    return {"message": "تم حل التنبيه بنجاح"}


@router.post("/alerts/check")
async def trigger_stock_check(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_role(["admin", "warehouse_manager"]))
):
    """تشغيل فحص المخزون وإنشاء التنبيهات"""
    background_tasks.add_task(check_stock_alerts)
    return {"message": "تم بدء فحص المخزون"}


@router.delete("/alerts/{alert_id}")
async def delete_alert(
    alert_id: str,
    current_user: dict = Depends(require_role(["admin"]))
):
    """حذف تنبيه"""
    result = await db.stock_alerts.delete_one({"id": alert_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="التنبيه غير موجود")
    return {"message": "تم حذف التنبيه"}


# ==================== إعدادات التنبيهات ====================

@router.get("/alerts/settings")
async def get_alert_settings(current_user: dict = Depends(get_current_user)):
    """الحصول على إعدادات التنبيهات"""
    settings = await db.system_settings.find_one({"key": "warehouse_alert_settings"}, {"_id": 0})
    if not settings:
        return {
            "email_enabled": True,
            "sms_enabled": False,
            "in_app_enabled": True,
            "default_alert_recipients": []
        }
    return settings.get("value", {})


@router.put("/alerts/settings")
async def update_alert_settings(
    data: dict,
    current_user: dict = Depends(require_role(["admin"]))
):
    """تحديث إعدادات التنبيهات"""
    await db.system_settings.update_one(
        {"key": "warehouse_alert_settings"},
        {"$set": {"key": "warehouse_alert_settings", "value": data}},
        upsert=True
    )
    return {"message": "تم تحديث الإعدادات"}


@router.put("/warehouses/{warehouse_id}/alert-recipients")
async def update_warehouse_alert_recipients(
    warehouse_id: str,
    data: dict,
    current_user: dict = Depends(require_role(["admin", "warehouse_manager"]))
):
    """تحديث مستلمي التنبيهات للمخزن"""
    update_data = {}
    
    if "supervisor_id" in data:
        update_data["supervisor_id"] = data["supervisor_id"]
    if "supervisor_name" in data:
        update_data["supervisor_name"] = data["supervisor_name"]
    if "supervisor_email" in data:
        update_data["supervisor_email"] = data["supervisor_email"]
    if "supervisor_phone" in data:
        update_data["supervisor_phone"] = data["supervisor_phone"]
    if "warehouse_manager_id" in data:
        update_data["warehouse_manager_id"] = data["warehouse_manager_id"]
    if "warehouse_manager_name" in data:
        update_data["warehouse_manager_name"] = data["warehouse_manager_name"]
    if "warehouse_manager_email" in data:
        update_data["warehouse_manager_email"] = data["warehouse_manager_email"]
    if "warehouse_manager_phone" in data:
        update_data["warehouse_manager_phone"] = data["warehouse_manager_phone"]
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        result = await db.warehouses.update_one(
            {"id": warehouse_id},
            {"$set": update_data}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="المخزن غير موجود")
    
    return {"message": "تم تحديث مستلمي التنبيهات"}


# ==================== المراكز ====================

@router.get("/centers")
async def get_centers(current_user: dict = Depends(get_current_user)):
    """الحصول على قائمة المراكز"""
    return CENTERS


@router.get("/warehouse-categories")
async def get_warehouse_categories(current_user: dict = Depends(get_current_user)):
    """الحصول على تصنيفات المخازن"""
    return WAREHOUSE_CATEGORIES


# ==================== التكامل مع المبيعات ====================

@router.get("/stock/check-availability")
async def check_stock_availability(
    product_id: str,
    warehouse_id: Optional[str] = None,
    required_quantity: float = 1,
    current_user: dict = Depends(get_current_user)
):
    """
    التحقق من توفر الكمية المطلوبة
    يُستخدم قبل إتمام عملية البيع
    """
    query = {"product_id": product_id}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    
    stock_items = await db.product_stock.find(query, {"_id": 0}).to_list(100)
    
    if not stock_items:
        return {
            "available": False,
            "message": "المنتج غير موجود في المخزون",
            "total_available": 0,
            "required": required_quantity
        }
    
    total_available = sum(s.get("available_quantity", 0) for s in stock_items)
    
    return {
        "available": total_available >= required_quantity,
        "message": "الكمية متوفرة" if total_available >= required_quantity else "الكمية غير كافية",
        "total_available": total_available,
        "required": required_quantity,
        "stock_by_warehouse": [
            {
                "warehouse_id": s.get("warehouse_id"),
                "warehouse_name": s.get("warehouse_name"),
                "available_quantity": s.get("available_quantity", 0),
                "unit_price": s.get("unit_price", 0)
            }
            for s in stock_items if s.get("available_quantity", 0) > 0
        ]
    }


@router.post("/stock/reserve")
async def reserve_stock(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    حجز كمية من المخزون لعملية بيع
    يُستخدم عند إنشاء عرض سعر أو طلب بيع
    """
    product_id = data.get("product_id")
    warehouse_id = data.get("warehouse_id")
    quantity = data.get("quantity")
    reference_type = data.get("reference_type", "sales_order")  # sales_order, quotation
    reference_id = data.get("reference_id")
    
    stock = await db.product_stock.find_one({
        "product_id": product_id,
        "warehouse_id": warehouse_id
    }, {"_id": 0})
    
    if not stock:
        raise HTTPException(status_code=404, detail="المخزون غير موجود")
    
    if stock.get("available_quantity", 0) < quantity:
        raise HTTPException(status_code=400, detail="الكمية المطلوبة غير متوفرة")
    
    # تحديث الكمية المحجوزة
    new_reserved = stock.get("reserved_quantity", 0) + quantity
    new_available = stock.get("quantity", 0) - new_reserved
    
    await db.product_stock.update_one(
        {"id": stock["id"]},
        {"$set": {
            "reserved_quantity": new_reserved,
            "available_quantity": new_available,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # تسجيل الحجز
    reservation = {
        "id": str(uuid.uuid4()),
        "product_id": product_id,
        "warehouse_id": warehouse_id,
        "quantity": quantity,
        "reference_type": reference_type,
        "reference_id": reference_id,
        "status": "active",
        "created_by": current_user["id"],
        "created_by_name": current_user.get("full_name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.stock_reservations.insert_one(reservation)
    
    return {
        "message": "تم حجز الكمية بنجاح",
        "reservation_id": reservation["id"],
        "reserved_quantity": quantity,
        "available_quantity": new_available
    }


@router.post("/stock/release-reservation/{reservation_id}")
async def release_stock_reservation(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    إلغاء حجز الكمية
    يُستخدم عند إلغاء عرض السعر أو الطلب
    """
    reservation = await db.stock_reservations.find_one({"id": reservation_id, "status": "active"}, {"_id": 0})
    
    if not reservation:
        raise HTTPException(status_code=404, detail="الحجز غير موجود أو تم إلغاؤه")
    
    stock = await db.product_stock.find_one({
        "product_id": reservation["product_id"],
        "warehouse_id": reservation["warehouse_id"]
    }, {"_id": 0})
    
    if stock:
        new_reserved = max(0, stock.get("reserved_quantity", 0) - reservation["quantity"])
        new_available = stock.get("quantity", 0) - new_reserved
        
        await db.product_stock.update_one(
            {"id": stock["id"]},
            {"$set": {
                "reserved_quantity": new_reserved,
                "available_quantity": new_available,
                "last_updated": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    await db.stock_reservations.update_one(
        {"id": reservation_id},
        {"$set": {
            "status": "released",
            "released_at": datetime.now(timezone.utc).isoformat(),
            "released_by": current_user.get("full_name", "")
        }}
    )
    
    return {"message": "تم إلغاء الحجز بنجاح"}


@router.post("/stock/issue-from-sale")
async def issue_stock_from_sale(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    صرف مخزون لعملية بيع محددة
    يُنشئ حركة صرف + قيد تكلفة البضاعة المباعة آلياً
    """
    sale_id = data.get("sale_id")
    sale_number = data.get("sale_number")
    customer_id = data.get("customer_id")
    customer_name = data.get("customer_name")
    items = data.get("items", [])  # [{product_id, warehouse_id, quantity}]
    
    if not items:
        raise HTTPException(status_code=400, detail="لا توجد منتجات للصرف")
    
    results = []
    total_cost = 0
    
    for item in items:
        product_id = item.get("product_id")
        warehouse_id = item.get("warehouse_id")
        quantity = item.get("quantity")
        
        # صرف المنتج
        issue_result = await issue_stock(
            data={
                "product_id": product_id,
                "warehouse_id": warehouse_id,
                "quantity": quantity,
                "customer_id": customer_id,
                "customer_name": customer_name,
                "reference_number": sale_number,
                "issue_type": "sales",
                "notes": f"صرف لعملية البيع رقم {sale_number}"
            },
            current_user=current_user
        )
        
        results.append({
            "product_id": product_id,
            "quantity": quantity,
            "movement": issue_result.get("movement", {}).get("movement_number"),
            "journal_entry": issue_result.get("journal_entry"),
            "cost": issue_result.get("total_value", 0)
        })
        
        total_cost += issue_result.get("total_value", 0)
        
        # إلغاء الحجز إذا كان موجوداً
        reservation = await db.stock_reservations.find_one({
            "product_id": product_id,
            "warehouse_id": warehouse_id,
            "reference_id": sale_id,
            "status": "active"
        }, {"_id": 0})
        
        if reservation:
            await release_stock_reservation(reservation["id"], current_user)
    
    return {
        "message": f"تم صرف {len(results)} منتج بنجاح",
        "sale_id": sale_id,
        "total_cost": total_cost,
        "items": results
    }


# ==================== تقارير التكامل المالي ====================

@router.get("/finance/stock-value-report")
async def get_stock_value_report(
    warehouse_id: Optional[str] = None,
    center_name: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    تقرير قيمة المخزون
    يُستخدم لمطابقة حساب المخزون في الميزانية
    """
    query = {}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    
    stock_items = await db.product_stock.find(query, {"_id": 0}).to_list(5000)
    
    # تصفية حسب المركز إذا طُلب
    if center_name:
        warehouses = await db.warehouses.find({"center_name": center_name}, {"_id": 0, "id": 1}).to_list(100)
        warehouse_ids = [w["id"] for w in warehouses]
        stock_items = [s for s in stock_items if s.get("warehouse_id") in warehouse_ids]
    
    # حساب القيم
    total_value = 0
    by_warehouse = {}
    by_category = {}
    
    for item in stock_items:
        value = item.get("quantity", 0) * item.get("unit_price", 0)
        total_value += value
        
        wh_id = item.get("warehouse_id", "unknown")
        wh_name = item.get("warehouse_name", "غير محدد")
        
        if wh_id not in by_warehouse:
            by_warehouse[wh_id] = {"name": wh_name, "value": 0, "items": 0}
        by_warehouse[wh_id]["value"] += value
        by_warehouse[wh_id]["items"] += 1
    
    return {
        "total_value": round(total_value, 3),
        "total_items": len(stock_items),
        "by_warehouse": list(by_warehouse.values()),
        "report_date": datetime.now(timezone.utc).isoformat()
    }


@router.get("/finance/movements-summary")
async def get_movements_financial_summary(
    start_date: str = Query(default=None),
    end_date: str = Query(default=None),
    current_user: dict = Depends(get_current_user)
):
    """
    ملخص مالي لحركات المخزون
    يُستخدم للمراجعة المالية ومطابقة القيود
    """
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    
    movements = await db.stock_movements.find({
        "movement_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(5000)
    
    summary = {
        "period": {"start": start_date, "end": end_date},
        "receive": {"count": 0, "total_value": 0},
        "issue": {"count": 0, "total_value": 0, "by_type": {"sales": 0, "consumption": 0}},
        "transfer": {"count": 0},
        "adjust": {"count": 0, "value_change": 0}
    }
    
    for m in movements:
        mtype = m.get("movement_type")
        value = m.get("total_value", 0)
        
        if mtype == "receive":
            summary["receive"]["count"] += 1
            summary["receive"]["total_value"] += value
        elif mtype == "issue":
            summary["issue"]["count"] += 1
            summary["issue"]["total_value"] += value
            ref_type = m.get("reference_type", "consumption")
            if ref_type in summary["issue"]["by_type"]:
                summary["issue"]["by_type"][ref_type] += value
        elif mtype == "transfer":
            summary["transfer"]["count"] += 1
        elif mtype == "adjust":
            summary["adjust"]["count"] += 1
            summary["adjust"]["value_change"] += value
    
    summary["net_change"] = summary["receive"]["total_value"] - summary["issue"]["total_value"]
    
    return summary


# ==================== طلبات صرف المواد ====================

# ربط تصنيف المخزن بالصلاحية المطلوبة
CATEGORY_TO_PERMISSION = {
    "lab": "warehouse_issue_lab",
    "cleaning": "warehouse_issue_cleaning",
    "maintenance": "warehouse_issue_maintenance",
    "ppe": "warehouse_issue_ppe",
    "feed": "warehouse_issue_feed",
    "equipment": "warehouse_issue_equipment",
    "supplies": "warehouse_issue_supplies",
}


@router.get("/my-warehouses")
async def get_my_warehouses(current_user: dict = Depends(get_current_user)):
    """
    الحصول على المخازن التي يمكن للمستخدم الصرف منها
    بناءً على صلاحياته
    """
    user_permissions = current_user.get("permissions", [])
    user_role = current_user.get("role", "")
    
    # المدير أو من لديه صلاحية الصرف من الكل
    if user_role == "admin" or "warehouse_issue_all" in user_permissions or "warehouse_stock_issue" in user_permissions:
        warehouses = await db.warehouses.find({"status": "active"}, {"_id": 0}).to_list(200)
        return warehouses
    
    # تحديد تصنيفات المخازن المتاحة بناءً على الصلاحيات
    allowed_categories = []
    for category, permission in CATEGORY_TO_PERMISSION.items():
        if permission in user_permissions:
            allowed_categories.append(category)
    
    if not allowed_categories:
        return []
    
    # جلب المخازن حسب التصنيفات المتاحة
    warehouses = await db.warehouses.find({
        "status": "active",
        "warehouse_category": {"$in": allowed_categories}
    }, {"_id": 0}).to_list(200)
    
    return warehouses


@router.get("/my-stock")
async def get_my_stock(
    warehouse_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    الحصول على المخزون المتاح للمستخدم
    """
    # جلب المخازن المتاحة للمستخدم
    my_warehouses = await get_my_warehouses(current_user)
    warehouse_ids = [w["id"] for w in my_warehouses]
    
    if not warehouse_ids:
        return []
    
    query = {"warehouse_id": {"$in": warehouse_ids}}
    if warehouse_id and warehouse_id in warehouse_ids:
        query["warehouse_id"] = warehouse_id
    
    stock = await db.product_stock.find(query, {"_id": 0}).to_list(1000)
    
    # إضافة معلومات المخزن لكل عنصر
    warehouse_map = {w["id"]: w for w in my_warehouses}
    for item in stock:
        wh = warehouse_map.get(item.get("warehouse_id"), {})
        item["warehouse_category"] = wh.get("warehouse_category", "")
        item["center_name"] = wh.get("center_name", "")
    
    return stock


@router.post("/issue-request")
async def create_issue_request(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    إنشاء طلب صرف مواد
    يمكن للموظف إنشاء طلب والمشرف يوافق عليه
    أو الصرف المباشر إذا كان لديه الصلاحية
    """
    warehouse_id = data.get("warehouse_id")
    items = data.get("items", [])  # [{product_id, quantity, notes}]
    purpose = data.get("purpose", "")  # الغرض من الصرف
    direct_issue = data.get("direct_issue", False)  # صرف مباشر بدون موافقة
    
    if not warehouse_id or not items:
        raise HTTPException(status_code=400, detail="يجب تحديد المخزن والمنتجات")
    
    # التحقق من صلاحية المستخدم للصرف من هذا المخزن
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    if not warehouse:
        raise HTTPException(status_code=404, detail="المخزن غير موجود")
    
    category = warehouse.get("warehouse_category", "")
    required_permission = CATEGORY_TO_PERMISSION.get(category, "warehouse_stock_issue")
    user_permissions = current_user.get("permissions", [])
    user_role = current_user.get("role", "")
    
    has_permission = (
        user_role == "admin" or
        "warehouse_issue_all" in user_permissions or
        "warehouse_stock_issue" in user_permissions or
        required_permission in user_permissions
    )
    
    if not has_permission:
        raise HTTPException(status_code=403, detail=f"ليس لديك صلاحية الصرف من هذا المخزن ({warehouse.get('name')})")
    
    # التحقق من توفر الكميات
    for item in items:
        stock = await db.product_stock.find_one({
            "product_id": item["product_id"],
            "warehouse_id": warehouse_id
        }, {"_id": 0})
        
        if not stock:
            product = await db.warehouse_products.find_one({"id": item["product_id"]}, {"_id": 0})
            raise HTTPException(
                status_code=400, 
                detail=f"المنتج {product.get('name', item['product_id'])} غير متوفر في هذا المخزن"
            )
        
        if stock.get("available_quantity", 0) < item["quantity"]:
            raise HTTPException(
                status_code=400,
                detail=f"الكمية المطلوبة ({item['quantity']}) من {stock.get('product_name')} غير متوفرة (المتاح: {stock.get('available_quantity', 0)})"
            )
    
    # إنشاء طلب الصرف
    request_id = str(uuid.uuid4())
    request_number = f"ISS-REQ-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    issue_request = {
        "id": request_id,
        "request_number": request_number,
        "warehouse_id": warehouse_id,
        "warehouse_name": warehouse.get("name", ""),
        "warehouse_category": category,
        "center_name": warehouse.get("center_name", ""),
        "items": [],
        "purpose": purpose,
        "status": "approved" if direct_issue else "pending",  # pending, approved, rejected, completed
        "requested_by": current_user["id"],
        "requested_by_name": current_user.get("full_name", ""),
        "requested_at": datetime.now(timezone.utc).isoformat(),
        "approved_by": current_user["id"] if direct_issue else None,
        "approved_by_name": current_user.get("full_name", "") if direct_issue else None,
        "approved_at": datetime.now(timezone.utc).isoformat() if direct_issue else None,
        "total_items": len(items),
        "total_value": 0
    }
    
    total_value = 0
    for item in items:
        stock = await db.product_stock.find_one({
            "product_id": item["product_id"],
            "warehouse_id": warehouse_id
        }, {"_id": 0})
        
        product = await db.warehouse_products.find_one({"id": item["product_id"]}, {"_id": 0})
        item_value = item["quantity"] * stock.get("unit_price", 0)
        total_value += item_value
        
        issue_request["items"].append({
            "product_id": item["product_id"],
            "product_name": product.get("name", ""),
            "product_code": product.get("code", ""),
            "quantity": item["quantity"],
            "unit": product.get("unit", ""),
            "unit_price": stock.get("unit_price", 0),
            "total_value": item_value,
            "notes": item.get("notes", "")
        })
    
    issue_request["total_value"] = total_value
    
    await db.warehouse_issue_requests.insert_one(issue_request)
    
    # إذا كان صرف مباشر، ننفذ الصرف فوراً
    if direct_issue:
        for item in items:
            await issue_stock(
                data={
                    "product_id": item["product_id"],
                    "warehouse_id": warehouse_id,
                    "quantity": item["quantity"],
                    "issue_type": "consumption",
                    "reference_number": request_number,
                    "notes": f"{purpose} - {item.get('notes', '')}",
                    "create_journal": True
                },
                current_user=current_user
            )
        
        # تحديث حالة الطلب
        await db.warehouse_issue_requests.update_one(
            {"id": request_id},
            {"$set": {
                "status": "completed",
                "completed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        await log_activity(
            user_id=current_user["id"],
            user_name=current_user.get("full_name", ""),
            action="direct_issue",
            entity_type="warehouse_issue",
            entity_id=request_id,
            entity_name=request_number,
            details=f"صرف مباشر من {warehouse.get('name')} - {len(items)} منتج بقيمة {total_value}"
        )
        
        return {
            "message": "تم الصرف بنجاح",
            "request": issue_request,
            "status": "completed"
        }
    
    # إذا كان طلب يحتاج موافقة
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", ""),
        action="create_issue_request",
        entity_type="warehouse_issue",
        entity_id=request_id,
        entity_name=request_number,
        details=f"طلب صرف من {warehouse.get('name')} - {len(items)} منتج"
    )
    
    return {
        "message": "تم إنشاء طلب الصرف بنجاح",
        "request": issue_request,
        "status": "pending"
    }


@router.get("/issue-requests")
async def get_issue_requests(
    status: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    my_requests: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """
    الحصول على طلبات الصرف
    """
    query = {}
    
    if status:
        query["status"] = status
    
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    
    if my_requests:
        query["requested_by"] = current_user["id"]
    
    requests = await db.warehouse_issue_requests.find(query, {"_id": 0}).sort("requested_at", -1).to_list(500)
    return requests


@router.post("/issue-requests/{request_id}/approve")
async def approve_issue_request(
    request_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    الموافقة على طلب صرف
    """
    # التحقق من صلاحية الموافقة
    user_permissions = current_user.get("permissions", [])
    user_role = current_user.get("role", "")
    
    if user_role != "admin" and "warehouse_approve_issue" not in user_permissions and "warehouse_issue_all" not in user_permissions:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية الموافقة على طلبات الصرف")
    
    request = await db.warehouse_issue_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if request["status"] != "pending":
        raise HTTPException(status_code=400, detail="الطلب ليس في حالة انتظار")
    
    # تنفيذ الصرف
    for item in request["items"]:
        await issue_stock(
            data={
                "product_id": item["product_id"],
                "warehouse_id": request["warehouse_id"],
                "quantity": item["quantity"],
                "issue_type": "consumption",
                "reference_number": request["request_number"],
                "notes": f"{request.get('purpose', '')} - {item.get('notes', '')}",
                "create_journal": True
            },
            current_user=current_user
        )
    
    # تحديث الطلب
    await db.warehouse_issue_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "completed",
            "approved_by": current_user["id"],
            "approved_by_name": current_user.get("full_name", ""),
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await log_activity(
        user_id=current_user["id"],
        user_name=current_user.get("full_name", ""),
        action="approve_issue_request",
        entity_type="warehouse_issue",
        entity_id=request_id,
        entity_name=request["request_number"],
        details=f"موافقة على صرف من {request.get('warehouse_name')}"
    )
    
    return {"message": "تم الموافقة والصرف بنجاح"}


@router.post("/issue-requests/{request_id}/reject")
async def reject_issue_request(
    request_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    رفض طلب صرف
    """
    reason = data.get("reason", "")
    
    user_permissions = current_user.get("permissions", [])
    user_role = current_user.get("role", "")
    
    if user_role != "admin" and "warehouse_approve_issue" not in user_permissions:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية رفض طلبات الصرف")
    
    request = await db.warehouse_issue_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    await db.warehouse_issue_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "rejected_by": current_user["id"],
            "rejected_by_name": current_user.get("full_name", ""),
            "rejected_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": reason
        }}
    )
    
    return {"message": "تم رفض الطلب"}


@router.get("/consumption-log")
async def get_consumption_log(
    warehouse_id: Optional[str] = None,
    category: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    سجل الاستهلاك - من صرف ماذا ومتى
    """
    query = {"movement_type": "issue"}
    
    if warehouse_id:
        query["from_warehouse_id"] = warehouse_id
    
    if start_date:
        query["movement_date"] = {"$gte": start_date}
    if end_date:
        if "movement_date" in query:
            query["movement_date"]["$lte"] = end_date
        else:
            query["movement_date"] = {"$lte": end_date}
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # إضافة تصنيف المخزن
    if category:
        warehouse_ids = [m.get("from_warehouse_id") for m in movements]
        warehouses = await db.warehouses.find(
            {"id": {"$in": warehouse_ids}, "warehouse_category": category}, 
            {"_id": 0, "id": 1}
        ).to_list(100)
        valid_warehouse_ids = [w["id"] for w in warehouses]
        movements = [m for m in movements if m.get("from_warehouse_id") in valid_warehouse_ids]
    
    return movements


# ==================== الأصول الثابتة ====================

@router.get("/fixed-assets")
async def get_fixed_assets(
    asset_type: Optional[str] = None,
    category: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    center_id: Optional[str] = None,
    status: Optional[str] = None,
    condition: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على قائمة الأصول الثابتة"""
    query = {}
    
    if asset_type:
        query["asset_type"] = asset_type
    if category:
        query["category"] = category
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    if center_id:
        query["center_id"] = center_id
    if status:
        query["status"] = status
    if condition:
        query["condition"] = condition
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"asset_code": {"$regex": search, "$options": "i"}},
            {"serial_number": {"$regex": search, "$options": "i"}}
        ]
    
    assets = await db.fixed_assets.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return assets


@router.get("/fixed-assets/stats")
async def get_fixed_assets_stats(
    current_user: dict = Depends(get_current_user)
):
    """إحصائيات الأصول الثابتة"""
    total = await db.fixed_assets.count_documents({})
    active = await db.fixed_assets.count_documents({"status": "active"})
    in_maintenance = await db.fixed_assets.count_documents({"status": "in_maintenance"})
    disposed = await db.fixed_assets.count_documents({"status": "disposed"})
    
    # إجمالي القيمة
    pipeline = [
        {"$match": {"status": "active"}},
        {"$group": {"_id": None, "total_value": {"$sum": "$current_value"}}}
    ]
    value_result = await db.fixed_assets.aggregate(pipeline).to_list(1)
    total_value = value_result[0]["total_value"] if value_result else 0
    
    # حسب النوع
    by_type = {}
    types = ["equipment", "vehicle", "machinery", "furniture", "electronics"]
    for t in types:
        by_type[t] = await db.fixed_assets.count_documents({"asset_type": t})
    
    # حسب الفئة
    by_category = {}
    categories = ["fixed_assets", "consumables", "spare_parts"]
    for c in categories:
        by_category[c] = await db.fixed_assets.count_documents({"category": c})
    
    # الأصول التي تحتاج صيانة قريباً
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    next_week = (datetime.now(timezone.utc) + timedelta(days=7)).strftime("%Y-%m-%d")
    needs_maintenance = await db.fixed_assets.count_documents({
        "next_maintenance_date": {"$gte": today, "$lte": next_week}
    })
    
    return {
        "total": total,
        "active": active,
        "in_maintenance": in_maintenance,
        "disposed": disposed,
        "total_value": total_value,
        "by_type": by_type,
        "by_category": by_category,
        "needs_maintenance": needs_maintenance
    }


@router.get("/fixed-assets/{asset_id}")
async def get_fixed_asset(
    asset_id: str,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على تفاصيل أصل"""
    asset = await db.fixed_assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="الأصل غير موجود")
    
    # الحصول على تاريخ الحركات
    movements = await db.asset_movements.find(
        {"asset_id": asset_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    asset["movements"] = movements
    return asset


@router.post("/fixed-assets")
async def create_fixed_asset(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء أصل ثابت جديد"""
    from models.all_models import FixedAsset
    
    # إنشاء رمز الأصل
    count = await db.fixed_assets.count_documents({})
    asset_code = f"AST-{datetime.now().strftime('%Y%m')}-{str(count + 1).zfill(4)}"
    
    asset = FixedAsset(
        **data,
        asset_code=asset_code,
        created_by=current_user["id"],
        created_by_name=current_user.get("full_name", "")
    )
    
    await db.fixed_assets.insert_one(asset.model_dump())
    return asset.model_dump()


@router.put("/fixed-assets/{asset_id}")
async def update_fixed_asset(
    asset_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تحديث أصل ثابت"""
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.fixed_assets.update_one(
        {"id": asset_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الأصل غير موجود")
    
    asset = await db.fixed_assets.find_one({"id": asset_id}, {"_id": 0})
    return asset


@router.post("/fixed-assets/{asset_id}/transfer")
async def transfer_fixed_asset(
    asset_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تحويل أصل لموقع جديد"""
    from models.all_models import AssetMovement
    
    asset = await db.fixed_assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="الأصل غير موجود")
    
    to_warehouse_id = data.get("to_warehouse_id")
    to_location = data.get("to_location")
    reason = data.get("reason")
    
    # إنشاء حركة التحويل
    movement = AssetMovement(
        asset_id=asset_id,
        asset_name=asset["name"],
        asset_code=asset["asset_code"],
        movement_type="transfer",
        from_location=asset.get("location_details"),
        to_location=to_location,
        from_warehouse_id=asset.get("warehouse_id"),
        to_warehouse_id=to_warehouse_id,
        reason=reason,
        performed_by=current_user["id"],
        performed_by_name=current_user.get("full_name", ""),
        movement_number=f"MOV-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    )
    
    await db.asset_movements.insert_one(movement.model_dump())
    
    # تحديث موقع الأصل
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if to_warehouse_id:
        warehouse = await db.warehouses.find_one({"id": to_warehouse_id}, {"_id": 0})
        if warehouse:
            update_data["warehouse_id"] = to_warehouse_id
            update_data["warehouse_name"] = warehouse.get("name")
            update_data["center_id"] = warehouse.get("center_id")
            update_data["center_name"] = warehouse.get("center_name")
    
    if to_location:
        update_data["location_details"] = to_location
    
    await db.fixed_assets.update_one({"id": asset_id}, {"$set": update_data})
    
    return {"message": "تم تحويل الأصل بنجاح", "movement": movement.model_dump()}


@router.get("/fixed-assets/movements/all")
async def get_all_asset_movements(
    asset_id: Optional[str] = None,
    movement_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على جميع حركات الأصول"""
    query = {}
    
    if asset_id:
        query["asset_id"] = asset_id
    if movement_type:
        query["movement_type"] = movement_type
    if start_date:
        query.setdefault("created_at", {})["$gte"] = start_date
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date + "T23:59:59"
    
    movements = await db.asset_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return movements


@router.delete("/fixed-assets/{asset_id}")
async def delete_fixed_asset(
    asset_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف/إتلاف أصل"""
    asset = await db.fixed_assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="الأصل غير موجود")
    
    # تغيير الحالة بدلاً من الحذف الفعلي
    await db.fixed_assets.update_one(
        {"id": asset_id},
        {"$set": {
            "status": "disposed",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "تم إتلاف الأصل"}


# ==================== فئات المخازن ====================

@router.get("/warehouse-categories")
async def get_warehouse_categories(
    current_user: dict = Depends(get_current_user)
):
    """الحصول على فئات المخازن"""
    from models.all_models import PRODUCT_CATEGORIES, WAREHOUSE_TYPES
    return {
        "product_categories": PRODUCT_CATEGORIES,
        "warehouse_types": WAREHOUSE_TYPES
    }


@router.get("/warehouse-hierarchy/{center_id}")
async def get_warehouse_hierarchy(
    center_id: str,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على الهيكل الهرمي للمخازن في مركز معين"""
    # المخازن الرئيسية (الخارجية)
    external_warehouses = await db.warehouses.find({
        "center_id": center_id,
        "warehouse_type": {"$in": ["external", "main"]},
        "parent_warehouse_id": None
    }, {"_id": 0}).to_list(50)
    
    hierarchy = []
    for ext_wh in external_warehouses:
        # المخازن الداخلية التابعة
        internal_warehouses = await db.warehouses.find({
            "center_id": center_id,
            "parent_warehouse_id": ext_wh["id"]
        }, {"_id": 0}).to_list(50)
        
        children = []
        for int_wh in internal_warehouses:
            # المخازن الفرعية
            sub_warehouses = await db.warehouses.find({
                "center_id": center_id,
                "parent_warehouse_id": int_wh["id"]
            }, {"_id": 0}).to_list(50)
            
            int_wh["children"] = sub_warehouses
            children.append(int_wh)
        
        ext_wh["children"] = children
        hierarchy.append(ext_wh)
    
    return hierarchy


# ==================== إعدادات المخزون والميزات المتقدمة ====================

@router.get("/settings")
async def get_inventory_settings(current_user: dict = Depends(get_current_user)):
    """الحصول على إعدادات المخزون"""
    settings = await db.inventory_settings.find_one({"id": "inventory_settings"}, {"_id": 0})
    if not settings:
        # إنشاء الإعدادات الافتراضية
        default_settings = {
            "id": "inventory_settings",
            "auto_reorder_enabled": False,
            "auto_reorder_check_interval": 24,
            "auto_reorder_notification_email": None,
            "auto_reorder_notification_phone": None,
            "abc_a_percentage": 80,
            "abc_b_percentage": 15,
            "abc_auto_calculate": True,
            "abc_calculation_period_months": 12,
            "barcode_auto_generate": True,
            "barcode_prefix": "PRD",
            "qr_code_include_price": False,
            "qr_code_include_expiry": True,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.inventory_settings.insert_one(default_settings)
        settings = default_settings
    
    return settings


@router.put("/settings")
async def update_inventory_settings(
    settings_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تحديث إعدادات المخزون"""
    settings_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    settings_data["updated_by"] = current_user.get("user_id")
    
    await db.inventory_settings.update_one(
        {"id": "inventory_settings"},
        {"$set": settings_data},
        upsert=True
    )
    
    return {"message": "تم تحديث الإعدادات بنجاح"}


# ==================== الباركود و QR Code ====================

@router.post("/products/{product_id}/generate-barcode")
async def generate_product_barcode(
    product_id: str,
    barcode_type: str = "EAN13",
    current_user: dict = Depends(get_current_user)
):
    """توليد باركود للمنتج"""
    product = await db.warehouse_products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    # توليد باركود فريد
    settings = await db.inventory_settings.find_one({"id": "inventory_settings"}, {"_id": 0})
    prefix = settings.get("barcode_prefix", "PRD") if settings else "PRD"
    
    # استخدام الوقت والـ ID لإنشاء باركود فريد
    timestamp = datetime.now().strftime("%y%m%d%H%M")
    barcode = f"{prefix}{timestamp}{str(uuid.uuid4())[:4].upper()}"
    
    # تحديث المنتج
    await db.warehouse_products.update_one(
        {"id": product_id},
        {"$set": {
            "barcode": barcode,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"barcode": barcode, "product_id": product_id}


@router.get("/products/scan/{barcode}")
async def scan_barcode(
    barcode: str,
    current_user: dict = Depends(get_current_user)
):
    """البحث عن منتج بالباركود"""
    # البحث في barcode أو code
    product = await db.warehouse_products.find_one(
        {"$or": [{"barcode": barcode}, {"code": barcode}]},
        {"_id": 0}
    )
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    # جلب معلومات المخزون
    stocks = await db.warehouse_stock.find({"product_id": product["id"]}, {"_id": 0}).to_list(100)
    total_quantity = sum(s.get("quantity", 0) for s in stocks)
    
    return {
        "product": product,
        "stocks": stocks,
        "total_quantity": total_quantity
    }


@router.post("/products/{product_id}/generate-qr")
async def generate_product_qr(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """توليد رمز QR للمنتج"""
    product = await db.warehouse_products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    settings = await db.inventory_settings.find_one({"id": "inventory_settings"}, {"_id": 0})
    
    # بناء محتوى QR
    qr_data = {
        "id": product["id"],
        "code": product.get("code", ""),
        "name": product.get("name", "")
    }
    
    if settings and settings.get("qr_code_include_price"):
        qr_data["price"] = product.get("unit_price", 0)
    
    import json
    qr_content = json.dumps(qr_data, ensure_ascii=False)
    
    # تحديث المنتج
    await db.warehouse_products.update_one(
        {"id": product_id},
        {"$set": {
            "qr_code": qr_content,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"qr_code": qr_content, "product_id": product_id}


# ==================== تصنيف ABC ====================

@router.post("/abc/calculate")
async def calculate_abc_classification(
    current_user: dict = Depends(get_current_user)
):
    """حساب تصنيف ABC لجميع المنتجات"""
    settings = await db.inventory_settings.find_one({"id": "inventory_settings"}, {"_id": 0})
    a_pct = settings.get("abc_a_percentage", 80) if settings else 80
    b_pct = settings.get("abc_b_percentage", 15) if settings else 15
    months = settings.get("abc_calculation_period_months", 12) if settings else 12
    
    # حساب فترة التحليل
    start_date = (datetime.now(timezone.utc) - timedelta(days=months * 30)).isoformat()
    
    # جلب جميع حركات الصرف في الفترة المحددة
    movements = await db.warehouse_movements.find({
        "movement_type": "issue",
        "movement_date": {"$gte": start_date}
    }, {"_id": 0}).to_list(10000)
    
    # حساب القيمة السنوية لكل منتج
    product_values = {}
    for mov in movements:
        pid = mov.get("product_id")
        value = mov.get("quantity", 0) * mov.get("unit_price", 0)
        product_values[pid] = product_values.get(pid, 0) + value
    
    # ترتيب المنتجات حسب القيمة
    sorted_products = sorted(product_values.items(), key=lambda x: x[1], reverse=True)
    total_value = sum(product_values.values())
    
    if total_value == 0:
        return {"message": "لا توجد حركات صرف لحساب ABC", "updated": 0}
    
    # تصنيف المنتجات
    cumulative_value = 0
    updated_count = 0
    
    for product_id, annual_value in sorted_products:
        cumulative_value += annual_value
        cumulative_pct = (cumulative_value / total_value) * 100
        
        if cumulative_pct <= a_pct:
            classification = "A"
        elif cumulative_pct <= (a_pct + b_pct):
            classification = "B"
        else:
            classification = "C"
        
        # تحديث المنتج
        await db.warehouse_products.update_one(
            {"id": product_id},
            {"$set": {
                "abc_classification": classification,
                "annual_value": annual_value,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        updated_count += 1
    
    return {
        "message": f"تم تحديث تصنيف ABC لـ {updated_count} منتج",
        "total_value": total_value,
        "a_count": len([p for p in sorted_products if product_values.get(p[0], 0) / total_value * 100 <= a_pct]),
        "b_count": len([p for p in sorted_products if a_pct < product_values.get(p[0], 0) / total_value * 100 <= a_pct + b_pct]),
        "c_count": len([p for p in sorted_products if product_values.get(p[0], 0) / total_value * 100 > a_pct + b_pct])
    }


@router.get("/abc/summary")
async def get_abc_summary(current_user: dict = Depends(get_current_user)):
    """ملخص تصنيف ABC"""
    products = await db.warehouse_products.find({"status": "active"}, {"_id": 0}).to_list(10000)
    
    summary = {
        "A": {"count": 0, "value": 0, "products": []},
        "B": {"count": 0, "value": 0, "products": []},
        "C": {"count": 0, "value": 0, "products": []}
    }
    
    for product in products:
        classification = product.get("abc_classification", "C")
        annual_value = product.get("annual_value", 0)
        
        if classification in summary:
            summary[classification]["count"] += 1
            summary[classification]["value"] += annual_value
            if len(summary[classification]["products"]) < 10:  # أعلى 10 منتجات فقط
                summary[classification]["products"].append({
                    "id": product["id"],
                    "name": product.get("name"),
                    "code": product.get("code"),
                    "annual_value": annual_value
                })
    
    # ترتيب المنتجات حسب القيمة
    for cls in summary:
        summary[cls]["products"] = sorted(
            summary[cls]["products"],
            key=lambda x: x["annual_value"],
            reverse=True
        )
    
    return summary


# ==================== إعادة الطلب التلقائي ====================

@router.post("/auto-reorder/check")
async def check_auto_reorder(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """فحص المنتجات التي تحتاج إعادة طلب"""
    settings = await db.inventory_settings.find_one({"id": "inventory_settings"}, {"_id": 0})
    
    if not settings or not settings.get("auto_reorder_enabled"):
        return {"message": "إعادة الطلب التلقائي غير مفعلة", "requests": []}
    
    # جلب المنتجات المفعل بها إعادة الطلب التلقائي
    products = await db.warehouse_products.find({
        "auto_reorder_enabled": True,
        "status": "active"
    }, {"_id": 0}).to_list(1000)
    
    reorder_requests = []
    
    for product in products:
        # حساب الكمية الإجمالية المتاحة
        stocks = await db.warehouse_stock.find({"product_id": product["id"]}, {"_id": 0}).to_list(100)
        total_quantity = sum(s.get("quantity", 0) for s in stocks)
        
        reorder_point = product.get("reorder_point", 0)
        
        if total_quantity <= reorder_point and reorder_point > 0:
            # إنشاء طلب إعادة طلب
            request_number = f"RO-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"
            
            reorder_request = {
                "id": str(uuid.uuid4()),
                "request_number": request_number,
                "product_id": product["id"],
                "product_name": product.get("name"),
                "product_code": product.get("code"),
                "current_quantity": total_quantity,
                "reorder_point": reorder_point,
                "reorder_quantity": product.get("reorder_quantity", reorder_point * 2),
                "warehouse_id": stocks[0].get("warehouse_id") if stocks else None,
                "warehouse_name": stocks[0].get("warehouse_name") if stocks else None,
                "estimated_cost": product.get("cost_price", 0) * product.get("reorder_quantity", reorder_point * 2),
                "status": "pending",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # التحقق من عدم وجود طلب معلق لنفس المنتج
            existing = await db.auto_reorder_requests.find_one({
                "product_id": product["id"],
                "status": "pending"
            })
            
            if not existing:
                await db.auto_reorder_requests.insert_one(reorder_request)
                reorder_requests.append(reorder_request)
    
    return {
        "message": f"تم إنشاء {len(reorder_requests)} طلب إعادة طلب",
        "requests": reorder_requests
    }


@router.get("/auto-reorder/requests")
async def get_auto_reorder_requests(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على طلبات إعادة الطلب"""
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.auto_reorder_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return requests


@router.put("/auto-reorder/requests/{request_id}/approve")
async def approve_reorder_request(
    request_id: str,
    current_user: dict = Depends(get_current_user)
):
    """الموافقة على طلب إعادة الطلب"""
    request = await db.auto_reorder_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    await db.auto_reorder_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user.get("user_id"),
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "تمت الموافقة على الطلب"}


@router.put("/auto-reorder/requests/{request_id}/reject")
async def reject_reorder_request(
    request_id: str,
    reason: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """رفض طلب إعادة الطلب"""
    await db.auto_reorder_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "rejection_reason": reason,
            "approved_by": current_user.get("user_id"),
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "تم رفض الطلب"}


@router.get("/products/low-stock")
async def get_low_stock_products(current_user: dict = Depends(get_current_user)):
    """المنتجات التي وصلت لنقطة إعادة الطلب أو أقل"""
    products = await db.warehouse_products.find({
        "status": "active",
        "reorder_point": {"$gt": 0}
    }, {"_id": 0}).to_list(1000)
    
    low_stock = []
    
    for product in products:
        stocks = await db.warehouse_stock.find({"product_id": product["id"]}, {"_id": 0}).to_list(100)
        total_quantity = sum(s.get("quantity", 0) for s in stocks)
        reorder_point = product.get("reorder_point", 0)
        min_quantity = product.get("min_quantity", 0)
        
        if total_quantity <= reorder_point:
            status = "critical" if total_quantity <= min_quantity else "warning"
            low_stock.append({
                "product": product,
                "total_quantity": total_quantity,
                "reorder_point": reorder_point,
                "min_quantity": min_quantity,
                "shortage": reorder_point - total_quantity,
                "status": status,
                "auto_reorder_enabled": product.get("auto_reorder_enabled", False)
            })
    
    # ترتيب حسب الأهمية (critical أولاً)
    low_stock.sort(key=lambda x: (0 if x["status"] == "critical" else 1, -x["shortage"]))
    
    return low_stock

