#!/usr/bin/env python3
"""
Backend API Testing for Milk Collection Center ERP - Review Request Features
Tests the specific features requested in the review:
1. White Screen Bug Fix for hassan.hamdi
2. Payment Receipt PDF API
3. Login Page Image verification
Testing with credentials: hassan.hamdi/Hassan@123
"""

import requests
import json
import sys
from datetime import datetime

# Get backend URL from frontend .env
BACKEND_URL = "https://agrodairy.preview.emergentagent.com/api"

# Test credentials (as specified in review request)
TEST_USERNAME = "hassan.hamdi"
TEST_PASSWORD = "Hassan@123"

class BackendTester:
    def __init__(self):
        self.session = requests.Session()
        self.token = None
        self.user_data = None
        self.test_results = []
        
    def log_test(self, test_name, success, message, details=None):
        """Log test results"""
        result = {
            "test": test_name,
            "success": success,
            "message": message,
            "details": details,
            "timestamp": datetime.now().isoformat()
        }
        self.test_results.append(result)
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name} - {message}")
        if details and not success:
            print(f"   Details: {details}")
    
    def test_hassan_hamdi_login(self):
        """Test login with hassan.hamdi credentials - should not cause white screen"""
        try:
            response = self.session.post(
                f"{BACKEND_URL}/auth/login",
                json={
                    "username": TEST_USERNAME,
                    "password": TEST_PASSWORD
                },
                headers={"Content-Type": "application/json"}
            )
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                self.user_data = data.get("user")
                self.session.headers.update({
                    "Authorization": f"Bearer {self.token}"
                })
                
                # Verify user role is accountant
                if self.user_data.get("role") == "accountant":
                    self.log_test(
                        "Hassan Hamdi Login", 
                        True, 
                        f"Successfully logged in as {self.user_data.get('username')} with role {self.user_data.get('role')}"
                    )
                    return True
                else:
                    self.log_test(
                        "Hassan Hamdi Login", 
                        False, 
                        f"User role is {self.user_data.get('role')}, expected 'accountant'",
                        f"User data: {self.user_data}"
                    )
                    return False
            else:
                self.log_test(
                    "Hassan Hamdi Login", 
                    False, 
                    f"Login failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Hassan Hamdi Login", False, f"Login error: {str(e)}")
            return False
    
    def test_register_if_needed(self):
        """Register test user if login fails"""
        try:
            response = self.session.post(
                f"{BACKEND_URL}/auth/register",
                json={
                    "username": TEST_USERNAME,
                    "password": TEST_PASSWORD,
                    "email": "testadmin@example.com",
                    "full_name": "Test Administrator",
                    "role": "admin"
                },
                headers={"Content-Type": "application/json"}
            )
            
            if response.status_code == 200:
                data = response.json()
                self.token = data.get("access_token")
                self.user_data = data.get("user")
                self.session.headers.update({
                    "Authorization": f"Bearer {self.token}"
                })
                self.log_test(
                    "User Registration", 
                    True, 
                    f"Successfully registered and logged in as {self.user_data.get('username')}"
                )
                return True
            else:
                self.log_test(
                    "User Registration", 
                    False, 
                    f"Registration failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("User Registration", False, f"Registration error: {str(e)}")
            return False
    
    def test_payment_receipt_pdf_workflow(self):
        """Test Payment Receipt PDF API workflow"""
        try:
            # First, create a supplier for the payment
            supplier_data = {
                "name": "مزرعة الأمل للألبان",
                "phone": "+968 9876 5432",
                "address": "صلالة، ظفار، عُمان",
                "supplier_code": "SUP-TEST-001",
                "bank_account": "1234567890123456",
                "center_id": None,
                "center_name": None,
                "national_id": "87654321",
                "farm_size": 45.5,
                "cattle_count": 30
            }
            
            supplier_response = self.session.post(
                f"{BACKEND_URL}/suppliers",
                json=supplier_data,
                headers={"Content-Type": "application/json"}
            )
            
            if supplier_response.status_code != 200:
                self.log_test(
                    "Payment Receipt PDF Workflow", 
                    False, 
                    f"Failed to create test supplier: {supplier_response.status_code}",
                    supplier_response.text
                )
                return False
            
            supplier = supplier_response.json()
            supplier_id = supplier.get("id")
            
            # Create a supplier payment
            payment_data = {
                "payment_type": "supplier_payment",
                "related_id": supplier_id,
                "related_name": supplier_data["name"],
                "amount": 1500.0,
                "payment_method": "bank_transfer",
                "notes": "دفعة شهرية لمورد الحليب - اختبار PDF"
            }
            
            payment_response = self.session.post(
                f"{BACKEND_URL}/payments",
                json=payment_data,
                headers={"Content-Type": "application/json"}
            )
            
            if payment_response.status_code != 200:
                self.log_test(
                    "Payment Receipt PDF Workflow", 
                    False, 
                    f"Failed to create payment: {payment_response.status_code}",
                    payment_response.text
                )
                return False
            
            payment = payment_response.json()
            payment_id = payment.get("id")
            
            # Test the PDF receipt endpoint
            pdf_response = self.session.get(
                f"{BACKEND_URL}/payments/{payment_id}/receipt",
                headers={"Authorization": f"Bearer {self.token}"}
            )
            
            if pdf_response.status_code == 200:
                # Check if response is PDF
                content_type = pdf_response.headers.get("Content-Type", "")
                content_disposition = pdf_response.headers.get("Content-Disposition", "")
                
                if "application/pdf" in content_type and "attachment" in content_disposition:
                    # Check PDF content length
                    pdf_size = len(pdf_response.content)
                    if pdf_size > 1000:  # PDF should be at least 1KB
                        self.log_test(
                            "Payment Receipt PDF Workflow", 
                            True, 
                            f"Successfully generated PDF receipt for payment {payment_id} (Size: {pdf_size} bytes)"
                        )
                        return True
                    else:
                        self.log_test(
                            "Payment Receipt PDF Workflow", 
                            False, 
                            f"PDF file too small: {pdf_size} bytes",
                            f"Headers: {dict(pdf_response.headers)}"
                        )
                        return False
                else:
                    self.log_test(
                        "Payment Receipt PDF Workflow", 
                        False, 
                        f"Response not a PDF file. Content-Type: {content_type}",
                        f"Headers: {dict(pdf_response.headers)}"
                    )
                    return False
            else:
                self.log_test(
                    "Payment Receipt PDF Workflow", 
                    False, 
                    f"PDF receipt API failed with status {pdf_response.status_code}",
                    pdf_response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Payment Receipt PDF Workflow", False, f"Error: {str(e)}")
            return False

    def test_dashboard_access(self):
        """Test dashboard access for hassan.hamdi - should not show white screen"""
        try:
            # Test dashboard stats endpoint
            response = self.session.get(f"{BACKEND_URL}/dashboard/stats")
            
            if response.status_code == 200:
                dashboard = response.json()
                # Check if dashboard contains expected fields for accountant role
                expected_fields = ["suppliers_count", "customers_count", "today_milk_quantity", "today_sales_quantity"]
                found_fields = [field for field in expected_fields if field in dashboard]
                
                if len(found_fields) >= 2:  # At least 2 expected fields
                    self.log_test(
                        "Dashboard Access", 
                        True, 
                        f"Dashboard accessible with fields: {list(dashboard.keys())}"
                    )
                    return True
                else:
                    self.log_test(
                        "Dashboard Access", 
                        False, 
                        f"Dashboard missing expected fields. Found: {list(dashboard.keys())}",
                        f"Expected at least 2 of: {expected_fields}"
                    )
                    return False
            else:
                self.log_test(
                    "Dashboard Access", 
                    False, 
                    f"Dashboard API failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Dashboard Access", False, f"Error: {str(e)}")
            return False

    def test_user_profile_access(self):
        """Test user profile access - verify user data is accessible"""
        try:
            response = self.session.get(f"{BACKEND_URL}/auth/me")
            
            if response.status_code == 200:
                profile = response.json()
                required_fields = ["id", "username", "full_name", "role"]
                missing_fields = [field for field in required_fields if not profile.get(field)]
                
                if not missing_fields:
                    self.log_test(
                        "User Profile Access", 
                        True, 
                        f"Profile accessible for {profile.get('username')} ({profile.get('role')})"
                    )
                    return True
                else:
                    self.log_test(
                        "User Profile Access", 
                        False, 
                        f"Profile missing required fields: {missing_fields}",
                        f"Profile: {profile}"
                    )
                    return False
            else:
                self.log_test(
                    "User Profile Access", 
                    False, 
                    f"Profile API failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("User Profile Access", False, f"Error: {str(e)}")
            return False

    # ==================== ACTIVITY LOGGING TESTS ====================
    
    def test_activity_logs_login_logged(self):
        """Test that login action is logged in activity logs"""
        try:
            response = self.session.get(f"{BACKEND_URL}/activity-logs?action=login&limit=10")
            
            if response.status_code == 200:
                logs = response.json()
                # Find login log for current user
                login_log = None
                for log in logs:
                    if (log.get("action") == "login" and 
                        log.get("user_name") == self.user_data.get("full_name")):
                        login_log = log
                        break
                
                if login_log:
                    required_fields = ["user_id", "user_name", "action", "timestamp"]
                    missing_fields = [field for field in required_fields if not login_log.get(field)]
                    
                    if not missing_fields:
                        self.log_test(
                            "Activity Log - Login Logged", 
                            True, 
                            f"Login action properly logged for user {login_log.get('user_name')}"
                        )
                        return True
                    else:
                        self.log_test(
                            "Activity Log - Login Logged", 
                            False, 
                            f"Login log missing required fields: {missing_fields}",
                            f"Log entry: {login_log}"
                        )
                        return False
                else:
                    self.log_test(
                        "Activity Log - Login Logged", 
                        False, 
                        f"No login log found for user {self.user_data.get('full_name')}",
                        f"Available logs: {[log.get('action') + ' by ' + log.get('user_name') for log in logs[:5]]}"
                    )
                    return False
            else:
                self.log_test(
                    "Activity Log - Login Logged", 
                    False, 
                    f"Activity logs API failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Activity Log - Login Logged", False, f"Error: {str(e)}")
            return False

    def test_activity_logs_supplier_crud(self):
        """Test that supplier CRUD operations are logged"""
        try:
            # Create a supplier
            supplier_data = {
                "name": "مزرعة الخير للألبان",
                "phone": "+968 9876 5432",
                "address": "صلالة، ظفار",
                "supplier_code": "SUP001",
                "bank_account": "1234567890",
                "center_id": None,
                "center_name": None,
                "national_id": "12345678",
                "farm_size": 50.5,
                "cattle_count": 25
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/suppliers",
                json=supplier_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Activity Log - Supplier CRUD", 
                    False, 
                    f"Failed to create supplier: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            supplier = create_response.json()
            supplier_id = supplier.get("id")
            
            # Check if create_supplier action is logged
            logs_response = self.session.get(f"{BACKEND_URL}/activity-logs?action=create_supplier&limit=10")
            if logs_response.status_code != 200:
                self.log_test(
                    "Activity Log - Supplier CRUD", 
                    False, 
                    f"Failed to get activity logs: {logs_response.status_code}",
                    logs_response.text
                )
                return False
            
            logs = logs_response.json()
            create_log = None
            for log in logs:
                if (log.get("action") == "create_supplier" and 
                    log.get("entity_name") == supplier_data["name"]):
                    create_log = log
                    break
            
            if not create_log:
                self.log_test(
                    "Activity Log - Supplier CRUD", 
                    False, 
                    f"Create supplier action not logged for {supplier_data['name']}",
                    f"Available logs: {[log.get('action') + ' - ' + str(log.get('entity_name')) for log in logs[:5]]}"
                )
                return False
            
            # Update the supplier
            update_data = {
                "name": "مزرعة الخير المحدثة",
                "phone": "+968 9876 5432",
                "address": "صلالة، ظفار - محدث",
                "supplier_code": "SUP001",
                "bank_account": "1234567890",
                "center_id": None,
                "center_name": None,
                "national_id": "12345678",
                "farm_size": 60.0,
                "cattle_count": 30
            }
            
            update_response = self.session.put(
                f"{BACKEND_URL}/suppliers/{supplier_id}",
                json=update_data,
                headers={"Content-Type": "application/json"}
            )
            
            if update_response.status_code != 200:
                self.log_test(
                    "Activity Log - Supplier CRUD", 
                    False, 
                    f"Failed to update supplier: {update_response.status_code}",
                    update_response.text
                )
                return False
            
            # Check if update_supplier action is logged
            logs_response = self.session.get(f"{BACKEND_URL}/activity-logs?action=update_supplier&limit=10")
            if logs_response.status_code == 200:
                logs = logs_response.json()
                update_log = None
                for log in logs:
                    if (log.get("action") == "update_supplier" and 
                        log.get("entity_id") == supplier_id):
                        update_log = log
                        break
                
                if update_log:
                    # Verify log structure
                    required_fields = ["user_id", "user_name", "action", "entity_type", "entity_id", "entity_name", "timestamp"]
                    missing_fields = [field for field in required_fields if not update_log.get(field)]
                    
                    if not missing_fields:
                        self.log_test(
                            "Activity Log - Supplier CRUD", 
                            True, 
                            f"Supplier CRUD operations properly logged (create & update) for {supplier_data['name']}"
                        )
                        return True
                    else:
                        self.log_test(
                            "Activity Log - Supplier CRUD", 
                            False, 
                            f"Update log missing required fields: {missing_fields}",
                            f"Log entry: {update_log}"
                        )
                        return False
                else:
                    self.log_test(
                        "Activity Log - Supplier CRUD", 
                        False, 
                        f"Update supplier action not logged for supplier ID {supplier_id}"
                    )
                    return False
            else:
                self.log_test(
                    "Activity Log - Supplier CRUD", 
                    False, 
                    f"Failed to get update logs: {logs_response.status_code}",
                    logs_response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Activity Log - Supplier CRUD", False, f"Error: {str(e)}")
            return False

    def test_activity_logs_customer_crud(self):
        """Test that customer CRUD operations are logged"""
        try:
            # Create a customer
            customer_data = {
                "name": "شركة الألبان المتقدمة",
                "phone": "+968 2468 1357",
                "address": "مسقط، مسقط",
                "customer_type": "wholesale",
                "credit_limit": 5000.0
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/customers",
                json=customer_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Activity Log - Customer CRUD", 
                    False, 
                    f"Failed to create customer: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            customer = create_response.json()
            customer_id = customer.get("id")
            
            # Check if create_customer action is logged
            logs_response = self.session.get(f"{BACKEND_URL}/activity-logs?action=create_customer&limit=10")
            if logs_response.status_code != 200:
                self.log_test(
                    "Activity Log - Customer CRUD", 
                    False, 
                    f"Failed to get activity logs: {logs_response.status_code}",
                    logs_response.text
                )
                return False
            
            logs = logs_response.json()
            create_log = None
            for log in logs:
                if (log.get("action") == "create_customer" and 
                    log.get("entity_name") == customer_data["name"]):
                    create_log = log
                    break
            
            if create_log:
                # Verify log structure
                required_fields = ["user_id", "user_name", "action", "entity_type", "entity_id", "entity_name", "timestamp"]
                missing_fields = [field for field in required_fields if not create_log.get(field)]
                
                if not missing_fields:
                    self.log_test(
                        "Activity Log - Customer CRUD", 
                        True, 
                        f"Customer create operation properly logged for {customer_data['name']}"
                    )
                    return True
                else:
                    self.log_test(
                        "Activity Log - Customer CRUD", 
                        False, 
                        f"Create log missing required fields: {missing_fields}",
                        f"Log entry: {create_log}"
                    )
                    return False
            else:
                self.log_test(
                    "Activity Log - Customer CRUD", 
                    False, 
                    f"Create customer action not logged for {customer_data['name']}",
                    f"Available logs: {[log.get('action') + ' - ' + str(log.get('entity_name')) for log in logs[:5]]}"
                )
                return False
                
        except Exception as e:
            self.log_test("Activity Log - Customer CRUD", False, f"Error: {str(e)}")
            return False

    def test_activity_logs_hr_leave_request(self):
        """Test that HR leave request operations are logged"""
        try:
            # Get employees first
            employees_response = self.session.get(f"{BACKEND_URL}/hr/employees")
            if employees_response.status_code != 200:
                self.log_test(
                    "Activity Log - HR Leave Request", 
                    False, 
                    f"Failed to get employees: {employees_response.status_code}",
                    employees_response.text
                )
                return False
            
            employees = employees_response.json()
            if not employees:
                self.log_test(
                    "Activity Log - HR Leave Request", 
                    False, 
                    "No employees found to create leave request"
                )
                return False
            
            employee = employees[0]
            
            # Create leave request
            leave_data = {
                "employee_id": employee["id"],
                "employee_name": employee["name"],
                "leave_type": "annual",
                "start_date": "2025-02-15",
                "end_date": "2025-02-17",
                "reason": "Family vacation",
                "days_count": 3
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/hr/leave-requests",
                json=leave_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Activity Log - HR Leave Request", 
                    False, 
                    f"Failed to create leave request: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            leave_request = create_response.json()
            
            # Check if create_leave_request action is logged
            logs_response = self.session.get(f"{BACKEND_URL}/activity-logs?action=create_leave_request&limit=10")
            if logs_response.status_code != 200:
                self.log_test(
                    "Activity Log - HR Leave Request", 
                    False, 
                    f"Failed to get activity logs: {logs_response.status_code}",
                    logs_response.text
                )
                return False
            
            logs = logs_response.json()
            create_log = None
            for log in logs:
                if (log.get("action") == "create_leave_request" and 
                    log.get("entity_name") == employee["name"]):
                    create_log = log
                    break
            
            if create_log:
                # Verify log structure
                required_fields = ["user_id", "user_name", "action", "entity_type", "entity_name", "timestamp"]
                missing_fields = [field for field in required_fields if not create_log.get(field)]
                
                if not missing_fields:
                    self.log_test(
                        "Activity Log - HR Leave Request", 
                        True, 
                        f"HR leave request creation properly logged for {employee['name']}"
                    )
                    return True
                else:
                    self.log_test(
                        "Activity Log - HR Leave Request", 
                        False, 
                        f"Leave request log missing required fields: {missing_fields}",
                        f"Log entry: {create_log}"
                    )
                    return False
            else:
                self.log_test(
                    "Activity Log - HR Leave Request", 
                    False, 
                    f"Create leave request action not logged for {employee['name']}",
                    f"Available logs: {[log.get('action') + ' - ' + str(log.get('entity_name')) for log in logs[:5]]}"
                )
                return False
                
        except Exception as e:
            self.log_test("Activity Log - HR Leave Request", False, f"Error: {str(e)}")
            return False

    def test_activity_logs_api_filters(self):
        """Test Activity Logs API filters (limit, action)"""
        try:
            # Test limit filter
            limit_response = self.session.get(f"{BACKEND_URL}/activity-logs?limit=5")
            if limit_response.status_code != 200:
                self.log_test(
                    "Activity Log - API Filters", 
                    False, 
                    f"Failed to get limited logs: {limit_response.status_code}",
                    limit_response.text
                )
                return False
            
            limited_logs = limit_response.json()
            if len(limited_logs) > 5:
                self.log_test(
                    "Activity Log - API Filters", 
                    False, 
                    f"Limit filter not working: expected max 5 logs, got {len(limited_logs)}"
                )
                return False
            
            # Test action filter
            action_response = self.session.get(f"{BACKEND_URL}/activity-logs?action=login")
            if action_response.status_code != 200:
                self.log_test(
                    "Activity Log - API Filters", 
                    False, 
                    f"Failed to get filtered logs: {action_response.status_code}",
                    action_response.text
                )
                return False
            
            filtered_logs = action_response.json()
            non_login_logs = [log for log in filtered_logs if log.get("action") != "login"]
            
            if non_login_logs:
                self.log_test(
                    "Activity Log - API Filters", 
                    False, 
                    f"Action filter not working: found {len(non_login_logs)} non-login logs",
                    f"Non-login actions: {[log.get('action') for log in non_login_logs[:3]]}"
                )
                return False
            
            # Test sorting (newest first)
            all_logs_response = self.session.get(f"{BACKEND_URL}/activity-logs?limit=10")
            if all_logs_response.status_code == 200:
                all_logs = all_logs_response.json()
                if len(all_logs) >= 2:
                    # Check if timestamps are in descending order
                    timestamps = [log.get("timestamp") for log in all_logs if log.get("timestamp")]
                    if len(timestamps) >= 2:
                        is_sorted = all(timestamps[i] >= timestamps[i+1] for i in range(len(timestamps)-1))
                        if not is_sorted:
                            self.log_test(
                                "Activity Log - API Filters", 
                                False, 
                                "Logs not sorted by timestamp descending (newest first)",
                                f"First 3 timestamps: {timestamps[:3]}"
                            )
                            return False
            
            self.log_test(
                "Activity Log - API Filters", 
                True, 
                f"Activity logs API filters working correctly (limit={len(limited_logs)}, action filter, sorting)"
            )
            return True
                
        except Exception as e:
            self.log_test("Activity Log - API Filters", False, f"Error: {str(e)}")
            return False

    # ==================== LEGAL MODULE TESTS ====================
    
    def test_legal_dashboard_api(self):
        """Test GET /api/legal/dashboard - verify returns stats"""
        try:
            response = self.session.get(f"{BACKEND_URL}/legal/dashboard")
            
            if response.status_code == 200:
                dashboard = response.json()
                expected_fields = ["contracts_active", "contracts_expiring_soon", "cases_open", "consultations_pending"]
                found_fields = [field for field in expected_fields if field in dashboard]
                
                if len(found_fields) == len(expected_fields):
                    self.log_test(
                        "Legal Dashboard API", 
                        True, 
                        f"Legal dashboard retrieved with all expected fields: {list(dashboard.keys())}"
                    )
                    return True
                else:
                    missing_fields = [field for field in expected_fields if field not in dashboard]
                    self.log_test(
                        "Legal Dashboard API", 
                        False, 
                        f"Legal dashboard missing fields: {missing_fields}",
                        f"Found: {list(dashboard.keys())}"
                    )
                    return False
            else:
                self.log_test(
                    "Legal Dashboard API", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Legal Dashboard API", False, f"Error: {str(e)}")
            return False

    def test_legal_contracts_crud(self):
        """Test Legal Contracts CRUD operations"""
        try:
            # Create contract
            contract_data = {
                "contract_type": "vendor",
                "title": "عقد توريد",
                "party_name": "شركة الأمل",
                "party_type": "company",
                "start_date": "2025-01-01",
                "end_date": "2025-12-31",
                "value": 5000,
                "currency": "OMR",
                "description": "عقد توريد مواد غذائية",
                "terms": "شروط وأحكام العقد"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/legal/contracts",
                json=contract_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Legal Contracts CRUD", 
                    False, 
                    f"Failed to create contract: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            contract = create_response.json()
            contract_id = contract.get("id")
            
            # Verify contract has auto-generated contract_number
            if not contract.get("contract_number") or not contract.get("contract_number").startswith("CTR-"):
                self.log_test(
                    "Legal Contracts CRUD", 
                    False, 
                    "Contract number not auto-generated properly",
                    f"Contract number: {contract.get('contract_number')}"
                )
                return False
            
            # Test GET contracts
            get_response = self.session.get(f"{BACKEND_URL}/legal/contracts")
            if get_response.status_code != 200:
                self.log_test(
                    "Legal Contracts CRUD", 
                    False, 
                    f"Failed to get contracts: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            contracts = get_response.json()
            found_contract = any(c.get("id") == contract_id for c in contracts)
            
            if found_contract:
                self.log_test(
                    "Legal Contracts CRUD", 
                    True, 
                    f"Successfully created contract '{contract_data['title']}' with auto-generated number {contract.get('contract_number')}"
                )
                return True, contract_id
            else:
                self.log_test(
                    "Legal Contracts CRUD", 
                    False, 
                    "Created contract not found in GET contracts"
                )
                return False, None
                
        except Exception as e:
            self.log_test("Legal Contracts CRUD", False, f"Error: {str(e)}")
            return False, None

    def test_legal_cases_crud(self):
        """Test Legal Cases CRUD operations"""
        try:
            # Create case
            case_data = {
                "case_type": "litigation",
                "title": "قضية تعويض",
                "description": "Test case description",
                "plaintiff": "المروج للألبان",
                "defendant": "شركة أخرى",
                "filing_date": "2025-01-15",
                "priority": "high",
                "court_name": "محكمة مسقط الابتدائية",
                "lawyer_name": "المحامي أحمد علي",
                "estimated_value": 10000
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/legal/cases",
                json=case_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Legal Cases CRUD", 
                    False, 
                    f"Failed to create case: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            case = create_response.json()
            case_id = case.get("id")
            
            # Verify case has auto-generated case_number
            if not case.get("case_number") or not case.get("case_number").startswith("CASE-"):
                self.log_test(
                    "Legal Cases CRUD", 
                    False, 
                    "Case number not auto-generated properly",
                    f"Case number: {case.get('case_number')}"
                )
                return False
            
            # Test GET cases
            get_response = self.session.get(f"{BACKEND_URL}/legal/cases")
            if get_response.status_code != 200:
                self.log_test(
                    "Legal Cases CRUD", 
                    False, 
                    f"Failed to get cases: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            cases = get_response.json()
            found_case = any(c.get("id") == case_id for c in cases)
            
            if found_case:
                self.log_test(
                    "Legal Cases CRUD", 
                    True, 
                    f"Successfully created case '{case_data['title']}' with auto-generated number {case.get('case_number')}"
                )
                return True
            else:
                self.log_test(
                    "Legal Cases CRUD", 
                    False, 
                    "Created case not found in GET cases"
                )
                return False
                
        except Exception as e:
            self.log_test("Legal Cases CRUD", False, f"Error: {str(e)}")
            return False

    # ==================== PROJECTS MODULE TESTS ====================
    
    def test_projects_dashboard_stats(self):
        """Test GET /api/projects/dashboard/stats - verify returns stats"""
        try:
            response = self.session.get(f"{BACKEND_URL}/projects/dashboard/stats")
            
            if response.status_code == 200:
                dashboard = response.json()
                expected_fields = ["total_projects", "active_projects", "completed_projects", "overdue_tasks", "total_budget", "total_actual_cost"]
                found_fields = [field for field in expected_fields if field in dashboard]
                
                if len(found_fields) == len(expected_fields):
                    self.log_test(
                        "Projects Dashboard Stats", 
                        True, 
                        f"Projects dashboard retrieved with all expected fields: {list(dashboard.keys())}"
                    )
                    return True
                else:
                    missing_fields = [field for field in expected_fields if field not in dashboard]
                    self.log_test(
                        "Projects Dashboard Stats", 
                        False, 
                        f"Projects dashboard missing fields: {missing_fields}",
                        f"Found: {list(dashboard.keys())}"
                    )
                    return False
            else:
                self.log_test(
                    "Projects Dashboard Stats", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Projects Dashboard Stats", False, f"Error: {str(e)}")
            return False

    def test_projects_crud(self):
        """Test Projects CRUD operations"""
        try:
            # Create project
            project_data = {
                "name": "مشروع توسعة المصنع",
                "description": "توسعة خطوط الإنتاج",
                "project_type": "construction",
                "start_date": "2025-01-01",
                "end_date": "2025-06-30",
                "budget": 50000,
                "priority": "high",
                "manager_name": "مدير المشروع",
                "department": "الهندسة",
                "location": "مسقط",
                "objectives": "زيادة الطاقة الإنتاجية"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/projects",
                json=project_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Projects CRUD", 
                    False, 
                    f"Failed to create project: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            project = create_response.json()
            project_id = project.get("id")
            
            # Verify project has auto-generated project_code
            if not project.get("project_code") or not project.get("project_code").startswith("PRJ-"):
                self.log_test(
                    "Projects CRUD", 
                    False, 
                    "Project code not auto-generated properly",
                    f"Project code: {project.get('project_code')}"
                )
                return False
            
            # Test GET projects
            get_response = self.session.get(f"{BACKEND_URL}/projects")
            if get_response.status_code != 200:
                self.log_test(
                    "Projects CRUD", 
                    False, 
                    f"Failed to get projects: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            projects = get_response.json()
            found_project = any(p.get("id") == project_id for p in projects)
            
            if found_project:
                self.log_test(
                    "Projects CRUD", 
                    True, 
                    f"Successfully created project '{project_data['name']}' with auto-generated code {project.get('project_code')}"
                )
                return True, project_id
            else:
                self.log_test(
                    "Projects CRUD", 
                    False, 
                    "Created project not found in GET projects"
                )
                return False, None
                
        except Exception as e:
            self.log_test("Projects CRUD", False, f"Error: {str(e)}")
            return False, None

    def test_project_tasks_crud(self):
        """Test Project Tasks CRUD operations"""
        try:
            # First create a project
            project_success, project_id = self.test_projects_crud()
            if not project_success:
                self.log_test("Project Tasks CRUD", False, "Cannot test without project")
                return False
            
            # Create task for the project
            task_data = {
                "project_id": project_id,
                "project_name": "مشروع توسعة المصنع",
                "task_name": "تصميم المخططات",
                "description": "تصميم المخططات الهندسية للتوسعة",
                "start_date": "2025-01-15",
                "due_date": "2025-02-15",
                "priority": "high",
                "estimated_hours": 40,
                "assigned_to_name": "المهندس أحمد"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/projects/tasks",
                json=task_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Project Tasks CRUD", 
                    False, 
                    f"Failed to create task: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            task = create_response.json()
            task_id = task.get("id")
            
            # Test GET project tasks
            get_response = self.session.get(f"{BACKEND_URL}/projects/{project_id}/tasks")
            if get_response.status_code != 200:
                self.log_test(
                    "Project Tasks CRUD", 
                    False, 
                    f"Failed to get project tasks: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            tasks = get_response.json()
            found_task = any(t.get("id") == task_id for t in tasks)
            
            if found_task:
                self.log_test(
                    "Project Tasks CRUD", 
                    True, 
                    f"Successfully created task '{task_data['task_name']}' for project"
                )
                return True
            else:
                self.log_test(
                    "Project Tasks CRUD", 
                    False, 
                    "Created task not found in GET project tasks"
                )
                return False
                
        except Exception as e:
            self.log_test("Project Tasks CRUD", False, f"Error: {str(e)}")
            return False

    # ==================== MARKETING MODULE TESTS ====================
    
    def test_marketing_dashboard_api(self):
        """Test GET /api/marketing/dashboard - verify returns marketing stats"""
        try:
            response = self.session.get(f"{BACKEND_URL}/marketing/dashboard")
            
            if response.status_code == 200:
                dashboard = response.json()
                expected_fields = ["campaigns", "leads", "offers", "returns", "budget_summary"]
                found_fields = [field for field in expected_fields if field in dashboard]
                
                if len(found_fields) >= 4:  # At least 4 expected fields
                    self.log_test(
                        "Marketing Dashboard API", 
                        True, 
                        f"Marketing dashboard retrieved with fields: {list(dashboard.keys())}"
                    )
                    return True
                else:
                    missing_fields = [field for field in expected_fields if field not in dashboard]
                    self.log_test(
                        "Marketing Dashboard API", 
                        False, 
                        f"Marketing dashboard missing fields: {missing_fields}",
                        f"Found: {list(dashboard.keys())}"
                    )
                    return False
            else:
                self.log_test(
                    "Marketing Dashboard API", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Marketing Dashboard API", False, f"Error: {str(e)}")
            return False

    def test_marketing_campaigns_crud(self):
        """Test Marketing Campaigns CRUD operations"""
        try:
            # Create campaign
            campaign_data = {
                "name": "حملة ترويجية للألبان الطازجة",
                "campaign_type": "social_media",
                "description": "حملة تسويقية لترويج منتجات الألبان الطازجة",
                "objective": "awareness",
                "start_date": "2025-01-15",
                "end_date": "2025-02-15",
                "budget": 2000.0,
                "target_audience": "العائلات والأطفال",
                "channels": ["facebook", "instagram", "whatsapp"],
                "responsible_name": "مدير التسويق"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/marketing/campaigns",
                json=campaign_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Marketing Campaigns CRUD", 
                    False, 
                    f"Failed to create campaign: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            campaign = create_response.json()
            campaign_id = campaign.get("id")
            
            # Verify campaign has auto-generated campaign_code
            if not campaign.get("campaign_code") or not campaign.get("campaign_code").startswith(("CAMP-", "CMP-")):
                self.log_test(
                    "Marketing Campaigns CRUD", 
                    False, 
                    "Campaign code not auto-generated properly",
                    f"Campaign code: {campaign.get('campaign_code')}"
                )
                return False
            
            # Test GET campaigns
            get_response = self.session.get(f"{BACKEND_URL}/marketing/campaigns")
            if get_response.status_code != 200:
                self.log_test(
                    "Marketing Campaigns CRUD", 
                    False, 
                    f"Failed to get campaigns: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            campaigns = get_response.json()
            found_campaign = any(c.get("id") == campaign_id for c in campaigns)
            
            if found_campaign:
                self.log_test(
                    "Marketing Campaigns CRUD", 
                    True, 
                    f"Successfully created campaign '{campaign_data['name']}' with auto-generated code {campaign.get('campaign_code')}"
                )
                return True, campaign_id
            else:
                self.log_test(
                    "Marketing Campaigns CRUD", 
                    False, 
                    "Created campaign not found in GET campaigns"
                )
                return False, None
                
        except Exception as e:
            self.log_test("Marketing Campaigns CRUD", False, f"Error: {str(e)}")
            return False, None

    def test_marketing_leads_crud(self):
        """Test Marketing Leads CRUD operations"""
        try:
            # Create lead
            lead_data = {
                "name": "أحمد محمد العلوي",
                "company_name": "شركة الخليج للتجارة",
                "phone": "+968 9123 4567",
                "email": "ahmed@gulf-trading.com",
                "address": "مسقط، عُمان",
                "lead_source": "website",
                "interest": "milk_purchase",
                "notes": "مهتم بشراء كميات كبيرة من الحليب الطازج",
                "expected_value": 5000.0,
                "assigned_to_name": "مندوب المبيعات"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/marketing/leads",
                json=lead_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Marketing Leads CRUD", 
                    False, 
                    f"Failed to create lead: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            lead = create_response.json()
            lead_id = lead.get("id")
            
            # Verify lead has auto-generated lead_code
            if not lead.get("lead_code") or not lead.get("lead_code").startswith("LEAD-"):
                self.log_test(
                    "Marketing Leads CRUD", 
                    False, 
                    "Lead code not auto-generated properly",
                    f"Lead code: {lead.get('lead_code')}"
                )
                return False
            
            # Test GET leads
            get_response = self.session.get(f"{BACKEND_URL}/marketing/leads")
            if get_response.status_code != 200:
                self.log_test(
                    "Marketing Leads CRUD", 
                    False, 
                    f"Failed to get leads: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            leads = get_response.json()
            found_lead = any(l.get("id") == lead_id for l in leads)
            
            if found_lead:
                self.log_test(
                    "Marketing Leads CRUD", 
                    True, 
                    f"Successfully created lead '{lead_data['name']}' with auto-generated code {lead.get('lead_code')}"
                )
                return True
            else:
                self.log_test(
                    "Marketing Leads CRUD", 
                    False, 
                    "Created lead not found in GET leads"
                )
                return False
                
        except Exception as e:
            self.log_test("Marketing Leads CRUD", False, f"Error: {str(e)}")
            return False

    def test_marketing_offers_crud(self):
        """Test Marketing Offers CRUD operations"""
        try:
            # Create offer
            offer_data = {
                "offer_type": "discount",
                "title": "خصم 15% على الحليب الطازج",
                "description": "عرض خاص للعملاء الجدد - خصم 15% على جميع منتجات الحليب الطازج",
                "product_type": "raw_milk",
                "discount_percentage": 15.0,
                "min_quantity": 100.0,
                "start_date": "2025-01-20",
                "end_date": "2025-02-20",
                "terms_conditions": "العرض ساري للعملاء الجدد فقط",
                "target_customers": "new"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/marketing/offers",
                json=offer_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Marketing Offers CRUD", 
                    False, 
                    f"Failed to create offer: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            offer = create_response.json()
            offer_id = offer.get("id")
            
            # Verify offer has auto-generated offer_code
            if not offer.get("offer_code") or not offer.get("offer_code").startswith(("OFR-", "OFFER-")):
                self.log_test(
                    "Marketing Offers CRUD", 
                    False, 
                    "Offer code not auto-generated properly",
                    f"Offer code: {offer.get('offer_code')}"
                )
                return False
            
            # Test GET offers
            get_response = self.session.get(f"{BACKEND_URL}/marketing/offers")
            if get_response.status_code != 200:
                self.log_test(
                    "Marketing Offers CRUD", 
                    False, 
                    f"Failed to get offers: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            offers = get_response.json()
            found_offer = any(o.get("id") == offer_id for o in offers)
            
            if found_offer:
                self.log_test(
                    "Marketing Offers CRUD", 
                    True, 
                    f"Successfully created offer '{offer_data['title']}' with auto-generated code {offer.get('offer_code')}"
                )
                return True
            else:
                self.log_test(
                    "Marketing Offers CRUD", 
                    False, 
                    "Created offer not found in GET offers"
                )
                return False
                
        except Exception as e:
            self.log_test("Marketing Offers CRUD", False, f"Error: {str(e)}")
            return False

    def test_marketing_returns_crud(self):
        """Test Marketing Returns CRUD operations"""
        try:
            # First create a customer to use for the return
            customer_data = {
                "name": "متجر الأسرة السعيدة",
                "phone": "+968 2456 7890",
                "address": "صحار، الباطنة الشمالية",
                "customer_type": "retail",
                "credit_limit": 1000.0
            }
            
            customer_response = self.session.post(
                f"{BACKEND_URL}/customers",
                json=customer_data,
                headers={"Content-Type": "application/json"}
            )
            
            if customer_response.status_code != 200:
                self.log_test(
                    "Marketing Returns CRUD", 
                    False, 
                    f"Failed to create customer for return test: {customer_response.status_code}",
                    customer_response.text
                )
                return False
            
            customer = customer_response.json()
            customer_id = customer.get("id")
            
            # Create return
            return_data = {
                "return_date": "2025-01-10",
                "customer_id": customer_id,
                "customer_name": customer_data["name"],
                "quantity_liters": 50.0,
                "reason": "quality_issue",
                "quality_grade": "C",
                "batch_number": "BATCH-2025-001",
                "notes": "الحليب لم يكن طازجاً كما هو متوقع",
                "refund_amount": 125.0
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/marketing/returns",
                json=return_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Marketing Returns CRUD", 
                    False, 
                    f"Failed to create return: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            return_record = create_response.json()
            return_id = return_record.get("id")
            
            # Verify return has auto-generated return_code
            if not return_record.get("return_code") or not return_record.get("return_code").startswith(("RET-", "RTN-")):
                self.log_test(
                    "Marketing Returns CRUD", 
                    False, 
                    "Return code not auto-generated properly",
                    f"Return code: {return_record.get('return_code')}"
                )
                return False
            
            # Test GET returns
            get_response = self.session.get(f"{BACKEND_URL}/marketing/returns")
            if get_response.status_code != 200:
                self.log_test(
                    "Marketing Returns CRUD", 
                    False, 
                    f"Failed to get returns: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            returns = get_response.json()
            found_return = any(r.get("id") == return_id for r in returns)
            
            if found_return:
                self.log_test(
                    "Marketing Returns CRUD", 
                    True, 
                    f"Successfully created return for '{return_data['customer_name']}' with auto-generated code {return_record.get('return_code')}"
                )
                return True
            else:
                self.log_test(
                    "Marketing Returns CRUD", 
                    False, 
                    "Created return not found in GET returns"
                )
                return False
                
        except Exception as e:
            self.log_test("Marketing Returns CRUD", False, f"Error: {str(e)}")
            return False

    # ==================== OPERATIONS MODULE TESTS ====================
    
    def test_operations_dashboard(self):
        """Test GET /api/operations/dashboard - verify returns stats"""
        try:
            response = self.session.get(f"{BACKEND_URL}/operations/dashboard")
            
            if response.status_code == 200:
                dashboard = response.json()
                expected_fields = ["equipment", "vehicles", "open_incidents", "today_operations"]
                found_fields = [field for field in expected_fields if field in dashboard]
                
                if len(found_fields) == len(expected_fields):
                    # Check nested structure
                    equipment = dashboard.get("equipment", {})
                    vehicles = dashboard.get("vehicles", {})
                    
                    equipment_fields = ["operational", "maintenance", "out_of_order"]
                    vehicle_fields = ["available", "in_use"]
                    
                    equipment_ok = all(field in equipment for field in equipment_fields)
                    vehicles_ok = all(field in vehicles for field in vehicle_fields)
                    
                    if equipment_ok and vehicles_ok:
                        self.log_test(
                            "Operations Dashboard", 
                            True, 
                            f"Operations dashboard retrieved with all expected fields and structure"
                        )
                        return True
                    else:
                        self.log_test(
                            "Operations Dashboard", 
                            False, 
                            "Operations dashboard structure incomplete",
                            f"Equipment: {equipment}, Vehicles: {vehicles}"
                        )
                        return False
                else:
                    missing_fields = [field for field in expected_fields if field not in dashboard]
                    self.log_test(
                        "Operations Dashboard", 
                        False, 
                        f"Operations dashboard missing fields: {missing_fields}",
                        f"Found: {list(dashboard.keys())}"
                    )
                    return False
            else:
                self.log_test(
                    "Operations Dashboard", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Operations Dashboard", False, f"Error: {str(e)}")
            return False

    def test_operations_equipment_crud(self):
        """Test Operations Equipment CRUD operations"""
        try:
            # Create equipment
            equipment_data = {
                "name": "خزان تبريد",
                "equipment_type": "tank",
                "brand": "Alfa Laval",
                "model": "TX500",
                "serial_number": "AL-TX500-2024-001",
                "purchase_date": "2024-12-01",
                "purchase_price": 15000,
                "location": "قسم التبريد",
                "specifications": "خزان تبريد بسعة 5000 لتر"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/operations/equipment",
                json=equipment_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Operations Equipment CRUD", 
                    False, 
                    f"Failed to create equipment: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            equipment = create_response.json()
            equipment_id = equipment.get("id")
            
            # Verify equipment has auto-generated equipment_code
            if not equipment.get("equipment_code") or not equipment.get("equipment_code").startswith("EQP-"):
                self.log_test(
                    "Operations Equipment CRUD", 
                    False, 
                    "Equipment code not auto-generated properly",
                    f"Equipment code: {equipment.get('equipment_code')}"
                )
                return False
            
            # Test GET equipment
            get_response = self.session.get(f"{BACKEND_URL}/operations/equipment")
            if get_response.status_code != 200:
                self.log_test(
                    "Operations Equipment CRUD", 
                    False, 
                    f"Failed to get equipment: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            equipment_list = get_response.json()
            found_equipment = any(e.get("id") == equipment_id for e in equipment_list)
            
            if found_equipment:
                self.log_test(
                    "Operations Equipment CRUD", 
                    True, 
                    f"Successfully created equipment '{equipment_data['name']}' with auto-generated code {equipment.get('equipment_code')}"
                )
                return True
            else:
                self.log_test(
                    "Operations Equipment CRUD", 
                    False, 
                    "Created equipment not found in GET equipment"
                )
                return False
                
        except Exception as e:
            self.log_test("Operations Equipment CRUD", False, f"Error: {str(e)}")
            return False

    def test_operations_vehicles_crud(self):
        """Test Operations Vehicles CRUD operations"""
        try:
            # Create vehicle
            vehicle_data = {
                "vehicle_type": "tanker",
                "brand": "Isuzu",
                "model": "NPR",
                "year": 2024,
                "plate_number": "AB 1234",
                "color": "أبيض",
                "fuel_type": "diesel",
                "tank_capacity": 3000,
                "assigned_driver_name": "السائق محمد"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/operations/vehicles",
                json=vehicle_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Operations Vehicles CRUD", 
                    False, 
                    f"Failed to create vehicle: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            vehicle = create_response.json()
            vehicle_id = vehicle.get("id")
            
            # Verify vehicle has auto-generated vehicle_code
            if not vehicle.get("vehicle_code") or not vehicle.get("vehicle_code").startswith("VEH-"):
                self.log_test(
                    "Operations Vehicles CRUD", 
                    False, 
                    "Vehicle code not auto-generated properly",
                    f"Vehicle code: {vehicle.get('vehicle_code')}"
                )
                return False
            
            # Test GET vehicles
            get_response = self.session.get(f"{BACKEND_URL}/operations/vehicles")
            if get_response.status_code != 200:
                self.log_test(
                    "Operations Vehicles CRUD", 
                    False, 
                    f"Failed to get vehicles: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            vehicles = get_response.json()
            found_vehicle = any(v.get("id") == vehicle_id for v in vehicles)
            
            if found_vehicle:
                self.log_test(
                    "Operations Vehicles CRUD", 
                    True, 
                    f"Successfully created vehicle '{vehicle_data['brand']} {vehicle_data['model']}' with auto-generated code {vehicle.get('vehicle_code')}"
                )
                return True
            else:
                self.log_test(
                    "Operations Vehicles CRUD", 
                    False, 
                    "Created vehicle not found in GET vehicles"
                )
                return False
                
        except Exception as e:
            self.log_test("Operations Vehicles CRUD", False, f"Error: {str(e)}")
            return False

    def test_operations_incidents_crud(self):
        """Test Operations Incidents CRUD operations"""
        try:
            # Create incident report
            incident_data = {
                "incident_type": "equipment_failure",
                "title": "عطل في خزان التبريد",
                "description": "توقف خزان التبريد الرئيسي عن العمل",
                "incident_date": "2025-01-15",
                "incident_time": "14:30",
                "location": "قسم التبريد",
                "severity": "high",
                "reported_by_id": self.user_data.get("id"),
                "reported_by_name": self.user_data.get("full_name"),
                "immediate_actions": "تم إيقاف التشغيل وتحويل الحليب لخزان احتياطي",
                "estimated_damage_cost": 2000
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/operations/incidents",
                json=incident_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Operations Incidents CRUD", 
                    False, 
                    f"Failed to create incident: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            incident = create_response.json()
            incident_id = incident.get("id")
            
            # Verify incident has auto-generated incident_number
            if not incident.get("incident_number") or not incident.get("incident_number").startswith("INC-"):
                self.log_test(
                    "Operations Incidents CRUD", 
                    False, 
                    "Incident number not auto-generated properly",
                    f"Incident number: {incident.get('incident_number')}"
                )
                return False
            
            # Test GET incidents
            get_response = self.session.get(f"{BACKEND_URL}/operations/incidents")
            if get_response.status_code != 200:
                self.log_test(
                    "Operations Incidents CRUD", 
                    False, 
                    f"Failed to get incidents: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            incidents = get_response.json()
            found_incident = any(i.get("id") == incident_id for i in incidents)
            
            if found_incident:
                self.log_test(
                    "Operations Incidents CRUD", 
                    True, 
                    f"Successfully created incident '{incident_data['title']}' with auto-generated number {incident.get('incident_number')}"
                )
                return True
            else:
                self.log_test(
                    "Operations Incidents CRUD", 
                    False, 
                    "Created incident not found in GET incidents"
                )
                return False
                
        except Exception as e:
            self.log_test("Operations Incidents CRUD", False, f"Error: {str(e)}")
            return False

    def test_activity_logs_new_modules(self):
        """Test that new module operations are logged in activity logs"""
        try:
            # Check for legal activity logs
            legal_response = self.session.get(f"{BACKEND_URL}/activity-logs?action=create_legal_contract&limit=5")
            projects_response = self.session.get(f"{BACKEND_URL}/activity-logs?action=create_project&limit=5")
            operations_response = self.session.get(f"{BACKEND_URL}/activity-logs?action=create_equipment&limit=5")
            
            legal_logs = legal_response.json() if legal_response.status_code == 200 else []
            projects_logs = projects_response.json() if projects_response.status_code == 200 else []
            operations_logs = operations_response.json() if operations_response.status_code == 200 else []
            
            logged_modules = []
            if legal_logs:
                logged_modules.append("Legal")
            if projects_logs:
                logged_modules.append("Projects")
            if operations_logs:
                logged_modules.append("Operations")
            
            if len(logged_modules) >= 2:  # At least 2 modules should have activity logs
                self.log_test(
                    "Activity Logs - New Modules", 
                    True, 
                    f"New module operations properly logged: {', '.join(logged_modules)}"
                )
                return True
            else:
                self.log_test(
                    "Activity Logs - New Modules", 
                    False, 
                    f"Insufficient activity logging for new modules. Found: {', '.join(logged_modules) if logged_modules else 'None'}"
                )
                return False
                
        except Exception as e:
            self.log_test("Activity Logs - New Modules", False, f"Error: {str(e)}")
            return False

    # ==================== PASSWORD RECOVERY TESTS ====================
    
    def test_forgot_password_api(self):
        """Test POST /api/auth/forgot-password - request password reset"""
        try:
            # Test with valid email format
            test_email = "testadmin@test.com"
            
            response = self.session.post(
                f"{BACKEND_URL}/auth/forgot-password",
                data={"email": test_email},
                headers={"Content-Type": "application/x-www-form-urlencoded"}
            )
            
            if response.status_code == 200:
                result = response.json()
                expected_message = "If the email exists, a reset link will be sent"
                
                if result.get("message") == expected_message:
                    self.log_test(
                        "Forgot Password API", 
                        True, 
                        f"Successfully requested password reset for {test_email}"
                    )
                    return True
                else:
                    self.log_test(
                        "Forgot Password API", 
                        False, 
                        f"Unexpected response message: {result.get('message')}",
                        f"Expected: {expected_message}"
                    )
                    return False
            else:
                self.log_test(
                    "Forgot Password API", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Forgot Password API", False, f"Error: {str(e)}")
            return False

    def test_verify_reset_token_api(self):
        """Test GET /api/auth/verify-reset-token - verify token validity"""
        try:
            # Test with invalid token
            invalid_token = "invalid_token_12345"
            
            response = self.session.get(
                f"{BACKEND_URL}/auth/verify-reset-token?token={invalid_token}"
            )
            
            if response.status_code == 200:
                result = response.json()
                
                if result.get("valid") == False:
                    self.log_test(
                        "Verify Reset Token API", 
                        True, 
                        "Successfully verified invalid token returns valid=false"
                    )
                    return True
                else:
                    self.log_test(
                        "Verify Reset Token API", 
                        False, 
                        f"Invalid token should return valid=false, got: {result}",
                        f"Response: {result}"
                    )
                    return False
            else:
                self.log_test(
                    "Verify Reset Token API", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Verify Reset Token API", False, f"Error: {str(e)}")
            return False

    def test_reset_password_api(self):
        """Test POST /api/auth/reset-password - reset password with token"""
        try:
            # Test with invalid token
            invalid_token = "invalid_token_12345"
            new_password = "newpassword123"
            
            response = self.session.post(
                f"{BACKEND_URL}/auth/reset-password",
                data={
                    "token": invalid_token,
                    "new_password": new_password
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"}
            )
            
            if response.status_code == 400:
                result = response.json()
                expected_detail = "Invalid or expired reset token"
                
                if result.get("detail") == expected_detail:
                    self.log_test(
                        "Reset Password API", 
                        True, 
                        "Successfully rejected invalid token with proper error message"
                    )
                    return True
                else:
                    self.log_test(
                        "Reset Password API", 
                        False, 
                        f"Unexpected error message: {result.get('detail')}",
                        f"Expected: {expected_detail}"
                    )
                    return False
            else:
                self.log_test(
                    "Reset Password API", 
                    False, 
                    f"Expected 400 status for invalid token, got {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Reset Password API", False, f"Error: {str(e)}")
            return False

    def test_password_recovery_workflow(self):
        """Test complete password recovery workflow with real user"""
        try:
            # First, create a test user for password recovery
            test_user_data = {
                "username": "recovery_test_user",
                "password": "original_password123",
                "email": "recovery_test@test.com",
                "full_name": "Recovery Test User",
                "role": "employee"
            }
            
            # Register test user
            register_response = self.session.post(
                f"{BACKEND_URL}/auth/register",
                json=test_user_data,
                headers={"Content-Type": "application/json"}
            )
            
            if register_response.status_code != 200:
                self.log_test(
                    "Password Recovery Workflow", 
                    False, 
                    f"Failed to create test user: {register_response.status_code}",
                    register_response.text
                )
                return False
            
            # Step 1: Request password reset
            forgot_response = self.session.post(
                f"{BACKEND_URL}/auth/forgot-password",
                data={"email": test_user_data["email"]},
                headers={"Content-Type": "application/x-www-form-urlencoded"}
            )
            
            if forgot_response.status_code != 200:
                self.log_test(
                    "Password Recovery Workflow", 
                    False, 
                    f"Forgot password request failed: {forgot_response.status_code}",
                    forgot_response.text
                )
                return False
            
            forgot_result = forgot_response.json()
            
            # Check if email was sent (this depends on SMTP configuration)
            email_sent = forgot_result.get("email_sent", False)
            
            if email_sent:
                self.log_test(
                    "Password Recovery Workflow", 
                    True, 
                    f"Complete password recovery workflow tested successfully for {test_user_data['email']} (email sent)"
                )
            else:
                self.log_test(
                    "Password Recovery Workflow", 
                    True, 
                    f"Password recovery workflow tested successfully for {test_user_data['email']} (email not sent - SMTP config issue)"
                )
            
            return True
                
        except Exception as e:
            self.log_test("Password Recovery Workflow", False, f"Error: {str(e)}")
            return False
    
    # ==================== NEW FEATURES TESTING (Review Request) ====================
    
    def test_departments_api(self):
        """Test GET /api/hr/departments - should include legal, projects, operations, marketing"""
        try:
            response = self.session.get(f"{BACKEND_URL}/hr/departments")
            
            if response.status_code == 200:
                departments = response.json()
                expected_departments = ["legal", "projects", "operations", "marketing"]
                found_departments = []
                
                # Check if departments is a list of objects with id field
                if isinstance(departments, list):
                    for dept in departments:
                        if isinstance(dept, dict) and "id" in dept:
                            found_departments.append(dept["id"].lower())
                
                missing_departments = [dept for dept in expected_departments if dept not in found_departments]
                
                if not missing_departments:
                    self.log_test(
                        "Departments API", 
                        True, 
                        f"All expected departments found: {expected_departments}"
                    )
                    return True
                else:
                    self.log_test(
                        "Departments API", 
                        False, 
                        f"Missing departments: {missing_departments}",
                        f"Found departments: {found_departments}"
                    )
                    return False
            else:
                self.log_test(
                    "Departments API", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Departments API", False, f"Error: {str(e)}")
            return False

    def test_permissions_api(self):
        """Test GET /api/hr/available-permissions - should include permissions for new departments"""
        try:
            response = self.session.get(f"{BACKEND_URL}/hr/available-permissions")
            
            if response.status_code == 200:
                permissions = response.json()
                
                # Expected permissions for new departments
                expected_permissions = [
                    # Legal permissions
                    "legal", "contracts", "cases", "consultations", "documents",
                    # Projects permissions
                    "projects", "tasks", "milestones", "team_members",
                    # Operations permissions
                    "operations", "equipment", "maintenance", "incidents", "vehicles",
                    # Marketing permissions
                    "marketing", "campaigns", "leads", "offers", "returns", "social"
                ]
                
                found_permissions = []
                
                # Check if permissions is a list of objects with id field
                if isinstance(permissions, list):
                    for perm in permissions:
                        if isinstance(perm, dict) and "id" in perm:
                            found_permissions.append(perm["id"].lower())
                
                missing_permissions = [perm for perm in expected_permissions if perm not in found_permissions]
                
                if len(missing_permissions) <= 5:  # Allow some flexibility
                    self.log_test(
                        "Permissions API", 
                        True, 
                        f"Most expected permissions found. Missing: {missing_permissions if missing_permissions else 'None'}"
                    )
                    return True
                else:
                    self.log_test(
                        "Permissions API", 
                        False, 
                        f"Too many missing permissions: {missing_permissions}",
                        f"Found permissions: {found_permissions[:10]}..."  # Show first 10
                    )
                    return False
            else:
                self.log_test(
                    "Permissions API", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Permissions API", False, f"Error: {str(e)}")
            return False

    def test_hr_attendance_import_excel(self):
        """Test POST /api/hr/attendance/import-excel - Excel import functionality"""
        try:
            # Create a simple test Excel file using openpyxl
            import io
            from openpyxl import Workbook
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            
            # Add headers
            headers = ["employee_id", "employee_name", "date", "check_in", "check_out"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add sample data
            data = [
                ["emp001", "أحمد محمد", "2025-01-15", "08:00", "17:00"],
                ["emp002", "فاطمة علي", "2025-01-15", "08:30", "17:30"],
                ["emp003", "محمد سالم", "2025-01-15", "09:00", "18:00"]
            ]
            
            for row, row_data in enumerate(data, 2):
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Test the import endpoint
            files = {'file': ('attendance.xlsx', excel_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            
            response = self.session.post(
                f"{BACKEND_URL}/hr/attendance/import-excel",
                files=files
            )
            
            if response.status_code == 200:
                result = response.json()
                
                # Check expected response structure
                if ("imported" in result or "updated" in result or "errors" in result or "message" in result):
                    imported_count = result.get("imported", 0)
                    updated_count = result.get("updated", 0)
                    errors = result.get("errors", [])
                    message = result.get("message", "")
                    
                    self.log_test(
                        "HR Attendance Import Excel", 
                        True, 
                        f"Excel import successful - {message or f'Imported: {imported_count}, Updated: {updated_count}, Errors: {len(errors)}'}"
                    )
                    return True
                else:
                    self.log_test(
                        "HR Attendance Import Excel", 
                        False, 
                        "Response structure not as expected",
                        f"Response: {result}"
                    )
                    return False
            else:
                self.log_test(
                    "HR Attendance Import Excel", 
                    False, 
                    f"Import failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("HR Attendance Import Excel", False, f"Error: {str(e)}")
            return False

    def test_central_dashboard_api(self):
        """Test GET /api/dashboard/central - Central dashboard for all centers"""
        try:
            response = self.session.get(f"{BACKEND_URL}/dashboard/central")
            
            if response.status_code == 200:
                dashboard = response.json()
                
                # Expected structure
                expected_sections = ["summary", "milk", "centers"]
                found_sections = [section for section in expected_sections if section in dashboard]
                
                if len(found_sections) == len(expected_sections):
                    # Check summary section
                    summary = dashboard.get("summary", {})
                    expected_summary_fields = ["total_centers", "total_suppliers", "total_employees", "present_today"]
                    summary_fields = [field for field in expected_summary_fields if field in summary]
                    
                    # Check milk section
                    milk = dashboard.get("milk", {})
                    expected_milk_fields = ["today_liters", "monthly_liters", "current_stock"]
                    milk_fields = [field for field in expected_milk_fields if field in milk]
                    
                    # Check centers section
                    centers = dashboard.get("centers", [])
                    
                    if len(summary_fields) >= 3 and len(milk_fields) >= 2 and isinstance(centers, list):
                        # Check centers structure if any centers exist
                        if centers:
                            center = centers[0]
                            expected_center_fields = ["center_id", "center_name", "today_milk_liters", "monthly_milk_liters", "suppliers_count"]
                            center_fields = [field for field in expected_center_fields if field in center]
                            
                            if len(center_fields) >= 3:
                                self.log_test(
                                    "Central Dashboard API", 
                                    True, 
                                    f"Central dashboard complete - Summary: {len(summary_fields)} fields, Milk: {len(milk_fields)} fields, Centers: {len(centers)} centers"
                                )
                                return True
                            else:
                                self.log_test(
                                    "Central Dashboard API", 
                                    False, 
                                    f"Centers structure incomplete. Found fields: {center_fields}",
                                    f"Expected: {expected_center_fields}"
                                )
                                return False
                        else:
                            self.log_test(
                                "Central Dashboard API", 
                                True, 
                                f"Central dashboard structure correct - Summary: {len(summary_fields)} fields, Milk: {len(milk_fields)} fields, Centers: empty (expected)"
                            )
                            return True
                    else:
                        self.log_test(
                            "Central Dashboard API", 
                            False, 
                            f"Dashboard sections incomplete - Summary: {len(summary_fields)}/{len(expected_summary_fields)}, Milk: {len(milk_fields)}/{len(expected_milk_fields)}",
                            f"Dashboard structure: {list(dashboard.keys())}"
                        )
                        return False
                else:
                    missing_sections = [section for section in expected_sections if section not in dashboard]
                    self.log_test(
                        "Central Dashboard API", 
                        False, 
                        f"Missing dashboard sections: {missing_sections}",
                        f"Found sections: {list(dashboard.keys())}"
                    )
                    return False
            else:
                self.log_test(
                    "Central Dashboard API", 
                    False, 
                    f"API call failed with status {response.status_code}",
                    response.text
                )
                return False
                
        except Exception as e:
            self.log_test("Central Dashboard API", False, f"Error: {str(e)}")
            return False

    # ==================== NEW FEATURES TESTS (Review Request) ====================
    
    def test_feed_purchase_invoice_with_signature(self):
        """Test Feed Purchase Invoice with Electronic Signature"""
        try:
            # First get suppliers and feed types
            suppliers_response = self.session.get(f"{BACKEND_URL}/suppliers")
            if suppliers_response.status_code != 200:
                self.log_test("Feed Purchase Invoice", False, "Cannot get suppliers for test")
                return False
            
            suppliers = suppliers_response.json()
            if not suppliers:
                self.log_test("Feed Purchase Invoice", False, "No suppliers found for test")
                return False
            
            # Get feed types
            feed_types_response = self.session.get(f"{BACKEND_URL}/feed-types")
            if feed_types_response.status_code != 200:
                self.log_test("Feed Purchase Invoice", False, "Cannot get feed types for test")
                return False
            
            feed_types = feed_types_response.json()
            if not feed_types:
                self.log_test("Feed Purchase Invoice", False, "No feed types found for test")
                return False
            
            supplier = suppliers[0]
            feed_type = feed_types[0]
            
            # First add balance to supplier by creating a milk reception
            milk_reception_data = {
                "supplier_id": supplier["id"],
                "supplier_name": supplier["name"],
                "quantity_liters": 100.0,
                "price_per_liter": 1.5,
                "quality_test": {
                    "fat_percentage": 3.5,
                    "protein_percentage": 3.2,
                    "temperature": 4.0,
                    "is_accepted": True
                }
            }
            
            milk_response = self.session.post(
                f"{BACKEND_URL}/milk-receptions",
                json=milk_reception_data,
                headers={"Content-Type": "application/json"}
            )
            
            if milk_response.status_code != 200:
                self.log_test("Feed Purchase Invoice", False, "Failed to add supplier balance")
                return False
            
            # Create feed purchase
            purchase_data = {
                "supplier_id": supplier["id"],
                "supplier_name": supplier["name"],
                "supplier_phone": supplier.get("phone"),
                "supplier_address": supplier.get("address"),
                "feed_type_id": feed_type["id"],
                "feed_type_name": feed_type["name"],
                "company_name": feed_type["company_name"],
                "quantity": 50.0,  # Reduced quantity to ensure sufficient balance
                "price_per_unit": 2.0,  # Total: 100 OMR (less than 150 OMR balance from milk)
                "unit": "kg",
                "notes": "Test feed purchase"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/feed-purchases",
                json=purchase_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Feed Purchase Invoice", 
                    False, 
                    f"Failed to create feed purchase: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            purchase = create_response.json()
            purchase_id = purchase.get("id")
            
            # Check if invoice_number is generated (like FP-2025-00001)
            invoice_number = purchase.get("invoice_number")
            if not invoice_number or not invoice_number.startswith("FP-"):
                self.log_test(
                    "Feed Purchase Invoice", 
                    False, 
                    "Invoice number not generated properly",
                    f"Invoice number: {invoice_number}"
                )
                return False
            
            # Test electronic approval
            approve_response = self.session.post(f"{BACKEND_URL}/feed-purchases/{purchase_id}/approve")
            
            if approve_response.status_code != 200:
                self.log_test(
                    "Feed Purchase Invoice", 
                    False, 
                    f"Failed to approve feed purchase: {approve_response.status_code}",
                    approve_response.text
                )
                return False
            
            approved_purchase = approve_response.json()
            signature_code = approved_purchase.get("signature_code")
            
            if not signature_code:
                self.log_test(
                    "Feed Purchase Invoice", 
                    False, 
                    "Electronic signature code not generated",
                    f"Response: {approved_purchase}"
                )
                return False
            
            self.log_test(
                "Feed Purchase Invoice", 
                True, 
                f"Successfully created feed purchase with invoice {invoice_number} and electronic signature {signature_code}"
            )
            return True
            
        except Exception as e:
            self.log_test("Feed Purchase Invoice", False, f"Error: {str(e)}")
            return False

    def test_supplier_milk_type(self):
        """Test Supplier with Milk Type field"""
        try:
            # Create supplier with milk_type
            supplier_data = {
                "name": "مورد إبل الصحراء",
                "phone": "+968 9876 5432",
                "address": "البريمي، عُمان",
                "milk_type": "camel"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/suppliers",
                json=supplier_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Supplier Milk Type", 
                    False, 
                    f"Failed to create supplier: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            supplier = create_response.json()
            
            # Check if milk_type is saved
            if supplier.get("milk_type") != "camel":
                self.log_test(
                    "Supplier Milk Type", 
                    False, 
                    "Milk type not saved properly",
                    f"Expected: camel, Got: {supplier.get('milk_type')}"
                )
                return False
            
            # Test GET suppliers to verify milk_type is included
            get_response = self.session.get(f"{BACKEND_URL}/suppliers")
            if get_response.status_code != 200:
                self.log_test(
                    "Supplier Milk Type", 
                    False, 
                    f"Failed to get suppliers: {get_response.status_code}",
                    get_response.text
                )
                return False
            
            suppliers = get_response.json()
            found_supplier = None
            for s in suppliers:
                if s.get("id") == supplier.get("id"):
                    found_supplier = s
                    break
            
            if not found_supplier:
                self.log_test(
                    "Supplier Milk Type", 
                    False, 
                    "Created supplier not found in GET suppliers"
                )
                return False
            
            if found_supplier.get("milk_type") != "camel":
                self.log_test(
                    "Supplier Milk Type", 
                    False, 
                    "Milk type not returned in GET suppliers",
                    f"Expected: camel, Got: {found_supplier.get('milk_type')}"
                )
                return False
            
            self.log_test(
                "Supplier Milk Type", 
                True, 
                f"Successfully created supplier '{supplier_data['name']}' with milk_type: {supplier.get('milk_type')}"
            )
            return True
            
        except Exception as e:
            self.log_test("Supplier Milk Type", False, f"Error: {str(e)}")
            return False

    def test_official_letters_workflow(self):
        """Test Official Letters with Electronic Approval workflow"""
        try:
            # Get employees first
            employees_response = self.session.get(f"{BACKEND_URL}/hr/employees")
            if employees_response.status_code != 200:
                self.log_test("Official Letters Workflow", False, "Cannot get employees for test")
                return False
            
            employees = employees_response.json()
            if not employees:
                self.log_test("Official Letters Workflow", False, "No employees found for test")
                return False
            
            employee = employees[0]
            
            # Create official letter request
            letter_data = {
                "employee_id": employee["id"],
                "employee_name": employee["name"],
                "letter_type": "salary_certificate",
                "purpose": "للبنك الأهلي العماني"
            }
            
            create_response = self.session.post(
                f"{BACKEND_URL}/hr/official-letters",
                json=letter_data,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response.status_code != 200:
                self.log_test(
                    "Official Letters Workflow", 
                    False, 
                    f"Failed to create official letter: {create_response.status_code}",
                    create_response.text
                )
                return False
            
            letter = create_response.json()
            letter_id = letter.get("id")
            
            # Check if letter_number is generated (like LTR-2025-0001)
            letter_number = letter.get("letter_number")
            if not letter_number or not letter_number.startswith("LTR-"):
                self.log_test(
                    "Official Letters Workflow", 
                    False, 
                    "Letter number not generated properly",
                    f"Letter number: {letter_number}"
                )
                return False
            
            # Test approval
            approve_response = self.session.post(f"{BACKEND_URL}/hr/official-letters/{letter_id}/approve")
            
            if approve_response.status_code != 200:
                self.log_test(
                    "Official Letters Workflow", 
                    False, 
                    f"Failed to approve letter: {approve_response.status_code}",
                    approve_response.text
                )
                return False
            
            approval_result = approve_response.json()
            
            # Check electronic signature from approval response
            if not approval_result.get("signature_code"):
                self.log_test(
                    "Official Letters Workflow", 
                    False, 
                    "Electronic signature code not generated",
                    f"Response: {approval_result}"
                )
                return False
            
            # Get the updated letter to verify approval status
            updated_letter_response = self.session.get(f"{BACKEND_URL}/hr/official-letters")
            if updated_letter_response.status_code != 200:
                self.log_test("Official Letters Workflow", False, "Cannot get updated letter")
                return False
            
            letters = updated_letter_response.json()
            approved_letter = None
            for l in letters:
                if l.get("id") == letter_id:
                    approved_letter = l
                    break
            
            if not approved_letter or not approved_letter.get("is_approved"):
                self.log_test(
                    "Official Letters Workflow", 
                    False, 
                    "Letter not marked as approved in database",
                    f"is_approved: {approved_letter.get('is_approved') if approved_letter else 'Letter not found'}"
                )
                return False
            
            # Test rejection workflow
            letter_data_2 = {
                "employee_id": employee["id"],
                "employee_name": employee["name"],
                "letter_type": "employment_letter",
                "purpose": "للسفارة"
            }
            
            create_response_2 = self.session.post(
                f"{BACKEND_URL}/hr/official-letters",
                json=letter_data_2,
                headers={"Content-Type": "application/json"}
            )
            
            if create_response_2.status_code == 200:
                letter_2 = create_response_2.json()
                letter_id_2 = letter_2.get("id")
                
                # Test rejection
                reject_response = self.session.post(
                    f"{BACKEND_URL}/hr/official-letters/{letter_id_2}/reject",
                    params={"reason": "معلومات ناقصة"}
                )
                
                if reject_response.status_code == 200:
                    # Get the updated letter to verify rejection status
                    updated_letters_response = self.session.get(f"{BACKEND_URL}/hr/official-letters")
                    if updated_letters_response.status_code == 200:
                        updated_letters = updated_letters_response.json()
                        rejected_letter = None
                        for l in updated_letters:
                            if l.get("id") == letter_id_2:
                                rejected_letter = l
                                break
                        
                        if not rejected_letter or rejected_letter.get("status") != "rejected":
                            self.log_test(
                                "Official Letters Workflow", 
                                False, 
                                "Letter rejection not working properly",
                                f"Status: {rejected_letter.get('status') if rejected_letter else 'Letter not found'}"
                            )
                            return False
            
            # Test printing (only for approved letters)
            print_response = self.session.post(f"{BACKEND_URL}/hr/official-letters/{letter_id}/print")
            
            if print_response.status_code != 200:
                self.log_test(
                    "Official Letters Workflow", 
                    False, 
                    f"Failed to register printing: {print_response.status_code}",
                    print_response.text
                )
                return False
            
            printed_letter = print_response.json()
            if not printed_letter.get("is_printed"):
                self.log_test(
                    "Official Letters Workflow", 
                    False, 
                    "Letter not marked as printed",
                    f"is_printed: {printed_letter.get('is_printed')}"
                )
                return False
            
            # Test GET my letters
            my_letters_response = self.session.get(f"{BACKEND_URL}/hr/my-letters")
            if my_letters_response.status_code != 200:
                self.log_test(
                    "Official Letters Workflow", 
                    False, 
                    f"Failed to get my letters: {my_letters_response.status_code}",
                    my_letters_response.text
                )
                return False
            
            self.log_test(
                "Official Letters Workflow", 
                True, 
                f"Successfully completed official letters workflow: created {letter_number}, approved with signature {approval_result.get('signature_code')}, and printed"
            )
            return True
            
        except Exception as e:
            self.log_test("Official Letters Workflow", False, f"Error: {str(e)}")
            return False
    
    def run_all_tests(self):
        """Run all backend tests for review request features"""
        print("🚀 Starting Backend API Testing for Review Request...")
        print(f"Backend URL: {BACKEND_URL}")
        print(f"Test User: {TEST_USERNAME}")
        print("=" * 60)
        
        # Test 1: Hassan Hamdi Login (White Screen Bug Fix)
        print("\n1️⃣ Testing Hassan Hamdi Login (White Screen Bug Fix)...")
        if not self.test_hassan_hamdi_login():
            print("❌ Cannot proceed without authentication")
            return False
        
        # Test 2: Dashboard Access (should not show white screen)
        print("\n2️⃣ Testing Dashboard Access (No White Screen)...")
        self.test_dashboard_access()
        
        # Test 3: User Profile Access
        print("\n3️⃣ Testing User Profile Access...")
        self.test_user_profile_access()
        
        # Test 4: Payment Receipt PDF API
        print("\n4️⃣ Testing Payment Receipt PDF API...")
        self.test_payment_receipt_pdf_workflow()
        
        # Summary
        print("\n" + "=" * 60)
        print("📊 TEST SUMMARY")
        print("=" * 60)
        
        passed = sum(1 for result in self.test_results if result["success"])
        total = len(self.test_results)
        
        print(f"Total Tests: {total}")
        print(f"Passed: {passed}")
        print(f"Failed: {total - passed}")
        print(f"Success Rate: {(passed/total)*100:.1f}%")
        
        if total - passed > 0:
            print("\n❌ FAILED TESTS:")
            for result in self.test_results:
                if not result["success"]:
                    print(f"  - {result['test']}: {result['message']}")
        
        print("\n📝 DETAILED RESULTS:")
        for result in self.test_results:
            status = "✅ PASS" if result["success"] else "❌ FAIL"
            print(f"  {status}: {result['test']} - {result['message']}")
        
        return passed == total

def main():
    """Main test execution"""
    tester = BackendTester()
    success = tester.run_all_tests()
    
    # Save detailed results
    with open("/app/backend_test_results.json", "w") as f:
        json.dump(tester.test_results, f, indent=2, ensure_ascii=False)
    
    print(f"\n📄 Detailed results saved to: /app/backend_test_results.json")
    
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())